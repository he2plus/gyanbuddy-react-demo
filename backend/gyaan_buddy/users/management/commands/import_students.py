"""
Import students from an Excel file (URL or local path) into the database.

Each sheet in the Excel represents one class (e.g. 9A, 9B).
Supported columns: Sr.No., Reg No, Class Name, Student Name,
                   Father's Name, Mother Name, Father Mobile No., Mother Mobile No.

Usage:
    # Interactive — asks for school name and Excel URL/path at runtime:
    python manage.py import_students

    # Non-interactive:
    python manage.py import_students \\
        --school "Delhi Public School" \\
        --file "https://storage.googleapis.com/.../students.xlsx"

    # Dry-run (preview without saving):
    python manage.py import_students --dry-run

    # Skip existing students (by Reg No / admission number) instead of updating:
    python manage.py import_students --skip-existing
"""

import io
import re
import unicodedata

import openpyxl
import requests
from django.contrib.auth.hashers import make_password
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from gyaan_buddy.users.models import Account, Class, School, Student, StudentSubjectEnrollment, UserProfile


# Sheets whose names clearly are not student-data sheets
_NON_DATA_SHEETS = {"export summary", "summary", "index", "contents", "toc"}


def _normalise(text):
    """Return a clean string — strips newlines, extra spaces, normalises unicode."""
    if text is None:
        return ""
    text = str(text).strip()
    text = re.sub(r"[\r\n\t]+", " ", text)
    text = re.sub(r" {2,}", " ", text)
    return unicodedata.normalize("NFKC", text)


def _to_username(full_name, existing_usernames):
    """
    Convert 'AARUSH RAI' → 'aarush.rai'.
    If already taken, appends a number: 'aarush.rai2', 'aarush.rai3', …
    """
    base = re.sub(r"[^a-z0-9]+", ".", full_name.lower().strip()).strip(".")
    base = re.sub(r"\.{2,}", ".", base)
    candidate = base
    n = 2
    while candidate in existing_usernames:
        candidate = f"{base}{n}"
        n += 1
    existing_usernames.add(candidate)
    return candidate


def _is_data_sheet(sheet_name):
    return sheet_name.strip().lower() not in _NON_DATA_SHEETS


def _col_index(header_row, *candidates):
    """Return the 0-based index of the first matching header (case-insensitive)."""
    normalised = [_normalise(h).lower() for h in header_row]
    for candidate in candidates:
        c = candidate.lower()
        for i, h in enumerate(normalised):
            if c in h:
                return i
    return None


def _load_workbook_from_source(source):
    """Load openpyxl workbook from a URL or a local file path."""
    if source.startswith("http://") or source.startswith("https://"):
        response = requests.get(source, timeout=60)
        response.raise_for_status()
        return openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
    else:
        return openpyxl.load_workbook(source, data_only=True)


class Command(BaseCommand):
    help = "Import students from an Excel file (URL or local path) into the database."

    def add_arguments(self, parser):
        parser.add_argument(
            "--school", dest="school_name", default=None,
            help="School name (will be created if not found).",
        )
        parser.add_argument(
            "--file", dest="file_source", default=None,
            help="URL or local path to the Excel file.",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="Preview what would be imported without saving anything.",
        )
        parser.add_argument(
            "--skip-existing", action="store_true",
            help="Skip students whose Reg No already exists; default is to update.",
        )
        parser.add_argument(
            "--default-password", dest="default_password", default=None,
            help="Default password for new accounts. Defaults to 'Gyaan@<RegNo>'.",
        )

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _ask(self, prompt, default=None):
        suffix = f" [{default}]" if default else ""
        value = input(f"{prompt}{suffix}: ").strip()
        return value if value else default

    def _get_or_create_school(self, name, dry_run):
        school, created = School.objects.get_or_create(
            name__iexact=name,
            defaults={"name": name, "is_active": True},
        )
        if created:
            if dry_run:
                self.stdout.write(self.style.WARNING(f"  [dry-run] Would CREATE school: '{name}'"))
            else:
                self.stdout.write(self.style.SUCCESS(f"  Created school: '{name}'"))
        else:
            self.stdout.write(f"  Using existing school: '{school.name}'")
        return school

    def _get_or_create_class(self, class_name, school, dry_run):
        cls, created = Class.objects.get_or_create(
            name__iexact=class_name,
            school=school,
            defaults={"name": class_name, "school": school, "is_active": True},
        )
        if created:
            if dry_run:
                self.stdout.write(self.style.WARNING(f"    [dry-run] Would CREATE class: '{class_name}'"))
            else:
                self.stdout.write(self.style.SUCCESS(f"    Created class: '{class_name}'"))
        return cls

    # ------------------------------------------------------------------
    # Main handle
    # ------------------------------------------------------------------

    def handle(self, *args, **options):
        dry_run = options["dry_run"]
        skip_existing = options["skip_existing"]
        default_password = options["default_password"]

        if dry_run:
            self.stdout.write(self.style.WARNING("\n*** DRY-RUN MODE — nothing will be saved ***\n"))

        # --- School ---
        school_name = options["school_name"]
        if not school_name:
            school_name = self._ask("Enter school name")
        if not school_name:
            raise CommandError("School name is required.")

        # --- Excel source ---
        file_source = options["file_source"]
        if not file_source:
            file_source = self._ask(
                "Enter Excel URL or local file path",
                default="https://storage.googleapis.com/gyaanbuddy-media/Class_IX_Separate_Sheets%20(1)%202.xlsx",
            )
        if not file_source:
            raise CommandError("Excel file source is required.")

        # --- Load workbook ---
        self.stdout.write(f"\nLoading Excel from: {file_source}")
        try:
            wb = _load_workbook_from_source(file_source)
        except Exception as exc:
            raise CommandError(f"Failed to load Excel: {exc}") from exc

        data_sheets = [s for s in wb.sheetnames if _is_data_sheet(s)]
        self.stdout.write(f"Found {len(data_sheets)} data sheet(s): {', '.join(data_sheets)}\n")

        # --- School lookup/create ---
        with transaction.atomic():
            school = self._get_or_create_school(school_name, dry_run)

            # Collect existing usernames once to avoid per-row DB hits
            existing_usernames = set(Account.objects.values_list("username", flat=True))

            total_created = 0
            total_updated = 0
            total_skipped = 0
            total_errors = 0

            for sheet_name in data_sheets:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))

                # Find header row (first row with at least 3 non-null cells)
                header_row_idx = None
                for idx, row in enumerate(rows):
                    non_null = sum(1 for v in row if v is not None)
                    if non_null >= 3:
                        header_row_idx = idx
                        break

                if header_row_idx is None:
                    self.stdout.write(self.style.WARNING(f"  Sheet '{sheet_name}': no header found, skipping."))
                    continue

                header = [_normalise(h) for h in rows[header_row_idx]]
                data_rows = rows[header_row_idx + 1:]

                # Map columns
                col_sr      = _col_index(header, "sr.no", "sr no", "s.no", "sno")
                col_reg     = _col_index(header, "reg no", "reg. no", "regno", "registration")
                col_class   = _col_index(header, "class name", "class")
                col_name    = _col_index(header, "student name", "name")
                col_father  = _col_index(header, "father's name", "father name", "father")
                col_mother  = _col_index(header, "mother name", "mother")
                col_father_mob = _col_index(header, "father mobile", "father mob")
                col_mother_mob = _col_index(header, "mother mobile", "mother mob")

                if col_name is None:
                    self.stdout.write(self.style.WARNING(
                        f"  Sheet '{sheet_name}': 'Student Name' column not found, skipping."
                    ))
                    continue

                self.stdout.write(f"\n--- Sheet: {sheet_name} ({len(data_rows)} rows) ---")

                sheet_created = sheet_updated = sheet_skipped = sheet_errors = 0
                # Track reg numbers seen in this import run to catch in-sheet duplicates
                seen_reg_nos = set()

                for row in data_rows:
                    # Skip completely empty rows
                    if all(v is None for v in row):
                        continue

                    student_name = _normalise(row[col_name]) if col_name is not None else ""
                    if not student_name:
                        continue

                    reg_no = _normalise(row[col_reg]) if col_reg is not None else ""

                    # Skip duplicate rows within the same Excel (same reg no seen twice)
                    if reg_no and reg_no in seen_reg_nos:
                        self.stdout.write(self.style.WARNING(
                            f"    Duplicate reg no {reg_no!r} for {student_name!r} — skipping."
                        ))
                        sheet_skipped += 1
                        continue
                    if reg_no:
                        seen_reg_nos.add(reg_no)
                    class_name = _normalise(row[col_class]) if col_class is not None else sheet_name
                    father_name = _normalise(row[col_father]) if col_father is not None else ""
                    mother_name = _normalise(row[col_mother]) if col_mother is not None else ""
                    father_mob = _normalise(row[col_father_mob]) if col_father_mob is not None else ""
                    mother_mob = _normalise(row[col_mother_mob]) if col_mother_mob is not None else ""

                    # Clean up "---" placeholders
                    father_mob = "" if father_mob == "---" else father_mob
                    mother_mob = "" if mother_mob == "---" else mother_mob
                    phone = father_mob or mother_mob

                    parent_name = father_name or mother_name

                    # Resolve admission number (Reg No)
                    try:
                        admission_number = int(reg_no) if reg_no else None
                    except (ValueError, TypeError):
                        admission_number = None

                    # Determine password
                    if default_password:
                        password = default_password
                    elif reg_no:
                        password = f"Gyaan@{reg_no}"
                    else:
                        password = "Gyaan@1234"

                    # Build name parts
                    name_parts = student_name.split()
                    first_name = name_parts[0].title() if name_parts else student_name.title()
                    last_name = " ".join(name_parts[1:]).title() if len(name_parts) > 1 else ""

                    if dry_run:
                        username = _to_username(student_name, existing_usernames)
                        self.stdout.write(
                            f"    [dry-run] {student_name!r:30s} | class={class_name} | "
                            f"reg={reg_no} | user={username}"
                        )
                        sheet_created += 1
                        continue

                    try:
                        with transaction.atomic():
                            # --- Get or create Class ---
                            cls = self._get_or_create_class(class_name, school, dry_run)

                            # --- Check if student already exists by admission number ---
                            existing_student = None
                            if admission_number:
                                existing_student = Student.objects.filter(
                                    admission_number=admission_number,
                                    user_profile__school=school,
                                ).select_related("user_profile__account").first()

                            if existing_student and skip_existing:
                                sheet_skipped += 1
                                continue

                            if existing_student:
                                # Update existing
                                account = existing_student.user_profile.account
                                account.first_name = first_name
                                account.last_name = last_name
                                account.save(update_fields=["first_name", "last_name"])

                                profile = existing_student.user_profile
                                if phone:
                                    profile.phone_number = phone
                                    profile.save(update_fields=["phone_number"])

                                existing_student.class_instance = cls
                                existing_student.parent_name = parent_name
                                existing_student.save(update_fields=["class_instance", "parent_name"])
                                sheet_updated += 1
                            else:
                                # Create new account (get_or_create guards against
                                # any edge-case username collision)
                                username = _to_username(student_name, existing_usernames)
                                account, acct_created = Account.objects.get_or_create(
                                    username=username,
                                    defaults={
                                        "first_name": first_name,
                                        "last_name": last_name,
                                        "email": f"{username}@school.gyaanbuddy.com",
                                        "password": make_password(password),
                                        "is_active": True,
                                    },
                                )

                                # UserProfile is auto-created by signal; fetch and update it.
                                profile = account.profile
                                profile.school = school
                                profile.user_type = "student"
                                if phone:
                                    profile.phone_number = phone
                                profile.save(update_fields=["school", "user_type", "phone_number"])

                                # Student record — get_or_create on user_profile
                                student, _ = Student.objects.get_or_create(
                                    user_profile=profile,
                                    defaults={
                                        "admission_number": admission_number or 0,
                                        "class_instance": cls,
                                        "parent_name": parent_name,
                                        "total_exp": 0,
                                        "rewards": 0,
                                    },
                                )

                                # Enroll student in all subjects linked to the class (school-scoped)
                                for subject in cls.subjects.filter(is_active=True, school=school):
                                    StudentSubjectEnrollment.objects.get_or_create(
                                        student=student,
                                        subject=subject,
                                        defaults={"is_active": True},
                                    )

                                sheet_created += 1

                    except Exception as exc:
                        self.stdout.write(self.style.ERROR(
                            f"    ERROR for {student_name!r} (reg={reg_no}): {exc}"
                        ))
                        sheet_errors += 1

                self.stdout.write(
                    f"  Sheet '{sheet_name}': "
                    f"created={sheet_created}, updated={sheet_updated}, "
                    f"skipped={sheet_skipped}, errors={sheet_errors}"
                )
                total_created += sheet_created
                total_updated += sheet_updated
                total_skipped += sheet_skipped
                total_errors += sheet_errors

            if dry_run:
                self.stdout.write(self.style.WARNING("\n*** Dry-run complete — no data was saved. ***"))
            else:
                self.stdout.write(self.style.SUCCESS(
                    f"\nImport complete: "
                    f"created={total_created}, updated={total_updated}, "
                    f"skipped={total_skipped}, errors={total_errors}"
                ))

            if dry_run:
                # Roll back the transaction entirely in dry-run mode
                transaction.set_rollback(True)
