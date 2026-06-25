import logging
import re
import uuid
from collections import defaultdict
from rest_framework import viewsets, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django.contrib.auth import login, logout
from django.db.models import Q, Avg, Max, Count, F
from django.db import models, transaction
from rest_framework_simplejwt.tokens import RefreshToken
from .models import Account, UserProfile, Student, Class, School, Grade, Mission, Competition, UserMissionProgress, UserCompetitionProgress, Teacher, TeacherProfile, MissionQuestion, Test, TestQuestion, UserTestProgress
from gyaan_buddy.subjects.models import Subject, Module, Question, Option, ModuleChapter, ModuleContent, Answer
from gyaan_buddy.subjects.helpers import normalize_question_type
from .serializers import (
    UserSerializer, UserCreateSerializer, UserUpdateSerializer,
    UserPasswordChangeSerializer, UserLoginSerializer, UserExpSerializer,
    UserListSerializer, ClassSerializer, ClassListSerializer,
    SchoolSerializer, SchoolListSerializer, GradeSerializer, GradeListSerializer,
    MissionSerializer, MissionCreateSerializer,
    CompetitionSerializer, CompetitionCreateSerializer, CompetitionJoinSerializer,
    UserMissionProgressSerializer, UserCompetitionProgressSerializer,
    TestSerializer, TestCreateSerializer, TestUpdateSerializer, TestListSerializer, TestWithProgressSerializer,
)
from gyaan_buddy.utils.response_utils import (
    success, error, created, not_found, forbidden,
    validation_error, Messages, StatusCodes
)
from django.utils import timezone
from .models import UserModuleProgress, UserChapterProgress
from .serializers import UserModuleProgressSerializer, UserChapterProgressSerializer
from .helpers import complete_mission

logger = logging.getLogger('gyaan_buddy.users')
auth_logger = logging.getLogger('gyaan_buddy.auth')
api_logger = logging.getLogger('gyaan_buddy.api')


class TeacherAdminPermission(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            logger.warning(f"Permission denied: User not authenticated for {request.path}")
            return False
        user_type = request.user.profile.user_type if hasattr(request.user, 'profile') else None
        is_allowed = user_type in ['teacher', 'administrator']
        if not is_allowed:
            logger.warning(f"Permission denied: User {request.user.username} ({user_type}) not authorized for {request.path}")
        return is_allowed


class StudentViewSet(viewsets.ModelViewSet):
    """
    ViewSet dedicated to Student operations.
    Clean, focused queryset for student-specific operations.
    """
    serializer_class = UserListSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None  # Return all students; teachers only see their own class anyway
    
    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return UserUpdateSerializer
        elif self.action == 'retrieve':
            return UserSerializer
        return UserListSerializer
    
    def get_queryset(self):
        """Clean queryset - only students, with proper prefetching."""
        queryset = Account.objects.select_related(
            'profile__school',
            'profile__student__level',
            'profile__student__class_instance',
            'profile__student__class_instance__grade',
        ).prefetch_related(
            'module_progress',
            'profile__student__class_instance__subjects',
            'profile__student__subject_enrollments__subject',
        ).filter(
            profile__user_type='student',
            is_active=True
        )

        # Scope to requester's school
        requester_profile = getattr(self.request.user, 'profile', None)
        requester_type = getattr(requester_profile, 'user_type', None)
        requester_school = getattr(requester_profile, 'school', None)
        if requester_school:
            queryset = queryset.filter(profile__school=requester_school)
        else:
            return queryset.none()

        # Teachers are further scoped to their assigned classes
        if requester_type == 'teacher':
            teacher_profile = getattr(requester_profile, 'teacher_profile', None)
            if teacher_profile:
                teacher_classes = Teacher.objects.filter(
                    teacher=teacher_profile
                ).values_list('class_instance', flat=True).distinct()
                queryset = queryset.filter(profile__student__class_instance__id__in=teacher_classes)
            else:
                return queryset.none()

        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(username__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(profile__student__roll_number__icontains=search)
            )
        
        class_id = self.request.query_params.get('class_id', None)
        if class_id:
            queryset = queryset.filter(profile__student__class_instance__id=class_id)
        
        class_param = self.request.query_params.get('class', None)
        if class_param:
            try:
                uuid.UUID(class_param)
                queryset = queryset.filter(profile__student__class_instance__id=class_param)
            except (ValueError, TypeError):
                queryset = queryset.filter(profile__student__class_instance__name__icontains=class_param)
        
        grade = self.request.query_params.get('grade', None)
        if grade:
            queryset = queryset.filter(profile__student__class_instance__grade__name__icontains=grade)
        
        subject = self.request.query_params.get('subject', None)
        if subject:
            queryset = queryset.filter(
                Q(profile__student__subject_enrollments__subject__name__icontains=subject) |
                Q(profile__student__subject_enrollments__subject__id=subject) |
                Q(profile__student__class_instance__subjects__name__icontains=subject) |
                Q(profile__student__class_instance__subjects__id=subject)
            ).distinct()
        
        teacher_id = self.request.query_params.get('teacher', None)
        if teacher_id:
            try:
                teacher_uuid = uuid.UUID(teacher_id)
                user_profile = UserProfile.objects.filter(
                    account__id=teacher_uuid,
                    user_type='teacher'
                ).select_related('teacher_profile').first()
                
                if user_profile and hasattr(user_profile, 'teacher_profile'):
                    teacher_classes = Teacher.objects.filter(
                        teacher=user_profile.teacher_profile
                    ).values_list('class_instance', flat=True).distinct()
                    queryset = queryset.filter(profile__student__class_instance__id__in=teacher_classes)
                else:
                    queryset = queryset.none()
            except (ValueError, TypeError):
                queryset = queryset.none()
        
        return queryset
    
    def get_permissions(self):
        if self.action in ['create', 'destroy']:
            return [TeacherAdminPermission()]
        return [permissions.IsAuthenticated()]
    
    def list(self, request, *args, **kwargs):
        api_logger.info(f"Students list requested by {request.user.username}")
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return success(data=serializer.data, message="Students retrieved successfully")
        serializer = self.get_serializer(queryset, many=True)
        return success(data=serializer.data, message="Students retrieved successfully")

    @action(detail=False, methods=['get'], url_path='attempt-rates')
    def attempt_rates(self, request):
        """
        Return student-wise attempt rates for the current principal/teacher student scope.
        Formula matches the teacher dashboard pattern:
        due chapters attempted / total due chapters * 100.
        """
        api_logger.info(f"Student attempt rates requested by {request.user.username}")

        queryset = self.filter_queryset(self.get_queryset()).select_related(
            'profile__student__class_instance'
        ).prefetch_related(
            'profile__student__subject_enrollments__subject',
            'profile__student__class_instance__subjects',
        )

        student_rows = []
        class_ids = set()
        subject_ids_by_account = {}

        subject_param = request.query_params.get('subject', None)
        resolved_subject_id = None
        if subject_param:
            try:
                resolved_subject_id = str(Subject.objects.get(id=subject_param).id)
            except (Subject.DoesNotExist, ValueError, TypeError):
                resolved_subject = Subject.objects.filter(
                    name__iexact=str(subject_param).strip(),
                    is_active=True,
                ).values_list('id', flat=True).first()
                resolved_subject_id = str(resolved_subject) if resolved_subject else None

        for account in queryset:
            student = getattr(account.profile, 'student', None) if hasattr(account, 'profile') else None
            class_instance = getattr(student, 'class_instance', None)
            if not student or not class_instance:
                continue

            account_subject_ids = []
            if resolved_subject_id:
                account_subject_ids = [resolved_subject_id]
            else:
                enrollments = getattr(student, 'subject_enrollments', None)
                if enrollments is not None:
                    account_subject_ids = [
                        str(enrollment.subject_id)
                        for enrollment in enrollments.all()
                        if getattr(enrollment, 'is_active', True) and enrollment.subject_id
                    ]
                if not account_subject_ids:
                    account_subject_ids = [
                        str(subject.id)
                        for subject in class_instance.subjects.filter(is_active=True)
                    ]

            account_subject_ids = sorted(set(account_subject_ids))
            student_rows.append({
                'account_id': str(account.id),
                'class_id': str(class_instance.id),
            })
            class_ids.add(class_instance.id)
            subject_ids_by_account[str(account.id)] = account_subject_ids

        if not student_rows:
            return success(
                data={
                    'overall_attempt_rate': 0,
                    'students': {},
                },
                message="Student attempt rates retrieved successfully",
            )

        due_chapters_qs = ModuleChapter.objects.filter(
            module__is_active=True,
            is_deleted=False,
            due_date__isnull=False,
            module__class_instance_id__in=class_ids,
        )
        if resolved_subject_id:
            due_chapters_qs = due_chapters_qs.filter(module__subject_id=resolved_subject_id)

        due_chapter_rows = list(
            due_chapters_qs.values_list('id', 'module__class_instance_id', 'module__subject_id').distinct()
        )

        due_chapters_by_scope = defaultdict(set)
        relevant_chapter_ids = set()
        for chapter_id, class_id, subject_id in due_chapter_rows:
            due_chapters_by_scope[(str(class_id), str(subject_id))].add(chapter_id)
            relevant_chapter_ids.add(chapter_id)

        if not relevant_chapter_ids:
            zero_map = {row['account_id']: 0 for row in student_rows}
            return success(
                data={
                    'overall_attempt_rate': 0,
                    'students': zero_map,
                },
                message="Student attempt rates retrieved successfully",
            )

        account_ids = [row['account_id'] for row in student_rows]
        attempted_pairs = set(
            Answer.objects.filter(
                user_id__in=account_ids,
                chapter_id__in=relevant_chapter_ids,
            ).values_list('user_id', 'chapter_id').distinct()
        )

        attempt_rates = {}
        total_rate = 0
        for row in student_rows:
            account_id = row['account_id']
            class_id = row['class_id']
            scoped_chapter_ids = set()
            for subject_id in subject_ids_by_account.get(account_id, []):
                scoped_chapter_ids.update(due_chapters_by_scope.get((class_id, subject_id), set()))

            total_due_chapters = len(scoped_chapter_ids)
            if total_due_chapters == 0:
                attempt_rate = 0
            else:
                attempted_count = sum(
                    1
                    for chapter_id in scoped_chapter_ids
                    if (account_id, chapter_id) in attempted_pairs
                )
                attempt_rate = round((attempted_count / total_due_chapters) * 100)

            attempt_rates[account_id] = attempt_rate
            total_rate += attempt_rate

        overall_attempt_rate = round(total_rate / len(student_rows)) if student_rows else 0

        return success(
            data={
                'overall_attempt_rate': overall_attempt_rate,
                'students': attempt_rates,
            },
            message="Student attempt rates retrieved successfully",
        )
    
    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return success(data=serializer.data, message="Student retrieved successfully")
    
    def create(self, request, *args, **kwargs):
        api_logger.info(f"Student creation by {request.user.username}")
        data = request.data.copy()
        data['user_type'] = 'student'
        serializer = self.get_serializer(data=data, context={'request': request})
        if serializer.is_valid():
            user = serializer.save()
            return created(data=UserSerializer(user).data, message="Student created successfully")
        return validation_error(serializer.errors)
    
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            user = serializer.save()
            return success(data=UserSerializer(user).data, message="Student updated successfully")
        return validation_error(serializer.errors)
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get student statistics for the requesting teacher."""
        api_logger.info(f"Student stats requested by {request.user.username}")
        
        from django.db.models import Avg
        
        if hasattr(request.user, 'profile') and hasattr(request.user.profile, 'teacher_profile'):
            class_ids = request.user.profile.teacher_profile.teacher_assignments.values_list(
                'class_instance_id', flat=True
            ).distinct()
            students = Account.objects.filter(
                profile__user_type='student',
                profile__student__class_instance_id__in=class_ids,
                is_active=True
            )
        else:
            students = Account.objects.none()
        
        total_students = students.count()
        active_students = students.filter(is_active=True).count()
        average_exp = students.aggregate(avg=Avg('profile__student__total_exp'))['avg'] or 0
        
        class_distribution = {}
        for student in students.select_related('profile__student__class_instance'):
            if hasattr(student.profile, 'student') and student.profile.student.class_instance:
                class_name = student.profile.student.class_instance.name
                class_distribution[class_name] = class_distribution.get(class_name, 0) + 1
        
        thirty_days_ago = timezone.now() - timezone.timedelta(days=30)
        recent_logins = students.filter(last_login__gte=thirty_days_ago).count()
        attendance_rate = (recent_logins / max(total_students, 1)) * 100
        
        stats = {
            'totalStudents': total_students,
            'activeStudents': active_students,
            'averageScore': round(average_exp, 2),
            'averageAttendance': round(attendance_rate, 2),
            'classDistribution': class_distribution,
        }
        
        return success(data=stats, message="Student statistics retrieved successfully")
    
    @action(detail=False, methods=['post'], url_path='bulk-import', permission_classes=[TeacherAdminPermission])
    def bulk_import(self, request):
        """
        Bulk-import students from an Excel file uploaded by a principal/admin.
        Each sheet represents one class. School is taken from the requesting user's profile.
        Replicates the logic of the import_students management command.
        """
        import io
        import re
        import unicodedata
        import openpyxl
        from django.contrib.auth.hashers import make_password
        from gyaan_buddy.users.models import StudentSubjectEnrollment

        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return validation_error({'file': 'No file provided.'})

        profile = getattr(request.user, 'profile', None)
        school = getattr(profile, 'school', None) if profile else None
        if not school:
            return forbidden('Your account is not linked to a school.')

        # ── helpers (mirrors management command) ──────────────────────────
        _NON_DATA_SHEETS = {"export summary", "summary", "index", "contents", "toc"}

        def _normalise(text):
            if text is None:
                return ""
            text = str(text).strip()
            text = re.sub(r"[\r\n\t]+", " ", text)
            text = re.sub(r" {2,}", " ", text)
            return unicodedata.normalize("NFKC", text)

        school_initials = school.password_prefix or "".join(
            word[0].lower()
            for word in re.split(r"\s+", school.name.strip())
            if word
        )

        def _to_username(reg_no, full_name, existing_usernames):
            if reg_no:
                base = f"{reg_no}@{school_initials}"
            else:
                base = re.sub(r"[^a-z0-9]+", ".", full_name.lower().strip()).strip(".")
                base = re.sub(r"\.{2,}", ".", base)
            candidate = base
            n = 2
            while candidate in existing_usernames:
                candidate = f"{base}{n}"
                n += 1
            existing_usernames.add(candidate)
            return candidate

        def _col_index(header_row, *candidates):
            normalised = [_normalise(h).lower() for h in header_row]
            for candidate in candidates:
                c = candidate.lower()
                for i, h in enumerate(normalised):
                    if c in h:
                        return i
            return None

        # ── load workbook ─────────────────────────────────────────────────
        try:
            wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()), data_only=True)
        except Exception as exc:
            return validation_error({'file': f'Could not read Excel file: {exc}'})

        data_sheets = [s for s in wb.sheetnames if s.strip().lower() not in _NON_DATA_SHEETS]
        if not data_sheets:
            return validation_error({'file': 'No valid data sheets found in the Excel file.'})

        existing_usernames = set(Account.objects.values_list("username", flat=True))
        total_created = total_updated = total_skipped = total_errors = 0
        sheet_summaries = []
        row_errors = []
        dry_run = str(request.data.get('dry_run', '')).lower() in ('true', '1', 'yes')
        preview_rows = []

        from django.db import transaction as db_transaction
        from contextlib import nullcontext
        ctx = nullcontext() if dry_run else db_transaction.atomic()
        with ctx:
            for sheet_name in data_sheets:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))

                # Find header row (first row with ≥3 non-null cells)
                header_row_idx = None
                for idx, row in enumerate(rows):
                    if sum(1 for v in row if v is not None) >= 3:
                        header_row_idx = idx
                        break

                if header_row_idx is None:
                    sheet_summaries.append({'sheet': sheet_name, 'created': 0, 'updated': 0, 'skipped': 0, 'errors': 0, 'note': 'No header found'})
                    continue

                header = [_normalise(h) for h in rows[header_row_idx]]
                data_rows = rows[header_row_idx + 1:]

                col_reg      = _col_index(header, "reg no", "reg. no", "regno", "registration")
                col_class    = _col_index(header, "class name", "class")
                col_name     = _col_index(header, "student name", "name")
                col_dob      = _col_index(header, "d.o.b", "dob", "date of birth", "birth date")
                col_father   = _col_index(header, "father's name", "father name", "father")
                col_mother   = _col_index(header, "mother name", "mother")
                col_father_mob = _col_index(header, "father mobile", "father mob")
                col_mother_mob = _col_index(header, "mother mobile", "mother mob")

                if col_name is None:
                    sheet_summaries.append({'sheet': sheet_name, 'created': 0, 'updated': 0, 'skipped': 0, 'errors': 0, 'note': "No 'Student Name' column"})
                    continue

                sheet_created = sheet_updated = sheet_skipped = sheet_errors = 0
                seen_reg_nos = set()

                for row in data_rows:
                    if all(v is None for v in row):
                        continue

                    student_name = _normalise(row[col_name]) if col_name is not None else ""
                    if not student_name:
                        continue

                    reg_no = _normalise(row[col_reg]) if col_reg is not None else ""
                    # Excel stores numbers as floats — strip trailing .0 (e.g. "9396.0" → "9396")
                    if reg_no:
                        try:
                            reg_no = str(int(float(reg_no)))
                        except (ValueError, TypeError):
                            pass
                    if reg_no and reg_no in seen_reg_nos:
                        sheet_skipped += 1
                        continue
                    if reg_no:
                        seen_reg_nos.add(reg_no)

                    class_name   = _normalise(row[col_class]) if col_class is not None else sheet_name
                    father_name  = _normalise(row[col_father]) if col_father is not None else ""
                    mother_name  = _normalise(row[col_mother]) if col_mother is not None else ""
                    father_mob   = _normalise(row[col_father_mob]) if col_father_mob is not None else ""
                    mother_mob   = _normalise(row[col_mother_mob]) if col_mother_mob is not None else ""
                    father_mob   = "" if father_mob == "---" else father_mob
                    mother_mob   = "" if mother_mob == "---" else mother_mob
                    phone        = father_mob or mother_mob
                    parent_name  = father_name or mother_name

                    try:
                        admission_number = int(reg_no) if reg_no else None
                    except (ValueError, TypeError):
                        admission_number = None

                    name_parts = student_name.split()
                    first_name = name_parts[0].title() if name_parts else student_name.title()

                    # Parse DOB — store as a date object and extract dd/mm for password
                    import datetime as _dt
                    dob_raw = row[col_dob] if col_dob is not None else None
                    dob_date = None
                    if dob_raw is not None:
                        try:
                            if isinstance(dob_raw, (_dt.date, _dt.datetime)):
                                dob_date = dob_raw if isinstance(dob_raw, _dt.date) else dob_raw.date()
                            else:
                                dob_str = _normalise(str(dob_raw)).strip("'\"")
                                parts = re.split(r"[/\-.]", dob_str)
                                parts = [p.strip("'\"") for p in parts]
                                if len(parts) >= 3:
                                    dd, mm, yyyy = int(parts[0]), int(parts[1]), int(parts[2])
                                    dob_date = _dt.date(yyyy, mm, dd)
                                elif len(parts) == 2:
                                    dd, mm = int(parts[0]), int(parts[1])
                                    dob_date = _dt.date(2000, mm, dd)
                        except (ValueError, TypeError):
                            dob_date = None

                    # Build password: reg_no@first_name if reg_no available, else Gyan@first_name
                    if reg_no:
                        password = f"{reg_no}@{first_name.lower()}"
                    else:
                        password = f"Gyan@{first_name.lower()}"
                    last_name  = " ".join(name_parts[1:]).title() if len(name_parts) > 1 else ""

                    if dry_run:
                        existing = Student.objects.filter(
                            admission_number=admission_number,
                            user_profile__school=school,
                        ).first() if admission_number else None
                        if existing:
                            sheet_updated += 1
                            row_action = 'update'
                        else:
                            sheet_created += 1
                            row_action = 'create'
                        preview_rows.append({
                            'sheet': sheet_name,
                            'student_name': student_name,
                            'class_name': class_name,
                            'reg_no': reg_no,
                            'action': row_action,
                        })
                        continue

                    try:
                        with db_transaction.atomic():
                            cls, _ = Class.objects.get_or_create(
                                name__iexact=class_name,
                                school=school,
                                defaults={"name": class_name, "school": school, "is_active": True},
                            )

                            existing_student = None
                            if admission_number:
                                existing_student = Student.objects.filter(
                                    admission_number=admission_number,
                                    user_profile__school=school,
                                ).select_related("user_profile__account").first()

                            if existing_student:
                                acct = existing_student.user_profile.account
                                acct.first_name = first_name
                                acct.last_name  = last_name
                                acct.password   = make_password(password)
                                acct.save(update_fields=["first_name", "last_name", "password"])
                                prof = existing_student.user_profile
                                prof_fields = []
                                if phone:
                                    prof.phone_number = phone
                                    prof_fields.append("phone_number")
                                if dob_date:
                                    prof.date_of_birth = dob_date
                                    prof_fields.append("date_of_birth")
                                if prof_fields:
                                    prof.save(update_fields=prof_fields)
                                existing_student.class_instance = cls
                                existing_student.parent_name    = parent_name
                                existing_student.save(update_fields=["class_instance", "parent_name"])
                                sheet_updated += 1
                            else:
                                username = _to_username(reg_no, student_name, existing_usernames)
                                safe_username = re.sub(r"[^a-z0-9.]", ".", username)
                                email_addr    = f"{safe_username}@school.gyaanbuddy.com"
                                acct, _ = Account.objects.get_or_create(
                                    username=username,
                                    defaults={
                                        "first_name": first_name,
                                        "last_name":  last_name,
                                        "email":      email_addr,
                                        "password":   make_password(password),
                                        "is_active":  True,
                                    },
                                )
                                prof = acct.profile
                                prof.school     = school
                                prof.user_type  = "student"
                                prof_fields     = ["school", "user_type"]
                                if phone:
                                    prof.phone_number = phone
                                    prof_fields.append("phone_number")
                                if dob_date:
                                    prof.date_of_birth = dob_date
                                    prof_fields.append("date_of_birth")
                                prof.save(update_fields=prof_fields)

                                student, _ = Student.objects.get_or_create(
                                    user_profile=prof,
                                    defaults={
                                        "admission_number": admission_number or 0,
                                        "class_instance":   cls,
                                        "parent_name":      parent_name,
                                        "total_exp":        0,
                                        "rewards":          0,
                                    },
                                )
                                for subject in cls.subjects.filter(is_active=True, school=school):
                                    StudentSubjectEnrollment.objects.get_or_create(
                                        student=student,
                                        subject=subject,
                                        defaults={"is_active": True},
                                    )
                                sheet_created += 1

                    except Exception as exc:
                        row_errors.append({'sheet': sheet_name, 'student': student_name, 'error': str(exc)})
                        sheet_errors += 1

                sheet_summaries.append({
                    'sheet': sheet_name,
                    'created': sheet_created,
                    'updated': sheet_updated,
                    'skipped': sheet_skipped,
                    'errors':  sheet_errors,
                })
                total_created += sheet_created
                total_updated += sheet_updated
                total_skipped += sheet_skipped
                total_errors  += sheet_errors

        if dry_run:
            return success(data={
                'dry_run': True,
                'would_create': total_created,
                'would_update': total_updated,
                'would_skip':   total_skipped,
                'total_rows':   total_created + total_updated + total_skipped + total_errors,
                'sheets': [
                    {**s, 'created': s['created'], 'updated': s['updated'], 'skipped': s['skipped']}
                    for s in sheet_summaries
                ],
                'preview': preview_rows,
            }, message=f"Dry run: {total_created} would be created, {total_updated} updated, {total_skipped} skipped.")

        api_logger.info(
            f"Bulk import by {request.user.username}: "
            f"created={total_created}, updated={total_updated}, "
            f"skipped={total_skipped}, errors={total_errors}"
        )
        return success(data={
            'created': total_created,
            'updated': total_updated,
            'skipped': total_skipped,
            'errors':  total_errors,
            'sheets':  sheet_summaries,
            'rowErrors': row_errors,
        }, message=f"Import complete: {total_created} created, {total_updated} updated, {total_skipped} skipped, {total_errors} errors.")

    @action(detail=True, methods=['get'], url_path='recent-tests')
    def recent_tests(self, request, pk=None):
        """
        Get recent chapter activity for a specific student.
        Returns last 10 chapters accessed/completed, with score and status.
        """
        try:
            student = self.get_object()
        except Account.DoesNotExist:
            return not_found("Student not found")

        from gyaan_buddy.users.models import UserChapterProgress
        recent = (
            UserChapterProgress.objects
            .filter(account=student)
            .select_related('chapter__module__subject')
            .order_by('-last_accessed')[:10]
        )
        results = []
        for cp in recent:
            chapter = cp.chapter
            module = chapter.module if chapter else None
            subject = module.subject if module else None
            results.append({
                'chapter_id': str(chapter.id) if chapter else None,
                'chapter_title': chapter.title if chapter else '',
                'module_name': module.name if module else '',
                'subject_name': subject.name if subject else '',
                'score': cp.percentage,
                'status': cp.status,
                'last_accessed': cp.last_accessed.isoformat() if cp.last_accessed else None,
                'completed_at': cp.completed_at.isoformat() if cp.completed_at else None,
            })
        return success(data=results, message="Recent tests retrieved successfully")

    @action(detail=True, methods=['get'], url_path='progress-trends')
    def progress_trends(self, request, pk=None):
        """
        Get progress trends for a specific student.
        Subject-wise: (correct answers / total questions of due module chapters for all due modules) * 100.
        """
        try:
            student = self.get_object()
        except Account.DoesNotExist:
            return not_found("Student not found")
        
        period_label = 'Due modules'
        trends_data = {}
        due_modules = Module.objects.filter(
            chapters__due_date__isnull=False,
            chapters__is_deleted=False,
            is_active=True,
        ).distinct().select_related('subject')
        if not due_modules.exists():
            return success(data={}, message="No progress data available")
        subjects_seen = set()
        for subject in Subject.objects.filter(is_active=True):
            subject_name = subject.name
            if subject_name in subjects_seen:
                continue
            subject_due_modules = due_modules.filter(subject=subject)
            if not subject_due_modules.exists():
                continue
            subjects_seen.add(subject_name)
            total_questions = ModuleContent.objects.filter(
                chapter__module__in=subject_due_modules,
                content_type='question',
                is_deleted=False
            ).values('question').distinct().count()
            if total_questions == 0:
                trends_data[subject_name.lower()] = {
                    'score': 0,
                    'change': 0,
                    'period': period_label
                }
                continue
            question_ids = ModuleContent.objects.filter(
                chapter__module__in=subject_due_modules,
                content_type='question',
                is_deleted=False
            ).values_list('question_id', flat=True).distinct()
            correct_count = Answer.objects.filter(
                user=student,
                is_correct=True,
                question_id__in=question_ids
            ).values('question').distinct().count()
            score_pct = round((correct_count / total_questions) * 100)
            trends_data[subject_name.lower()] = {
                'score': score_pct,
                'change': 0,
                'period': period_label
            }
        if not trends_data:
            return success(data={}, message="No progress data available")
        return success(data=trends_data, message="Progress trends retrieved successfully")


class TeacherViewSet(viewsets.ModelViewSet):
    """
    ViewSet dedicated to Teacher operations.
    Clean, focused queryset for teacher-specific operations.
    """
    serializer_class = UserListSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None
    
    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return UserUpdateSerializer
        elif self.action == 'retrieve':
            return UserSerializer
        return UserListSerializer
    
    def get_queryset(self):
        """Clean queryset - only teachers, with proper prefetching."""
        queryset = Account.objects.select_related(
            'profile__school',
            'profile__teacher_profile',
        ).prefetch_related(
            'profile__teacher_profile__teacher_assignments',
            'profile__teacher_profile__teacher_assignments__class_instance',
            'profile__teacher_profile__teacher_assignments__subject',
            'created_modules__subject',
            'created_subjects',
        ).filter(
            profile__user_type='teacher',
            is_active=True
        )

        # Scope to requester's school (teacher, admin/principal all see their school only)
        requester_profile = getattr(self.request.user, 'profile', None)
        requester_school = getattr(requester_profile, 'school', None)
        if requester_school:
            queryset = queryset.filter(profile__school=requester_school)
        else:
            return queryset.none()

        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(username__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(profile__teacher_profile__employee_id__icontains=search)
            )

        subject = self.request.query_params.get('subject', None)
        if subject:
            try:
                uuid.UUID(str(subject))
                queryset = queryset.filter(
                    profile__teacher_profile__teacher_assignments__subject__id=subject
                ).distinct()
            except (ValueError, TypeError):
                queryset = queryset.filter(
                    profile__teacher_profile__teacher_assignments__subject__name__icontains=subject
                ).distinct()

        class_id = self.request.query_params.get('class_id', None)
        if class_id:
            queryset = queryset.filter(
                profile__teacher_profile__teacher_assignments__class_instance__id=class_id
            ).distinct()

        return queryset
    
    def get_permissions(self):
        if self.action in ['create', 'destroy']:
            return [TeacherAdminPermission()]
        return [permissions.IsAuthenticated()]

    def _calculate_teacher_attempt_rate(self, teacher, subject_filter=None):
        """Mirror teacher-dashboard attempt-rate logic for one teacher row."""
        profile = getattr(teacher, 'profile', None)
        teacher_profile = getattr(profile, 'teacher_profile', None) if profile else None
        if not teacher_profile:
            return 0

        assignments_qs = teacher_profile.teacher_assignments.filter(
            is_deleted=False
        ).select_related('class_instance', 'subject')

        if subject_filter:
            try:
                uuid.UUID(str(subject_filter))
                assignments_qs = assignments_qs.filter(subject_id=subject_filter)
            except (ValueError, TypeError):
                assignments_qs = assignments_qs.filter(subject__name__icontains=subject_filter)

        assignments = list(assignments_qs)
        if not assignments:
            return 0

        assignment_subjects_by_class = defaultdict(set)
        class_ids = set()
        subject_ids = set()
        for assignment in assignments:
            class_id = str(assignment.class_instance_id)
            subject_id = str(assignment.subject_id)
            assignment_subjects_by_class[class_id].add(subject_id)
            class_ids.add(assignment.class_instance_id)
            subject_ids.add(assignment.subject_id)

        students_qs = Account.objects.filter(
            is_active=True,
            is_deleted=False,
            profile__is_deleted=False,
            profile__user_type='student',
            profile__student__is_deleted=False,
            profile__student__class_instance_id__in=class_ids,
            profile__student__subject_enrollments__subject_id__in=subject_ids,
            profile__student__subject_enrollments__is_active=True,
        ).select_related(
            'profile__student__class_instance'
        ).prefetch_related(
            'profile__student__subject_enrollments'
        ).distinct()

        if not students_qs.exists():
            students_qs = Account.objects.filter(
                is_active=True,
                is_deleted=False,
                profile__is_deleted=False,
                profile__user_type='student',
                profile__student__is_deleted=False,
                profile__student__class_instance_id__in=class_ids,
            ).select_related(
                'profile__student__class_instance'
            ).prefetch_related(
                'profile__student__subject_enrollments'
            ).distinct()

        students = list(students_qs)
        if not students:
            return 0

        due_scope_rows = list(
            ModuleChapter.objects.filter(
                module__class_instance_id__in=class_ids,
                module__subject_id__in=subject_ids,
                module__is_active=True,
                is_deleted=False,
                due_date__isnull=False,
            ).values_list('id', 'module__class_instance_id', 'module__subject_id').distinct()
        )

        due_chapters_by_scope = defaultdict(set)
        chapter_ids = set()
        for chapter_id, class_id, subject_id in due_scope_rows:
            due_chapters_by_scope[(str(class_id), str(subject_id))].add(chapter_id)
            chapter_ids.add(chapter_id)

        if not chapter_ids:
            return 0

        attempted_pairs = {
            (str(user_id), chapter_id)
            for user_id, chapter_id in Answer.objects.filter(
                user_id__in=[student.id for student in students],
                chapter_id__in=chapter_ids,
            ).values_list('user_id', 'chapter_id').distinct()
        }

        total_rate = 0
        for student in students:
            student_profile = getattr(student, 'profile', None)
            student_obj = getattr(student_profile, 'student', None) if student_profile else None
            class_id = str(getattr(student_obj, 'class_instance_id', '') or '')
            if not class_id:
                continue

            enrolled_subject_ids = {
                str(enrollment.subject_id)
                for enrollment in student_obj.subject_enrollments.all()
                if getattr(enrollment, 'is_active', True) and enrollment.subject_id
            } if student_obj else set()

            assigned_subject_ids = assignment_subjects_by_class.get(class_id, set())
            subject_ids_for_student = enrolled_subject_ids.intersection(assigned_subject_ids) or assigned_subject_ids

            scoped_chapter_ids = set()
            for subject_id in subject_ids_for_student:
                scoped_chapter_ids.update(due_chapters_by_scope.get((class_id, subject_id), set()))

            total_due_chapters = len(scoped_chapter_ids)
            if total_due_chapters == 0:
                continue

            attempted_count = sum(
                1 for chapter_id in scoped_chapter_ids
                if (str(student.id), chapter_id) in attempted_pairs
            )
            total_rate += round((attempted_count / total_due_chapters) * 100)

        return round(total_rate / len(students)) if students else 0
    
    def list(self, request, *args, **kwargs):
        api_logger.info(f"Teachers list requested by {request.user.username}")
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            teacher_objects = list(page)
            teachers_data = list(serializer.data)
        else:
            teacher_objects = list(queryset)
            serializer = self.get_serializer(teacher_objects, many=True)
            teachers_data = list(serializer.data)

        subject_filter = request.query_params.get('subject', None)
        enriched_teachers = []
        for teacher_obj, teacher_data in zip(teacher_objects, teachers_data):
            teacher_profile = getattr(getattr(teacher_obj, 'profile', None), 'teacher_profile', None)
            assignments = list(
                teacher_profile.teacher_assignments.filter(is_deleted=False).select_related('class_instance', 'subject')
            ) if teacher_profile else []

            class_list = sorted({
                assignment.class_instance.name
                for assignment in assignments
                if assignment.class_instance and assignment.class_instance.name
            })
            subject_list = sorted({
                assignment.subject.name
                for assignment in assignments
                if assignment.subject and assignment.subject.name
            })

            enriched_teachers.append({
                **teacher_data,
                'class_list': class_list,
                'subject_list': subject_list,
                'attempt_rate': self._calculate_teacher_attempt_rate(teacher_obj, subject_filter),
            })

        return success(data=enriched_teachers, message="Teachers retrieved successfully")
    
    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return success(data=serializer.data, message="Teacher retrieved successfully")
    
    def create(self, request, *args, **kwargs):
        api_logger.info(f"Teacher creation by {request.user.username}")
        data = request.data.copy()
        data['user_type'] = 'teacher'
        serializer = self.get_serializer(data=data, context={'request': request})
        if serializer.is_valid():
            user = serializer.save()
            return created(data=UserSerializer(user).data, message="Teacher created successfully")
        return validation_error(serializer.errors)
    
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            user = serializer.save()
            return success(data=UserSerializer(user).data, message="Teacher updated successfully")
        return validation_error(serializer.errors)
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get teacher statistics."""
        api_logger.info(f"Teacher stats requested by {request.user.username}")
        
        from django.db.models import Avg, Count
        
        teachers = self.get_queryset()
        total_teachers = teachers.count()
        
        teachers_with_content = teachers.annotate(
            content_count=Count('created_modules', distinct=True) + 
                         Count('created_subjects', distinct=True) +
                         Count('created_questions', distinct=True)
        )
        avg_content = teachers_with_content.aggregate(avg=Avg('content_count'))['avg'] or 0
        
        subject_distribution = {}
        for teacher in teachers:
            if hasattr(teacher.profile, 'teacher_profile'):
                for assignment in teacher.profile.teacher_profile.teacher_assignments.all():
                    subject_name = assignment.subject.name
                    subject_distribution[subject_name] = subject_distribution.get(subject_name, 0) + 1
        
        thirty_days_ago = timezone.now() - timezone.timedelta(days=30)
        recent_logins = teachers.filter(last_login__gte=thirty_days_ago).count()
        usage_rate = (recent_logins / max(total_teachers, 1)) * 100
        
        stats = {
            'totalTeachers': total_teachers,
            'activeTeachers': recent_logins,
            'averageContentCreated': round(avg_content, 2),
            'averageUsage': round(usage_rate, 2),
            'subjectDistribution': subject_distribution,
        }
        
        return success(data=stats, message="Teacher statistics retrieved successfully")
    
    @action(detail=True, methods=['get'])
    def assignments(self, request, pk=None):
        """Get all class-subject assignments for a teacher."""
        teacher = self.get_object()

        if not hasattr(teacher.profile, 'teacher_profile'):
            return success(data=[], message="No assignments found")

        assignments = teacher.profile.teacher_profile.teacher_assignments.select_related(
            'class_instance', 'subject'
        ).all()

        data = [{
            'class': {
                'id': str(a.class_instance.id),
                'name': a.class_instance.name
            },
            'subject': {
                'id': str(a.subject.id),
                'name': a.subject.name
            },
            'is_class_teacher': a.class_instance.class_teacher == a.teacher if a.class_instance.class_teacher else False
        } for a in assignments]

        return success(data=data, message="Teacher assignments retrieved successfully")

    @action(detail=False, methods=['post'], permission_classes=[TeacherAdminPermission])
    def bulk_import(self, request):
        """
        Bulk-import teachers from an Excel file.

        Expected columns (order-insensitive, case-insensitive):
            first_name, last_name, school, class_name, subject, employee_id

        Optional query param:
            ?dry_run=true  — preview without saving
        """
        import io, re
        import openpyxl
        from django.contrib.auth.hashers import make_password
        from gyaan_buddy.subjects.models import Subject

        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return validation_error({'file': 'No file provided.'})

        dry_run = request.query_params.get('dry_run', '').lower() == 'true'

        # ── Load workbook ────────────────────────────────────────────────
        try:
            wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()), data_only=True)
        except Exception as exc:
            return validation_error({'file': f'Could not read Excel file: {exc}'})

        ws = wb.active

        # ── Parse header row ─────────────────────────────────────────────
        FIELD_ALIASES = {
            'first_name':  ['first_name', 'first name', 'firstname'],
            'last_name':   ['last_name',  'last name',  'lastname', 'surname'],
            'school':      ['school', 'school_name'],
            'class_name':  ['class_name', 'class', 'class name'],
            'subject':     ['subject', 'subject_name'],
            'employee_id': ['employee_id', 'employee id', 'emp_id', 'empid'],
            'mobile_no':   ['mobile_no', 'mobile no', 'mobile', 'phone', 'phone_no', 'phone no', 'contact'],
        }

        header_row = [str(c.value).strip().lower() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_index = {}
        for field, aliases in FIELD_ALIASES.items():
            for alias in aliases:
                if alias in header_row:
                    col_index[field] = header_row.index(alias)
                    break

        required_fields = ['first_name', 'school', 'class_name', 'subject']
        missing = [f for f in required_fields if f not in col_index]
        if missing:
            return validation_error({'file': f'Missing required columns: {", ".join(missing)}'})

        # ── Helper: username generator ───────────────────────────────────
        existing_usernames = set(Account.objects.values_list('username', flat=True))

        def _to_username(first, last):
            parts = []
            for p in [first, last]:
                slug = re.sub(r'[^a-z0-9]+', '.', p.lower().strip()).strip('.')
                if slug:
                    parts.append(slug)
            base = '.'.join(parts) or 'teacher'
            candidate = base
            n = 2
            while candidate in existing_usernames:
                candidate = f'{base}{n}'
                n += 1
            existing_usernames.add(candidate)
            return candidate

        # ── Process rows ─────────────────────────────────────────────────
        results = []
        errors = []

        def _cell(row, field):
            idx = col_index.get(field)
            if idx is None:
                return ''
            val = row[idx].value
            if val is None:
                return ''
            if isinstance(val, float) and val.is_integer():
                val = int(val)
            return str(val).strip()

        rows = list(ws.iter_rows(min_row=2))

        # ── Validate all rows first (no DB writes) ───────────────────────
        parsed_rows = []
        for row_num, row in enumerate(rows, start=2):
            first_name   = _cell(row, 'first_name')
            last_name    = _cell(row, 'last_name')
            school_name  = _cell(row, 'school')
            class_name   = _cell(row, 'class_name')
            subject_name = _cell(row, 'subject')
            employee_id  = _cell(row, 'employee_id')
            mobile_no    = _cell(row, 'mobile_no')

            if not any([first_name, school_name, class_name, subject_name]):
                continue  # skip blank rows

            row_errors = []
            if not first_name:
                row_errors.append('first_name is required')
            if not school_name:
                row_errors.append('school is required')
            if not class_name:
                row_errors.append('class_name is required')
            if not subject_name:
                row_errors.append('subject is required')

            if row_errors:
                errors.append({'row': row_num, 'errors': row_errors})
            else:
                parsed_rows.append({
                    'row': row_num, 'first_name': first_name, 'last_name': last_name,
                    'school_name': school_name, 'class_name': class_name,
                    'subject_name': subject_name, 'employee_id': employee_id,
                    'mobile_no': mobile_no,
                })

        if errors:
            return validation_error({'errors': errors, 'message': 'Fix the errors above before importing.'})

        if dry_run:
            return success(
                data={
                    'dry_run': True,
                    'total_rows': len(rows),
                    'valid_rows': len(parsed_rows),
                    'preview': parsed_rows,
                },
                message='Dry-run preview complete — nothing saved.',
            )

        # ── Import inside a single transaction ───────────────────────────
        with transaction.atomic():
            for item in parsed_rows:
                row_num      = item['row']
                first_name   = item['first_name']
                last_name    = item['last_name']
                school_name  = item['school_name']
                class_name   = item['class_name']
                subject_name = item['subject_name']
                employee_id  = item['employee_id']
                mobile_no    = item['mobile_no']

                # School
                school, _ = School.objects.get_or_create(
                    name__iexact=school_name,
                    defaults={'name': school_name, 'is_active': True},
                )

                # Class
                cls, _ = Class.objects.get_or_create(
                    name__iexact=class_name,
                    school=school,
                    defaults={'name': class_name, 'school': school, 'is_active': True},
                )

                # Subject — find or create within school
                subject = (
                    Subject.objects.filter(name__iexact=subject_name, school=school).first()
                    or Subject.objects.filter(code__iexact=subject_name, school=school).first()
                )
                if subject is None:
                    code = subject_name.strip()[:10].upper()
                    subject, _ = Subject.objects.get_or_create(
                        code=code,
                        school=school,
                        defaults={'name': subject_name.strip(), 'is_active': True},
                    )

                # Account — find by mobile/username, employee_id, or name+school
                account = None
                if mobile_no:
                    account = Account.objects.filter(username=mobile_no).first()

                if account is None and employee_id:
                    tp_match = TeacherProfile.objects.filter(
                        employee_id=employee_id
                    ).select_related('user_profile__account').first()
                    if tp_match:
                        account = tp_match.user_profile.account

                if account is None:
                    profile_match = UserProfile.objects.filter(
                        school=school,
                        user_type='teacher',
                        account__first_name__iexact=first_name.strip(),
                        account__last_name__iexact=last_name.strip(),
                    ).select_related('account').first()
                    if profile_match:
                        account = profile_match.account

                account_created = False
                password = None
                if account is None:
                    username = mobile_no if mobile_no else _to_username(first_name, last_name)
                    school_prefix = school.password_prefix or ''.join(w[0] for w in school_name.split() if w).lower()
                    password = f'{school_prefix}.{first_name.strip().lower()}'
                    account = Account.objects.create(
                        username=username,
                        first_name=first_name.title(),
                        last_name=last_name.title(),
                        email=f'{username}@school.gyaanbuddy.com',
                        password=make_password(password),
                        is_active=True,
                    )
                    account_created = True

                # UserProfile
                profile = account.profile
                profile.school = school
                profile.user_type = 'teacher'
                update_fields = ['school', 'user_type']
                if mobile_no:
                    profile.phone_number = mobile_no
                    update_fields.append('phone_number')
                profile.save(update_fields=update_fields)

                # TeacherProfile
                teacher_profile, _ = TeacherProfile.objects.get_or_create(
                    user_profile=profile,
                    defaults={'employee_id': employee_id or None},
                )
                if employee_id and not teacher_profile.employee_id:
                    teacher_profile.employee_id = employee_id
                    teacher_profile.save(update_fields=['employee_id'])

                # Link subject to class
                cls.subjects.add(subject)

                # Teacher assignment
                _, assign_created = Teacher.objects.get_or_create(
                    teacher=teacher_profile,
                    class_instance=cls,
                    subject=subject,
                )

                row_result = {
                    'row': row_num,
                    'status': 'created' if account_created else 'existing',
                    'username': account.username,
                    'school': school.name,
                    'class': cls.name,
                    'subject': subject.name,
                    'assignment': 'new' if assign_created else 'already_exists',
                }
                if password:
                    row_result['password'] = password
                results.append(row_result)

        return success(
            data={
                'dry_run': False,
                'total_rows': len(rows),
                'imported': len(results),
                'results': results,
            },
            message=f'{len(results)} teacher(s) imported successfully',
        )

    @action(detail=True, methods=['get'])
    def performance(self, request, pk=None):
        """Get average student performance across all classes assigned to this teacher."""
        teacher = self.get_object()

        if not hasattr(teacher.profile, 'teacher_profile'):
            return success(
                data={'average_score': 0, 'improvement': 0, 'class_breakdown': []},
                message="Teacher performance retrieved successfully"
            )

        assignments = teacher.profile.teacher_profile.teacher_assignments.select_related(
            'class_instance', 'subject'
        ).all()

        class_ids = list(assignments.values_list('class_instance_id', flat=True).distinct())

        students = Account.objects.filter(
            profile__user_type='student',
            profile__student__class_instance_id__in=class_ids,
            is_active=True,
        ).select_related('profile__student__class_instance')

        class_breakdown = []
        total_scores = []

        class_map = {}
        for s in students:
            cls = s.profile.student.class_instance
            if cls:
                class_map.setdefault(str(cls.id), {'name': cls.name, 'students': []})['students'].append(s)

        for class_id, info in class_map.items():
            class_scores = []
            for student in info['students']:
                total = Answer.objects.filter(user=student).count()
                if total > 0:
                    correct = Answer.objects.filter(user=student, is_correct=True).count()
                    class_scores.append(round((correct / total) * 100))
                elif hasattr(student.profile, 'student') and student.profile.student.total_exp:
                    class_scores.append(min(100, round(student.profile.student.total_exp / 10)))

            avg = round(sum(class_scores) / len(class_scores)) if class_scores else 0
            total_scores.extend(class_scores)
            class_breakdown.append({'class': info['name'], 'average_score': avg})

        overall_avg = round(sum(total_scores) / len(total_scores)) if total_scores else 0

        return success(
            data={
                'average_score': overall_avg,
                'improvement': 0,
                'class_breakdown': class_breakdown,
            },
            message="Teacher performance retrieved successfully"
        )


def _build_subject_progress(user):
    from django.db.models import Avg
    from gyaan_buddy.subjects.models import Subject, Answer
    from gyaan_buddy.users.models import StudentSubjectEnrollment

    subject_progress = []

    profile = getattr(user, 'profile', None)
    if profile and profile.user_type == 'student':
        student = getattr(profile, 'student', None)
        if student:
            enrolled_ids = StudentSubjectEnrollment.objects.filter(
                student=student,
                is_active=True,
            ).values_list('subject_id', flat=True)
            subjects = Subject.objects.filter(id__in=enrolled_ids, is_active=True)
        elif profile.school:
            subjects = Subject.objects.filter(school=profile.school, is_active=True)
        else:
            subjects = Subject.objects.none()
    elif profile and profile.school:
        subjects = Subject.objects.filter(school=profile.school, is_active=True)
    else:
        subjects = Subject.objects.filter(is_active=True)

    answers = Answer.objects.filter(user=user)

    for subject in subjects:
        subj_answers = answers.filter(
            question__module_contents__chapter__module__subject=subject
        ).distinct()
        s_total = subj_answers.count()
        s_correct = subj_answers.filter(is_correct=True).count()

        avg_tries = subj_answers.aggregate(avg=Avg('tries'))['avg'] or 0
        first_try_correct = subj_answers.filter(is_correct=True, tries=1).count()
        first_try_accuracy = round((first_try_correct / s_total) * 100, 2) if s_total > 0 else 0.0

        subj_module_progress = UserModuleProgress.objects.filter(
            account=user,
            module__subject=subject
        )
        avg_progress = subj_module_progress.aggregate(avg=Avg('percentage'))['avg'] or 0

        attempted_module_ids = subj_module_progress.values_list('module_id', flat=True)

        total_chapters_in_attempted_modules = ModuleChapter.objects.filter(
            module_id__in=attempted_module_ids,
            is_enabled=True,
            is_deleted=False
        ).count()

        chapters_completed = UserChapterProgress.objects.filter(
            account=user,
            chapter__module_id__in=attempted_module_ids,
            status='completed'
        ).count()

        chapter_completion_rate = round(
            (chapters_completed / total_chapters_in_attempted_modules) * 100, 2
        ) if total_chapters_in_attempted_modules > 0 else 0.0

        subject_progress.append({
            'subject_id': str(subject.id),
            'subject_name': subject.name,
            'subject_code': subject.code,
            'color': subject.color,
            'questions_attempted': s_total,
            'correct_answers': s_correct,
            'accuracy': round((s_correct / s_total) * 100, 2) if s_total > 0 else 0.0,
            'avg_tries': round(avg_tries, 2),
            'first_try_correct': first_try_correct,
            'first_try_accuracy': first_try_accuracy,
            'avg_completion': round(avg_progress, 2),
            'chapters_completed': chapters_completed,
            'total_chapters_in_attempted_modules': total_chapters_in_attempted_modules,
            'chapter_completion_rate': chapter_completion_rate,
        })

    return subject_progress


class AuthViewSet(viewsets.ViewSet):
    permission_classes = [permissions.AllowAny]
    
    @action(detail=False, methods=['get'])
    def test(self, request):
        auth_logger.info(f"Test endpoint accessed from {request.META.get('REMOTE_ADDR', 'unknown')}")
        return Response({
            "success": True,
            "message": "Auth endpoint is working",
            "data": {"status": "ok"}
        })
    
    @action(detail=False, methods=['post'])
    def login(self, request):
        auth_logger.info(f"Login attempt from {request.META.get('REMOTE_ADDR', 'unknown')} - Data: {request.data}")
        
        serializer = UserLoginSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.validated_data['user']
            login_type = serializer.validated_data.get('type', 'mobile')
            fcm_token = serializer.validated_data.get('fcm_token', '')
            
            if fcm_token and fcm_token.strip():
                user.fcm_token = fcm_token.strip()
                user.save(update_fields=['fcm_token'])
                auth_logger.info(f"FCM token updated for user {user.username}")
            
            refresh = RefreshToken.for_user(user)
            access_token = refresh.access_token
            login(request, user)
            
            is_first_login = not user.logged_in_once
            user_data = dict(UserSerializer(user).data)
            user_data['subject_progress'] = _build_subject_progress(user)
            if is_first_login:
                user.logged_in_once = True
                user.save(update_fields=['logged_in_once'])

            auth_logger.info(f"Successful login for user {user.username} (ID: {user.id}) from {request.META.get('REMOTE_ADDR', 'unknown')}")

            return Response({
                "success": True,
                "message": "Login successful",
                "data": {
                    "user": user_data,
                    "portal_type": login_type,
                    "tokens": {
                        "access": str(access_token),
                        "refresh": str(refresh),
                        "access_token_expires": access_token.current_time + access_token.lifetime,
                        "refresh_token_expires": refresh.current_time + refresh.lifetime
                    }
                }
            })
        else:
            auth_logger.warning(f"Login failed for data: {request.data} - Errors: {serializer.errors}")
            return Response({
                "success": False,
                "message": "Login failed",
                "errors": serializer.errors
            }, status=400)
    
    @action(detail=False, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def logout(self, request):
        user = request.user
        auth_logger.info(f"Logout for user {user.username} (ID: {user.id}) from {request.META.get('REMOTE_ADDR', 'unknown')}")
        user.fcm_token = None
        user.save(update_fields=['fcm_token'])
        auth_logger.info(f"FCM token cleared for user {user.username} on logout")
        logout(request)
        return Response({
            "success": True,
            "message": "Logout successful"
        })
    
    @action(detail=False, methods=['post'])
    def refresh(self, request):
        auth_logger.info(f"Token refresh attempt from {request.META.get('REMOTE_ADDR', 'unknown')}")
        
        try:
            refresh_token = request.data.get('refresh')
            if not refresh_token:
                auth_logger.warning("Token refresh failed: No refresh token provided")
                return Response({
                    "success": False,
                    "message": "Refresh token is required"
                }, status=400)
            
            refresh = RefreshToken(refresh_token)
            access_token = refresh.access_token
            
            auth_logger.info(f"Token refresh successful for user ID: {refresh.payload.get('user_id', 'unknown')}")
            
            return Response({
                "success": True,
                "message": "Token refreshed successfully",
                "data": {
                    "access": str(access_token),
                    "access_token_expires": access_token.current_time + access_token.lifetime
                }
            })
        except Exception as e:
            auth_logger.error(f"Token refresh failed: {str(e)}")
            return Response({
                "success": False,
                "message": "Invalid refresh token",
                "errors": {"refresh": "Token is invalid or expired"}
            }, status=400)
    
    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def me(self, request):
        auth_logger.info(f"Current user profile requested by {request.user.username} (ID: {request.user.id})")
        
        user = Account.objects.select_related(
            'profile__school',
            'profile__student__level',
            'profile__student__class_instance',
            'profile__student__class_instance__grade',
            'profile__teacher_profile',
        ).prefetch_related(
            'profile__teacher_profile__teacher_assignments',
            'profile__teacher_profile__teacher_assignments__class_instance',
            'profile__teacher_profile__teacher_assignments__class_instance__grade',
            'profile__teacher_profile__teacher_assignments__subject',
        ).get(id=request.user.id)

        response_data = dict(UserSerializer(user).data)
        response_data['subject_progress'] = _build_subject_progress(user)

        return Response({
            "success": True,
            "message": "Profile retrieved successfully",
            "data": response_data
        })


class UserViewSet(viewsets.ModelViewSet):
    """
    General User ViewSet for admin operations.
    For student/teacher specific operations, use StudentViewSet or TeacherViewSet.
    """
    queryset = Account.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return UserUpdateSerializer
        elif self.action == 'list':
            return UserListSerializer
        return UserSerializer
    
    def get_queryset(self):
        """Simplified queryset - for general user operations."""
        queryset = Account.objects.select_related(
            'profile__school',
            'profile__student__level',
            'profile__student__class_instance',
            'profile__student__class_instance__grade',
            'profile__teacher_profile',
        ).prefetch_related(
            'profile__teacher_profile__teacher_assignments',
            'profile__teacher_profile__teacher_assignments__class_instance',
            'profile__teacher_profile__teacher_assignments__subject',
        ).filter(is_active=True)

        requester_profile = getattr(self.request.user, 'profile', None)
        requester_type = getattr(requester_profile, 'user_type', None)

        user_type = self.request.query_params.get('user_type', None)
        if user_type:
            queryset = queryset.filter(profile__user_type=user_type)

        # Scope by school — all roles are school-scoped
        # Teachers are additionally scoped to their assigned classes when viewing students
        requester_school = getattr(requester_profile, 'school', None)
        if user_type == 'student':
            if requester_school:
                queryset = queryset.filter(profile__school=requester_school)
            else:
                return queryset.none()

            if requester_type == 'teacher':
                teacher_profile = getattr(requester_profile, 'teacher_profile', None)
                if teacher_profile:
                    teacher_classes = Teacher.objects.filter(
                        teacher=teacher_profile
                    ).values_list('class_instance', flat=True).distinct()
                    queryset = queryset.filter(profile__student__class_instance__id__in=teacher_classes)
                else:
                    return queryset.none()

        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(username__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search)
            )

        is_active = self.request.query_params.get('is_active', None)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')

        return queryset
    
    def get_permissions(self):
        if self.action in ['create', 'list', 'destroy']:
            return [TeacherAdminPermission()]
        elif self.action in ['update', 'partial_update']:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated()]
    
    def perform_update(self, serializer):
        return serializer.save()
    
    def list(self, request, *args, **kwargs):
        api_logger.info(f"User list requested by {request.user.username} (ID: {request.user.id}) from {request.META.get('REMOTE_ADDR', 'unknown')}")
        
        queryset = self.filter_queryset(self.get_queryset())
        # page = self.paginate_queryset(queryset)
        # if page is not None:
        #     serializer = self.get_serializer(page, many=True)
        #     api_logger.info(f"User list returned {len(serializer.data)} users")
        #     return success(
        #         data=serializer.data,
        #         message="Users retrieved successfully"
        #     )
        serializer = self.get_serializer(queryset, many=True)
        api_logger.info(f"User list returned {len(serializer.data)} users")
        return success(
            data=serializer.data,
            message="Users retrieved successfully"
        )
    
    def retrieve(self, request, *args, **kwargs):
        user_id = kwargs.get('pk')
        api_logger.info(f"User retrieve requested by {request.user.username} (ID: {request.user.id}) for user ID: {user_id}")
        
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return success(
            data=serializer.data,
            message="User retrieved successfully"
        )
    
    def create(self, request, *args, **kwargs):
        import traceback
        api_logger.info(f"User creation requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")

        serializer = self.get_serializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            try:
                user = serializer.save()
                user_serializer = UserSerializer(user)
                api_logger.info(f"User created successfully: {user.username} (ID: {user.id}) by {request.user.username}")
                return created(
                    data=user_serializer.data,
                    message=Messages.USER_CREATED
                )
            except Exception as e:
                api_logger.error(f"User creation exception: {e}\n{traceback.format_exc()}")
                return validation_error({"error": str(e)})
        api_logger.warning(f"User creation failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)
    
    def update(self, request, *args, **kwargs):
        import traceback
        user_id = kwargs.get('pk')
        api_logger.info(f"User update requested by {request.user.username} (ID: {request.user.id}) for user ID: {user_id} - Data: {request.data}")

        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            try:
                user = serializer.save()
                user_serializer = UserSerializer(user)
                api_logger.info(f"User updated successfully: {user.username} (ID: {user.id}) by {request.user.username}")
                return success(
                    data=user_serializer.data,
                    message=Messages.USER_UPDATED
                )
            except Exception as e:
                api_logger.error(f"User update exception: {e}\n{traceback.format_exc()}")
                return validation_error({"error": str(e)})
        api_logger.warning(f"User update failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)
    
    def destroy(self, request, *args, **kwargs):
        user_id = kwargs.get('pk')
        api_logger.info(f"User soft delete requested by {request.user.username} (ID: {request.user.id}) for user ID: {user_id}")
        
        instance = self.get_object()
        instance.soft_delete()
        api_logger.info(f"User soft deleted successfully: {instance.username} (ID: {instance.id}) by {request.user.username}")
        return success(
            message="User soft deleted successfully",
            status_code=StatusCodes.NO_CONTENT
        )
    
    @action(detail=True, methods=['post'], permission_classes=[TeacherAdminPermission])
    def restore(self, request, pk=None):
        api_logger.info(f"User restore requested by {request.user.username} (ID: {request.user.id}) for user ID: {pk}")
        
        instance = self.get_object()
        if not instance.is_deleted:
            api_logger.warning(f"User restore failed: User {instance.username} (ID: {instance.id}) is not deleted")
            return error(message="User is not deleted", status_code=StatusCodes.BAD_REQUEST)
        
        instance.restore()
        api_logger.info(f"User restored successfully: {instance.username} (ID: {instance.id}) by {request.user.username}")
        return success(
            data=UserSerializer(instance).data,
            message="User restored successfully"
        )
    
    @action(detail=True, methods=['post'])
    def add_exp(self, request, pk=None):
        user_id = pk
        api_logger.info(f"Add experience points requested by {request.user.username} (ID: {request.user.id}) for user ID: {user_id} - Data: {request.data}")
        
        user = self.get_object()
        serializer = UserExpSerializer(data=request.data, context={'request': request})
        
        if serializer.is_valid():
            exp_points = serializer.validated_data['exp_points']
            if hasattr(user, 'profile'):
                user.profile.add_exp(exp_points)
            api_logger.info(f"Experience points added successfully: {exp_points} points to {user.username} (ID: {user.id}) by {request.user.username}")
            user.refresh_from_db()
            return success(
                data={
                    'total_exp': user.profile.total_exp if hasattr(user, 'profile') else 0,
                    'level': user.profile.get_level() if hasattr(user, 'profile') else None,
                    'exp_to_next_level': user.profile.get_exp_to_next_level() if hasattr(user, 'profile') else 0
                },
                message=f"Added {exp_points} experience points"
            )
        api_logger.warning(f"Add experience points failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)
    
    @action(detail=True, methods=['get'])
    def profile(self, request, pk=None):
        user_id = pk
        api_logger.info(f"User profile requested by {request.user.username} (ID: {request.user.id}) for user ID: {user_id}")
        
        user = self.get_object()
        serializer = UserSerializer(user)
        return success(
            data=serializer.data,
            message="User profile retrieved successfully"
        )
    
    @action(detail=False, methods=['get', 'put', 'patch'])
    def me(self, request):
        api_logger.info(f"Current user profile requested by {request.user.username} (ID: {request.user.id})")

        if request.method in ('PUT', 'PATCH'):
            partial = request.method == 'PATCH'
            serializer = UserSerializer(request.user, data=request.data, partial=partial, context={'request': request})
            if serializer.is_valid():
                serializer.save()
                api_logger.info(f"Profile updated for user {request.user.username} (ID: {request.user.id})")
                return success(data=serializer.data, message="Profile updated successfully")
            return validation_error(serializer.errors)

        user = Account.objects.select_related(
            'profile__school',
            'profile__student__level',
            'profile__student__class_instance',
            'profile__student__class_instance__grade',
            'profile__teacher_profile',
        ).prefetch_related(
            'profile__teacher_profile__teacher_assignments',
            'profile__teacher_profile__teacher_assignments__class_instance',
            'profile__teacher_profile__teacher_assignments__class_instance__grade',
            'profile__teacher_profile__teacher_assignments__subject',
        ).get(id=request.user.id)
        user.last_login = timezone.now()
        user.save(update_fields=['last_login'])

        response_data = dict(UserSerializer(user).data)
        response_data['subject_progress'] = _build_subject_progress(user)

        return success(
            data=response_data,
            message="Profile retrieved successfully"
        )
    
    @action(detail=False, methods=['post'])
    def change_password(self, request):
        api_logger.info(f"Password change requested by {request.user.username} (ID: {request.user.id})")
        
        serializer = UserPasswordChangeSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            api_logger.info(f"Password changed successfully for user {request.user.username} (ID: {request.user.id})")
            return success(message=Messages.PASSWORD_CHANGED)
        api_logger.warning(f"Password change failed for user {request.user.username} (ID: {request.user.id}) - Errors: {serializer.errors}")
        return validation_error(serializer.errors)
    
    @action(detail=False, methods=['get'])
    def leaderboard(self, request):
        api_logger.info(f"Leaderboard requested by {request.user.username} (ID: {request.user.id}) from {request.META.get('REMOTE_ADDR', 'unknown')}")
        
        from django.db.models import Avg, F, Case, When, IntegerField, Q
        from datetime import timedelta

        queryset = Account.objects.select_related(
            'profile__school',
            'profile__student__level',
            'profile__student__class_instance',
            'profile__student__class_instance__grade',
        ).prefetch_related(
            'missions__progress',
            'competition_progress',
        ).filter(
            is_active=True,
            profile__user_type='student',
        )
        
        user_type = request.query_params.get('user_type', None)
        if user_type:
            queryset = queryset.filter(profile__user_type=user_type)
        
        school_id = request.query_params.get('school_id', None)
        if school_id:
            queryset = queryset.filter(profile__school_id=school_id)
        
        class_id = request.query_params.get('class_id', None)
        if class_id:
            queryset = queryset.filter(profile__student__class_instance_id=class_id)
        
        grade = request.query_params.get('grade', None)
        if grade:
            queryset = queryset.filter(profile__student__class_instance__grade__name__icontains=grade)
        
        requester_profile = getattr(request.user, 'profile', None)
        requester_type = getattr(requester_profile, 'user_type', None)

        # Always scope by school (teacher, admin/principal, student all see their school only)
        requester_school = getattr(requester_profile, 'school', None)
        if requester_school:
            queryset = queryset.filter(profile__school=requester_school)
        else:
            queryset = queryset.none()

        if not class_id and not grade:
            if requester_type == 'teacher':
                # Teachers are further scoped to their assigned classes
                teacher_profile = getattr(requester_profile, 'teacher_profile', None)
                if teacher_profile:
                    teacher_classes = Teacher.objects.filter(
                        teacher=teacher_profile
                    ).values_list('class_instance', flat=True).distinct()
                    queryset = queryset.filter(profile__student__class_instance__id__in=teacher_classes)
                else:
                    queryset = queryset.none()
            elif requester_type == 'student':
                # Students see their own class leaderboard
                try:
                    student = request.user.profile.student
                    if student and student.class_instance:
                        queryset = queryset.filter(profile__student__class_instance=student.class_instance)
                except (UserProfile.DoesNotExist, Student.DoesNotExist, AttributeError):
                    pass
            # admin/principal: no further scoping — sees all students in their school
        
        subject_id = request.query_params.get('subject_id', None)
        if subject_id:
            user_ids_with_subject_progress = UserModuleProgress.objects.filter(
                module__subject_id=subject_id
            ).values_list('account_id', flat=True).distinct()
            queryset = queryset.filter(id__in=user_ids_with_subject_progress)
        
        min_xp = request.query_params.get('min_xp', None)
        if min_xp:
            try:
                queryset = queryset.filter(profile__student__total_exp__gte=int(min_xp))
            except (ValueError, TypeError):
                pass
        
        max_xp = request.query_params.get('max_xp', None)
        if max_xp:
            try:
                queryset = queryset.filter(profile__student__total_exp__lte=int(max_xp))
            except (ValueError, TypeError):
                pass
        
        queryset = queryset.annotate(
            avg_mission_score=Avg(
                Case(
                    When(
                        missions__progress__status='completed',
                        then=F('missions__progress__exp_earned')
                    ),
                    default=None,
                    output_field=IntegerField()
                )
            )
        ).annotate(
            avg_competition_score=Avg(
                Case(
                    When(
                        competition_progress__status='completed',
                        then=F('competition_progress__score')
                    ),
                    default=None,
                    output_field=IntegerField()
                )
            )
        )
        
        from django.db.models import FloatField
        queryset = queryset.annotate(
            average_score=Case(
                When(
                    avg_mission_score__isnull=False,
                    avg_competition_score__isnull=False,
                    then=(F('avg_mission_score') + F('avg_competition_score')) / 2.0
                ),
                When(
                    avg_mission_score__isnull=False,
                    then=F('avg_mission_score')
                ),
                When(
                    avg_competition_score__isnull=False,
                    then=F('avg_competition_score')
                ),
                default=0.0,
                output_field=FloatField()
            )
        )
        
        min_score = request.query_params.get('min_score', None)
        if min_score:
            try:
                queryset = queryset.filter(average_score__gte=int(min_score))
            except (ValueError, TypeError):
                pass
        
        max_score = request.query_params.get('max_score', None)
        if max_score:
            try:
                queryset = queryset.filter(average_score__lte=int(max_score))
            except (ValueError, TypeError):
                pass
        
        sort_by = request.query_params.get('sort_by', 'xp')
        if sort_by == 'score':
            queryset = queryset.order_by('-average_score', '-profile__student__total_exp')
        else:
            queryset = queryset.order_by('-profile__student__total_exp', '-average_score')
        
        stats_queryset = queryset
        
        highest_xp_user = stats_queryset.order_by('-profile__student__total_exp').first()
        highest_xp = highest_xp_user.profile.total_exp if (highest_xp_user and hasattr(highest_xp_user, 'profile')) else 0
        
        best_score_user = stats_queryset.order_by('-average_score').first()
        best_average_score = best_score_user.average_score if best_score_user and best_score_user.average_score else 0
        
        thirty_days_ago = timezone.now() - timedelta(days=30)
        active_students = stats_queryset.filter(
            Q(missions__progress__completed_at__gte=thirty_days_ago) |
            Q(competition_progress__completed_at__gte=thirty_days_ago) |
            Q(missions__progress__started_at__gte=thirty_days_ago) |
            Q(competition_progress__started_at__gte=thirty_days_ago)
        ).distinct().count()
        
        class_active_students = None
        subject_active_students = None
        
        if class_id:
            class_active_students = stats_queryset.filter(
                profile__student__class_instance_id=class_id
            ).filter(
                Q(missions__progress__completed_at__gte=thirty_days_ago) |
                Q(competition_progress__completed_at__gte=thirty_days_ago) |
                Q(missions__progress__started_at__gte=thirty_days_ago) |
                Q(competition_progress__started_at__gte=thirty_days_ago)
            ).distinct().count()
        
        if subject_id:
            subject_user_ids = UserModuleProgress.objects.filter(
                module__subject_id=subject_id
            ).values_list('account_id', flat=True).distinct()
            subject_active_students = stats_queryset.filter(
                id__in=subject_user_ids
            ).filter(
                Q(missions__progress__completed_at__gte=thirty_days_ago) |
                Q(competition_progress__completed_at__gte=thirty_days_ago) |
                Q(missions__progress__started_at__gte=thirty_days_ago) |
                Q(competition_progress__started_at__gte=thirty_days_ago)
            ).distinct().count()
        
        class_name = None
        grade_name = None
        
        if class_id:
            try:
                class_instance = Class.objects.select_related('grade').get(id=class_id)
                class_name = class_instance.name
                if class_instance.grade:
                    grade_name = class_instance.grade.name
            except (Class.DoesNotExist, ValueError, TypeError):
                pass
        else:
            if requester_type == 'student':
                try:
                    student = request.user.profile.student
                    if student and student.class_instance:
                        class_instance = student.class_instance
                        class_name = class_instance.name
                        if class_instance.grade:
                            grade_name = class_instance.grade.name
                except (UserProfile.DoesNotExist, Student.DoesNotExist, AttributeError):
                    pass
        
        serializer = UserListSerializer(queryset[:1000], many=True)
        api_logger.info(f"Leaderboard returned {len(serializer.data)} users")
        
        response_data = {
            'results': serializer.data,
            'statistics': {
                'highest_xp': highest_xp,
                'best_average_score': round(best_average_score, 2) if best_average_score else 0,
                'active_students': active_students,
                'class_active_students': class_active_students,
                'subject_active_students': subject_active_students,
            }
        }
        
        return Response({
            'data': response_data,
            'message': 'Leaderboard retrieved successfully',
            'class_name': class_name,
            'grade_name': grade_name
        })
    
class DashboardViewSet(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def _calculate_teacher_style_subject_metrics(self, student_ids, subject_id):
        """Mirror teacher dashboard formulas for one subject and one scoped student set."""
        if not student_ids or not subject_id:
            return {
                'proficiency': 0,
                'attempt_rate': 0,
                'weak_topic_count': 0,
            }

        due_chapter_ids = list(
            ModuleChapter.objects.filter(
                module__subject_id=subject_id,
                module__is_active=True,
                is_deleted=False,
                due_date__isnull=False,
            ).values_list('id', flat=True)
        )
        total_due_chapters = len(due_chapter_ids)

        attempt_rate = 0
        if total_due_chapters > 0:
            attempted_pairs = set(
                Answer.objects.filter(
                    user_id__in=student_ids,
                    chapter_id__in=due_chapter_ids,
                ).values_list('user_id', 'chapter_id').distinct()
            )
            sum_student_pct = sum(
                (
                    sum(1 for ch_id in due_chapter_ids if (uid, ch_id) in attempted_pairs) / total_due_chapters
                ) * 100
                for uid in student_ids
            )
            attempt_rate = round(sum_student_pct / len(student_ids))

        weak_topic_count = 0
        chapter_weak_stats = (
            Answer.objects.filter(
                chapter_id__in=due_chapter_ids,
                user_id__in=student_ids,
            )
            .values('chapter_id')
            .annotate(
                total=Count('id'),
                correct=Count('id', filter=Q(is_correct=True)),
            )
            .filter(total__gt=0)
        )
        weak_topic_count = len({
            row['chapter_id'] for row in chapter_weak_stats
            if round((row['correct'] / row['total']) * 100) < 50
        })

        due_modules = Module.objects.filter(
            subject_id=subject_id,
            is_active=True,
            chapters__due_date__isnull=False,
            chapters__is_deleted=False,
        ).distinct().order_by('order').prefetch_related('chapters')

        module_scores = []
        for mod in due_modules:
            chapter_scores = []
            for ch in mod.chapters.filter(is_deleted=False, due_date__isnull=False).order_by('order'):
                q_count = ModuleContent.objects.filter(
                    chapter=ch,
                    content_type='question',
                    is_deleted=False,
                ).values('question_id').distinct().count()
                if q_count == 0:
                    continue

                correct_per_user = dict(
                    Answer.objects.filter(
                        chapter=ch,
                        user_id__in=student_ids,
                        is_correct=True,
                    )
                    .values('user_id')
                    .annotate(c=Count('question_id', distinct=True))
                    .values_list('user_id', 'c')
                )
                sum_pct = sum(
                    (correct_per_user.get(uid, 0) / q_count) * 100
                    for uid in student_ids
                )
                chapter_scores.append(sum_pct / len(student_ids))

            if chapter_scores:
                module_scores.append(sum(chapter_scores) / len(chapter_scores))

        proficiency = round(sum(module_scores) / len(module_scores)) if module_scores else 0

        return {
            'proficiency': proficiency,
            'attempt_rate': attempt_rate,
            'weak_topic_count': weak_topic_count,
        }
    
    @action(detail=False, methods=['get'], url_path='metrics')
    def metrics(self, request):
        api_logger.info(f"Dashboard metrics requested by {request.user.username} (ID: {request.user.id})")

        role = request.query_params.get('role', request.user.profile.user_type if hasattr(request.user, 'profile') else None)

        if role == 'principal':
            user_school = request.user.profile.school if hasattr(request.user, 'profile') else None
            class_filter = request.query_params.get('class', '').strip()
            subject_filter = request.query_params.get('subject', '').strip()

            school_classes = Class.objects.filter(school=user_school, is_active=True)
            if class_filter:
                school_classes = school_classes.filter(name__icontains=class_filter)
            class_ids = list(school_classes.values_list('id', flat=True))

            school_students = Account.objects.filter(
                profile__user_type='student',
                is_active=True
            ).filter(
                Q(profile__school=user_school) |
                Q(profile__student__class_instance__school=user_school)
            ).distinct()
            if class_ids:
                school_students = school_students.filter(profile__student__class_instance_id__in=class_ids)
            student_ids = list(school_students.values_list('id', flat=True))

            school_teachers = Account.objects.filter(
                profile__user_type='teacher',
                profile__school=user_school,
                is_active=True
            )
            school_subjects = Subject.objects.filter(
                is_active=True,
                school=user_school
            ).distinct().order_by('name')
            if not school_subjects.exists():
                school_subjects = Subject.objects.filter(
                    is_active=True,
                    classes__school=user_school
                ).distinct().order_by('name')
            if not school_subjects.exists():
                school_subjects = Subject.objects.filter(
                    is_active=True,
                    modules__is_active=True
                ).distinct().order_by('name')
            if subject_filter:
                school_subjects = school_subjects.filter(id=subject_filter)

            total_students = school_students.count()
            total_teachers = school_teachers.count()
            total_classes = school_classes.count()
            total_subjects = school_subjects.count()

            avg_student_proficiency = school_students.aggregate(avg=Avg('profile__student__total_exp'))['avg'] or 0

            teachers_with_content = school_teachers.filter(
                Q(created_questions__isnull=False) |
                Q(created_subjects__isnull=False) |
                Q(created_modules__isnull=False)
            ).distinct().count()
            teacher_engagement = (teachers_with_content / max(total_teachers, 1)) * 100

            avg_class_score = school_students.aggregate(avg=Avg('profile__student__total_exp'))['avg'] or 0

            last_month = timezone.now() - timezone.timedelta(days=30)

            last_month_students = school_students.filter(date_joined__lt=last_month)
            last_month_avg_proficiency = last_month_students.aggregate(avg=Avg('profile__student__total_exp'))['avg'] or 0
            proficiency_change = self._calculate_percentage_change(avg_student_proficiency, last_month_avg_proficiency)

            last_month_teachers = school_teachers.filter(date_joined__lt=last_month)
            last_month_teachers_with_content = last_month_teachers.filter(
                Q(created_questions__isnull=False) |
                Q(created_subjects__isnull=False) |
                Q(created_modules__isnull=False)
            ).distinct().count()
            last_month_engagement = (last_month_teachers_with_content / max(last_month_teachers.count(), 1)) * 100
            engagement_change = self._calculate_percentage_change(teacher_engagement, last_month_engagement)

            class_score_change = self._calculate_percentage_change(avg_class_score, last_month_avg_proficiency)

            last_month_students_count = last_month_students.count()
            students_change = total_students - last_month_students_count

            metrics = [
                {
                    "title": "Overall Student Proficiency",
                    "value": f"{avg_student_proficiency:.0f}",
                    "change": f"{proficiency_change['change']}% vs last month",
                    "changeType": proficiency_change['type'],
                    "trend": proficiency_change['trend']
                },
                {
                    "title": "Teacher Engagement",
                    "value": f"{teacher_engagement:.0f}%",
                    "change": f"{engagement_change['change']}% vs last month",
                    "changeType": engagement_change['type'],
                    "trend": engagement_change['trend']
                },
                {
                    "title": "Average Class Score",
                    "value": f"{avg_class_score:.0f}",
                    "change": f"{class_score_change['change']}% vs last month",
                    "changeType": class_score_change['type'],
                    "trend": class_score_change['trend']
                },
                {
                    "title": "Active Students",
                    "value": str(total_students),
                    "change": f"{students_change:+d} vs last month",
                    "changeType": "positive" if students_change >= 0 else "negative",
                    "trend": "up" if students_change >= 0 else "down"
                }
            ]

            quick_summary = [
                {"label": "Active Teachers", "value": str(total_teachers)},
                {"label": "Active Students", "value": str(total_students)},
                {"label": "Classes in Session", "value": str(total_classes)},
                {"label": "Subjects Covered", "value": str(total_subjects)}
            ]

            subject_proficiency_data = []
            attempt_rate_data = []
            weak_topic_count_data = []
            for subject in school_subjects:
                enrolled_students_qs = school_students.filter(
                    profile__student__subject_enrollments__subject=subject,
                    profile__student__subject_enrollments__is_active=True,
                ).distinct()
                if not enrolled_students_qs.exists():
                    enrolled_students_qs = school_students.filter(
                        profile__student__class_instance__subjects=subject
                    ).distinct()
                if not enrolled_students_qs.exists():
                    enrolled_students_qs = school_students
                enrolled_student_ids = list(enrolled_students_qs.values_list('id', flat=True))
                teacher_style_metrics = self._calculate_teacher_style_subject_metrics(
                    enrolled_student_ids,
                    subject.id,
                )
                proficiency = teacher_style_metrics['proficiency']
                attempt_rate = teacher_style_metrics['attempt_rate']
                weak_topic_count = teacher_style_metrics['weak_topic_count']

                subject_proficiency_data.append({
                    'subjectId': str(subject.id),
                    'subject': subject.name,
                    'value': proficiency,
                })
                attempt_rate_data.append({
                    'subjectId': str(subject.id),
                    'subject': subject.name,
                    'value': attempt_rate,
                })
                weak_topic_count_data.append({
                    'subjectId': str(subject.id),
                    'subject': subject.name,
                    'value': weak_topic_count,
                })

            teacher_proficiency_data = []
            teacher_assignments = Teacher.objects.filter(
                is_deleted=False,
                class_instance__in=school_classes
            )
            if subject_filter:
                teacher_assignments = teacher_assignments.filter(subject_id=subject_filter)

            for teacher_account in school_teachers.select_related('profile__teacher_profile'):
                teacher_profile = getattr(getattr(teacher_account, 'profile', None), 'teacher_profile', None)
                if not teacher_profile:
                    continue
                assignments = teacher_assignments.filter(teacher=teacher_profile)
                if not assignments.exists():
                    continue

                assigned_class_ids = list(assignments.values_list('class_instance_id', flat=True).distinct())
                assigned_subject_ids = list(assignments.values_list('subject_id', flat=True).distinct())
                teacher_student_ids = list(Account.objects.filter(
                    profile__user_type='student',
                    profile__school=user_school,
                    profile__student__class_instance_id__in=assigned_class_ids,
                    profile__student__subject_enrollments__subject_id__in=assigned_subject_ids,
                    profile__student__subject_enrollments__is_active=True,
                    is_active=True
                ).distinct().values_list('id', flat=True))

                teacher_answers = Answer.objects.filter(
                    user_id__in=teacher_student_ids,
                    chapter__module__subject_id__in=assigned_subject_ids,
                    chapter__module__is_active=True,
                    chapter__is_deleted=False,
                )
                teacher_total = teacher_answers.count()
                teacher_correct = teacher_answers.filter(is_correct=True).count()
                teacher_score = round((teacher_correct / teacher_total) * 100) if teacher_total else 0
                teacher_name = f"{teacher_account.first_name} {teacher_account.last_name}".strip() or teacher_account.username
                teacher_proficiency_data.append({
                    'teacher': teacher_name,
                    'value': teacher_score,
                })

            selected_subject_proficiency = next(
                (item['value'] for item in subject_proficiency_data if item.get('subjectId') == subject_filter),
                None
            )
            selected_subject_attempt_rate = next(
                (item['value'] for item in attempt_rate_data if item.get('subjectId') == subject_filter),
                None
            )
            selected_subject_weak_topic_count = next(
                (item['value'] for item in weak_topic_count_data if item.get('subjectId') == subject_filter),
                None
            )

            principal_charts = {
                'summary': {
                    'studentProficiency': selected_subject_proficiency if selected_subject_proficiency is not None else (
                        round(
                            sum(item['value'] for item in subject_proficiency_data) / max(len(subject_proficiency_data), 1)
                        ) if subject_proficiency_data else 0
                    ),
                    'attemptRate': selected_subject_attempt_rate if selected_subject_attempt_rate is not None else (
                        round(
                            sum(item['value'] for item in attempt_rate_data) / max(len(attempt_rate_data), 1)
                        ) if attempt_rate_data else 0
                    ),
                    'weakTopicCount': selected_subject_weak_topic_count if selected_subject_weak_topic_count is not None else sum(item['value'] for item in weak_topic_count_data),
                },
                'subjectProficiency': subject_proficiency_data,
                'attemptRate': attempt_rate_data,
                'teacherProficiency': teacher_proficiency_data,
                'weakTopicCount': weak_topic_count_data,
                'classes': list(school_classes.values_list('name', flat=True)),
                'subjects': [{'id': str(subject.id), 'name': subject.name} for subject in school_subjects],
            }
            if not principal_charts['classes']:
                principal_charts['classes'] = list(
                    Class.objects.filter(school=user_school, is_active=True).values_list('name', flat=True)
                )

        elif role == 'teacher':
            user_school = request.user.profile.school if hasattr(request.user, 'profile') else None
            class_filter = request.query_params.get('class', '').strip() or request.query_params.get('class_id', '').strip()
            subject_filter = request.query_params.get('subject', '').strip()

            # ── Step 1: Resolve teacher's assignments ─────────────────────────
            # All scoping flows from the Teacher (TeacherAssignment) table.
            # teacher_profile → Teacher rows → class_instance + subject
            teacher_profile = getattr(request.user.profile, 'teacher_profile', None)
            teacher_assignments = Teacher.objects.none()
            if teacher_profile:
                teacher_assignments = Teacher.objects.filter(
                    teacher=teacher_profile,
                    is_deleted=False,
                ).select_related('class_instance', 'subject')

            assigned_class_ids = teacher_assignments.values_list('class_instance_id', flat=True).distinct()
            teacher_subject_ids = teacher_assignments.values_list('subject_id', flat=True).distinct()

            teacher_classes = Class.objects.filter(
                id__in=assigned_class_ids,
                school=user_school,
                is_active=True,
            )
            if class_filter:
                try:
                    uuid.UUID(class_filter)
                    teacher_classes = teacher_classes.filter(id=class_filter)
                except (ValueError, TypeError):
                    teacher_classes = teacher_classes.filter(name__icontains=class_filter)

            classes_count = teacher_classes.count()

            # Re-derive subject IDs scoped to the filtered classes so that all
            # downstream queries (students, modules, chapters, weak topics) are
            # restricted to only the subjects the teacher teaches in those classes.
            teacher_subject_ids = teacher_assignments.filter(
                class_instance__in=teacher_classes,
            ).values_list('subject_id', flat=True).distinct()

            # ── Step 2: Base student queryset ─────────────────────────────────
            # Students in the teacher's classes who are enrolled (via
            # StudentSubjectEnrollment) in at least one of the teacher's subjects.
            # This is the source-of-truth for "which students does this teacher own."
            teacher_students_base = Account.objects.filter(
                is_active=True,
                is_deleted=False,
                profile__is_deleted=False,
                profile__user_type='student',
                profile__student__is_deleted=False,
                profile__student__class_instance__in=teacher_classes,
                profile__student__subject_enrollments__subject_id__in=teacher_subject_ids,
                profile__student__subject_enrollments__is_active=True,
            ).distinct()

            # ── Step 3: Narrow to a specific subject when filter is provided ──
            # Students must BOTH be in a class where the teacher teaches that
            # subject AND be individually enrolled in that subject via SSE.
            teacher_students = teacher_students_base
            if subject_filter:
                subject_assigned_class_ids = teacher_assignments.filter(
                    subject_id=subject_filter,
                    class_instance__in=teacher_classes,
                ).values_list('class_instance_id', flat=True)
                teacher_students = Account.objects.filter(
                    is_active=True,
                    is_deleted=False,
                    profile__is_deleted=False,
                    profile__user_type='student',
                    profile__student__is_deleted=False,
                    profile__student__class_instance__in=subject_assigned_class_ids,
                    profile__student__subject_enrollments__subject_id=subject_filter,
                    profile__student__subject_enrollments__is_active=True,
                ).distinct()

            students_count = teacher_students.count()
            teacher_student_ids = list(teacher_students.values_list('id', flat=True))

            # ── Fallback: if SSE-based count is 0, fall back to class membership ──
            # This handles legacy students who were created before StudentSubjectEnrollment
            # was introduced, or students whose SSE records were not populated.
            if students_count == 0 and teacher_classes.exists():
                _fallback_base = dict(
                    is_active=True,
                    is_deleted=False,
                    profile__is_deleted=False,
                    profile__user_type='student',
                    profile__student__is_deleted=False,
                )
                if subject_filter:
                    subject_assigned_class_ids = teacher_assignments.filter(
                        subject_id=subject_filter,
                        class_instance__in=teacher_classes,
                    ).values_list('class_instance_id', flat=True)
                    fallback_students = Account.objects.filter(
                        **_fallback_base,
                        profile__student__class_instance__in=subject_assigned_class_ids,
                    ).distinct()
                else:
                    fallback_students = Account.objects.filter(
                        **_fallback_base,
                        profile__student__class_instance__in=teacher_classes,
                    ).distinct()
                if fallback_students.exists():
                    api_logger.info(
                        f"[Teacher dashboard] SSE-based student count was 0; "
                        f"falling back to class membership for teacher {request.user.username}"
                    )
                    students_count = fallback_students.count()
                    teacher_student_ids = list(fallback_students.values_list('id', flat=True))

            questions_created = Question.objects.filter(created_by=request.user).count()

            # Modules covered = modules (in teacher's subjects) where every
            # chapter has a due_date set.
            modules_covered = Module.objects.filter(
                is_active=True,
                subject_id__in=teacher_subject_ids,
                class_instance__in=teacher_classes,
            ).annotate(
                total_chapters=Count('chapters', filter=Q(chapters__is_deleted=False)),
                chapters_with_due=Count(
                    'chapters',
                    filter=Q(chapters__is_deleted=False, chapters__due_date__isnull=False),
                ),
            ).filter(
                total_chapters__gt=0,
                total_chapters=F('chapters_with_due'),
            ).count()

            # Average Score = accuracy % (correct answers / total answers * 100).
            # Raw total_exp is meaningless as a "score" — use Answer table instead.
            last_month = timezone.now() - timezone.timedelta(days=30)
            _ans_stats = Answer.objects.filter(
                user_id__in=teacher_student_ids,
                chapter__module__subject_id__in=teacher_subject_ids,
            ).aggregate(
                total=Count('id'),
                correct=Count('id', filter=Q(is_correct=True)),
                total_last_month=Count('id', filter=Q(created_at__lt=last_month)),
                correct_last_month=Count('id', filter=Q(is_correct=True, created_at__lt=last_month)),
            )
            _total = _ans_stats['total'] or 0
            _correct = _ans_stats['correct'] or 0
            _total_lm = _ans_stats['total_last_month'] or 0
            _correct_lm = _ans_stats['correct_last_month'] or 0
            avg_student_score = round((_correct / _total) * 100) if _total > 0 else 0
            last_month_avg_score = round((_correct_lm / _total_lm) * 100) if _total_lm > 0 else 0

            score_change = self._calculate_percentage_change(avg_student_score, last_month_avg_score)
            last_month_students_count = Account.objects.filter(
                id__in=teacher_student_ids, date_joined__lt=last_month
            ).count()
            students_change = students_count - last_month_students_count

            last_month_questions = Question.objects.filter(
                created_by=request.user, created_at__lt=last_month,
            ).count()
            questions_change = questions_created - last_month_questions

            metrics = [
                {
                    "title": "My Students' Average Score",
                    "value": f"{avg_student_score}%",
                    "change": f"{score_change['change']}% vs last month",
                    "changeType": score_change['type'],
                    "trend": score_change['trend']
                },
                {
                    "title": "Students Engaged",
                    "value": str(students_count),
                    "change": f"{students_change:+d} vs last month",
                    "changeType": "positive" if students_change >= 0 else "negative",
                    "trend": "up" if students_change >= 0 else "down"
                },
                {
                    "title": "Questions Created",
                    "value": str(questions_created),
                    "change": f"{questions_change:+d} vs last month",
                    "changeType": "positive" if questions_change >= 0 else "negative",
                    "trend": "up" if questions_change >= 0 else "down"
                },
                {
                    "title": "Modules covered",
                    "value": str(modules_covered),
                    "change": "Modules with all chapters due",
                    "changeType": "positive",
                    "trend": "up"
                }
            ]

            # ── Weak Topics ───────────────────────────────────────────────────
            # Only chapters with a due_date set are considered.
            # A chapter is "weak" when (correct / total answers)*100 < 50.
            # Chapters with zero answers are ignored (not counted as weak).
            if teacher_student_ids:
                if subject_filter:
                    _scope_chapter_ids = list(ModuleChapter.objects.filter(
                        module__subject_id=subject_filter,
                        module__is_active=True,
                        is_deleted=False,
                        due_date__isnull=False,
                    ).values_list('id', flat=True))
                else:
                    _scope_chapter_ids = list(ModuleChapter.objects.filter(
                        module__subject_id__in=teacher_subject_ids,
                        module__is_active=True,
                        is_deleted=False,
                        due_date__isnull=False,
                    ).values_list('id', flat=True))
                ch_weak_stats = (
                    Answer.objects.filter(
                        chapter_id__in=_scope_chapter_ids,
                        user_id__in=teacher_student_ids,
                    )
                    .values('chapter_id')
                    .annotate(
                        total=Count('id'),
                        correct=Count('id', filter=Q(is_correct=True)),
                    )
                    .filter(total__gt=0)
                )
                weak_topic_count = len({
                    row['chapter_id'] for row in ch_weak_stats
                    if round((row['correct'] / row['total']) * 100) < 50
                })
            else:
                weak_topic_count = 0

            # ── Chapters Covered ──────────────────────────────────────────────
            if subject_filter:
                modules_for_chapters = Module.objects.filter(
                    subject_id=subject_filter, is_active=True,
                    class_instance__in=teacher_classes,
                )
                total_chapters = ModuleChapter.objects.filter(
                    module__in=modules_for_chapters, is_deleted=False
                ).count()
                chapters_with_due = ModuleChapter.objects.filter(
                    module__in=modules_for_chapters,
                    is_deleted=False,
                    due_date__isnull=False,
                ).count()
                chapter_covered_value = f"{chapters_with_due}/{total_chapters}" if total_chapters else "0/0"
            else:
                total_chapters = 0
                chapters_with_due = 0
                chapter_covered_value = "—"

            # ── Last Assignment Attempt Rate ───────────────────────────────────
            # Find the most recently due chapter for this subject, then compute
            # % of teacher's students who have at least one Answer for it.
            # Module.due_date has been removed; we order chapters by their own due_date.
            last_assignment_attempt_rate_value = "—"
            if subject_filter and teacher_student_ids:
                last_due_chapter = (
                    ModuleChapter.objects.filter(
                        module__subject_id=subject_filter,
                        module__is_active=True,
                        is_deleted=False,
                        due_date__isnull=False,
                    )
                    .order_by('due_date')
                    .last()
                )
                if last_due_chapter:
                    # Answer.chapter FK: direct lookup, no ModuleContent join
                    students_attempted = (
                        Answer.objects.filter(
                            user_id__in=teacher_student_ids,
                            chapter=last_due_chapter,
                        )
                        .values('user_id')
                        .distinct()
                        .count()
                    )
                    pct = round(students_attempted / len(teacher_student_ids) * 100)
                    last_assignment_attempt_rate_value = f"{pct}%"
                    api_logger.info(
                        f"[Teacher dashboard] Last assignment: chapter={last_due_chapter.id} "
                        f"({last_due_chapter.title}), attempted={students_attempted}/{len(teacher_student_ids)}, "
                        f"rate={pct}%"
                    )

            # ── Subject-wise stats ────────────────────────────────────────────
            # Pre-aggregate all per-subject data in bulk (3 queries total)
            # instead of 6 DB queries per subject (N+1).
            subject_wise_list = []
            teacher_own_subjects = list(Subject.objects.filter(
                id__in=teacher_subject_ids,
                is_active=True,
            ).order_by('name'))

            if teacher_own_subjects:
                subj_ids = [s.id for s in teacher_own_subjects]

                # ① Chapter totals + due counts per subject (1 query)
                ch_stats_qs = ModuleChapter.objects.filter(
                    module__subject_id__in=subj_ids,
                    module__is_active=True,
                    module__class_instance__in=teacher_classes,
                    is_deleted=False,
                ).values('module__subject_id').annotate(
                    total_ch=Count('id'),
                    ch_with_due=Count('id', filter=Q(due_date__isnull=False)),
                )
                ch_stats = {
                    row['module__subject_id']: row
                    for row in ch_stats_qs
                }

                # ② Weak topic counts per subject (1 query)
                # Aggregate answer accuracy per chapter, grouped by subject.
                weak_by_subj = {s.id: 0 for s in teacher_own_subjects}
                if teacher_student_ids:
                    ch_accuracy_qs = (
                        Answer.objects.filter(
                            user_id__in=teacher_student_ids,
                            chapter__isnull=False,
                            chapter__is_deleted=False,
                            chapter__due_date__isnull=False,
                            chapter__module__subject_id__in=subj_ids,
                            chapter__module__is_active=True,
                        )
                        .values('chapter_id', 'chapter__module__subject_id')
                        .annotate(
                            total=Count('id'),
                            correct=Count('id', filter=Q(is_correct=True)),
                        )
                        .filter(total__gt=0)
                    )
                    for row in ch_accuracy_qs:
                        if round((row['correct'] / row['total']) * 100) < 50:
                            weak_by_subj[row['chapter__module__subject_id']] += 1

                # ③ Questions created per subject by this teacher (1 query)
                q_counts_qs = Question.objects.filter(
                    created_by=request.user,
                    module_contents__chapter__module__subject_id__in=subj_ids,
                    module_contents__is_deleted=False,
                ).values('module_contents__chapter__module__subject_id').annotate(
                    q_count=Count('id', distinct=True),
                )
                q_by_subj = {
                    row['module_contents__chapter__module__subject_id']: row['q_count']
                    for row in q_counts_qs
                }

                # ④ Student counts per subject (1 query via annotation on assignments)
                # Group teacher assignments by subject → distinct student count per subject
                # Intersect with teacher_classes to respect any active class filter.
                students_by_subj = {}
                student_ids_by_subj = {}
                for subj in teacher_own_subjects:
                    subj_class_ids = teacher_assignments.filter(
                        subject=subj,
                        class_instance__in=teacher_classes,
                    ).values_list('class_instance_id', flat=True)
                    subj_student_ids = list(Account.objects.filter(
                        is_active=True,
                        is_deleted=False,
                        profile__is_deleted=False,
                        profile__user_type='student',
                        profile__student__is_deleted=False,
                        profile__student__class_instance__in=subj_class_ids,
                        profile__student__subject_enrollments__subject=subj,
                        profile__student__subject_enrollments__is_active=True,
                    ).distinct().values_list('id', flat=True))
                    students_by_subj[subj.id] = len(subj_student_ids)
                    student_ids_by_subj[subj.id] = subj_student_ids

                # ⑤ Attempt rate per subject (due chapters attempted / total due chapters)
                attempt_rate_by_subj = {}
                if teacher_student_ids:
                    subj_due_ch_map = {}
                    for row in ModuleChapter.objects.filter(
                        module__subject_id__in=subj_ids,
                        module__is_active=True,
                        is_deleted=False,
                        due_date__isnull=False,
                    ).values('id', 'module__subject_id'):
                        subj_due_ch_map.setdefault(row['module__subject_id'], []).append(row['id'])

                    for subj in teacher_own_subjects:
                        ch_ids = subj_due_ch_map.get(subj.id, [])
                        total_due = len(ch_ids)
                        subj_sids = student_ids_by_subj.get(subj.id) or teacher_student_ids
                        if total_due == 0 or not subj_sids:
                            attempt_rate_by_subj[subj.id] = 0
                            continue
                        pairs = set(
                            Answer.objects.filter(
                                user_id__in=subj_sids,
                                chapter_id__in=ch_ids,
                            ).values_list('user_id', 'chapter_id').distinct()
                        )
                        sum_pct = sum(
                            sum(1 for ch_id in ch_ids if (uid, ch_id) in pairs) / total_due * 100
                            for uid in subj_sids
                        )
                        attempt_rate_by_subj[subj.id] = round(sum_pct / len(subj_sids))

                for subj in teacher_own_subjects:
                    s_ch = ch_stats.get(subj.id, {})
                    total_ch = s_ch.get('total_ch', 0)
                    ch_due = s_ch.get('ch_with_due', 0)
                    if total_ch == 0:
                        continue  # subject has no modules/chapters yet

                    subject_wise_list.append({
                        'subjectId': str(subj.id),
                        'subjectName': subj.name,
                        'totalChapters': total_ch,
                        'chaptersWithDue': ch_due,
                        'chapterCoveredValue': f"{ch_due}/{total_ch}",
                        'weakTopicCount': weak_by_subj.get(subj.id, 0),
                        'questionsCreated': q_by_subj.get(subj.id, 0),
                        'studentsCount': students_by_subj.get(subj.id, 0),
                        'attemptRate': attempt_rate_by_subj.get(subj.id, 0),
                    })

            # ── Overall Student Percentage ─────────────────────────────────────
            # Per due chapter: avg across students of (correct / total_questions)*100
            # Per module: avg of chapter scores.  Overall: avg of module scores.
            # Uses Answer.chapter for direct lookup (no ModuleContent join).
            overall_student_percentage_value = "—"
            if subject_filter and teacher_student_ids:
                due_modules = Module.objects.filter(
                    subject_id=subject_filter,
                    is_active=True,
                    chapters__due_date__isnull=False,
                    chapters__is_deleted=False,
                ).distinct().order_by('order').prefetch_related('chapters')

                module_scores = []
                for mod in due_modules:
                    due_chapters = mod.chapters.filter(
                        is_deleted=False, due_date__isnull=False
                    ).order_by('order')
                    chapter_scores = []
                    for ch in due_chapters:
                        q_count = ModuleContent.objects.filter(
                            chapter=ch, content_type='question', is_deleted=False,
                        ).values('question_id').distinct().count()
                        if q_count == 0:
                            continue
                        # Count distinct correct questions per student using Answer.chapter
                        correct_per_user = dict(
                            Answer.objects.filter(
                                chapter=ch,
                                user_id__in=teacher_student_ids,
                                is_correct=True,
                            )
                            .values('user_id')
                            .annotate(c=Count('question_id', distinct=True))
                            .values_list('user_id', 'c')
                        )
                        sum_pct = sum(
                            (correct_per_user.get(uid, 0) / q_count) * 100
                            for uid in teacher_student_ids
                        )
                        chapter_scores.append(sum_pct / len(teacher_student_ids))
                    if chapter_scores:
                        module_scores.append(sum(chapter_scores) / len(chapter_scores))

                if module_scores:
                    overall_pct = round(sum(module_scores) / len(module_scores))
                    overall_student_percentage_value = f"{overall_pct}%"

            # ── Attempt Rate ──────────────────────────────────────────────────
            # For each student: (due chapters attempted / total due chapters)*100.
            # Uses Answer.chapter for O(1) lookup per chapter.
            # When subject_filter is set, scope to that subject; otherwise use all
            # teacher subjects so the card shows a value even without a filter.
            attempt_rate_value = "—"
            if teacher_student_ids:
                _ar_chapter_filter = dict(
                    module__is_active=True,
                    is_deleted=False,
                    due_date__isnull=False,
                )
                if subject_filter:
                    _ar_chapter_filter['module__subject_id'] = subject_filter
                else:
                    _ar_chapter_filter['module__subject_id__in'] = list(teacher_subject_ids)

                due_chapter_ids = list(
                    ModuleChapter.objects.filter(**_ar_chapter_filter).values_list('id', flat=True)
                )
                total_due_chapters = len(due_chapter_ids)
                if total_due_chapters > 0:
                    attempted_pairs = set(
                        Answer.objects.filter(
                            user_id__in=teacher_student_ids,
                            chapter_id__in=due_chapter_ids,
                        )
                        .values_list('user_id', 'chapter_id')
                        .distinct()
                    )
                    sum_student_pct = sum(
                        (
                            sum(
                                1 for ch_id in due_chapter_ids
                                if (uid, ch_id) in attempted_pairs
                            ) / total_due_chapters
                        ) * 100
                        for uid in teacher_student_ids
                    )
                    attempt_rate_value = f"{round(sum_student_pct / len(teacher_student_ids))}%"

            # ── Per-class chart data ───────────────────────────────────────────
            # Uses the same formulas as attempt_rate_value and
            # overall_student_percentage_value, scoped per class.
            class_wise_chart = []
            if subject_filter and teacher_classes.exists():
                _due_ch_ids = list(
                    ModuleChapter.objects.filter(
                        module__subject_id=subject_filter,
                        module__is_active=True,
                        is_deleted=False,
                        due_date__isnull=False,
                    ).values_list('id', flat=True)
                )
                _total_due = len(_due_ch_ids)
                _due_mods = list(
                    Module.objects.filter(
                        subject_id=subject_filter,
                        is_active=True,
                        chapters__due_date__isnull=False,
                        chapters__is_deleted=False,
                    ).distinct().order_by('order').prefetch_related('chapters')
                )
                # Pre-compute question counts per due chapter (shared across classes)
                _ch_q_counts = {}
                for _mod in _due_mods:
                    for _ch in _mod.chapters.filter(is_deleted=False, due_date__isnull=False):
                        _ch_q_counts[_ch.id] = ModuleContent.objects.filter(
                            chapter=_ch, content_type='question', is_deleted=False,
                        ).values('question_id').distinct().count()

                for cls in teacher_classes.order_by('name'):
                    cls_student_ids = list(Account.objects.filter(
                        is_active=True,
                        is_deleted=False,
                        profile__user_type='student',
                        profile__student__is_deleted=False,
                        profile__student__class_instance=cls,
                        profile__student__subject_enrollments__subject_id=subject_filter,
                        profile__student__subject_enrollments__is_active=True,
                    ).distinct().values_list('id', flat=True))
                    if not cls_student_ids:
                        continue

                    # Attempt rate (same as attempt_rate_value)
                    cls_attempt_rate = 0
                    if _total_due > 0:
                        cls_pairs = set(
                            Answer.objects.filter(
                                user_id__in=cls_student_ids,
                                chapter_id__in=_due_ch_ids,
                            ).values_list('user_id', 'chapter_id').distinct()
                        )
                        cls_attempt_rate = round(
                            sum(
                                sum(1 for ch_id in _due_ch_ids if (uid, ch_id) in cls_pairs) / _total_due * 100
                                for uid in cls_student_ids
                            ) / len(cls_student_ids)
                        )

                    # Performance (same as overall_student_percentage_value)
                    cls_mod_scores = []
                    for _mod in _due_mods:
                        cls_ch_scores = []
                        for _ch in _mod.chapters.filter(is_deleted=False, due_date__isnull=False):
                            q_count = _ch_q_counts.get(_ch.id, 0)
                            if q_count == 0:
                                continue
                            correct_per_user = dict(
                                Answer.objects.filter(
                                    chapter=_ch,
                                    user_id__in=cls_student_ids,
                                    is_correct=True,
                                ).values('user_id').annotate(c=Count('question_id', distinct=True)).values_list('user_id', 'c')
                            )
                            sum_ch_pct = sum(
                                (correct_per_user.get(uid, 0) / q_count) * 100
                                for uid in cls_student_ids
                            )
                            cls_ch_scores.append(sum_ch_pct / len(cls_student_ids))
                        if cls_ch_scores:
                            cls_mod_scores.append(sum(cls_ch_scores) / len(cls_ch_scores))
                    cls_performance = round(sum(cls_mod_scores) / len(cls_mod_scores)) if cls_mod_scores else 0

                    class_wise_chart.append({
                        'className': cls.name,
                        'attemptRate': cls_attempt_rate,
                        'performance': cls_performance,
                    })

            quick_summary = [
                {"label": "My Students", "value": str(students_count)},
                {"label": "Classes Teaching", "value": str(classes_count)},
                {"label": "Questions Created", "value": str(questions_created)},
                {"label": "Chapters covered", "value": str(modules_covered)},
                {"label": "Weak Topics", "value": str(weak_topic_count)},
                {"label": "Topics Covered", "value": chapter_covered_value},
                {"label": "Last Assignment Attempt Rate", "value": last_assignment_attempt_rate_value},
                {"label": "Overall Student Percentage", "value": overall_student_percentage_value},
                {"label": "Attempt Rate", "value": attempt_rate_value},
            ]
            
        else:
            student_modules_started = UserModuleProgress.objects.filter(account=request.user).count()

            questions_solved = (request.user.profile.total_exp if hasattr(request.user, 'profile') else 0) // 10
            
            # Modules covered = modules where all chapters (is_deleted=False) have due_date set
            modules_covered = Module.objects.filter(is_active=True).annotate(
                total_chapters=Count('chapters', filter=Q(chapters__is_deleted=False)),
                chapters_with_due=Count('chapters', filter=Q(chapters__is_deleted=False) & Q(chapters__due_date__isnull=False)),
            ).filter(total_chapters__gt=0, total_chapters=F('chapters_with_due')).count()

            last_week_exp = (request.user.profile.total_exp if hasattr(request.user, 'profile') else 0) - 50
            progress_change = (request.user.profile.total_exp if hasattr(request.user, 'profile') else 0) - last_week_exp
            
            current_level = request.user.profile.get_level() if hasattr(request.user, 'profile') else None
            level_change_text = "Level up!" if current_level > 1 else "New to platform"
            
            last_week_questions = questions_solved - 5
            questions_change = questions_solved - last_week_questions

            metrics = [
                {
                    "title": "My Progress",
                    "value": f"{request.user.profile.total_exp if hasattr(request.user, 'profile') else 0}",
                    "change": f"{progress_change:+d} vs last week",
                    "changeType": "positive" if progress_change >= 0 else "negative",
                    "trend": "up" if progress_change >= 0 else "down"
                },
                {
                    "title": "Current Level",
                    "value": f"Level {current_level}",
                    "change": level_change_text,
                    "changeType": "positive",
                    "trend": "up"
                },
                {
                    "title": "Questions Solved",
                    "value": str(questions_solved),
                    "change": f"{questions_change:+d} vs last week",
                    "changeType": "positive" if questions_change >= 0 else "negative",
                    "trend": "up" if questions_change >= 0 else "down"
                },
                {
                    "title": "Modules covered",
                    "value": str(modules_covered),
                    "change": "Modules with due date set",
                    "changeType": "positive",
                    "trend": "up"
                }
            ]
            
            quick_summary = [
                {"label": "Total Experience", "value": str(request.user.profile.total_exp if hasattr(request.user, 'profile') else 0)},
                {"label": "Current Level", "value": f"Level {request.user.profile.get_level() if hasattr(request.user, 'profile') else 0}"},
                {"label": "Rewards", "value": str(request.user.profile.rewards if hasattr(request.user, 'profile') else 0)},
                {"label": "Modules Started", "value": str(student_modules_started)}
            ]
        
        api_logger.info(f"Dashboard metrics returned for {role} role")
        response_data = {
            'metrics': metrics,
            'quickSummary': quick_summary
        }
        if role == 'principal':
            response_data['principalCharts'] = principal_charts
        if role == 'teacher':
            response_data['weakTopicCount'] = weak_topic_count
            response_data['subjectWise'] = subject_wise_list
            response_data['classWiseChart'] = class_wise_chart
        return success(
            data=response_data,
            message="Dashboard metrics retrieved successfully"
        )
    
    @action(detail=False, methods=['get'], url_path='progress-trends')
    def progress_trends(self, request):
        """Get progress trends data for charts."""
        api_logger.info(f"Progress trends requested by {request.user.username} (ID: {request.user.id})")

        from datetime import timedelta

        # ── FIX #8: Scope progress trends to the requesting user's school only.
        # Old code had no school filter, returning data across all schools.
        user_school = request.user.profile.school if hasattr(request.user, 'profile') else None

        weeks = []
        student_progress = []
        class_average = []

        for i in range(4):
            week_start = timezone.now() - timedelta(weeks=i+1)
            week_end = timezone.now() - timedelta(weeks=i)

            week_students = Account.objects.filter(
                profile__user_type='student',
                profile__school=user_school,
                is_active=True,
                last_login__gte=week_start,
                last_login__lt=week_end,
            )
            
            if week_students.exists():
                avg_exp = week_students.aggregate(avg=Avg('profile__student__total_exp'))['avg'] or 0
                student_progress.insert(0, round(avg_exp, 1))
                class_average.insert(0, round(avg_exp * 0.9, 1))
            else:
                student_progress.insert(0, 0)
                class_average.insert(0, 0)
            
            weeks.insert(0, f"Week {4-i}")
        
        trends_data = {
            'labels': weeks,
            'datasets': [
                {
                    'label': 'Student Progress',
                    'data': student_progress,
                    'borderColor': 'rgb(75, 192, 192)',
                    'backgroundColor': 'rgba(75, 192, 192, 0.2)'
                },
                {
                    'label': 'Class Average',
                    'data': class_average,
                    'borderColor': 'rgb(255, 99, 132)',
                    'backgroundColor': 'rgba(255, 99, 132, 0.2)'
                }
            ]
        }
        
        api_logger.info(f"Progress trends data returned")
        return success(
            data=trends_data,
            message="Progress trends retrieved successfully"
        )
    
    @action(detail=False, methods=['get'], url_path='subject-performance')
    def subject_performance(self, request):
        """Get subject performance data."""
        api_logger.info(f"Subject performance requested by {request.user.username} (ID: {request.user.id})")

        # ── FIX #9: Scope subject performance to the requesting user's school only.
        # Old code returned all active subjects with no school boundary.
        user_school = request.user.profile.school if hasattr(request.user, 'profile') else None
        subjects = Subject.objects.filter(
            is_active=True,
            classes__school=user_school,
        ).distinct()
        performance_data = []

        for subject in subjects:
            subject_students = Account.objects.filter(
                profile__user_type='student',
                profile__school=user_school,
                is_active=True,
                profile__student__class_instance__subjects=subject,
            )
            
            if subject_students.exists():
                avg_performance = subject_students.aggregate(
                    avg_exp=Avg('profile__student__total_exp')
                )['avg_exp'] or 0
                student_count = subject_students.count()
            else:
                avg_performance = 0
                student_count = 0
            
            performance_data.append({
                'subject': subject.name,
                'performance': round(avg_performance, 2),
                'students': student_count,
                'trend': 'up' if avg_performance > 70 else 'down'
            })
        
        api_logger.info(f"Subject performance data returned for {len(performance_data)} subjects")
        return success(
            data=performance_data,
            message="Subject performance retrieved successfully"
        )
    
    @action(detail=False, methods=['get'], url_path='class-distribution')
    def class_distribution(self, request):
        """Get class distribution data."""
        api_logger.info(f"Class distribution requested by {request.user.username} (ID: {request.user.id})")
        
        user_school = request.user.profile.school if hasattr(request.user, 'profile') else None
        classes = Class.objects.filter(school=user_school, is_active=True)
        distribution_data = []
        
        # ── FIX #10: Pre-compute the school-scoped student total once so the
        # percentage denominator is correct and we don't run a global COUNT
        # inside the loop. Old code used all students across all schools.
        school_total_students = Account.objects.filter(
            profile__user_type='student',
            profile__school=user_school,
            is_active=True,
        ).count()

        for class_obj in classes:
            student_count = Account.objects.filter(
                profile__user_type='student',
                profile__student__class_instance=class_obj,
                is_active=True,
            ).count()

            distribution_data.append({
                'class': class_obj.name,
                'students': student_count,
                'percentage': round(
                    (student_count / max(school_total_students, 1)) * 100, 1
                ),
            })
        
        api_logger.info(f"Class distribution data returned for {len(distribution_data)} classes")
        return success(
            data=distribution_data,
            message="Class distribution retrieved successfully"
        )
    
    @action(detail=False, methods=['get'], url_path='alerts')
    def alerts(self, request):
        """Get alerts and announcements."""
        api_logger.info(f"Dashboard alerts requested by {request.user.username} (ID: {request.user.id})")
        
        alerts_data = []
        
        user_type = request.user.profile.user_type if hasattr(request.user, 'profile') else None
        if user_type == 'student':
            new_missions = Mission.objects.filter(
                class_group=request.user.profile.student.class_instance if hasattr(request.user, 'profile') and hasattr(request.user.profile, 'student') and request.user.profile.student.class_instance else None,
                is_active=True,
                mission_date__gte=timezone.now().date()
            ).exclude(
                user_progress__account=request.user,
                user_progress__status__in=['in_progress', 'completed']
            )
            
            for mission in new_missions[:2]:
                alerts_data.append({
                    'id': f"mission_{mission.id}",
                    'type': 'info',
                    'title': 'New Mission Available',
                    'message': f'A new mission "{mission.title}" has been assigned to your class.',
                    'timestamp': mission.created_at.isoformat(),
                    'read': False
                })
        
        overdue_modules = UserModuleProgress.objects.filter(
            account=request.user,
            status='in_progress',
            started_at__lt=timezone.now() - timezone.timedelta(days=7)
        )
        
        for module_progress in overdue_modules[:2]:
            alerts_data.append({
                'id': f"overdue_{module_progress.id}",
                'type': 'warning',
                'title': 'Assignment Overdue',
                'message': f'Module "{module_progress.module.name}" is overdue.',
                'timestamp': module_progress.started_at.isoformat(),
                'read': False
            })
        
        user_profile = getattr(request.user, 'profile', None)
        if user_profile and user_profile.level and user_profile.level.name > 1:
            alerts_data.append({
                'id': f"levelup_{request.user.id}",
                'type': 'success',
                'title': 'Level Up!',
                'message': f'Congratulations! You have reached Level {user_profile.level.name}.',
                'timestamp': request.user.last_login.isoformat() if request.user.last_login else timezone.now().isoformat(),
                'read': True
            })
        
        if not alerts_data:
            alerts_data.append({
                'id': 'system_welcome',
                'type': 'info',
                'title': 'Welcome!',
                'message': 'Welcome to Gyaan Buddy! Start exploring modules and missions.',
                'timestamp': timezone.now().isoformat(),
                'read': False
            })
        
        api_logger.info(f"Dashboard alerts returned {len(alerts_data)} alerts")
        return success(
            data=alerts_data,
            message="Dashboard alerts retrieved successfully"
        )
    
    def _calculate_percentage_change(self, current_value, previous_value):
        """Calculate percentage change between current and previous values."""
        if previous_value == 0:
            if current_value > 0:
                return {'change': 100, 'type': 'positive', 'trend': 'up'}
            else:
                return {'change': 0, 'type': 'neutral', 'trend': 'stable'}
        
        change_percentage = ((current_value - previous_value) / previous_value) * 100
        
        if change_percentage > 0:
            return {'change': round(change_percentage, 1), 'type': 'positive', 'trend': 'up'}
        elif change_percentage < 0:
            return {'change': round(abs(change_percentage), 1), 'type': 'negative', 'trend': 'down'}
        else:
            return {'change': 0, 'type': 'neutral', 'trend': 'stable'}


class ReportsViewSet(viewsets.ViewSet):
    """ViewSet for Reports operations."""
    permission_classes = [permissions.IsAuthenticated]
    
    @action(detail=False, methods=['get'], url_path='student-performance')
    def student_performance(self, request):
        """Get student performance reports."""
        api_logger.info(f"Student performance report requested by {request.user.username} (ID: {request.user.id})")
        
        class_name = request.query_params.get('class', None)
        student_id = request.query_params.get('studentId', None)

        students = Account.objects.filter(profile__user_type='student', is_active=True)

        if class_name:
            students = students.filter(profile__student__class_instance__name__icontains=class_name)

        if student_id:
            students = students.filter(id=student_id)
        
        performance_data = []
        for student in students:
            performance_data.append({
                'student_id': student.id,
                'student_name': f"{student.first_name} {student.last_name}",
                'class': (student.profile.student.class_instance.name if hasattr(student, 'profile') and hasattr(student.profile, 'student') and student.profile.student.class_instance else 'No Class'),
                'total_exp': student.profile.total_exp if hasattr(student, 'profile') else 0,
                'level': student.profile.get_level() if hasattr(student, 'profile') else None,
                'rewards': student.profile.rewards if hasattr(student, 'profile') else 0,
                'modules_completed': UserModuleProgress.objects.filter(account=student, status='completed').count(),
                'chapters_completed': UserChapterProgress.objects.filter(account=student, status='completed').count(),
                'last_login': student.last_login.isoformat() if student.last_login else None
            })
        
        api_logger.info(f"Student performance report returned {len(performance_data)} students")
        return success(
            data=performance_data,
            message="Student performance report retrieved successfully"
        )
    
    @action(detail=False, methods=['get'], url_path='progress-over-time')
    def progress_over_time(self, request):
        """Get progress over time data."""
        api_logger.info(f"Progress over time report requested by {request.user.username} (ID: {request.user.id})")
        
        from datetime import timedelta

        months = []
        student_progress = []
        class_average = []
        
        for i in range(6):
            month_start = timezone.now() - timedelta(days=30*(i+1))
            month_end = timezone.now() - timedelta(days=30*i)
            
            month_students = Account.objects.filter(
                profile__user_type='student',
                is_active=True,
                last_login__gte=month_start,
                last_login__lt=month_end
            )
            
            if month_students.exists():
                avg_exp = month_students.aggregate(avg=Avg('profile__student__total_exp'))['avg'] or 0
                student_progress.insert(0, round(avg_exp, 1))
                class_average.insert(0, round(avg_exp * 0.9, 1))
            else:
                student_progress.insert(0, 0)
                class_average.insert(0, 0)
            
            month_name = month_start.strftime('%b')
            months.insert(0, month_name)
        
        progress_data = {
            'labels': months,
            'datasets': [
                {
                    'label': 'Student Progress',
                    'data': student_progress,
                    'borderColor': 'rgb(75, 192, 192)',
                    'backgroundColor': 'rgba(75, 192, 192, 0.2)'
                },
                {
                    'label': 'Class Average',
                    'data': class_average,
                    'borderColor': 'rgb(255, 99, 132)',
                    'backgroundColor': 'rgba(255, 99, 132, 0.2)'
                }
            ]
        }
        
        api_logger.info(f"Progress over time data returned")
        return success(
            data=progress_data,
            message="Progress over time report retrieved successfully"
        )
    
    @action(detail=False, methods=['get'], url_path='quiz-assignment-summaries')
    def quiz_assignment_summaries(self, request):
        """Get quiz and assignment summaries."""
        api_logger.info(f"Quiz assignment summaries requested by {request.user.username} (ID: {request.user.id})")
        
        missions = Mission.objects.filter(is_deleted=False)
        competitions = Competition.objects.filter(is_active=True, is_deleted=False)
        
        summaries = []
        
        for mission in missions:
            mission_progress = UserMissionProgress.objects.filter(mission=mission)
            participants = mission_progress.count()
            completed = mission_progress.filter(status='completed').count()
            
            if completed > 0:
                avg_exp = mission_progress.filter(status='completed').aggregate(
                    avg=Avg('exp_earned')
                )['avg'] or 0
                average_score = round(avg_exp, 1)
            else:
                average_score = 0
            
            summaries.append({
                'id': mission.id,
                'title': mission.title,
                'type': 'mission',
                'date': mission.mission_date.isoformat(),
                'participants': participants,
                'completed': completed,
                'average_score': average_score,
                'status': 'active' if not mission.is_deleted else 'inactive'
            })
        
        for competition in competitions:
            comp_progress = UserCompetitionProgress.objects.filter(competition=competition)
            participants = comp_progress.count()
            completed = comp_progress.filter(status='completed').count()
            
            if completed > 0:
                avg_score = comp_progress.filter(status='completed').aggregate(
                    avg=Avg('score')
                )['avg'] or 0
                average_score = round(avg_score, 1)
            else:
                average_score = 0
            
            summaries.append({
                'id': competition.id,
                'title': competition.title,
                'type': 'competition',
                'date': competition.created_at.date().isoformat(),
                'participants': participants,
                'completed': completed,
                'average_score': average_score,
                'status': 'active' if competition.is_active else 'inactive'
            })
        
        api_logger.info(f"Quiz assignment summaries returned {len(summaries)} items")
        return success(
            data=summaries,
            message="Quiz assignment summaries retrieved successfully"
        )
    
    @action(detail=False, methods=['get'], url_path='ai-insights')
    def ai_insights(self, request):
        """Get AI insights reports."""
        api_logger.info(f"AI insights report requested by {request.user.username} (ID: {request.user.id})")
        
        insights_data = []
        
        subject_performance = {}
        for subject in Subject.objects.filter(is_active=True):
            subject_students = Account.objects.filter(
                profile__user_type='student',
                is_active=True,
                profile__student__class_instance__school=request.user.profile.school if hasattr(request.user, 'profile') else None
            )
            if subject_students.exists():
                avg_exp = subject_students.aggregate(avg=Avg('profile__student__total_exp'))['avg'] or 0
                subject_performance[subject.name] = avg_exp
        
        if subject_performance:
            best_subject = max(subject_performance, key=subject_performance.get)
            worst_subject = min(subject_performance, key=subject_performance.get)
            
            insights_data.append({
                'id': 1,
                'title': 'Subject Performance Analysis',
                'description': f'Students show highest performance in {best_subject} and lowest in {worst_subject}.',
                'confidence': 0.85,
                'recommendations': [
                    f'Increase {worst_subject} content difficulty gradually',
                    f'Apply {best_subject} teaching methods to other subjects'
                ],
                'impact': 'high',
                'created_at': timezone.now().isoformat()
            })
        
        total_students = Account.objects.filter(profile__user_type='student', is_active=True).count()
        active_students = Account.objects.filter(
            profile__user_type='student',
            is_active=True,
            last_login__gte=timezone.now() - timezone.timedelta(days=7)
        ).count()
        
        engagement_rate = (active_students / max(total_students, 1)) * 100
        
        if engagement_rate < 70:
            insights_data.append({
                'id': 2,
                'title': 'Student Engagement Analysis',
                'description': f'Student engagement is at {engagement_rate:.1f}%. Consider implementing more interactive content.',
                'confidence': 0.90,
                'recommendations': [
                    'Add more interactive modules',
                    'Implement gamification elements',
                    'Send engagement reminders'
                ],
                'impact': 'high',
                'created_at': timezone.now().isoformat()
            })
        
        total_modules = Module.objects.filter(is_active=True).count()
        completed_modules = UserModuleProgress.objects.filter(status='completed').count()
        completion_rate = (completed_modules / max(total_modules, 1)) * 100
        
        if completion_rate < 50:
            insights_data.append({
                'id': 3,
                'title': 'Module Completion Analysis',
                'description': f'Module completion rate is {completion_rate:.1f}%. Students may need additional support.',
                'confidence': 0.88,
                'recommendations': [
                    'Break down complex modules into smaller chunks',
                    'Add progress tracking and reminders',
                    'Provide additional learning resources'
                ],
                'impact': 'medium',
                'created_at': timezone.now().isoformat()
            })
        
        if not insights_data:
            insights_data.append({
                'id': 1,
                'title': 'System Analysis',
                'description': 'System is running smoothly. Continue monitoring student progress.',
                'confidence': 0.75,
                'recommendations': [
                    'Regular progress monitoring',
                    'Student feedback collection'
                ],
                'impact': 'low',
                'created_at': timezone.now().isoformat()
            })
        
        api_logger.info(f"AI insights report returned {len(insights_data)} insights")
        return success(
            data=insights_data,
            message="AI insights report retrieved successfully"
        )
    
    @action(detail=False, methods=['get'], url_path='analytics')
    def analytics(self, request):
        """Get analytics data."""
        api_logger.info(f"Analytics report requested by {request.user.username} (ID: {request.user.id})")
        
        total_students = Account.objects.filter(profile__user_type='student', is_active=True).count()
        total_teachers = Account.objects.filter(profile__user_type='teacher', is_active=True).count()
        total_questions = Question.objects.filter(is_active=True).count()
        total_modules = Module.objects.filter(is_active=True).count()
        total_subjects = Subject.objects.filter(is_active=True).count()
        
        analytics_data = {
            'overview': {
                'total_students': total_students,
                'total_teachers': total_teachers,
                'total_questions': total_questions,
                'total_modules': total_modules,
                'total_subjects': total_subjects,
                'active_classes': Class.objects.filter(is_active=True).count(),
                'completed_missions': UserMissionProgress.objects.filter(status='completed').count()
            },
            'performance_metrics': {
                'average_student_score': Account.objects.filter(profile__user_type='student').aggregate(avg=Avg('profile__student__total_exp'))['avg'] or 0,
                'completion_rate': (UserModuleProgress.objects.filter(status='completed').count() / max(UserModuleProgress.objects.count(), 1)) * 100,
                'engagement_rate': self._calculate_engagement_rate(),
                'satisfaction_score': self._calculate_satisfaction_score()
            },
            'trends': {
                'student_growth': '+12%',
                'performance_improvement': '+8%',
                'engagement_increase': '+5%',
                'content_creation': '+15%'
            }
        }
        
        api_logger.info(f"Analytics report returned")
        return success(
            data=analytics_data,
            message="Analytics report retrieved successfully"
        )
    
    @action(detail=False, methods=['get'], url_path='reports-analytics')
    def reports_analytics(self, request):
        """
        Get full reports & analytics payload for the Reports Analytics screen.
        Returns summary, filter options, module/chapter proficiency, analytics stats,
        and optional student proficiency list when filters are applied.
        Query params: period (7|30|90|180|365), class (class name), subject (subject id),
        module (module id), chapter (chapter id).
        """
        from datetime import timedelta

        def _time_ago(dt):
            if not dt:
                return ''
            delta = timezone.now() - dt
            if delta.days > 0:
                return f'{delta.days} days ago'
            hours = int(delta.seconds // 3600)
            if hours > 0:
                return f'{hours} hours ago'
            mins = int((delta.seconds % 3600) // 60)
            return f'{mins} minutes ago' if mins > 0 else 'Just now'

        api_logger.info(f"Reports analytics requested by {request.user.username}")

        user_school = None
        if hasattr(request.user, 'profile') and request.user.profile:
            profile = request.user.profile
            user_school = getattr(profile, 'school', None)
            if not user_school and hasattr(profile, 'student') and profile.student and getattr(profile.student, 'class_instance', None):
                user_school = profile.student.class_instance.school
            if not user_school and hasattr(profile, 'teacher_profile') and profile.teacher_profile:
                first_assignment = Teacher.objects.filter(teacher=profile.teacher_profile).select_related('class_instance__school').first()
                if first_assignment and first_assignment.class_instance:
                    user_school = first_assignment.class_instance.school
        if not user_school:
            empty_response = {
                'summary': {'totalStudents': 0, 'completionRate': 0, 'averageScore': 0, 'weakTopicCount': 0},
                'filterOptions': {'classes': [], 'subjects': [], 'modules': [], 'chapters': []},
                'moduleProficiencyData': [],
                'reportsData': {'summary': {'totalStudents': 0, 'completionRate': 0, 'averageScore': 0, 'weakTopicCount': 0}, 'chapterProficiency': []},
                'analyticsData': {'overallStats': {'totalStudents': 0, 'averageScore': 0, 'completionRate': 0, 'activeStudents': 0}, 'subjectPerformance': [], 'classPerformance': [], 'recentActivity': []},
                'sectionWisePerformance': [],
                'studentProficiencyData': [],
            }
            return success(data=empty_response, message="Reports analytics retrieved successfully")

        period = request.query_params.get('period', '30')
        try:
            period_days = int(period)
        except ValueError:
            period_days = 30
        since = timezone.now() - timedelta(days=period_days)

        class_filter = request.query_params.get('class', '').strip()
        subject_id = request.query_params.get('subject', '').strip()
        module_id = request.query_params.get('module', '').strip()
        topic_id = request.query_params.get('topic', '').strip()
        chapter_id = request.query_params.get('chapter', '').strip() or topic_id
        excluded_topic_labels = {
            # 'previous knowledge testing',
            # 'competency based questions',
            # 'competancy based questions',
            # 'competency-based questions',
            # 'summary',
        }

        def _clean_topic_name(name):
            return re.sub(r'\s+', ' ', str(name or '').strip())

        def _is_displayable_topic_name(name):
            cleaned_name = _clean_topic_name(name)
            return bool(cleaned_name) and cleaned_name.lower() not in excluded_topic_labels

        def _build_classes_with_subjects(classes_q, restricted_subj_ids=None, teacher_cls_subj_map=None):
            cls_list = list(classes_q.order_by('name'))
            if not cls_list:
                return []
            if teacher_cls_subj_map is not None:
                # For teachers: use exact per-class assignment rows so 10-A shows only
                # the subjects Naresh teaches there, not everything he teaches anywhere.
                all_taught_ids = {sid for sids in teacher_cls_subj_map.values() for sid in sids}
                subj_lookup = {
                    s.id: {'id': str(s.id), 'name': s.name}
                    for s in Subject.objects.filter(is_active=True, school=user_school, id__in=all_taught_ids)
                }
                return [
                    {
                        'id': str(c.id),
                        'name': c.name,
                        'subjects': [subj_lookup[sid] for sid in teacher_cls_subj_map.get(c.id, set()) if sid in subj_lookup],
                    }
                    for c in cls_list
                ]
            # Non-teacher (principal / admin): derive via Module.class_instance FK,
            # falling back to Subject.classes M2M for classes with no active modules.
            base_subj_q = Subject.objects.filter(is_active=True, school=user_school)
            if restricted_subj_ids is not None:
                base_subj_q = base_subj_q.filter(id__in=restricted_subj_ids)
            subj_ids = list(base_subj_q.values_list('id', flat=True))
            subj_lookup = {s.id: {'id': str(s.id), 'name': s.name} for s in base_subj_q}
            cls_ids = [c.id for c in cls_list]
            cls_subj = {}
            for cid, sid in Module.objects.filter(
                is_active=True, class_instance_id__in=cls_ids, subject_id__in=subj_ids
            ).values_list('class_instance_id', 'subject_id').distinct():
                cls_subj.setdefault(cid, set()).add(sid)
            no_mod_cls_ids = [cid for cid in cls_ids if cid not in cls_subj]
            if no_mod_cls_ids and subj_ids:
                for sid, cid in Subject.objects.filter(
                    id__in=subj_ids, classes__id__in=no_mod_cls_ids
                ).values_list('id', 'classes__id').distinct():
                    cls_subj.setdefault(cid, set()).add(sid)
            return [
                {
                    'id': str(c.id),
                    'name': c.name,
                    'subjects': [subj_lookup[sid] for sid in cls_subj.get(c.id, []) if sid in subj_lookup],
                }
                for c in cls_list
            ]

        # Teacher isolation: derive assigned class/subject IDs once, apply everywhere.
        teacher_class_ids = None
        teacher_subject_ids = None
        teacher_cls_subj_map = None  # {class_id: set(subject_ids)} — exact per-class assignments
        profile = getattr(request.user, 'profile', None)
        if (profile and profile.user_type == 'teacher'
                and hasattr(profile, 'teacher_profile')):
            _assignments = list(Teacher.objects.filter(
                teacher=profile.teacher_profile,
                is_deleted=False,
            ).values('class_instance_id', 'subject_id'))
            teacher_class_ids = list({a['class_instance_id'] for a in _assignments})
            teacher_subject_ids = list({a['subject_id'] for a in _assignments})
            teacher_cls_subj_map = {}
            for a in _assignments:
                teacher_cls_subj_map.setdefault(a['class_instance_id'], set()).add(a['subject_id'])

        students_qs = Account.objects.filter(
            profile__user_type='student',
            is_active=True,
            is_deleted=False
        )
        if user_school:
            students_qs = students_qs.filter(profile__student__class_instance__school=user_school)
        if teacher_class_ids is not None:
            students_qs = students_qs.filter(profile__student__class_instance_id__in=teacher_class_ids)
        if class_filter:
            students_qs = students_qs.filter(profile__student__class_instance__name__icontains=class_filter)
        student_ids = list(students_qs.values_list('id', flat=True))

        total_students = len(student_ids)
        if total_students == 0:
            _classes_qs = Class.objects.filter(is_active=True)
            if user_school:
                _classes_qs = _classes_qs.filter(school=user_school)
            if teacher_class_ids is not None:
                _classes_qs = _classes_qs.filter(id__in=teacher_class_ids)
            _matched = list(_classes_qs.filter(name__icontains=class_filter).distinct()) if class_filter else []
            if class_filter and not _matched:
                _matched = list(_classes_qs.filter(name=class_filter))
            _subjects_qs = Subject.objects.filter(is_active=True, school=user_school)
            if teacher_subject_ids is not None:
                _subjects_qs = _subjects_qs.filter(id__in=teacher_subject_ids)
            if _matched:
                _matched_ids = [c.id for c in _matched]
                _via_mods = _subjects_qs.filter(modules__class_instance_id__in=_matched_ids, modules__is_active=True).distinct()
                if _via_mods.exists():
                    _subjects_qs = _via_mods
                else:
                    _via_m2m = _subjects_qs.filter(classes__in=_matched).distinct()
                    if _via_m2m.exists():
                        _subjects_qs = _via_m2m
            _subjects_empty = [{'id': str(s.id), 'name': s.name} for s in _subjects_qs]
            empty_response = {
                'summary': {'totalStudents': 0, 'completionRate': 0, 'averageScore': 0, 'weakTopicCount': 0},
                'filterOptions': {
                    'classes': _build_classes_with_subjects(_classes_qs, teacher_subject_ids, teacher_cls_subj_map),
                    'subjects': _subjects_empty,
                    'modules': [],
                    'chapters': [],
                },
                'moduleProficiencyData': [],
                'reportsData': {
                    'summary': {'totalStudents': 0, 'completionRate': 0, 'averageScore': 0, 'weakTopicCount': 0},
                    'chapterProficiency': [],
                },
                'analyticsData': {
                    'overallStats': {'totalStudents': 0, 'averageScore': 0, 'completionRate': 0, 'activeStudents': 0},
                    'subjectPerformance': [],
                    'classPerformance': [],
                    'recentActivity': [],
                },
                'sectionWisePerformance': [],
                'studentProficiencyData': [],
            }
            return success(data=empty_response, message="Reports analytics retrieved successfully")

        # Find subjects for this school using direct school FK.
        _direct_school_subjects_ids = list(
            Subject.objects.filter(is_active=True, school=user_school).values_list('id', flat=True)
        )
        school_subject_ids_qs = Subject.objects.filter(is_active=True, school=user_school)
        if teacher_subject_ids is not None:
            school_subject_ids_qs = school_subject_ids_qs.filter(id__in=teacher_subject_ids)
        school_subject_ids = list(school_subject_ids_qs.values_list('id', flat=True))
        total_modules = Module.objects.filter(is_active=True, subject_id__in=school_subject_ids).count() if school_subject_ids else 0
        completed_modules_count = UserModuleProgress.objects.filter(
            account_id__in=student_ids,
            status='completed'
        ).count()
        # Denominator must be total_modules × total_students (possible completions),
        # not just total_modules, because completed_modules_count spans all students.
        completion_rate = min(100, round((completed_modules_count / max(total_modules * max(total_students, 1), 1)) * 100))

        avg_exp = students_qs.aggregate(avg=Avg('profile__student__total_exp'))['avg'] or 0
        average_score = min(100, round(avg_exp / 10)) if avg_exp else 0

        active_students = Account.objects.filter(
            profile__user_type='student',
            is_active=True,
            is_deleted=False,
            last_login__gte=since
        )
        if user_school:
            active_students = active_students.filter(profile__student__class_instance__school=user_school)
        if teacher_class_ids is not None:
            active_students = active_students.filter(profile__student__class_instance_id__in=teacher_class_ids)
        active_students_count = active_students.count()

        classes_qs = Class.objects.filter(is_active=True)
        if user_school:
            classes_qs = classes_qs.filter(school=user_school)
        if teacher_class_ids is not None:
            classes_qs = classes_qs.filter(id__in=teacher_class_ids)
        class_names = list(classes_qs.values_list('name', flat=True))

        # Resolve selected class(es) for subject filtering (same school / class subjects only)
        matched_classes = []
        if class_filter:
            matched_classes = list(classes_qs.filter(name__icontains=class_filter).distinct())
            if not matched_classes and class_filter:
                matched_classes = list(classes_qs.filter(name=class_filter))

        # Build subjects queryset scoped to the current school.
        subjects_qs = Subject.objects.filter(is_active=True, school=user_school)
        if teacher_subject_ids is not None:
            subjects_qs = subjects_qs.filter(id__in=teacher_subject_ids)
        if matched_classes:
            _matched_class_ids = [c.id for c in matched_classes]
            # Use Module.class_instance (explicit FK) as the authoritative source for class-subject mapping.
            # The Subject.classes M2M is often incomplete/unpopulated, leading to wrong results.
            _via_modules = subjects_qs.filter(modules__class_instance_id__in=_matched_class_ids, modules__is_active=True).distinct()
            if _via_modules.exists():
                subjects_qs = _via_modules
            else:
                # Fall back to M2M if no modules are assigned to these classes
                _via_class_filter = subjects_qs.filter(classes__in=matched_classes).distinct()
                if _via_class_filter.exists():
                    subjects_qs = _via_class_filter

        # Cascading filters: modules by subject (and class), topics (chapters) by module or subject (and class).
        matched_class_ids = [c.id for c in matched_classes] if matched_classes else None
        module_options = []
        if subject_id:
            modules_for_filter = Module.objects.filter(subject_id=subject_id, is_active=True).order_by('order')
            if matched_class_ids:
                modules_for_filter = modules_for_filter.filter(class_instance_id__in=matched_class_ids)
            module_options = [{'id': str(m.id), 'name': m.name} for m in modules_for_filter[:50]]
        chapter_options = []
        if module_id:
            chapters_in_module = ModuleChapter.objects.filter(module_id=module_id, is_deleted=False).order_by('order')
            if matched_class_ids:
                chapters_in_module = chapters_in_module.filter(module__class_instance_id__in=matched_class_ids)
            chapter_options = [
                {'id': str(c.id), 'name': _clean_topic_name(c.title)}
                for c in chapters_in_module[:50]
                if _is_displayable_topic_name(c.title)
            ]
        elif subject_id:
            chapters_for_subject = ModuleChapter.objects.filter(
                module__subject_id=subject_id,
                module__is_active=True,
                is_deleted=False
            ).order_by('module__order', 'order')
            if matched_class_ids:
                chapters_for_subject = chapters_for_subject.filter(module__class_instance_id__in=matched_class_ids)
            chapter_options = [
                {'id': str(c.id), 'name': _clean_topic_name(c.title)}
                for c in chapters_for_subject[:100]
                if _is_displayable_topic_name(c.title)
            ]
        subject_options = [{'id': str(s.id), 'name': s.name} for s in subjects_qs]

        filter_options = {
            'classes': _build_classes_with_subjects(classes_qs, teacher_subject_ids, teacher_cls_subj_map),
            'subjects': subject_options,
            'modules': module_options,
            'chapters': chapter_options,
        }

        scoped_question_content = ModuleContent.objects.filter(
            content_type='question',
            is_deleted=False,
            chapter__is_deleted=False,
            chapter__module__is_active=True,
        )
        if subject_id:
            scoped_question_content = scoped_question_content.filter(chapter__module__subject_id=subject_id)
        if module_id:
            scoped_question_content = scoped_question_content.filter(chapter__module_id=module_id)
        if chapter_id:
            scoped_question_content = scoped_question_content.filter(chapter_id=chapter_id)
        scoped_question_ids = set(scoped_question_content.values_list('question_id', flat=True))

        subject_performance = []
        for subj in subjects_qs:
            subj_students = students_qs.filter(
                profile__student__subject_enrollments__subject=subj,
                profile__student__subject_enrollments__is_active=True
            ).distinct()
            cnt = subj_students.count()
            if cnt == 0:
                # Fall back: students in classes that have this subject via M2M
                subj_students = students_qs.filter(
                    profile__student__class_instance__subjects=subj
                ).distinct()
                cnt = subj_students.count()
            if cnt == 0:
                continue
            subj_avg = subj_students.aggregate(avg=Avg('profile__student__total_exp'))['avg'] or 0
            subj_score = min(100, round((subj_avg / 100.0) * 10) if subj_avg else 0)
            subj_completed = UserModuleProgress.objects.filter(account_id__in=subj_students.values_list('id', flat=True), module__subject=subj, status='completed').count()
            subj_total_mods = Module.objects.filter(subject=subj, is_active=True).count()
            subj_compl_rate = round((subj_completed / max(subj_total_mods * max(cnt, 1), 1)) * 100) if subj_total_mods else 0
            subject_performance.append({
                'subject': subj.name,
                'averageScore': subj_score,
                'completionRate': min(100, subj_compl_rate),
                'students': cnt,
            })

        ordered_classes = list(classes_qs.order_by('name'))
        class_ids = [cls.id for cls in ordered_classes]

        def _teacher_display_name(teacher_profile):
            if not teacher_profile or teacher_profile.is_deleted:
                return ''
            user_profile = getattr(teacher_profile, 'user_profile', None)
            account = getattr(user_profile, 'account', None)
            if not account:
                return ''
            return account.get_full_name().strip() or account.username

        class_teacher_names = {
            cls.id: _teacher_display_name(cls.class_teacher)
            for cls in ordered_classes
        }

        subject_teacher_names = {}
        if class_ids:
            subject_teacher_assignments = Teacher.objects.filter(
                is_deleted=False,
                class_instance_id__in=class_ids,
            ).select_related('teacher__user_profile__account')
            if subject_id:
                subject_teacher_assignments = subject_teacher_assignments.filter(subject_id=subject_id)

            for assignment in subject_teacher_assignments:
                teacher_name = _teacher_display_name(assignment.teacher)
                if teacher_name and assignment.class_instance_id not in subject_teacher_names:
                    subject_teacher_names[assignment.class_instance_id] = teacher_name

        any_teacher_names = {}
        if class_ids:
            any_teacher_assignments = Teacher.objects.filter(
                is_deleted=False,
                class_instance_id__in=class_ids,
            ).select_related('teacher__user_profile__account')

            for assignment in any_teacher_assignments:
                teacher_name = _teacher_display_name(assignment.teacher)
                if teacher_name and assignment.class_instance_id not in any_teacher_names:
                    any_teacher_names[assignment.class_instance_id] = teacher_name

        section_wise_performance = []
        for cls in ordered_classes:
            cls_students_qs = students_qs.filter(profile__student__class_instance=cls)
            if subject_id:
                cls_students_qs = cls_students_qs.filter(
                    profile__student__subject_enrollments__subject_id=subject_id,
                    profile__student__subject_enrollments__is_active=True,
                ).distinct()
            cls_student_ids = list(cls_students_qs.values_list('id', flat=True))
            students_count = len(cls_student_ids)
            if students_count == 0:
                continue

            cls_answers = Answer.objects.filter(
                user_id__in=cls_student_ids,
                created_at__gte=since,
            )
            if scoped_question_ids:
                cls_answers = cls_answers.filter(question_id__in=scoped_question_ids)
            elif subject_id or module_id or chapter_id:
                cls_answers = cls_answers.none()

            total_answers = cls_answers.count()
            correct_answers = cls_answers.filter(is_correct=True).count()
            attempted_students = cls_answers.values('user_id').distinct().count()
            proficiency = round((correct_answers / total_answers) * 100) if total_answers else 0
            attempt_rate = round((attempted_students / students_count) * 100) if students_count else 0
            difficulty_label = '—'
            difficulty_counts = (
                cls_answers
                .exclude(question__level__isnull=True)
                .values('question__level')
                .annotate(total=Count('id'))
                .order_by('-total', 'question__level')
            )
            if difficulty_counts:
                top_level = difficulty_counts[0].get('question__level')
                if top_level is not None:
                    difficulty_label = str(top_level)

            chapter_stats = cls_answers.filter(chapter_id__isnull=False).values(
                'chapter_id',
                'chapter__title'
            ).annotate(
                total=Count('id'),
                correct=Count('id', filter=Q(is_correct=True)),
            )
            chapter_scores = []
            for chapter_stat in chapter_stats:
                total = chapter_stat['total'] or 0
                chapter_prof = round((chapter_stat['correct'] / total) * 100) if total else 0
                topic_name = _clean_topic_name(chapter_stat['chapter__title'] or 'Untitled Topic')
                if not _is_displayable_topic_name(topic_name):
                    continue
                chapter_scores.append({
                    'name': topic_name,
                    'proficiency': chapter_prof,
                })

            ranked_topics = sorted(chapter_scores, key=lambda item: item['proficiency'], reverse=True)
            good_topics = [topic['name'] for topic in ranked_topics[:3]]
            struggling_topics = [topic['name'] for topic in sorted(chapter_scores, key=lambda item: item['proficiency'])[:3]]
            teacher_name = (
                subject_teacher_names.get(cls.id)
                or class_teacher_names.get(cls.id)
                or any_teacher_names.get(cls.id)
                or '—'
            )

            section_wise_performance.append({
                'className': cls.name,
                'teacherName': teacher_name,
                'students': students_count,
                'averageScore': proficiency,
                'attemptRate': attempt_rate,
                'completionRate': attempt_rate,
                'difficulty': difficulty_label,
                'goodTopics': good_topics,
                'strugglingTopics': struggling_topics,
            })

        class_performance = [
            {
                'class': row['className'],
                'averageScore': row['averageScore'],
                'completionRate': row['completionRate'],
                'students': row['students'],
            }
            for row in section_wise_performance
        ]

        recent_activity = []
        recent_completions = UserModuleProgress.objects.filter(
            account_id__in=student_ids,
            status='completed',
            completed_at__isnull=False
        ).order_by('-completed_at').select_related('account', 'module__subject')[:10]
        for r in recent_completions:
            student_name = f"{r.account.first_name} {r.account.last_name}".strip() or r.account.username
            recent_activity.append({
                'action': 'Module Completed',
                'student': student_name,
                'subject': r.module.subject.name if r.module else '—',
                'time': _time_ago(r.completed_at),
                'score': r.percentage or 0,
            })

        modules_for_proficiency = Module.objects.filter(is_active=True, subject__in=subjects_qs).order_by('subject', 'order')
        if matched_class_ids:
            modules_for_proficiency = modules_for_proficiency.filter(class_instance_id__in=matched_class_ids)
        if subject_id:
            modules_for_proficiency = modules_for_proficiency.filter(subject_id=subject_id)
        if module_id:
            modules_for_proficiency = modules_for_proficiency.filter(id=module_id)
        elif chapter_id:
            chapter_module_id = ModuleChapter.objects.filter(id=chapter_id).values_list('module_id', flat=True).first()
            if chapter_module_id:
                modules_for_proficiency = modules_for_proficiency.filter(id=chapter_module_id)

        student_name_map = {
            s.id: f"{s.first_name} {s.last_name}".strip() or s.username
            for s in students_qs.only('id', 'first_name', 'last_name', 'username')[:200]
        }

        module_proficiency_data = []
        chapter_proficiency_list = []
        weak_topic_count = 0  # Module chapters where all students' avg answer % < 50%
        for idx, mod in enumerate(modules_for_proficiency.select_related('subject').prefetch_related('chapters'), start=1):
            chapters_data = []
            weak_subtopics = []
            weak_levels_set = set()
            chapters_qs = mod.chapters.filter(is_deleted=False).order_by('order')
            if chapter_id:
                chapters_qs = chapters_qs.filter(id=chapter_id)
            for ch in chapters_qs:
                topic_name = _clean_topic_name(ch.title)
                if not _is_displayable_topic_name(topic_name):
                    continue
                question_ids = set(ModuleContent.objects.filter(
                    chapter=ch,
                    content_type='question',
                    is_deleted=False
                ).values_list('question_id', flat=True))
                if not question_ids:
                    prof = 0
                    attempt_rate = 0
                else:
                    ans = Answer.objects.filter(question_id__in=question_ids, user_id__in=student_ids, created_at__gte=since)
                    total_a = ans.count()
                    correct_a = ans.filter(is_correct=True).count()
                    prof = round((correct_a / total_a) * 100) if total_a else 0
                    attempted_students = ans.values('user_id').distinct().count()
                    attempt_rate = round((attempted_students / len(student_ids)) * 100) if student_ids else 0
                ch_weak = []
                # Weak topic: module chapter where all students' average answer % for its questions is < 50%
                if prof < 50:
                    ch_weak = [1, 2, 3, 4, 5]
                    weak_subtopics.append(topic_name)
                    weak_topic_count += 1
                elif prof < 70:
                    ch_weak = [3, 4, 5]
                elif prof < 80:
                    ch_weak = [4, 5]
                elif prof < 95:
                    ch_weak = [5]
                else:
                    ch_weak = []
                weak_levels_set.update(ch_weak)
                chapter_students = []
                if question_ids and student_name_map:
                    stu_ans = Answer.objects.filter(
                        question_id__in=question_ids,
                        user_id__in=list(student_name_map.keys()),
                        created_at__gte=since,
                    ).values('user_id').annotate(
                        total=Count('id'),
                        correct=Count('id', filter=Q(is_correct=True)),
                    )
                    stu_ans_map = {r['user_id']: r for r in stu_ans}
                    for sid, sname in student_name_map.items():
                        r = stu_ans_map.get(sid, {})
                        t, c = r.get('total', 0), r.get('correct', 0)
                        chapter_students.append({
                            'id': sid,
                            'name': sname,
                            'proficient': min(100, round((c / t) * 100)) if t else 0,
                        })
                    chapter_students.sort(key=lambda x: (-x['proficient'], x['name'].lower()))

                chapters_data.append({
                    'id': ch.id,
                    'name': topic_name,
                    'proficiency': prof,
                    'weakLevels': ch_weak,
                    'attemptRate': attempt_rate,
                    'students': chapter_students,
                })
                sat = max(0, 75 - prof) if prof < 75 else 0
                need = max(0, 100 - prof - sat)
                chapter_proficiency_list.append({
                    'id': str(ch.id),
                    'chapterName': topic_name,
                    'proficient': min(100, prof),
                    'satisfactory': sat,
                    'needsImprovement': need,
                })

            mod_weak_levels = sorted(weak_levels_set)
            mod_attempt_rate = round(sum(c['attemptRate'] for c in chapters_data) / len(chapters_data)) if chapters_data else 0
            if chapters_data:
                module_proficiency_data.append({
                    'id': idx,
                    'moduleName': mod.name,
                    'weakSubtopics': weak_subtopics[:5],
                    'weakLevels': mod_weak_levels,
                    'attemptRate': mod_attempt_rate,
                    'chapters': chapters_data,
                })

        student_proficiency_data = []
        if class_filter or subject_id or module_id or chapter_id:
            if subject_id or module_id or chapter_id:
                # Scope: questions in selected chapter, or selected module's chapters, or subject's modules
                if chapter_id:
                    question_ids_scope = set(ModuleContent.objects.filter(
                        chapter_id=chapter_id,
                        content_type='question',
                        is_deleted=False
                    ).values_list('question_id', flat=True))
                elif module_id:
                    question_ids_scope = set(ModuleContent.objects.filter(
                        chapter__module_id=module_id,
                        chapter__is_deleted=False,
                        content_type='question',
                        is_deleted=False
                    ).values_list('question_id', flat=True))
                else:
                    question_ids_scope = set(ModuleContent.objects.filter(
                        chapter__module__subject_id=subject_id,
                        chapter__module__is_active=True,
                        chapter__is_deleted=False,
                        content_type='question',
                        is_deleted=False
                    ).values_list('question_id', flat=True))
                for student_id in student_ids[:100]:
                    student = students_qs.filter(id=student_id).select_related('profile__student').first()
                    if not student:
                        continue
                    name = f"{student.first_name} {student.last_name}".strip() or student.username
                    if not question_ids_scope:
                        proficient_pct = 0
                    else:
                        ans = Answer.objects.filter(question_id__in=question_ids_scope, user_id=student_id, created_at__gte=since)
                        total_a = ans.count()
                        correct_a = ans.filter(is_correct=True).count()
                        proficient_pct = round((correct_a / total_a) * 100) if total_a else 0
                        proficient_pct = min(100, proficient_pct)
                    student_proficiency_data.append({
                        'id': student.id,
                        'name': name,
                        'proficient': proficient_pct,
                    })
                student_proficiency_data.sort(key=lambda x: (-x['proficient'], x['name'].split()[0].lower() if x['name'] else ''))
            else:
                # Only class filter: use overall total_exp-based proficiency
                for student in students_qs.select_related('profile__student')[:100]:
                    name = f"{student.first_name} {student.last_name}".strip() or student.username
                    total_exp = getattr(student.profile.student, 'total_exp', 0) if hasattr(student, 'profile') and getattr(student.profile, 'student', None) else 0
                    proficient_pct = min(100, round((total_exp / 100.0) * 10) if total_exp else 0)
                    if proficient_pct == 0 and total_exp > 0:
                        proficient_pct = min(100, round(total_exp / 10))
                    student_proficiency_data.append({
                        'id': student.id,
                        'name': name,
                        'proficient': proficient_pct,
                    })
                student_proficiency_data.sort(key=lambda x: (-x['proficient'], x['name'].split()[0].lower() if x['name'] else ''))

        def _overall_stats():
            return {
                'totalStudents': total_students,
                'averageScore': average_score,
                'completionRate': completion_rate,
                'activeStudents': active_students_count,
            }

        response_data = {
            'summary': {
                'totalStudents': total_students,
                'completionRate': completion_rate,
                'averageScore': average_score,
                'weakTopicCount': weak_topic_count,
            },
            'filterOptions': filter_options,
            'moduleProficiencyData': module_proficiency_data,
            'reportsData': {
                'summary': {
                    'totalStudents': total_students,
                    'completionRate': completion_rate,
                    'averageScore': average_score,
                    'weakTopicCount': weak_topic_count,
                },
                'chapterProficiency': chapter_proficiency_list[:20],
            },
            'analyticsData': {
                'overallStats': _overall_stats(),
                'subjectPerformance': subject_performance,
                'classPerformance': class_performance,
                'recentActivity': recent_activity,
            },
            'sectionWisePerformance': section_wise_performance,
            'studentProficiencyData': student_proficiency_data,
        }
        return success(data=response_data, message="Reports analytics retrieved successfully")
    
    @action(detail=False, methods=['get'], url_path='chapter-student-proficiency')
    def chapter_student_proficiency(self, request):
        """
        Return per-student proficiency for a specific module chapter.
        Query params: chapter_id (required), period (days, default 30), class (class name filter).
        """
        from datetime import timedelta

        chapter_id = request.query_params.get('chapter_id', '').strip()
        if not chapter_id:
            return error(message="chapter_id is required", status_code=400)

        try:
            period_days = int(request.query_params.get('period', 30))
        except ValueError:
            period_days = 30
        since = timezone.now() - timedelta(days=period_days)
        class_filter = request.query_params.get('class', '').strip()

        # Resolve school from the requesting user
        user_school = None
        profile = getattr(request.user, 'profile', None)
        if profile:
            user_school = getattr(profile, 'school', None)
            if not user_school and getattr(profile, 'student', None) and getattr(profile.student, 'class_instance', None):
                user_school = profile.student.class_instance.school
            if not user_school and hasattr(profile, 'teacher_profile') and profile.teacher_profile:
                first_assignment = Teacher.objects.filter(teacher=profile.teacher_profile).select_related('class_instance__school').first()
                if first_assignment and first_assignment.class_instance:
                    user_school = first_assignment.class_instance.school

        if not user_school:
            return success(data=[], message="No school found for user")

        # Teacher isolation
        teacher_class_ids = None
        if profile and profile.user_type == 'teacher' and hasattr(profile, 'teacher_profile'):
            _assignments = Teacher.objects.filter(teacher=profile.teacher_profile, is_deleted=False)
            teacher_class_ids = list(_assignments.values_list('class_instance_id', flat=True).distinct())

        students_qs = Account.objects.filter(
            profile__user_type='student',
            is_active=True,
            is_deleted=False,
            profile__student__class_instance__school=user_school,
        )
        if teacher_class_ids is not None:
            students_qs = students_qs.filter(profile__student__class_instance_id__in=teacher_class_ids)
        if class_filter:
            students_qs = students_qs.filter(profile__student__class_instance__name__icontains=class_filter)

        student_ids = list(students_qs.values_list('id', flat=True))
        if not student_ids:
            return success(data=[], message="No students found")

        question_ids = set(ModuleContent.objects.filter(
            chapter_id=chapter_id,
            content_type='question',
            is_deleted=False,
        ).values_list('question_id', flat=True))

        result = []
        for student in students_qs.select_related('profile__student')[:200]:
            name = f"{student.first_name} {student.last_name}".strip() or student.username
            if not question_ids:
                proficient_pct = 0
            else:
                ans = Answer.objects.filter(question_id__in=question_ids, user_id=student.id, created_at__gte=since)
                total_a = ans.count()
                correct_a = ans.filter(is_correct=True).count()
                proficient_pct = min(100, round((correct_a / total_a) * 100)) if total_a else 0
            result.append({'id': student.id, 'name': name, 'proficient': proficient_pct})

        result.sort(key=lambda x: (-x['proficient'], (x['name'].split()[0].lower() if x['name'] else '')))
        return success(data=result, message="Chapter student proficiency retrieved successfully")

    @action(detail=False, methods=['get'], url_path='chapter-student-details')
    def chapter_student_details(self, request):
        """
        Return per-student detail with question answers for a specific module chapter.
        Query params: chapter_id (required), period (days, default 30), class (class name filter).
        """
        from datetime import timedelta
        from collections import defaultdict

        chapter_id = request.query_params.get('chapter_id', '').strip()
        if not chapter_id:
            return error(message="chapter_id is required", status_code=400)

        try:
            period_days = int(request.query_params.get('period', 30))
        except ValueError:
            period_days = 30
        since = timezone.now() - timedelta(days=period_days)
        class_filter = request.query_params.get('class', '').strip()

        try:
            chapter_obj = ModuleChapter.objects.get(id=chapter_id)
        except ModuleChapter.DoesNotExist:
            return error(message="Chapter not found", status_code=404)

        # Resolve school from the requesting user
        user_school = None
        profile = getattr(request.user, 'profile', None)
        if profile:
            user_school = getattr(profile, 'school', None)
            if not user_school and getattr(profile, 'student', None) and getattr(profile.student, 'class_instance', None):
                user_school = profile.student.class_instance.school
            if not user_school and hasattr(profile, 'teacher_profile') and profile.teacher_profile:
                first_assignment = Teacher.objects.filter(teacher=profile.teacher_profile).select_related('class_instance__school').first()
                if first_assignment and first_assignment.class_instance:
                    user_school = first_assignment.class_instance.school

        empty_response = {
            'chapter_name': chapter_obj.title,
            'total_questions': 0,
            'average_percentage': 0,
            'students': [],
            'questions': [],
        }
        if not user_school:
            return success(data=empty_response, message="No school found for user")

        # Teacher isolation
        teacher_class_ids = None
        if profile and profile.user_type == 'teacher' and hasattr(profile, 'teacher_profile'):
            _assignments = Teacher.objects.filter(teacher=profile.teacher_profile, is_deleted=False)
            teacher_class_ids = list(_assignments.values_list('class_instance_id', flat=True).distinct())

        students_qs = Account.objects.filter(
            profile__user_type='student',
            is_active=True,
            is_deleted=False,
            profile__student__class_instance__school=user_school,
        )
        if teacher_class_ids is not None:
            students_qs = students_qs.filter(profile__student__class_instance_id__in=teacher_class_ids)
        if class_filter:
            students_qs = students_qs.filter(profile__student__class_instance__name__icontains=class_filter)

        students_list = list(students_qs.select_related('profile__student').order_by('first_name', 'last_name')[:200])
        student_ids = [s.id for s in students_list]

        if not student_ids:
            return success(data=empty_response, message="No students found")

        # Get ordered questions for this chapter
        module_contents = list(
            ModuleContent.objects.filter(
                chapter_id=chapter_id,
                content_type='question',
                is_deleted=False,
            ).select_related('question').order_by('order')
        )
        questions_list = [(mc.question_id, mc.question, mc.order) for mc in module_contents if mc.question_id]
        question_ids = [q[0] for q in questions_list]
        total_questions = len(question_ids)

        # Fetch latest answer per (user, question) in one query
        all_answers = Answer.objects.filter(
            question_id__in=question_ids,
            user_id__in=student_ids,
            created_at__gte=since,
        ).order_by('user_id', 'question_id', '-created_at').values('user_id', 'question_id', 'is_correct')

        # Build lookup: (user_id, question_id) -> is_correct (latest only)
        latest_answer_map = {}
        correct_by_question = defaultdict(list)
        seen = set()
        for ans in all_answers:
            pair = (ans['user_id'], ans['question_id'])
            if pair not in seen:
                seen.add(pair)
                latest_answer_map[pair] = ans['is_correct']
                correct_by_question[ans['question_id']].append(ans['is_correct'])

        # Build per-student results
        student_results = []
        percentage_sum = 0
        attempted_count = 0
        for student in students_list:
            name = f"{student.first_name} {student.last_name}".strip() or student.username
            answers = []
            correct = 0
            attempted = 0
            for qid, question, order in questions_list:
                is_correct = latest_answer_map.get((student.id, qid))
                if is_correct is None:
                    continue
                attempted += 1
                if is_correct:
                    correct += 1
                answers.append({
                    'question_title': (question.question_text[:200] if question else '—'),
                    'is_correct': is_correct,
                })
            wrong = attempted - correct
            pct = round((correct / total_questions) * 100) if total_questions and attempted > 0 else 0
            if attempted > 0:
                percentage_sum += pct
                attempted_count += 1
            student_results.append({
                'user_id': student.id,
                'user_name': name,
                'percentage': pct,
                'questions_attempted': attempted,
                'correct_answer_count': correct,
                'wrong_answer_count': wrong,
                'status': 'attempted' if attempted > 0 else 'not_started',
                'answers': answers,
            })

        student_results.sort(key=lambda x: (-x['percentage'], x['user_name'].lower()))
        avg_pct = round(percentage_sum / attempted_count) if attempted_count else 0

        # Question-wise analysis
        question_analysis = []
        for idx, (qid, question, order) in enumerate(questions_list):
            ans_list = correct_by_question.get(qid, [])
            total_ans = len(ans_list)
            correct_ans = sum(1 for a in ans_list if a)
            correct_pct = round((correct_ans / total_ans) * 100) if total_ans else 0
            question_analysis.append({
                'question_id': str(qid),
                'question_title': (question.question_text[:200] if question else '—'),
                'correct_percentage': correct_pct,
                'order': order or idx + 1,
            })

        return success(data={
            'chapter_name': chapter_obj.title,
            'total_questions': total_questions,
            'average_percentage': avg_pct,
            'students': student_results,
            'questions': question_analysis,
        }, message="Chapter student details retrieved successfully")

    def _calculate_engagement_rate(self):
        """Calculate engagement rate based on recent activity."""
        recent_students = Account.objects.filter(
            profile__user_type='student',
            is_active=True,
            last_login__gte=timezone.now() - timezone.timedelta(days=30)
        ).count()
        
        total_students = Account.objects.filter(profile__user_type='student', is_active=True).count()
        
        if total_students > 0:
            return round((recent_students / total_students) * 100, 1)
        return 0
    
    def _calculate_satisfaction_score(self):
        """Calculate satisfaction score based on completion rates and performance."""
        total_modules = Module.objects.filter(is_active=True).count()
        completed_modules = UserModuleProgress.objects.filter(status='completed').count()
        
        if total_modules > 0:
            completion_rate = (completed_modules / total_modules) * 100
            
            if completion_rate >= 80:
                return 4.5
            elif completion_rate >= 60:
                return 4.0
            elif completion_rate >= 40:
                return 3.5
            elif completion_rate >= 20:
                return 3.0
            else:
                return 2.5
        
        return 3.0
    
class AIServiceViewSet(viewsets.ViewSet):
    """ViewSet for AI Service operations."""
    permission_classes = [permissions.IsAuthenticated]
    
    @action(detail=False, methods=['get'], url_path='suggestions')
    def suggestions(self, request):
        """Get AI suggestions."""
        api_logger.info(f"AI suggestions requested by {request.user.username} (ID: {request.user.id})")
        
        category = request.query_params.get('category', None)
        priority = request.query_params.get('priority', None)
        suggestions_data = []
        
        students = Account.objects.filter(profile__user_type='student', is_active=True)
        user_type = request.user.profile.user_type if hasattr(request.user, 'profile') else None
        if user_type == 'teacher':
            students = students.filter(class_instance__school=request.user.profile.school if hasattr(request.user, 'profile') else None)
        
        subject_performance = {}
        for subject in Subject.objects.filter(is_active=True):
            subject_avg_exp = students.aggregate(avg=Avg('profile__student__total_exp'))['avg'] or 0
            subject_performance[subject.name] = subject_avg_exp
        
        if subject_performance:
            worst_subject = min(subject_performance, key=subject_performance.get)
            worst_performance = subject_performance[worst_subject]
            
            if worst_performance < 50:
                suggestions_data.append({
                    'id': 1,
                    'category': 'content',
                    'title': f'Create Additional Content for {worst_subject}',
                    'description': f'Students are struggling with {worst_subject} (avg score: {worst_performance:.0f}). Create more interactive content and practice questions.',
                    'priority': 'high',
                    'impact': f'Expected to improve {worst_subject} performance by 25%',
                    'timeEstimate': '45 minutes',
                    'studentsAffected': students.count(),
                    'confidence': 0.85,
                    'createdAt': timezone.now().isoformat(),
                    'status': 'pending'
                })
        
        total_modules = Module.objects.filter(is_active=True).count()
        completed_modules = UserModuleProgress.objects.filter(
            account__in=students,
            status='completed'
        ).count()
        completion_rate = (completed_modules / max(total_modules, 1)) * 100
        
        if completion_rate < 60:
            suggestions_data.append({
                'id': 2,
                'category': 'assessment',
                'title': 'Implement Adaptive Assessment Strategy',
                'description': f'Module completion rate is {completion_rate:.1f}%. Implement adaptive difficulty and personalized assessments.',
                'priority': 'medium',
                'impact': 'Expected 20% improvement in completion rates',
                'timeEstimate': '60 minutes',
                'studentsAffected': students.count(),
                'confidence': 0.80,
                'createdAt': timezone.now().isoformat(),
                'status': 'pending'
            })
        
        recent_active_students = students.filter(
            last_login__gte=timezone.now() - timezone.timedelta(days=7)
        ).count()
        total_students = students.count()
        engagement_rate = (recent_active_students / max(total_students, 1)) * 100
        
        if engagement_rate < 70:
            suggestions_data.append({
                'id': 3,
                'category': 'engagement',
                'title': 'Boost Student Engagement',
                'description': f'Student engagement is {engagement_rate:.1f}%. Add gamification elements and interactive features.',
                'priority': 'high',
                'impact': 'Expected 30% increase in daily active users',
                'timeEstimate': '30 minutes',
                'studentsAffected': total_students,
                'confidence': 0.75,
                'createdAt': timezone.now().isoformat(),
                'status': 'pending'
            })
        
        if not suggestions_data:
            suggestions_data.append({
                'id': 1,
                'category': 'content',
                'title': 'Create New Practice Questions',
                'description': 'Add more practice questions to help students reinforce their learning.',
                'priority': 'medium',
                'impact': 'Improved learning retention',
                'timeEstimate': '30 minutes',
                'studentsAffected': total_students,
                'confidence': 0.70,
                'createdAt': timezone.now().isoformat(),
                'status': 'pending'
            })
        
        if category:
            suggestions_data = [s for s in suggestions_data if s['category'] == category]
        
        if priority:
            suggestions_data = [s for s in suggestions_data if s['priority'] == priority]
        
        categories = [
            {'id': 'content', 'name': 'Content Creation', 'count': len([s for s in suggestions_data if s['category'] == 'content'])},
            {'id': 'assessment', 'name': 'Assessment', 'count': len([s for s in suggestions_data if s['category'] == 'assessment'])},
            {'id': 'engagement', 'name': 'Student Engagement', 'count': len([s for s in suggestions_data if s['category'] == 'engagement'])}
        ]
        
        priorities = [
            {'id': 'high', 'name': 'High Priority', 'count': len([s for s in suggestions_data if s['priority'] == 'high'])},
            {'id': 'medium', 'name': 'Medium Priority', 'count': len([s for s in suggestions_data if s['priority'] == 'medium'])},
            {'id': 'low', 'name': 'Low Priority', 'count': len([s for s in suggestions_data if s['priority'] == 'low'])}
        ]
        
        api_logger.info(f"AI suggestions returned {len(suggestions_data)} suggestions")
        return success(
            data={
                'suggestions': suggestions_data,
                'categories': categories,
                'priorities': priorities
            },
            message="AI suggestions retrieved successfully"
        )
    
    @action(detail=False, methods=['get'], url_path='insights')
    def insights(self, request):
        """Get AI insights."""
        api_logger.info(f"AI insights requested by {request.user.username} (ID: {request.user.id})")
        
        insights_data = []
        
        students = Account.objects.filter(profile__user_type='student', is_active=True)
        user_type = request.user.profile.user_type if hasattr(request.user, 'profile') else None
        if user_type == 'teacher':
            students = students.filter(class_instance__school=request.user.profile.school if hasattr(request.user, 'profile') else None)
        
        morning_students = students.filter(
            last_login__hour__gte=6,
            last_login__hour__lt=12
        ).count()
        afternoon_students = students.filter(
            last_login__hour__gte=12,
            last_login__hour__lt=18
        ).count()
        
        if morning_students > 0 and afternoon_students > 0:
            morning_performance = students.filter(
                last_login__hour__gte=6,
                last_login__hour__lt=12
            ).aggregate(avg=Avg('profile__student__total_exp'))['avg'] or 0
            afternoon_performance = students.filter(
                last_login__hour__gte=12,
                last_login__hour__lt=18
            ).aggregate(avg=Avg('profile__student__total_exp'))['avg'] or 0
            
            if morning_performance > afternoon_performance:
                performance_diff = ((morning_performance - afternoon_performance) / max(afternoon_performance, 1)) * 100
                insights_data.append({
                    'id': 1,
                    'title': 'Learning Pattern Analysis',
                    'description': f'Students show {performance_diff:.0f}% better performance in morning sessions compared to afternoon.',
                    'confidence': 0.85,
                    'category': 'learning_patterns',
                    'actionable': True,
                    'created_at': timezone.now().isoformat()
                })
        
        subject_performance = {}
        for subject in Subject.objects.filter(is_active=True):
            subject_avg_exp = students.aggregate(avg=Avg('profile__student__total_exp'))['avg'] or 0
            subject_performance[subject.name] = subject_avg_exp
        
        if len(subject_performance) >= 2:
            best_subject = max(subject_performance, key=subject_performance.get)
            second_best = sorted(subject_performance, key=subject_performance.get, reverse=True)[1]
            
            insights_data.append({
                'id': 2,
                'title': 'Subject Performance Correlation',
                'description': f'{best_subject} and {second_best} show strong positive correlation in student performance.',
                'confidence': 0.80,
                'category': 'subject_analysis',
                'actionable': True,
                'created_at': timezone.now().isoformat()
            })
        
        total_students = students.count()
        recent_active = students.filter(
            last_login__gte=timezone.now() - timezone.timedelta(days=7)
        ).count()
        engagement_rate = (recent_active / max(total_students, 1)) * 100
        
        if engagement_rate < 80:
            insights_data.append({
                'id': 3,
                'title': 'Engagement Analysis',
                'description': f'Student engagement is {engagement_rate:.1f}%. Consider implementing engagement strategies.',
                'confidence': 0.90,
                'category': 'engagement',
                'actionable': True,
                'created_at': timezone.now().isoformat()
            })
        
        if not insights_data:
            insights_data.append({
                'id': 1,
                'title': 'System Performance',
                'description': 'System is performing well. Continue monitoring student progress.',
                'confidence': 0.75,
                'category': 'system_analysis',
                'actionable': False,
                'created_at': timezone.now().isoformat()
            })
        
        api_logger.info(f"AI insights returned {len(insights_data)} insights")
        return success(
            data=insights_data,
            message="AI insights retrieved successfully"
        )
    
    @action(detail=False, methods=['post'], url_path='generate')
    def generate(self, request):
        """Generate AI content."""
        api_logger.info(f"AI content generation requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")
        
        content_type = request.data.get('type', 'questions')
        subject = request.data.get('subject', 'Mathematics')
        grade = request.data.get('grade', '10')
        count = request.data.get('count', 5)
        
        generated_content = []
        
        subject_obj = Subject.objects.filter(name__icontains=subject, is_active=True).first()
        modules = Module.objects.filter(subject=subject_obj, is_active=True) if subject_obj else []
        
        if content_type == 'questions':
            for i in range(count):
                module = modules[i % len(modules)] if modules else None
                question_text = f"What is the main concept covered in {module.name if module else subject}?"
                
                generated_content.append({
                    'id': i + 1,
                    'question_text': question_text,
                    'difficulty': 'medium',
                    'subject': subject,
                    'grade': grade,
                    'module_id': module.id if module else None,
                    'options': [
                        {'text': 'Option A - Correct Answer', 'correct': True},
                        {'text': 'Option B - Incorrect', 'correct': False},
                        {'text': 'Option C - Incorrect', 'correct': False},
                        {'text': 'Option D - Incorrect', 'correct': False}
                    ],
                    'explanation': f"This question tests understanding of {module.name if module else subject} concepts for Grade {grade} students."
                })
        elif content_type == 'explanations':
            for i in range(count):
                module = modules[i % len(modules)] if modules else None
                generated_content.append({
                    'id': i + 1,
                    'title': f"Explanation: {module.name if module else subject} - Topic {i+1}",
                    'content': f"This explanation covers key concepts in {module.name if module else subject} for Grade {grade} students. It includes detailed examples and step-by-step solutions.",
                    'difficulty': 'medium',
                    'subject': subject,
                    'grade': grade,
                    'module_id': module.id if module else None
                })
        
        api_logger.info(f"AI generated {len(generated_content)} {content_type} for {request.user.username}")
        return success(
            data={
                'content': generated_content,
                'metadata': {
                    'type': content_type,
                    'subject': subject,
                    'grade': grade,
                    'count': count,
                    'generated_at': timezone.now().isoformat(),
                    'based_on_modules': len(modules) > 0
                }
            },
            message=f"Successfully generated {count} AI {content_type}"
        )
    
    @action(detail=False, methods=['get'], url_path='recommendations/(?P<type>[^/.]+)')
    def recommendations(self, request, type=None):
        """Get AI recommendations by type."""
        api_logger.info(f"AI recommendations requested by {request.user.username} (ID: {request.user.id}) for type: {type}")
        
        recommendations_data = []
        
        students = Account.objects.filter(profile__user_type='student', is_active=True)
        user_type = request.user.profile.user_type if hasattr(request.user, 'profile') else None
        if user_type == 'teacher':
            students = students.filter(class_instance__school=request.user.profile.school if hasattr(request.user, 'profile') else None)
        
        if type == 'content':
            subject_performance = {}
            for subject in Subject.objects.filter(is_active=True):
                subject_avg_exp = students.aggregate(avg=Avg('profile__student__total_exp'))['avg'] or 0
                subject_performance[subject.name] = subject_avg_exp
            
            if subject_performance:
                worst_subject = min(subject_performance, key=subject_performance.get)
                worst_performance = subject_performance[worst_subject]
                
                recommendations_data.append({
                    'id': 1,
                    'title': f'Create Additional Content for {worst_subject}',
                    'description': f'Students are struggling with {worst_subject} (avg score: {worst_performance:.0f}). Create more interactive content.',
                    'priority': 'high',
                    'estimated_impact': f'25% improvement in {worst_subject} performance'
                })
            
            total_modules = Module.objects.filter(is_active=True).count()
            completed_modules = UserModuleProgress.objects.filter(
                account__in=students,
                status='completed'
            ).count()
            completion_rate = (completed_modules / max(total_modules, 1)) * 100
            
            if completion_rate < 60:
                recommendations_data.append({
                    'id': 2,
                    'title': 'Add Interactive Learning Elements',
                    'description': f'Module completion rate is {completion_rate:.1f}%. Add interactive elements to improve engagement.',
                    'priority': 'medium',
                    'estimated_impact': '30% increase in completion rates'
                })
                
        elif type == 'assessment':
            avg_student_score = students.aggregate(avg=Avg('profile__student__total_exp'))['avg'] or 0
            
            if avg_student_score < 50:
                recommendations_data.append({
                    'id': 1,
                    'title': 'Implement Adaptive Assessment',
                    'description': f'Average student score is {avg_student_score:.0f}. Implement adaptive difficulty assessment.',
                    'priority': 'high',
                    'estimated_impact': '35% more accurate assessment'
                })
            
            total_questions = Question.objects.filter(is_active=True).count()
            if total_questions < 50:
                recommendations_data.append({
                    'id': 2,
                    'title': 'Add More Practice Questions',
                    'description': f'Only {total_questions} questions available. Add more practice questions for better assessment.',
                    'priority': 'medium',
                    'estimated_impact': '20% improvement in assessment accuracy'
                })
                
        elif type == 'engagement':
            recent_active = students.filter(
                last_login__gte=timezone.now() - timezone.timedelta(days=7)
            ).count()
            total_students = students.count()
            engagement_rate = (recent_active / max(total_students, 1)) * 100
            
            if engagement_rate < 70:
                recommendations_data.append({
                    'id': 1,
                    'title': 'Implement Gamification Elements',
                    'description': f'Student engagement is {engagement_rate:.1f}%. Add points, badges, and leaderboards.',
                    'priority': 'high',
                    'estimated_impact': '40% increase in daily active users'
                })
            
            total_missions = Mission.objects.filter(is_active=True).count()
            mission_participants = UserMissionProgress.objects.filter(
                account__in=students
            ).values('mission').distinct().count()
            
            if total_missions > 0 and (mission_participants / total_missions) < 0.5:
                recommendations_data.append({
                    'id': 2,
                    'title': 'Create Collaborative Learning Groups',
                    'description': 'Low mission participation. Create study groups and collaborative activities.',
                    'priority': 'medium',
                    'estimated_impact': '25% increase in collaborative learning'
                })
        else:
            recommendations_data.append({
                'id': 1,
                'title': 'Monitor Student Progress',
                'description': 'Continue monitoring student progress and adjust content accordingly.',
                'priority': 'low',
                'estimated_impact': '10% overall improvement'
            })
        
        api_logger.info(f"AI recommendations returned {len(recommendations_data)} recommendations for type: {type}")
        return success(
            data=recommendations_data,
            message=f"AI recommendations for {type} retrieved successfully"
        )


class ClassViewSet(viewsets.ModelViewSet):
    """ViewSet for Class model."""
    queryset = Class.objects.all()
    serializer_class = ClassSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Return optimized queryset with related data, filtered by user's school.
        Teachers see only classes they are assigned to via Teacher assignments."""
        user = self.request.user
        queryset = Class.objects.select_related(
            'school',
        ).prefetch_related(
            'enrolled_students__user_profile__account',
        ).filter(is_active=True)

        if hasattr(user, 'profile') and user.profile.school:
            queryset = queryset.filter(school=user.profile.school)

            if (user.profile.user_type == 'teacher'
                    and hasattr(user.profile, 'teacher_profile')):
                assigned_class_ids = Teacher.objects.filter(
                    teacher=user.profile.teacher_profile,
                    is_deleted=False,
                ).values_list('class_instance_id', flat=True).distinct()
                queryset = queryset.filter(id__in=assigned_class_ids)

        return queryset
    
    def get_serializer_class(self):
        """Return appropriate serializer class based on action."""
        if self.action == 'list':
            return ClassListSerializer
        return ClassSerializer
    
    def list(self, request, *args, **kwargs):
        """List classes with logging."""
        api_logger.info(f"Class list requested by {request.user.username} (ID: {request.user.id}) from {request.META.get('REMOTE_ADDR', 'unknown')}")
        
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        api_logger.info(f"Class list returned {len(serializer.data)} classes")
        
        return success(
            data=serializer.data,
            message="Classes retrieved successfully"
        )
    
    def create(self, request, *args, **kwargs):
        """Create a class with logging."""
        api_logger.info(f"Class creation requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")
        
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            class_instance = serializer.save()
            api_logger.info(f"Class created successfully: {class_instance.name} (ID: {class_instance.id}) by {request.user.username} in school {class_instance.school.name}")
            return created(
                data=serializer.data,
                message="Class created successfully"
            )
        api_logger.warning(f"Class creation failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)
    
    def update(self, request, *args, **kwargs):
        """Update a class with logging."""
        class_id = kwargs.get('pk')
        api_logger.info(f"Class update requested by {request.user.username} (ID: {request.user.id}) for class ID: {class_id} - Data: {request.data}")
        
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            class_instance = serializer.save()
            api_logger.info(f"Class updated successfully: {class_instance.name} (ID: {class_instance.id}) by {request.user.username}")
            return success(
                data=serializer.data,
                message="Class updated successfully"
            )
        api_logger.warning(f"Class update failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)
    
    def destroy(self, request, *args, **kwargs):
        """Delete a class with logging."""
        class_id = kwargs.get('pk')
        api_logger.info(f"Class delete requested by {request.user.username} (ID: {request.user.id}) for class ID: {class_id}")
        
        instance = self.get_object()
        instance.delete()
        api_logger.info(f"Class deleted successfully: {instance.name} (ID: {instance.id}) by {request.user.username}")
        return success(
            message="Class deleted successfully"
        )


class SchoolViewSet(viewsets.ModelViewSet):
    """ViewSet for School model."""
    queryset = School.objects.all()
    serializer_class = SchoolSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Return optimized queryset with related data, filtered by user's school."""
        user = self.request.user
        queryset = School.objects.prefetch_related(
            'users',
            'classes',
        ).filter(is_active=True, is_deleted=False)
        
        if hasattr(user, 'profile') and user.profile.school:
            queryset = queryset.filter(id=user.profile.school.id)
        
        return queryset
    
    def get_serializer_class(self):
        """Return appropriate serializer class based on action."""
        if self.action == 'list':
            return SchoolListSerializer
        return SchoolSerializer
    
    def list(self, request, *args, **kwargs):
        """List schools with logging."""
        api_logger.info(f"School list requested by {request.user.username} (ID: {request.user.id}) from {request.META.get('REMOTE_ADDR', 'unknown')}")
        
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        api_logger.info(f"School list returned {len(serializer.data)} schools")
        
        return success(
            data=serializer.data,
            message="Schools retrieved successfully"
        )
    
    def create(self, request, *args, **kwargs):
        """Create a school with logging."""
        api_logger.info(f"School creation requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")
        
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            school_instance = serializer.save()
            api_logger.info(f"School created successfully: {school_instance.name} (ID: {school_instance.id}) by {request.user.username}")
            return created(
                data=serializer.data,
                message="School created successfully"
            )
        api_logger.warning(f"School creation failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)
    
    def update(self, request, *args, **kwargs):
        """Update a school with logging."""
        school_id = kwargs.get('pk')
        api_logger.info(f"School update requested by {request.user.username} (ID: {request.user.id}) for school ID: {school_id} - Data: {request.data}")
        
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            school_instance = serializer.save()
            api_logger.info(f"School updated successfully: {school_instance.name} (ID: {school_instance.id}) by {request.user.username}")
            return success(
                data=serializer.data,
                message="School updated successfully"
            )
        api_logger.warning(f"School update failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)
    
    def destroy(self, request, *args, **kwargs):
        """Soft delete a school with logging."""
        school_id = kwargs.get('pk')
        api_logger.info(f"School soft delete requested by {request.user.username} (ID: {request.user.id}) for school ID: {school_id}")

        instance = self.get_object()
        instance.soft_delete()
        api_logger.info(f"School soft deleted successfully: {instance.name} (ID: {instance.id}) by {request.user.username}")
        return success(
            message="School soft deleted successfully"
        )

    @action(detail=False, methods=['get'], url_path='list_all', permission_classes=[TeacherAdminPermission])
    def list_all(self, request):
        """List all active schools (for admin use - source/target selection in clone)."""
        queryset = School.objects.filter(is_active=True, is_deleted=False).order_by('name')
        serializer = SchoolListSerializer(queryset, many=True)
        return success(data=serializer.data, message="All schools retrieved successfully")

    @action(detail=False, methods=['post'], url_path='clone', permission_classes=[TeacherAdminPermission])
    def clone_school_data(self, request):
        """
        Clone all academic content (subjects, modules, chapters, questions, theories)
        from one school to another. Student/teacher data is NOT cloned.

        Request body:
            source_school_id (str): ID of the source school
            target_school_id (str): ID of the destination school
            dry_run (bool): If true, preview counts without saving (default: false)
        """
        from gyaan_buddy.subjects.models import (
            Subject, Module, ModuleChapter, ModuleContent,
            Question, Option, Theory, ChapterHOTS,
        )
        from django.db import transaction

        source_id = request.data.get('source_school_id')
        target_id = request.data.get('target_school_id')
        dry_run = request.data.get('dry_run', False)

        if not source_id or not target_id:
            return validation_error({'error': 'source_school_id and target_school_id are required.'})

        if str(source_id) == str(target_id):
            return validation_error({'error': 'Source and destination school must be different.'})

        try:
            old_school = School.objects.get(id=source_id, is_active=True, is_deleted=False)
        except School.DoesNotExist:
            return validation_error({'error': f'Source school not found.'})

        try:
            new_school = School.objects.get(id=target_id, is_active=True, is_deleted=False)
        except School.DoesNotExist:
            return validation_error({'error': f'Target school not found.'})

        api_logger.info(
            f"Clone school data requested by {request.user.username}: "
            f"'{old_school.name}' → '{new_school.name}' (dry_run={dry_run})"
        )

        old_subjects = list(old_school.subjects.all())
        old_classes = list(old_school.classes.all())

        # --- dry-run summary ---
        if dry_run:
            total_modules = total_chapters = total_questions = total_theories = 0
            for subj in old_subjects:
                for mod in subj.modules.all():
                    total_modules += 1
                    for chap in mod.chapters.filter(is_deleted=False):
                        total_chapters += 1
                        for content in chap.contents.filter(is_deleted=False):
                            if content.content_type == 'question':
                                total_questions += 1
                            else:
                                total_theories += 1
            return success(data={
                'dry_run': True,
                'source_school': old_school.name,
                'target_school': new_school.name,
                'subjects': len(old_subjects),
                'classes': len(old_classes),
                'modules': total_modules,
                'chapters': total_chapters,
                'questions': total_questions,
                'theories': total_theories,
            }, message="Dry run complete — nothing was saved.")

        def _clone_question(old_q):
            new_q = Question.objects.create(
                question_text=old_q.question_text,
                question_type=old_q.question_type,
                exp_points=old_q.exp_points,
                difficulty_level=old_q.difficulty_level,
                explanation=old_q.explanation,
                hint=old_q.hint,
                is_active=old_q.is_active,
                is_hots=old_q.is_hots,
                ai_generated=old_q.ai_generated,
                level=old_q.level,
            )
            for old_opt in old_q.options.all().order_by('order'):
                Option.objects.create(
                    question=new_q,
                    option_text=old_opt.option_text,
                    is_correct=old_opt.is_correct,
                    order=old_opt.order,
                )
            return new_q

        stats = {
            'classes_created': 0, 'classes_reused': 0,
            'subjects_created': 0, 'subjects_reused': 0,
            'modules_created': 0, 'modules_reused': 0,
            'chapters_created': 0, 'chapters_reused': 0,
            'questions': 0, 'theories': 0, 'hots': 0,
        }

        try:
            with transaction.atomic():
                # Step 1: Map classes
                class_map = {}
                for old_cls in old_classes:
                    new_cls, created = Class.objects.get_or_create(
                        name=old_cls.name,
                        school=new_school,
                        defaults={
                            'description': old_cls.description,
                            'is_active': old_cls.is_active,
                        }
                    )
                    class_map[old_cls.id] = new_cls
                    if created:
                        stats['classes_created'] += 1
                    else:
                        stats['classes_reused'] += 1

                # Step 2: Clone subjects → modules → chapters → contents
                for old_subj in old_subjects:
                    new_subj, created = Subject.objects.get_or_create(
                        name=old_subj.name,
                        school=new_school,
                        defaults={
                            'code': old_subj.code,
                            'description': old_subj.description,
                            'logo_url': old_subj.logo_url,
                            'color': old_subj.color,
                            'is_active': old_subj.is_active,
                            'order': old_subj.order,
                        }
                    )
                    if created:
                        stats['subjects_created'] += 1
                    else:
                        stats['subjects_reused'] += 1

                    for old_mod in old_subj.modules.all().order_by('order'):
                        new_cls = class_map.get(old_mod.class_instance_id)
                        if not new_cls:
                            continue

                        new_mod, created = Module.objects.get_or_create(
                            name=old_mod.name,
                            subject=new_subj,
                            class_instance=new_cls,
                            defaults={
                                'description': old_mod.description,
                                'order': old_mod.order,
                                'is_active': old_mod.is_active,
                                'is_enabled': old_mod.is_enabled,
                                'logo_url': old_mod.logo_url,
                            }
                        )
                        if created:
                            stats['modules_created'] += 1
                        else:
                            stats['modules_reused'] += 1

                        question_map = {}
                        theory_map = {}

                        for old_chap in old_mod.chapters.filter(is_deleted=False).order_by('order'):
                            new_chap, created = ModuleChapter.objects.get_or_create(
                                module=new_mod,
                                order=old_chap.order,
                                defaults={
                                    'title': old_chap.title,
                                    'description': old_chap.description,
                                    'is_enabled': old_chap.is_enabled,
                                    'is_important': old_chap.is_important,
                                    'has_hots': old_chap.has_hots,
                                    'max_questions': old_chap.max_questions,
                                    'theory': old_chap.theory,
                                }
                            )
                            if created:
                                stats['chapters_created'] += 1
                            else:
                                stats['chapters_reused'] += 1

                            for old_content in old_chap.contents.filter(is_deleted=False).order_by('order'):
                                if old_content.content_type == 'question' and old_content.question:
                                    old_q = old_content.question
                                    if old_q.id not in question_map:
                                        question_map[old_q.id] = _clone_question(old_q)
                                        stats['questions'] += 1
                                    ModuleContent.objects.get_or_create(
                                        chapter=new_chap,
                                        order=old_content.order,
                                        defaults={'content_type': 'question', 'question': question_map[old_q.id]}
                                    )
                                elif old_content.content_type == 'theory' and old_content.theory:
                                    old_t = old_content.theory
                                    if old_t.id not in theory_map:
                                        theory_map[old_t.id] = Theory.objects.create(
                                            title=old_t.title,
                                            description=old_t.description,
                                        )
                                        stats['theories'] += 1
                                    ModuleContent.objects.get_or_create(
                                        chapter=new_chap,
                                        order=old_content.order,
                                        defaults={'content_type': 'theory', 'theory': theory_map[old_t.id]}
                                    )

                            for old_hots in old_chap.hots_questions.all().order_by('order'):
                                old_q = old_hots.question
                                if old_q.id not in question_map:
                                    question_map[old_q.id] = _clone_question(old_q)
                                    stats['questions'] += 1
                                ChapterHOTS.objects.get_or_create(
                                    chapter=new_chap,
                                    question=question_map[old_q.id],
                                    defaults={'order': old_hots.order}
                                )
                                stats['hots'] += 1

        except Exception as e:
            api_logger.error(f"Clone school data failed: {str(e)}", exc_info=True)
            return validation_error({'error': f'Clone failed: {str(e)}'})

        api_logger.info(
            f"Clone complete '{old_school.name}' → '{new_school.name}' by {request.user.username}: {stats}"
        )
        return success(data={
            'dry_run': False,
            'source_school': old_school.name,
            'target_school': new_school.name,
            'stats': stats,
        }, message=f"Successfully cloned '{old_school.name}' → '{new_school.name}'")


class GradeViewSet(viewsets.ModelViewSet):
    """ViewSet for Grade model."""
    queryset = Grade.objects.all()
    serializer_class = GradeSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Return optimized queryset with related data, filtered by user's school."""
        user = self.request.user
        queryset = Grade.objects.select_related(
            'school',
        ).prefetch_related(
            'classes',
        ).filter(is_active=True)
        
        if hasattr(user, 'profile') and user.profile.school:
            queryset = queryset.filter(school=user.profile.school)
        
        return queryset
    
    def get_serializer_class(self):
        """Return appropriate serializer class based on action."""
        if self.action == 'list':
            return GradeListSerializer
        return GradeSerializer
    
    def list(self, request, *args, **kwargs):
        """List grades with logging."""
        api_logger.info(f"Grade list requested by {request.user.username} (ID: {request.user.id}) from {request.META.get('REMOTE_ADDR', 'unknown')}")
        
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        api_logger.info(f"Grade list returned {len(serializer.data)} grades")
        
        return success(
            data=serializer.data,
            message="Grades retrieved successfully"
        )
    
    def create(self, request, *args, **kwargs):
        """Create a grade with logging."""
        api_logger.info(f"Grade creation requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")
        
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            grade_instance = serializer.save()
            api_logger.info(f"Grade created successfully: {grade_instance.name} (ID: {grade_instance.id}) by {request.user.username} in school {grade_instance.school.name}")
            return created(
                data=serializer.data,
                message="Grade created successfully"
            )
        api_logger.warning(f"Grade creation failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)
    
    def update(self, request, *args, **kwargs):
        """Update a grade with logging."""
        grade_id = kwargs.get('pk')
        api_logger.info(f"Grade update requested by {request.user.username} (ID: {request.user.id}) for grade ID: {grade_id} - Data: {request.data}")
        
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            grade_instance = serializer.save()
            api_logger.info(f"Grade updated successfully: {grade_instance.name} (ID: {grade_instance.id}) by {request.user.username}")
            return success(
                data=serializer.data,
                message="Grade updated successfully"
            )
        api_logger.warning(f"Grade update failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)
    
    def destroy(self, request, *args, **kwargs):
        """Delete a grade with logging."""
        grade_id = kwargs.get('pk')
        api_logger.info(f"Grade delete requested by {request.user.username} (ID: {request.user.id}) for grade ID: {grade_id}")
        
        instance = self.get_object()
        instance.delete()
        api_logger.info(f"Grade deleted successfully: {instance.name} (ID: {instance.id}) by {request.user.username}")
        return success(
            message="Grade deleted successfully"
        )


class MissionViewSet(viewsets.ModelViewSet):
    """ViewSet for Mission model."""
    queryset = Mission.objects.all()
    serializer_class = MissionSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Return optimized queryset with related data, filtered by current user."""
        user = self.request.user
        
        queryset = Mission.objects.select_related(
            'account',
            'subject',
        ).prefetch_related(
            'questions',
            'progress',
        ).filter(
            account=user,
            is_deleted=False,
        )
        api_logger.info(f"Missions for user {user.username}: {queryset.count()}")
        return queryset
    
    def get_serializer_class(self):
        """Return appropriate serializer class based on action."""
        if self.action == 'create':
            return MissionCreateSerializer
        return MissionSerializer
    
    def list(self, request, *args, **kwargs):
        """List all missions for the current user."""
        api_logger.info(f"Mission list requested by {request.user.username} (ID: {request.user.id}) from {request.META.get('REMOTE_ADDR', 'unknown')}")
        
        queryset = Mission.objects.select_related(
            'account',
            'subject',
        ).prefetch_related(
            'questions',
            'mission_questions',
            'mission_questions__question',
            'mission_questions__chapter',
            'progress',
        ).filter(
            account=request.user,
            is_deleted=False,
        ).order_by('-mission_date', '-created_at')
        
        serializer = self.get_serializer(queryset, many=True, context={'request': request})
        api_logger.info(f"Mission list returned {len(serializer.data)} missions for user {request.user.username}")
        
        return success(
            data=serializer.data,
            message="Missions retrieved successfully"
        )
    
    def create(self, request, *args, **kwargs):
        """Create a mission with logging."""
        api_logger.info(f"Mission creation requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")

        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            mission = serializer.save(account=request.user)
            api_logger.info(f"Mission created successfully: {mission.title} (ID: {mission.id}) by {request.user.username}")

            self._initialize_mission_progress(mission)

            try:
                from gyaan_buddy.utils.firebase_notifications import firebase_notification_service
                title = "🎯 Your next mission is ready"
                body = f"You have a new mission for {mission.subject.name}. Complete it to strengthen your knowledge!"
                data = {
                    'type': 'mission_created',
                    'mission_id': str(mission.id),
                    'subject_name': mission.subject.name,
                    'action': 'open_mission',
                }
                firebase_notification_service.send_notification_to_user(
                    user=request.user,
                    title=title,
                    body=body,
                    data=data,
                    notification_type='mission',
                    triggered_by='auto',
                )
                api_logger.info(f"Mission created notification sent to user {request.user.username} for mission {mission.id}")
            except Exception as notif_error:
                api_logger.error(f"Failed to send mission created notification to user {request.user.username}: {str(notif_error)}")

            return created(
                data=MissionSerializer(mission).data,
                message="Mission created successfully"
            )
        api_logger.warning(f"Mission creation failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)
    
    def _initialize_mission_progress(self, mission):
        """Create UserMissionProgress record for the mission (OneToOne relationship)."""
        try:
            _, was_created = UserMissionProgress.objects.get_or_create(
                mission=mission,
                defaults={
                    'status': 'not_started',
                    'exp_earned': 0,
                    'total_questions': mission.question_count
                }
            )
            
            if was_created:
                api_logger.info(
                    f"Initialized mission progress for mission '{mission.title}' (ID: {mission.id})"
                )
            else:
                api_logger.info(
                    f"Mission progress already exists for mission '{mission.title}' (ID: {mission.id})"
                )
        except Exception as e:
            api_logger.error(f"Error initializing mission progress for mission {mission.id}: {str(e)}", exc_info=True)
    
    @action(detail=True, methods=['post'])
    def start_mission(self, request, pk=None):
        """Start a mission for the current user."""
        try:
            mission = self.get_object()
            user = request.user
            
            progress, created = UserMissionProgress.objects.get_or_create(
                mission=mission,
                defaults={'status': 'in_progress', 'started_at': timezone.now()}
            )
            
            if not created:
                if progress.status == 'completed':
                    return validation_error({"error": "Mission already completed"})
                elif progress.status == 'in_progress':
                    return validation_error({"error": "Mission already in progress"})
                else:
                    progress.status = 'in_progress'
                    progress.started_at = timezone.now()
                    progress.save()
            
            api_logger.info(f"User {user.username} started mission {mission.title}")
            return success(
                data=UserMissionProgressSerializer(progress).data,
                message="Mission started successfully"
            )
        except Exception as e:
            api_logger.error(f"Error starting mission: {str(e)}")
            return validation_error({"error": "Failed to start mission"})

    @action(detail=True, methods=['post'])
    def complete_mission(self, request, pk=None):
        """Mark a mission as completed for the current user."""
        try:
            mission = self.get_object()
            user = request.user

            progress, created = UserMissionProgress.objects.get_or_create(
                mission=mission,
                defaults={
                    'status': 'completed', 
                    'started_at': timezone.now(), 
                    'completed_at': timezone.now()
                }
            )
            
            if not created:
                if progress.status != 'completed':
                    progress.status = 'completed'
                    progress.completed_at = timezone.now()
                    progress.save()

            api_logger.info(f"User {user.username} completed mission {mission.title}")
            return success(
                data=UserMissionProgressSerializer(progress).data,
                message="Mission completed successfully"
            )
        except Exception as e:
            api_logger.error(f"Error completing mission: {str(e)}")
            return validation_error({"error": "Failed to complete mission"})

    def retrieve(self, request, *args, **kwargs):
        """Retrieve a specific mission with logging."""
        try:
            mission = self.get_object()
            api_logger.info(f"Mission retrieve requested by {request.user.username} (ID: {request.user.id}) for mission '{mission.title}' (ID: {mission.id}) from {request.META.get('REMOTE_ADDR', 'unknown')}")
            
            serializer = self.get_serializer(mission, context={'request': request})
            api_logger.info(f"Mission '{mission.title}' retrieved successfully for user {request.user.username}")
            
            return success(
                data=serializer.data,
                message="Mission retrieved successfully"
            )
        except Exception as e:
            api_logger.error(f"Error retrieving mission: {str(e)}")
            return validation_error({"error": "Failed to retrieve mission"})

    @action(detail=True, methods=['get'], url_path='get_next_content')
    def get_next_content(self, request, pk=None):
        """Get the next question content for a specific mission."""
        try:
            mission = self.get_object()
            
            current_question_id = request.query_params.get('id')
            
            try:
                user_mission_progress = mission.progress
                if user_mission_progress and user_mission_progress.current_question:
                    current_question_id = user_mission_progress.current_question.id
            except UserMissionProgress.DoesNotExist:
                api_logger.info(f"No progress found for user {request.user.username} in mission {mission.id}")

            if current_question_id:
                try:
                    from .helpers import handle_next_question_request
                    return handle_next_question_request(
                        request, mission, current_question_id, 
                        api_logger, success, validation_error
                    )
                except Exception as e:
                    api_logger.error(f"Error in handle_next_question_request: {str(e)}")
                    return validation_error({"error": f"Failed to process next question request: {str(e)}"})
            else:
                try:
                    from .helpers import handle_first_question_request
                    return handle_first_question_request(
                        request, mission, api_logger, success, 
                        validation_error
                    )
                except Exception as e:
                    api_logger.error(f"Error in handle_first_question_request: {str(e)}")
                    return validation_error({"error": f"Failed to process first question request: {str(e)}"})
                    
        except Exception as e:
            api_logger.error(f"Error in mission get_next_content: {str(e)}")
            return validation_error({"error": f"Failed to process mission content request: {str(e)}"})
    
    @action(detail=True, methods=['get'], url_path='questions')
    def get_mission_questions(self, request, pk=None):
        """Get up to 15 questions for a specific mission.

        Priority: questions specifically added to the mission (wrong answers to reinforce)
        come first; the remainder is filled from the subject's chapters.
        """
        import random
        from gyaan_buddy.subjects.serializers import QuestionSerializer
        from gyaan_buddy.subjects.models import ModuleChapter

        try:
            mission = self.get_object()
            api_logger.info(f"Mission questions requested by {request.user.username} for mission '{mission.title}' (ID: {mission.id})")

            total_questions_needed = 10

            # --- Step 1: collect the mission's targeted questions (wrong answers) ---
            targeted_question_ids = set()
            targeted_questions = []  # list of (question, chapter)

            mission_qs = mission.mission_questions.select_related(
                'question', 'chapter'
            ).filter(
                question__is_active=True,
                question__is_deleted=False,
            ).order_by('order')

            for mq in mission_qs:
                targeted_questions.append((mq.question, mq.chapter))
                targeted_question_ids.add(mq.question.id)

            selected_questions = list(targeted_questions)

            # --- Step 2: fill remaining slots from subject chapters ---
            if len(selected_questions) < total_questions_needed:
                subject = mission.subject
                chapters = ModuleChapter.objects.filter(
                    module__subject=subject,
                    is_enabled=True,
                    is_deleted=False
                ).prefetch_related('contents__question__options')

                chapter_questions = {}
                for chapter in chapters:
                    questions = []
                    for content in chapter.contents.filter(
                        content_type='question',
                        is_deleted=False,
                        question__is_active=True,
                        question__is_deleted=False
                    ):
                        if content.question and content.question.id not in targeted_question_ids:
                            questions.append((content.question, chapter))
                    if questions:
                        chapter_questions[chapter.id] = {
                            'chapter': chapter,
                            'questions': questions
                        }

                needed = total_questions_needed - len(selected_questions)

                if chapter_questions:
                    if len(chapter_questions) == 1:
                        chapter_data = list(chapter_questions.values())[0]
                        pool = list(chapter_data['questions'])
                        random.shuffle(pool)
                        selected_questions.extend(pool[:needed])
                    else:
                        total_available = sum(len(d['questions']) for d in chapter_questions.values())
                        fill = []
                        for _, data in chapter_questions.items():
                            share = max(1, int((len(data['questions']) / total_available) * needed))
                            pool = list(data['questions'])
                            random.shuffle(pool)
                            fill.extend(pool[:share])

                        if len(fill) > needed:
                            random.shuffle(fill)
                            fill = fill[:needed]

                        if len(fill) < needed:
                            used_ids = {q.id for q, _ in selected_questions + fill}
                            remaining = []
                            for data in chapter_questions.values():
                                for q, ch in data['questions']:
                                    if q.id not in used_ids:
                                        remaining.append((q, ch))
                            random.shuffle(remaining)
                            fill.extend(remaining[:needed - len(fill)])

                        selected_questions.extend(fill)

            if len(selected_questions) == 0:
                return success(
                    data=[],
                    message="No questions available for this mission"
                )

            # Keep targeted questions at the front; shuffle only the filler portion
            filler_start = len(targeted_questions)
            filler = selected_questions[filler_start:]
            random.shuffle(filler)
            selected_questions = selected_questions[:filler_start] + filler

            questions_data = []
            for question, chapter in selected_questions:
                question_data = QuestionSerializer(question, context={'request': request}).data
                if chapter:
                    question_data['chapter_name'] = chapter.title
                    question_data['chapter_id'] = str(chapter.id)
                questions_data.append(question_data)

            api_logger.info(
                f"Returning {len(questions_data)} questions for mission '{mission.title}' "
                f"({len(targeted_questions)} targeted, {len(questions_data) - len(targeted_questions)} filler)"
            )

            return success(
                data=questions_data[:10],
                message=f"Retrieved {min(len(questions_data), 10)} questions for mission"
            )

        except Exception as e:
            api_logger.error(f"Error getting mission questions: {str(e)}", exc_info=True)
            return validation_error({"error": f"Failed to get mission questions: {str(e)}"})

    @action(detail=False, methods=['get'], url_path='all')
    def all(self, request):
        """Get all missions with past date created by the logged in teacher."""
        try:
            user = request.user
            today = timezone.now().date()
            
            api_logger.info(f"Missions/all requested by {user.username} (ID: {user.id}) from {request.META.get('REMOTE_ADDR', 'unknown')}")
            
            total_missions = Mission.objects.filter(created_by=user).count()
            total_not_deleted = Mission.objects.filter(created_by=user, is_deleted=False).count()
            total_past = Mission.objects.filter(created_by=user, mission_date__lt=today, is_deleted=False).count()
            
            api_logger.info(f"Debug - User {user.username}: Total missions={total_missions}, Not deleted={total_not_deleted}, Past dates={total_past}, Today={today}")
            
            queryset = Mission.objects.select_related(
                'created_by',
                'class_group',
                'subject'
            ).prefetch_related(
                'questions',
            ).filter(
                created_by=user,
                mission_date__lt=today,
                is_deleted=False
            ).order_by('-mission_date', '-created_at')
            
            serializer = self.get_serializer(queryset, many=True, context={'request': request})
            api_logger.info(f"Missions/all returned {len(serializer.data)} missions for teacher {user.username}")
            
            return success(
                data=serializer.data,
                message="Past missions retrieved successfully"
            )
        except Exception as e:
            api_logger.error(f"Error in missions/all: {str(e)}", exc_info=True)
            return validation_error({"error": f"Failed to retrieve past missions: {str(e)}"})
    
    @action(detail=True, methods=['get', 'post'], url_path='manage_questions')
    def manage_questions(self, request, pk=None):
        """Handle GET (retrieve all questions) and POST (create question) for a mission."""
        try:
            mission = Mission.objects.select_related(
                'account', 'subject'
            ).get(
                pk=pk,
                is_deleted=False
            )
        except Mission.DoesNotExist:
            return not_found({"error": "Mission not found"})
        
        if request.method == 'GET':
            """Get all questions for a specific mission."""
            try:
                api_logger.info(f"Mission questions requested by {request.user.username} (ID: {request.user.id}) for mission '{mission.title}' (ID: {mission.id})")
                
                mission_questions = MissionQuestion.objects.filter(
                    mission=mission
                ).select_related(
                    'question',
                    'chapter'
                ).prefetch_related(
                    'question__options'
                ).order_by('order')
                
                from gyaan_buddy.subjects.serializers import QuestionSerializer
                questions_data = []
                for mission_question in mission_questions:
                    question_data = QuestionSerializer(mission_question.question, context={'request': request}).data
                    question_data['order'] = mission_question.order
                    question_data['mission_question_id'] = str(mission_question.id)
                    question_data['chapter_name'] = mission_question.chapter.title if mission_question.chapter else None
                    question_data['chapter_id'] = str(mission_question.chapter.id) if mission_question.chapter else None
                    questions_data.append(question_data)
                
                api_logger.info(f"Returned {len(questions_data)} questions for mission '{mission.title}'")
                return success(
                    data=questions_data,
                    message="Mission questions retrieved successfully"
                )
            except Exception as e:
                api_logger.error(f"Error retrieving mission questions: {str(e)}", exc_info=True)
                return validation_error({"error": f"Failed to retrieve mission questions: {str(e)}"})
        
        elif request.method == 'POST':
            """Create a question and add it to the mission."""
            try:
                api_logger.info(f"Mission question creation requested by {request.user.username} (ID: {request.user.id}) for mission '{mission.title}' (ID: {mission.id})")
                
                question_data = request.data.copy()
                options_data = question_data.pop('options', [])
                
                required_fields = ['question_text', 'question_type']
                for field in required_fields:
                    if field not in question_data:
                        return validation_error({"error": f"Missing required field: {field}"})
                difficulty_level = question_data.get('difficulty_level', 'medium')
                if difficulty_level not in ['easy', 'medium', 'hard']:
                    difficulty_level = 'medium'
                question = Question.objects.create(
                    question_text=question_data.get('question_text'),
                    question_type=question_data.get('question_type'),
                    difficulty_level=difficulty_level,
                    exp_points=question_data.get('exp_points', 10),
                    explanation=question_data.get('explanation', ''),
                    is_active=question_data.get('is_active', True),
                    created_by=request.user
                )
                
                if options_data and question.question_type in ['mcq_single', 'mcq_multiple']:
                    for option_data in options_data:
                        Option.objects.create(
                            question=question,
                            option_text=option_data.get('option_text', ''),
                            is_correct=option_data.get('is_correct', False),
                            order=option_data.get('order', 1)
                        )
                
                max_order = MissionQuestion.objects.filter(mission=mission).aggregate(
                    max_order=Max('order')
                )['max_order'] or 0
                next_order = max_order + 1
                
                mission_question = MissionQuestion.objects.create(
                    mission=mission,
                    question=question,
                    order=next_order
                )
                
                api_logger.info(f"Question '{question.question_text[:50]}...' created and added to mission '{mission.title}' (Order: {next_order})")
                
                from gyaan_buddy.subjects.serializers import QuestionSerializer
                question_serializer = QuestionSerializer(question, context={'request': request})
                
                return created(
                    data={
                        'question': question_serializer.data,
                        'mission_question': {
                            'id': str(mission_question.id),
                            'mission_id': str(mission.id),
                            'question_id': str(question.id),
                            'order': mission_question.order
                        }
                    },
                    message="Question created and added to mission successfully"
                )
            except Exception as e:
                api_logger.error(f"Error creating mission question: {str(e)}", exc_info=True)
                return validation_error({"error": f"Failed to create question: {str(e)}"})

    @action(detail=True, methods=['get'], url_path='students_performance')
    def students_performance(self, request, pk=None):
        """
        Get the performance of all students who have taken a specific mission.
        
        Returns a list of students with their performance details including:
        - correct_answer_count
        - wrong_answer_count
        - percentage
        - user_id, user_name, class_name, subject_name
        - pass/fail status (pass if percentage > 40)
        - list of answers with question, answer, is_correct, correct_answer
        """
        try:
            try:
                mission = Mission.objects.select_related(
                    'account', 'subject'
                ).get(
                    pk=pk,
                    is_deleted=False
                )
            except Mission.DoesNotExist:
                return not_found({"error": "Mission not found"})
            
            api_logger.info(f"Students performance requested by {request.user.username} (ID: {request.user.id}) for mission '{mission.title}' (ID: {mission.id})")
            
            mission_questions = MissionQuestion.objects.filter(
                mission=mission
            ).select_related(
                'question'
            ).prefetch_related(
                'question__options'
            ).order_by('order')
            
            question_ids = [mq.question.id for mq in mission_questions]
            questions_map = {mq.question.id: mq.question for mq in mission_questions}
            
            if not question_ids:
                api_logger.info(f"No questions found in mission '{mission.title}'")
                return success(
                    data=[],
                    message="No questions found in this mission"
                )
            
            user_mission_progress = UserMissionProgress.objects.filter(
                mission=mission,
                status__in=['in_progress', 'completed']
            ).select_related(
                'account',
                'account__profile',
                'account__profile__student',
                'account__profile__student__class_instance'
            )
            
            students_performance_data = []
            
            for progress in user_mission_progress:
                user = progress.account
                
                from gyaan_buddy.subjects.models import Answer
                user_answers = Answer.objects.filter(
                    user=user,
                    question_id__in=question_ids
                ).select_related('question')
                
                answers_map = {answer.question_id: answer for answer in user_answers}
                
                correct_count = 0
                wrong_count = 0
                answers_list = []
                
                for question_id in question_ids:
                    question = questions_map.get(question_id)
                    if not question:
                        continue
                    
                    correct_options = question.options.filter(is_correct=True)
                    correct_answer_text = ", ".join([opt.option_text for opt in correct_options])
                    
                    user_answer = answers_map.get(question_id)
                    
                    if user_answer:
                        is_correct = user_answer.is_correct
                        if is_correct:
                            correct_count += 1
                        else:
                            wrong_count += 1
                        
                        answers_list.append({
                            "question": question.question_text,
                            "answer": user_answer.answer,
                            "is_correct": is_correct,
                            "correct_answer": correct_answer_text
                        })
                    else:
                        wrong_count += 1
                        answers_list.append({
                            "question": question.question_text,
                            "answer": None,
                            "is_correct": False,
                            "correct_answer": correct_answer_text
                        })
                
                total_questions = len(question_ids)
                percentage = round((correct_count / total_questions) * 100, 2) if total_questions > 0 else 0
                
                passed = percentage > 40
                
                user_name = user.get_full_name() or user.username
                class_name = None
                
                if hasattr(user, 'profile') and user.profile:
                    if hasattr(user.profile, 'student') and user.profile.student:
                        if user.profile.student.class_instance:
                            class_name = user.profile.student.class_instance.name
                
                subject_name = mission.subject.name if mission.subject else None
                
                student_data = {
                    "user_id": str(user.id),
                    "user_name": user_name,
                    "class_name": class_name,
                    "subject_name": subject_name,
                    "correct_answer_count": correct_count,
                    "wrong_answer_count": wrong_count,
                    "percentage": percentage,
                    "pass": passed,
                    "status": progress.status,
                    "answers": answers_list
                }
                
                students_performance_data.append(student_data)
            
            students_performance_data.sort(key=lambda x: x['percentage'], reverse=True)
            
            api_logger.info(f"Returned performance data for {len(students_performance_data)} students in mission '{mission.title}'")
            
            return success(
                data={
                    "mission_id": str(mission.id),
                    "mission_title": mission.title,
                    "total_questions": len(question_ids),
                    "total_students": len(students_performance_data),
                    "account_username": mission.account.username if mission.account else None,
                    "subject_name": mission.subject.name if mission.subject else None,
                    "students": students_performance_data
                },
                message="Students performance retrieved successfully"
            )
            
        except Exception as e:
            api_logger.error(f"Error retrieving students performance for mission {pk}: {str(e)}", exc_info=True)
            return validation_error({"error": f"Failed to retrieve students performance: {str(e)}"})

    @action(detail=False, methods=['post'], url_path='check_answer')
    def check_answer(self, request):
        """
        Check a user's answer for a mission question and calculate EXP earned.
        
        Request body:
        - tries: Number of attempts made
        - answer_id: The user's answer (option id or text)
        - mission_id: ID of the mission
        - is_correct: Whether the answer is correct
        - question_id: MissionQuestion ID
        """
        try:
            user = request.user
            
            tries = request.data.get('tries', 1)
            answer_id = request.data.get('answer_id')
            mission_id = request.data.get('mission_id')
            is_correct = request.data.get('is_correct', False)
            question_id = request.data.get('question_id')
            
            if not mission_id:
                return validation_error({"error": "mission_id is required"})
            if not question_id:
                return validation_error({"error": "question_id is required"})
            if answer_id is None:
                return validation_error({"error": "answer_id is required"})
            
            tries = max(1, int(tries))
            
            try:
                mission = Mission.objects.get(id=mission_id, is_deleted=False)
            except Mission.DoesNotExist:
                return not_found({"error": "Mission not found"})
            
            mission_question = None
            question = None
            try:
                mission_question = MissionQuestion.objects.select_related('question').get(id=question_id)
                question = mission_question.question
            except MissionQuestion.DoesNotExist:
                try:
                    mission_question = MissionQuestion.objects.select_related('question').get(
                        mission=mission,
                        question_id=question_id
                    )
                    question = mission_question.question
                except MissionQuestion.DoesNotExist:
                    return not_found({"error": "Mission question not found"})
            
            prev_exp = 0
            if hasattr(user, 'profile') and hasattr(user.profile, 'student'):
                prev_exp = user.profile.student.total_exp

            from gyaan_buddy.subjects.models import Answer, Option
            import uuid as uuid_module

            already_correct = Answer.objects.filter(
                user=user, question=question, is_correct=True
            ).exists()

            if not already_correct and is_correct and tries == 1:
                exp_earned = 2
            elif not already_correct and is_correct and tries == 2:
                exp_earned = 1
            else:
                exp_earned = 0

            current_exp = prev_exp + exp_earned

            option_ids = []
            if isinstance(answer_id, list):
                option_ids = [str(uid).strip() for uid in answer_id]
            elif isinstance(answer_id, str) and ',' in answer_id:
                option_ids = [uid.strip() for uid in answer_id.split(',')]
            else:
                option_ids = [str(answer_id).strip()] if answer_id else []

            api_logger.info(f"Processing answer_id: {answer_id} (type: {type(answer_id).__name__}), option_ids: {option_ids}")

            answer_text = ','.join(option_ids) if len(option_ids) > 1 else (option_ids[0] if option_ids else str(answer_id))

            try:
                valid_uuids = []
                for uid in option_ids:
                    try:
                        uuid_module.UUID(uid)
                        valid_uuids.append(uid)
                    except (ValueError, AttributeError):
                        pass

                if valid_uuids:
                    options = Option.objects.filter(id__in=valid_uuids)
                    if options.exists():
                        answer_text = ', '.join([opt.option_text for opt in options])
            except (Option.DoesNotExist, ValueError, AttributeError) as e:
                api_logger.warning(f"Could not get option text for answer_id {answer_id}: {str(e)}")
                pass

            api_logger.info(f"Creating/updating Answer: user={user.username}, question={question.id}, answer_text={answer_text[:100] if len(answer_text) > 100 else answer_text}")
            answer, created = Answer.objects.update_or_create(
                user=user,
                question=question,
                defaults={
                    'is_correct': is_correct,
                    'answer': answer_text,
                    'tries': tries,
                    'prev_exp': prev_exp,
                    'current_Exp': current_exp,
                    'from_mission': True
                }
            )
            api_logger.info(f"Answer {'created' if created else 'updated'}: id={answer.id}")

            if exp_earned > 0 and hasattr(user, 'profile') and hasattr(user.profile, 'student'):
                user.profile.student.total_exp = current_exp
                user.profile.student.save()
                api_logger.info(f"User {user.username} earned {exp_earned} exp from mission (tries: {tries})")
            
            mission_completed = False
            try:
                user_progress, _ = UserMissionProgress.objects.get_or_create(
                    account=user,
                    mission=mission,
                    defaults={'status': 'in_progress', 'started_at': timezone.now()}
                )
                user_progress.current_question = question
                if user_progress.status == 'not_started':
                    user_progress.status = 'in_progress'
                    user_progress.started_at = timezone.now()
                user_progress.save()
                
                if user_progress.status != 'completed':
                    mission_question_ids = list(
                        mission.mission_questions.filter(
                            question__is_active=True,
                            question__is_deleted=False
                        ).values_list('question_id', flat=True)
                    )
                    total_questions = len(mission_question_ids)
                    
                    answered_count = Answer.objects.filter(
                        user=user,
                        question_id__in=mission_question_ids
                    ).count()
                    
                    api_logger.info(f"Mission completion check: {answered_count}/{total_questions} questions answered for mission '{mission.title}' by user {user.username}")
                    
                    if answered_count >= total_questions and total_questions > 0:
                        complete_mission(user, mission)
                        mission_completed = True
                        api_logger.info(f"Mission '{mission.title}' completed by user {user.username}")
            except Exception as e:
                api_logger.error(f"Could not update mission progress: {str(e)}", exc_info=True)
            
            api_logger.info(f"User {user.username} answered question {question.id} in mission {mission.id} - Correct: {is_correct}, Tries: {tries}, EXP earned: {exp_earned}")
            
            return success(
                data={
                    "answer_id": str(answer.id),
                    "is_correct": is_correct,
                    "tries": tries,
                    "exp_earned": exp_earned,
                    "prev_exp": prev_exp,
                    "current_exp": current_exp,
                    "question_id": str(question.id),
                    "mission_id": str(mission.id),
                    "mission_completed": mission_completed
                },
                message="Mission completed successfully!" if mission_completed else "Answer checked successfully"
            )
            
        except Exception as e:
            api_logger.error(f"Error checking answer: {str(e)}", exc_info=True)
            return validation_error({"error": f"Failed to check answer: {str(e)}"})


class CompetitionViewSet(viewsets.ModelViewSet):
    """ViewSet for Competition model."""
    queryset = Competition.objects.all()
    serializer_class = CompetitionSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Return optimized queryset with related data, filtered by user's school."""
        user = self.request.user
        
        return Competition.objects.select_related(
            'created_by',
            'subject',
            'chapter',
        ).prefetch_related(
            'questions',
            'user_progress',
        ).filter(
            is_active=True,
            is_deleted=False,
            subject__classes__school=user.profile.school if hasattr(user, 'profile') else None
        ).distinct()
    
    def get_serializer_class(self):
        """Return appropriate serializer class based on action."""
        if self.action == 'create':
            return CompetitionCreateSerializer
        return CompetitionSerializer
    
    def list(self, request, *args, **kwargs):
        """List competitions with logging."""
        api_logger.info(f"Competition list requested by {request.user.username} (ID: {request.user.id}) from {request.META.get('REMOTE_ADDR', 'unknown')}")
        
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        api_logger.info(f"Competition list returned {len(serializer.data)} competitions")
        
        return success(
            data=serializer.data,
            message="Competitions retrieved successfully"
        )
    
    def create(self, request, *args, **kwargs):
        """Create a competition with logging."""
        api_logger.info(f"Competition creation requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")
        
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            competition = serializer.save(created_by=request.user)
            api_logger.info(f"Competition created successfully: {competition.title} (ID: {competition.id}) by {request.user.username}")
            return created(
                data=CompetitionSerializer(competition).data,
                message="Competition created successfully"
            )
        api_logger.warning(f"Competition creation failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)
    
    @action(detail=False, methods=['post'])
    def join(self, request):
        """Join a competition using a code."""
        serializer = CompetitionJoinSerializer(data=request.data)
        if serializer.is_valid():
            code = serializer.validated_data['code']
            user = request.user
            
            try:
                competition = Competition.objects.get(code=code, is_active=True, is_deleted=False)
                
                if UserCompetitionProgress.objects.filter(account=user, competition=competition).exists():
                    return validation_error({"error": "Already joined this competition"})
                
                progress = UserCompetitionProgress.objects.create(
                    account=user,
                    competition=competition,
                    status='not_started'
                )
                
                api_logger.info(f"User {user.username} joined competition {competition.title}")
                return success(
                    data=UserCompetitionProgressSerializer(progress).data,
                    message="Successfully joined competition"
                )
            except Competition.DoesNotExist:
                return validation_error({"error": "Competition not found"})
        
        return validation_error(serializer.errors)
    
    @action(detail=True, methods=['post'])
    def start_competition(self, request, pk=None):
        """Start a competition for the current user."""
        try:
            competition = self.get_object()
            user = request.user
            
            try:
                progress = UserCompetitionProgress.objects.get(account=user, competition=competition)
            except UserCompetitionProgress.DoesNotExist:
                return validation_error({"error": "You must join the competition first"})
            
            if progress.status == 'completed':
                return validation_error({"error": "Competition already completed"})
            elif progress.status == 'in_progress':
                return validation_error({"error": "Competition already in progress"})
            
            progress.status = 'in_progress'
            progress.started_at = timezone.now()
            progress.save()
            
            api_logger.info(f"User {user.username} started competition {competition.title}")
            return success(
                data=UserCompetitionProgressSerializer(progress).data,
                message="Competition started successfully"
            )
        except Exception as e:
            api_logger.error(f"Error starting competition: {str(e)}")
            return validation_error({"error": "Failed to start competition"})


class UserMissionProgressViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for UserMissionProgress model (read-only)."""
    serializer_class = UserMissionProgressSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Return user's own mission progress."""
        return UserMissionProgress.objects.select_related(
            'account', 'mission', 'current_question'
        ).filter(account=self.request.user)


class UserCompetitionProgressViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for UserCompetitionProgress model (read-only)."""
    serializer_class = UserCompetitionProgressSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Return user's own competition progress."""
        return UserCompetitionProgress.objects.select_related(
            'account', 'competition', 'current_question'
        ).filter(account=self.request.user)


class UserModuleProgressViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for UserModuleProgress model."""
    serializer_class = UserModuleProgressSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        queryset = UserModuleProgress.objects.filter(account=user).select_related(
            'account', 'module', 'module__subject', 'current_question'
        ).prefetch_related('module__chapters')
        return queryset
    
    def list(self, request, *args, **kwargs):
        """List user's module progress with filtering options."""
        queryset = self.get_queryset()
        
        status = request.query_params.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        subject_id = request.query_params.get('subject')
        if subject_id:
            queryset = queryset.filter(module__subject_id=subject_id)
        
        module_id = request.query_params.get('module')
        if module_id:
            queryset = queryset.filter(module_id=module_id)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'status': 'success',
            'data': serializer.data,
            'count': queryset.count()
        })


class UserChapterProgressViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for UserChapterProgress model."""
    serializer_class = UserChapterProgressSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        queryset = UserChapterProgress.objects.filter(account=user).select_related(
            'account', 'chapter', 'chapter__module', 'chapter__module__subject', 'current_question'
        )
        return queryset
    
    def list(self, request, *args, **kwargs):
        """List user's chapter progress with filtering options."""
        queryset = self.get_queryset()
        
        status = request.query_params.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        module_id = request.query_params.get('module')
        if module_id:
            queryset = queryset.filter(chapter__module_id=module_id)
        
        subject_id = request.query_params.get('subject')
        if subject_id:
            queryset = queryset.filter(chapter__module__subject_id=subject_id)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'status': 'success',
            'data': serializer.data,
            'count': queryset.count()
        })


class NotificationViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Notification model.
    
    Provides:
    - list: Returns all notifications for the current user
    - student_notifications: Returns notifications for all students in classes the teacher teaches
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        from .serializers import NotificationSerializer, NotificationCreateSerializer
        if self.action == 'create':
            return NotificationCreateSerializer
        return NotificationSerializer
    
    def get_queryset(self):
        """Return notifications for the current user."""
        from .models import Notification
        user = self.request.user
        return Notification.objects.filter(user=user).select_related('user').order_by('-created_at')
    
    def list(self, request, *args, **kwargs):
        """
        List all notifications for the current user.
        
        Query Parameters:
        - is_read: Filter by read status (true/false)
        - type: Filter by notification type (module, subject, user, mission, competition)
        - limit: Limit number of results
        """
        api_logger.info(f"Notifications list requested by {request.user.username} (ID: {request.user.id})")
        
        queryset = self.get_queryset()
        
        is_read = request.query_params.get('is_read')
        if is_read is not None:
            is_read_bool = is_read.lower() in ('true', '1', 'yes')
            queryset = queryset.filter(is_read=is_read_bool)
        
        notification_type = request.query_params.get('type')
        if notification_type:
            queryset = queryset.filter(type=notification_type)
        
        limit = request.query_params.get('limit')
        if limit:
            try:
                queryset = queryset[:int(limit)]
            except ValueError:
                pass
        
        serializer = self.get_serializer(queryset, many=True)
        
        from .models import Notification
        unread_count = Notification.objects.filter(user=request.user, is_read=False).count()
        
        api_logger.info(f"Notifications list returned {len(serializer.data)} notifications for {request.user.username}")
        
        return success(
            data={
                'notifications': serializer.data,
                'total_count': self.get_queryset().count(),
                'unread_count': unread_count
            },
            message="Notifications retrieved successfully"
        )
    
    def create(self, request, *args, **kwargs):
        """Create a new notification."""
        api_logger.info(f"Notification creation requested by {request.user.username}")
        
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            notification = serializer.save()
            from .serializers import NotificationSerializer
            api_logger.info(f"Notification created: {notification.notification_id}")
            return created(
                data=NotificationSerializer(notification).data,
                message="Notification created successfully"
            )
        
        api_logger.warning(f"Notification creation failed: {serializer.errors}")
        return validation_error(serializer.errors)
    
    @action(detail=True, methods=['post'], url_path='mark_read')
    def mark_read(self, request, pk=None):
        """Mark a notification as read."""
        try:
            notification = self.get_object()
            notification.mark_as_read()
            api_logger.info(f"Notification {notification.id} marked as read by {request.user.username}")
            return success(
                data=self.get_serializer(notification).data,
                message="Notification marked as read"
            )
        except Exception as e:
            api_logger.error(f"Error marking notification as read: {str(e)}")
            return validation_error({"error": "Failed to mark notification as read"})
    
    @action(detail=False, methods=['post'], url_path='mark_all_read')
    def mark_all_read(self, request):
        """Mark all notifications as read for the current user."""
        try:
            from .models import Notification
            updated_count = Notification.objects.filter(
                user=request.user, 
                is_read=False
            ).update(
                is_read=True, 
                read_at=timezone.now()
            )
            api_logger.info(f"Marked {updated_count} notifications as read for {request.user.username}")
            return success(
                data={'updated_count': updated_count},
                message=f"Marked {updated_count} notifications as read"
            )
        except Exception as e:
            api_logger.error(f"Error marking all notifications as read: {str(e)}")
            return validation_error({"error": "Failed to mark notifications as read"})
    
    @action(detail=False, methods=['get'], url_path='student_notifications')
    def student_notifications(self, request):
        """
        Get notifications for all students in classes that the current teacher teaches.
        
        This endpoint is for teachers to view notifications of students in their classes.
        
        Query Parameters:
        - class_id: Filter by specific class ID
        - is_read: Filter by read status (true/false)
        - type: Filter by notification type
        - limit: Limit number of results
        """
        from .models import Notification, Teacher, Student
        
        user = request.user
        api_logger.info(f"Student notifications requested by {user.username} (ID: {user.id})")
        
        if not hasattr(user, 'profile') or user.profile.user_type != 'teacher':
            api_logger.warning(f"Non-teacher user {user.username} attempted to access student notifications")
            return forbidden({"error": "Only teachers can access student notifications"})
        
        try:
            teacher_profile = user.profile.teacher_profile
        except Exception as e:
            api_logger.error(f"Error getting teacher profile: {str(e)}")
            return validation_error({"error": "Teacher profile not found"})
        
        teacher_assignments = Teacher.objects.filter(
            teacher=teacher_profile
        ).select_related('class_instance')
        
        class_ids = [ta.class_instance.id for ta in teacher_assignments]
        
        if not class_ids:
            api_logger.info(f"Teacher {user.username} has no class assignments")
            return success(
                data={
                    'notifications': [],
                    'total_count': 0,
                    'unread_count': 0,
                    'classes': []
                },
                message="No classes assigned to this teacher"
            )
        
        students = Student.objects.filter(
            class_instance_id__in=class_ids,
            is_deleted=False
        ).select_related('user_profile__account')
        
        student_user_ids = [s.user_profile.account.id for s in students]
        
        queryset = Notification.objects.filter(
            user_id__in=student_user_ids
        ).select_related('user').order_by('-created_at')
        
        class_filter = request.query_params.get('class_id')
        if class_filter:
            filtered_students = Student.objects.filter(
                class_instance_id=class_filter,
                is_deleted=False
            ).select_related('user_profile__account')
            filtered_user_ids = [s.user_profile.account.id for s in filtered_students]
            queryset = queryset.filter(user_id__in=filtered_user_ids)
        
        is_read = request.query_params.get('is_read')
        if is_read is not None:
            is_read_bool = is_read.lower() in ('true', '1', 'yes')
            queryset = queryset.filter(is_read=is_read_bool)
        
        notification_type = request.query_params.get('type')
        if notification_type:
            queryset = queryset.filter(type=notification_type)
        
        total_count = queryset.count()
        unread_count = queryset.filter(is_read=False).count()
        
        limit = request.query_params.get('limit')
        if limit:
            try:
                queryset = queryset[:int(limit)]
            except ValueError:
                pass
        
        serializer = self.get_serializer(queryset, many=True)
        
        classes_info = [
            {
                'id': str(ta.class_instance.id),
                'name': ta.class_instance.name,
                'subject': ta.subject.name if ta.subject else None
            }
            for ta in teacher_assignments
        ]
        
        api_logger.info(f"Student notifications returned {len(serializer.data)} notifications for teacher {user.username}")
        
        return success(
            data={
                'notifications': serializer.data,
                'total_count': total_count,
                'unread_count': unread_count,
                'classes': classes_info,
                'total_students': len(student_user_ids)
            },
            message="Student notifications retrieved successfully"
        )



class AnalyticsViewSet(viewsets.ViewSet):
    """
    ViewSet for comprehensive analytics across the platform.
    
    Provides analytics endpoints for:
    - Teachers: Class performance, student performance, subject-wise analytics
    - Students: Personal progress, subject performance, weak areas
    - Admins: School overview, grade-wise, class-wise analytics
    - Mission & Competition analytics
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def _get_user_type(self, user):
        """Get user type from profile."""
        if hasattr(user, 'profile'):
            return user.profile.user_type
        return None
    
    def _calculate_accuracy(self, correct, total):
        """Calculate accuracy percentage."""
        if total == 0:
            return 0.0
        return round((correct / total) * 100, 2)
    
    
    @action(detail=False, methods=['get'], url_path='teacher/overview')
    def teacher_overview(self, request):
        """
        Get overview analytics for a teacher.
        Returns summary of all classes and subjects assigned to the teacher.
        """
        try:
            user = request.user
            api_logger.info(f"Teacher overview analytics requested by {user.username}")
            
            user_type = self._get_user_type(user)
            if user_type not in ['teacher', 'administrator']:
                return forbidden({"error": "Only teachers can access this endpoint"})
            
            if not hasattr(user, 'profile') or not hasattr(user.profile, 'teacher_profile'):
                return validation_error({"error": "Teacher profile not found"})
            
            teacher_profile = user.profile.teacher_profile
            
            from .models import Teacher, Student
            from gyaan_buddy.subjects.models import Answer
            
            assignments = Teacher.objects.filter(
                teacher=teacher_profile
            ).select_related('class_instance', 'subject', 'class_instance__grade')
            
            classes_data = []
            total_students = 0
            total_questions_attempted = 0
            total_correct = 0
            
            for assignment in assignments:
                class_instance = assignment.class_instance
                subject = assignment.subject
                
                students = Student.objects.filter(
                    class_instance=class_instance,
                    is_deleted=False
                )
                student_count = students.count()
                total_students += student_count
                
                student_account_ids = list(students.values_list('user_profile__account_id', flat=True))
                
                answers = Answer.objects.filter(
                    user_id__in=student_account_ids,
                    question__module_contents__chapter__module__subject=subject
                )
                
                questions_attempted = answers.count()
                correct_answers = answers.filter(is_correct=True).count()
                
                total_questions_attempted += questions_attempted
                total_correct += correct_answers
                
                accuracy = self._calculate_accuracy(correct_answers, questions_attempted)
                
                modules_completed = UserModuleProgress.objects.filter(
                    account_id__in=student_account_ids,
                    module__subject=subject,
                    status='completed'
                ).count()
                
                classes_data.append({
                    'class_id': str(class_instance.id),
                    'class_name': class_instance.name,
                    'grade_name': class_instance.grade.name if class_instance.grade else None,
                    'subject_id': str(subject.id),
                    'subject_name': subject.name,
                    'student_count': student_count,
                    'questions_attempted': questions_attempted,
                    'correct_answers': correct_answers,
                    'accuracy_percentage': accuracy,
                    'modules_completed': modules_completed
                })
            
            overall_accuracy = self._calculate_accuracy(total_correct, total_questions_attempted)
            
            return success(
                data={
                    'teacher_id': str(teacher_profile.id),
                    'teacher_name': user.get_full_name() or user.username,
                    'total_classes': len(assignments),
                    'total_students': total_students,
                    'total_questions_attempted': total_questions_attempted,
                    'overall_accuracy': overall_accuracy,
                    'classes': classes_data
                },
                message="Teacher overview analytics retrieved successfully"
            )
            
        except Exception as e:
            api_logger.error(f"Error in teacher_overview: {str(e)}", exc_info=True)
            return validation_error({"error": f"Failed to retrieve analytics: {str(e)}"})
    
    @action(detail=False, methods=['get'], url_path='teacher/class/(?P<class_id>[^/.]+)/subject/(?P<subject_id>[^/.]+)')
    def teacher_class_subject(self, request, class_id=None, subject_id=None):
        """
        Get detailed analytics for a specific class and subject combination.
        """
        try:
            user = request.user
            api_logger.info(f"Teacher class-subject analytics requested by {user.username} for class {class_id}, subject {subject_id}")
            
            from gyaan_buddy.subjects.models import Answer, Module, ModuleChapter

            try:
                class_instance = Class.objects.get(id=class_id)
                subject = Subject.objects.get(id=subject_id)
            except (Class.DoesNotExist, Subject.DoesNotExist):
                return not_found({"error": "Class or Subject not found"})
            
            students = Student.objects.filter(
                class_instance=class_instance,
                is_deleted=False
            ).select_related('user_profile__account', 'level')
            
            student_account_ids = list(students.values_list('user_profile__account_id', flat=True))
            
            modules = Module.objects.filter(subject=subject, is_active=True)
            
            answers = Answer.objects.filter(
                user_id__in=student_account_ids,
                question__module_contents__chapter__module__subject=subject
            ).select_related('question')
            
            total_answers = answers.count()
            correct_answers = answers.filter(is_correct=True).count()
            overall_accuracy = self._calculate_accuracy(correct_answers, total_answers)
            
            difficulty_stats = {}
            for difficulty in ['easy', 'medium', 'hard']:
                diff_answers = answers.filter(question__difficulty_level=difficulty)
                diff_total = diff_answers.count()
                diff_correct = diff_answers.filter(is_correct=True).count()
                difficulty_stats[difficulty] = {
                    'total': diff_total,
                    'correct': diff_correct,
                    'accuracy': self._calculate_accuracy(diff_correct, diff_total)
                }

            due_chapter_ids = list(
                ModuleChapter.objects.filter(
                    module__subject=subject,
                    module__is_active=True,
                    is_deleted=False,
                    due_date__isnull=False
                ).values_list('id', flat=True)
            )
            total_due_chapters = len(due_chapter_ids)
            attempted_due_chapter_pairs = set()
            if total_due_chapters > 0 and student_account_ids:
                attempted_due_chapter_pairs = set(
                    Answer.objects.filter(
                        user_id__in=student_account_ids,
                        chapter_id__in=due_chapter_ids
                    )
                    .values_list('user_id', 'chapter_id')
                    .distinct()
                )
            
            module_progress = []
            for module in modules:
                module_answers = answers.filter(
                    question__module_contents__chapter__module=module
                )
                module_total = module_answers.count()
                module_correct = module_answers.filter(is_correct=True).count()
                
                progress_records = UserModuleProgress.objects.filter(
                    account_id__in=student_account_ids,
                    module=module
                )
                avg_progress = progress_records.aggregate(avg=Avg('percentage'))['avg'] or 0
                completed_count = progress_records.filter(status='completed').count()
                
                module_progress.append({
                    'module_id': str(module.id),
                    'module_name': module.name,
                    'questions_attempted': module_total,
                    'correct_answers': module_correct,
                    'accuracy': self._calculate_accuracy(module_correct, module_total),
                    'avg_completion': round(avg_progress, 2),
                    'students_completed': completed_count
                })
            
            student_performance = []
            for student in students:
                student_answers = answers.filter(user_id=student.user_profile.account_id)
                s_total = student_answers.count()
                s_correct = student_answers.filter(is_correct=True).count()
                due_chapters_attempted = 0
                attempt_rate = 0
                if total_due_chapters > 0:
                    due_chapters_attempted = sum(
                        1
                        for chapter_id in due_chapter_ids
                        if (student.user_profile.account_id, chapter_id) in attempted_due_chapter_pairs
                    )
                    attempt_rate = round((due_chapters_attempted / total_due_chapters) * 100)
                
                student_progress = UserModuleProgress.objects.filter(
                    account_id=student.user_profile.account_id,
                    module__subject=subject
                ).aggregate(avg=Avg('percentage'))['avg'] or 0
                
                student_performance.append({
                    'student_id': str(student.id),
                    'student_name': student.user_profile.account.get_full_name() or student.user_profile.account.username,
                    'username': student.user_profile.account.username,
                    'roll_number': student.roll_number,
                    'questions_attempted': s_total,
                    'correct_answers': s_correct,
                    'accuracy': self._calculate_accuracy(s_correct, s_total),
                    'due_chapters_attempted': due_chapters_attempted,
                    'total_due_chapters': total_due_chapters,
                    'attempt_rate': attempt_rate,
                    'avg_progress': round(student_progress, 2),
                    'total_exp': student.total_exp,
                    'level': student.level.name if student.level else 1
                })
            
            student_performance.sort(key=lambda x: x['accuracy'], reverse=True)
            class_attempt_rate = round(
                sum(student['attempt_rate'] for student in student_performance) / len(student_performance)
            ) if student_performance else 0
            
            struggling_students = [
                s for s in student_performance 
                if s['accuracy'] < 50 and s['questions_attempted'] > 0
            ]
            
            return success(
                data={
                    'class_id': str(class_instance.id),
                    'class_name': class_instance.name,
                    'subject_id': str(subject.id),
                    'subject_name': subject.name,
                    'total_students': len(students),
                    'overall_stats': {
                        'total_questions_attempted': total_answers,
                        'correct_answers': correct_answers,
                        'accuracy': overall_accuracy,
                        'total_due_chapters': total_due_chapters,
                        'attempt_rate': class_attempt_rate
                    },
                    'difficulty_breakdown': difficulty_stats,
                    'module_progress': module_progress,
                    'student_performance': student_performance,
                    'struggling_students': struggling_students
                },
                message="Class-subject analytics retrieved successfully"
            )
            
        except Exception as e:
            api_logger.error(f"Error in teacher_class_subject: {str(e)}", exc_info=True)
            return validation_error({"error": f"Failed to retrieve analytics: {str(e)}"})
    
    @action(detail=False, methods=['get'], url_path='teacher/student/(?P<student_id>[^/.]+)')
    def teacher_student_detail(self, request, student_id=None):
        """
        Get detailed analytics for a specific student.
        """
        try:
            user = request.user
            api_logger.info(f"Teacher student detail analytics requested by {user.username} for student {student_id}")
            
            from gyaan_buddy.subjects.models import Answer

            try:
                student = Student.objects.select_related(
                    'user_profile__account', 'level', 'class_instance', 'class_instance__grade'
                ).get(id=student_id, is_deleted=False)
            except Student.DoesNotExist:
                return not_found({"error": "Student not found"})
            
            account = student.user_profile.account
            
            answers = Answer.objects.filter(user=account).select_related('question')
            
            total_answers = answers.count()
            correct_answers = answers.filter(is_correct=True).count()
            first_try_correct = answers.filter(is_correct=True, tries=1).count()
            
            difficulty_stats = {}
            for difficulty in ['easy', 'medium', 'hard']:
                diff_answers = answers.filter(question__difficulty_level=difficulty)
                diff_total = diff_answers.count()
                diff_correct = diff_answers.filter(is_correct=True).count()
                difficulty_stats[difficulty] = {
                    'total': diff_total,
                    'correct': diff_correct,
                    'accuracy': self._calculate_accuracy(diff_correct, diff_total)
                }
            
            hots_answers = answers.filter(question__is_hots=True)
            hots_total = hots_answers.count()
            hots_correct = hots_answers.filter(is_correct=True).count()
            
            subject_performance = []
            if hasattr(account, 'profile') and account.profile and account.profile.school:
                subjects = Subject.objects.filter(
                    classes__school=account.profile.school,
                    is_active=True
                ).distinct()
            else:
                subjects = Subject.objects.filter(is_active=True)
            
            for subject in subjects:
                subj_answers = answers.filter(
                    Q(chapter__module__subject=subject) |
                    Q(question__module_contents__chapter__module__subject=subject)
                )
                s_total = subj_answers.count()
                s_correct = subj_answers.filter(is_correct=True).count()
                total_subject_questions = ModuleContent.objects.filter(
                    chapter__module__subject=subject,
                    chapter__module__is_active=True,
                    chapter__is_deleted=False,
                    content_type='question',
                    is_deleted=False,
                ).exclude(question__isnull=True).values('question_id').distinct().count()
                attempt_rate = round((s_total / total_subject_questions) * 100) if total_subject_questions else 0
                
                module_progress = UserModuleProgress.objects.filter(
                    account=account,
                    module__subject=subject
                )
                avg_progress = module_progress.aggregate(avg=Avg('percentage'))['avg'] or 0
                modules_completed = module_progress.filter(status='completed').count()
                total_modules = module_progress.count()
                
                chapter_progress = UserChapterProgress.objects.filter(
                    account=account,
                    chapter__module__subject=subject
                )
                chapters_completed = chapter_progress.filter(status='completed').count()
                
                subject_performance.append({
                    'subject_id': str(subject.id),
                    'subject_name': subject.name,
                    'questions_attempted': s_total,
                    'correct_answers': s_correct,
                    'accuracy': self._calculate_accuracy(s_correct, s_total),
                    'avg_module_progress': round(avg_progress, 2),
                    'modules_completed': modules_completed,
                    'total_modules': total_modules,
                    'chapters_completed': chapters_completed
                })
            
            mission_progress = UserMissionProgress.objects.filter(
                mission__account=account
            ).select_related('mission')
            
            missions_completed = mission_progress.filter(status='completed').count()
            total_missions = mission_progress.count()
            avg_mission_exp = mission_progress.aggregate(avg=Avg('exp_earned'))['avg'] or 0
            
            comp_progress = UserCompetitionProgress.objects.filter(
                account=account
            ).select_related('competition')
            
            competitions_participated = comp_progress.count()
            competitions_completed = comp_progress.filter(status='completed').count()
            avg_comp_score = comp_progress.aggregate(avg=Avg('score'))['avg'] or 0
            
            weak_areas = [
                {
                    'subject_id': s['subject_id'],
                    'subject_name': s['subject_name'],
                    'accuracy': s['accuracy'],
                    'recommendation': f"Focus on {s['subject_name']} - practice more questions"
                }
                for s in subject_performance
                if s['accuracy'] < 60 and s['questions_attempted'] > 0
            ]
            
            return success(
                data={
                    'student_id': str(student.id),
                    'student_name': account.get_full_name() or account.username,
                    'username': account.username,
                    'roll_number': student.roll_number,
                    'admission_number': student.admission_number,
                    'class_name': student.class_instance.name if student.class_instance else None,
                    'grade_name': student.class_instance.grade.name if student.class_instance and student.class_instance.grade else None,
                    'overall_stats': {
                        'total_questions_attempted': total_answers,
                        'correct_answers': correct_answers,
                        'accuracy': self._calculate_accuracy(correct_answers, total_answers),
                        'first_attempt_success_rate': self._calculate_accuracy(first_try_correct, total_answers),
                        'avg_attempts': round(answers.aggregate(avg=Avg('tries'))['avg'] or 0, 2)
                    },
                    'exp_and_level': {
                        'total_exp': student.total_exp,
                        'current_level': student.level.name if student.level else 1,
                        'rewards': student.rewards,
                        'exp_to_next_level': student.get_exp_to_next_level()
                    },
                    'difficulty_breakdown': difficulty_stats,
                    'hots_performance': {
                        'total': hots_total,
                        'correct': hots_correct,
                        'accuracy': self._calculate_accuracy(hots_correct, hots_total)
                    },
                    'subject_performance': subject_performance,
                    'mission_stats': {
                        'total_missions': total_missions,
                        'completed': missions_completed,
                        'avg_exp_earned': round(avg_mission_exp, 2)
                    },
                    'competition_stats': {
                        'participated': competitions_participated,
                        'completed': competitions_completed,
                        'avg_score': round(avg_comp_score, 2)
                    },
                    'weak_areas': weak_areas
                },
                message="Student detail analytics retrieved successfully"
            )
            
        except Exception as e:
            api_logger.error(f"Error in teacher_student_detail: {str(e)}", exc_info=True)
            return validation_error({"error": f"Failed to retrieve analytics: {str(e)}"})
    
    
    @action(detail=False, methods=['get'], url_path='student/my-progress')
    def student_my_progress(self, request):
        """
        Get personal progress analytics for the logged-in student.
        """
        try:
            user = request.user
            api_logger.info(f"Student my-progress analytics requested by {user.username}")

            from gyaan_buddy.subjects.models import Answer

            if not hasattr(user, 'profile') or not hasattr(user.profile, 'student'):
                return validation_error({"error": "Student profile not found"})
            
            student = user.profile.student
            
            answers = Answer.objects.filter(user=user).select_related('question')
            
            total_answers = answers.count()
            correct_answers = answers.filter(is_correct=True).count()
            
            module_progress = UserModuleProgress.objects.filter(account=user)
            modules_completed = module_progress.filter(status='completed').count()
            modules_in_progress = module_progress.filter(status='in_progress').count()
            total_modules = module_progress.count()
            avg_module_completion = module_progress.aggregate(avg=Avg('percentage'))['avg'] or 0
            
            chapter_progress = UserChapterProgress.objects.filter(account=user)
            chapters_completed = chapter_progress.filter(status='completed').count()
            chapters_in_progress = chapter_progress.filter(status='in_progress').count()
            
            subject_progress = []
            if hasattr(user, 'profile') and user.profile and user.profile.school:
                subjects = Subject.objects.filter(
                    classes__school=user.profile.school,
                    is_active=True
                ).distinct()
            else:
                subjects = Subject.objects.filter(is_active=True)
            
            for subject in subjects:
                subj_answers = answers.filter(
                    Q(chapter__module__subject=subject) |
                    Q(question__module_contents__chapter__module__subject=subject)
                )
                s_total = subj_answers.count()
                s_correct = subj_answers.filter(is_correct=True).count()
                total_subject_questions = ModuleContent.objects.filter(
                    chapter__module__subject=subject,
                    chapter__module__is_active=True,
                    chapter__is_deleted=False,
                    content_type='question',
                    is_deleted=False,
                ).exclude(question__isnull=True).values('question_id').distinct().count()
                attempt_rate = round((s_total / total_subject_questions) * 100) if total_subject_questions else 0
                
                avg_tries = subj_answers.aggregate(avg=Avg('tries'))['avg'] or 0
                first_try_correct = subj_answers.filter(is_correct=True, tries=1).count()
                first_try_accuracy = self._calculate_accuracy(first_try_correct, s_total)
                
                subj_module_progress = UserModuleProgress.objects.filter(
                    account=user,
                    module__subject=subject
                )
                avg_progress = subj_module_progress.aggregate(avg=Avg('percentage'))['avg'] or 0
                
                subject_progress.append({
                    'subject_id': str(subject.id),
                    'subject_name': subject.name,
                    'subject_code': subject.code,
                    'color': subject.color,
                    'questions_attempted': s_total,
                    'correct_answers': s_correct,
                    'accuracy': self._calculate_accuracy(s_correct, s_total),
                    'attempt_rate': attempt_rate,
                    'avg_tries': round(avg_tries, 2),
                    'first_try_correct': first_try_correct,
                    'first_try_accuracy': first_try_accuracy,
                    'avg_completion': round(avg_progress, 2)
                })
            
            mission_stats = UserMissionProgress.objects.filter(mission__account=user)
            comp_stats = UserCompetitionProgress.objects.filter(account=user)
            
            overall_avg_tries = answers.aggregate(avg=Avg('tries'))['avg'] or 0
            overall_first_try_correct = answers.filter(is_correct=True, tries=1).count()
            overall_first_try_accuracy = self._calculate_accuracy(overall_first_try_correct, total_answers)
            
            return success(
                data={
                    'student_id': str(student.id),
                    'student_name': user.get_full_name() or user.username,
                    'class_name': student.class_instance.name if student.class_instance else None,
                    'overall_stats': {
                        'total_questions_attempted': total_answers,
                        'correct_answers': correct_answers,
                        'wrong_answers': total_answers - correct_answers,
                        'accuracy': self._calculate_accuracy(correct_answers, total_answers),
                        'avg_tries': round(overall_avg_tries, 2),
                        'first_try_correct': overall_first_try_correct,
                        'first_try_accuracy': overall_first_try_accuracy
                    },
                    'exp_and_level': {
                        'total_exp': student.total_exp,
                        'current_level': student.level.name if student.level else 1,
                        'rewards': student.rewards,
                        'exp_to_next_level': student.get_exp_to_next_level()
                    },
                    'module_progress': {
                        'completed': modules_completed,
                        'in_progress': modules_in_progress,
                        'total': total_modules,
                        'avg_completion': round(avg_module_completion, 2)
                    },
                    'chapter_progress': {
                        'completed': chapters_completed,
                        'in_progress': chapters_in_progress
                    },
                    'subject_progress': subject_progress,
                    'mission_stats': {
                        'total': mission_stats.count(),
                        'completed': mission_stats.filter(status='completed').count(),
                        'total_exp_earned': mission_stats.aggregate(total=models.Sum('exp_earned'))['total'] or 0
                    },
                    'competition_stats': {
                        'participated': comp_stats.count(),
                        'completed': comp_stats.filter(status='completed').count(),
                        'total_exp_earned': comp_stats.aggregate(total=models.Sum('exp_earned'))['total'] or 0
                    }
                },
                message="Student progress analytics retrieved successfully"
            )
            
        except Exception as e:
            api_logger.error(f"Error in student_my_progress: {str(e)}", exc_info=True)
            return validation_error({"error": f"Failed to retrieve analytics: {str(e)}"})
    
    @action(detail=False, methods=['get'], url_path='student/subject/(?P<subject_id>[^/.]+)')
    def student_subject_performance(self, request, subject_id=None):
        """
        Get detailed performance analytics for a specific subject.
        """
        try:
            user = request.user
            api_logger.info(f"Student subject performance requested by {user.username} for subject {subject_id}")
            
            from gyaan_buddy.subjects.models import Answer, Module, ModuleChapter
            
            try:
                subject = Subject.objects.get(id=subject_id)
            except Subject.DoesNotExist:
                return not_found({"error": "Subject not found"})
            
            answers = Answer.objects.filter(
                user=user
            ).filter(
                Q(chapter__module__subject=subject) |
                Q(question__module_contents__chapter__module__subject=subject)
            ).distinct().select_related('question')
            
            total_answers = answers.count()
            correct_answers = answers.filter(is_correct=True).count()
            
            difficulty_stats = {}
            for difficulty in ['easy', 'medium', 'hard']:
                diff_answers = answers.filter(question__difficulty_level=difficulty)
                diff_total = diff_answers.count()
                diff_correct = diff_answers.filter(is_correct=True).count()
                difficulty_stats[difficulty] = {
                    'total': diff_total,
                    'correct': diff_correct,
                    'accuracy': self._calculate_accuracy(diff_correct, diff_total)
                }
            
            type_stats = {}
            for q_type in ['mcq_single', 'mcq_multiple', 'short_answer']:
                type_answers = answers.filter(question__question_type=q_type)
                type_total = type_answers.count()
                type_correct = type_answers.filter(is_correct=True).count()
                type_stats[q_type] = {
                    'total': type_total,
                    'correct': type_correct,
                    'accuracy': self._calculate_accuracy(type_correct, type_total)
                }
            
            modules = Module.objects.filter(subject=subject, is_active=True)
            module_stats = []
            topic_insights = []
            
            for module in modules:
                module_answers = answers.filter(
                    Q(chapter__module=module) |
                    Q(question__module_contents__chapter__module=module)
                ).distinct()
                m_total = module_answers.count()
                m_correct = module_answers.filter(is_correct=True).count()
                
                progress = UserModuleProgress.objects.filter(
                    account=user,
                    module=module
                ).first()
                
                chapters = ModuleChapter.objects.filter(module=module, is_deleted=False)
                chapter_data = []
                
                for chapter in chapters:
                    ch_answers = answers.filter(
                        Q(chapter=chapter) |
                        Q(question__module_contents__chapter=chapter)
                    ).distinct()
                    ch_total = ch_answers.count()
                    ch_correct = ch_answers.filter(is_correct=True).count()
                    chapter_question_count = ModuleContent.objects.filter(
                        chapter=chapter,
                        content_type='question',
                        is_deleted=False,
                    ).exclude(question__isnull=True).count()
                    chapter_attempt_rate = round((ch_total / chapter_question_count) * 100) if chapter_question_count else 0
                    chapter_accuracy = self._calculate_accuracy(ch_correct, ch_total)
                    
                    ch_progress = UserChapterProgress.objects.filter(
                        account=user,
                        chapter=chapter
                    ).first()
                    
                    chapter_data.append({
                        'chapter_id': str(chapter.id),
                        'chapter_title': chapter.title,
                        'questions_attempted': ch_total,
                        'correct': ch_correct,
                        'accuracy': chapter_accuracy,
                        'attempt_rate': chapter_attempt_rate,
                        'progress_percentage': ch_progress.percentage if ch_progress else 0,
                        'status': ch_progress.status if ch_progress else 'not_started'
                    })
                    topic_insights.append({
                        'topic_id': str(chapter.id),
                        'topic_name': chapter.title,
                        'attempt_rate': chapter_attempt_rate,
                        'proficiency': chapter_accuracy,
                    })
                
                module_stats.append({
                    'module_id': str(module.id),
                    'module_name': module.name,
                    'questions_attempted': m_total,
                    'correct': m_correct,
                    'accuracy': self._calculate_accuracy(m_correct, m_total),
                    'progress_percentage': progress.percentage if progress else 0,
                    'status': progress.status if progress else 'not_started',
                    'chapters': chapter_data
                })
            
            hots_answers = answers.filter(question__is_hots=True)
            hots_total = hots_answers.count()
            hots_correct = hots_answers.filter(is_correct=True).count()

            student_profile = getattr(getattr(user, 'profile', None), 'student', None)
            student_class = getattr(student_profile, 'class_instance', None)
            teacher_name = None
            if student_class:
                teacher_assignment = Teacher.objects.filter(
                    class_instance=student_class,
                    subject=subject,
                    is_deleted=False,
                ).select_related('teacher__user_profile__account').first()
                if teacher_assignment and teacher_assignment.teacher and teacher_assignment.teacher.user_profile:
                    teacher_account = teacher_assignment.teacher.user_profile.account
                    teacher_name = (teacher_account.get_full_name() or teacher_account.username) if teacher_account else None

            ranked_topics = sorted(topic_insights, key=lambda item: item['proficiency'], reverse=True)
            strong_topics = [item['topic_name'] for item in ranked_topics[:3]]
            weak_topics = [item['topic_name'] for item in sorted(topic_insights, key=lambda item: item['proficiency'])[:3]]

            level_stats = {}
            for level_num in [1, 2, 3, 4, 5]:
                level_answers = answers.filter(question__level=level_num)
                l_total = level_answers.count()
                l_correct = level_answers.filter(is_correct=True).count()
                level_stats[str(level_num)] = {
                    'total': l_total,
                    'correct': l_correct,
                    'accuracy': self._calculate_accuracy(l_correct, l_total),
                }

            attempted_level_stats = [
                (int(level), data['accuracy'])
                for level, data in level_stats.items()
                if data['total'] > 0
            ]
            strong_question_level = None
            weak_levels = []
            if attempted_level_stats:
                strong_question_level = max(attempted_level_stats, key=lambda item: item[1])[0]
                weak_levels = [level for level, accuracy in attempted_level_stats if accuracy < 50]
                if not weak_levels:
                    weak_levels = [item[0] for item in sorted(attempted_level_stats, key=lambda item: item[1])[:2]]

            return success(
                data={
                    'subject_id': str(subject.id),
                    'subject_name': subject.name,
                    'subject_code': subject.code,
                    'overall_stats': {
                        'total_questions': total_answers,
                        'correct_answers': correct_answers,
                        'accuracy': self._calculate_accuracy(correct_answers, total_answers),
                        'avg_attempts': round(answers.aggregate(avg=Avg('tries'))['avg'] or 0, 2)
                    },
                    'difficulty_breakdown': difficulty_stats,
                    'question_type_breakdown': type_stats,
                    'hots_performance': {
                        'total': hots_total,
                        'correct': hots_correct,
                        'accuracy': self._calculate_accuracy(hots_correct, hots_total)
                    },
                    'teacher_name': teacher_name,
                    'strong_topics': strong_topics,
                    'weak_topics': weak_topics,
                    'topics': topic_insights,
                    'strong_question_level': strong_question_level,
                    'weak_levels': weak_levels,
                    'level_breakdown': level_stats,
                    'modules': module_stats
                },
                message="Subject performance analytics retrieved successfully"
            )
            
        except Exception as e:
            api_logger.error(f"Error in student_subject_performance: {str(e)}", exc_info=True)
            return validation_error({"error": f"Failed to retrieve analytics: {str(e)}"})
    
    @action(detail=False, methods=['get'], url_path='student/weak-areas')
    def student_weak_areas(self, request):
        """
        Identify weak areas for the logged-in student.
        """
        try:
            user = request.user
            api_logger.info(f"Student weak areas requested by {user.username}")

            from gyaan_buddy.subjects.models import Answer

            if not hasattr(user, 'profile') or not hasattr(user.profile, 'student'):
                return validation_error({"error": "Student profile not found"})
            
            student = user.profile.student
            
            answers = Answer.objects.filter(user=user).select_related('question')
            
            weak_areas = []

            if hasattr(user, 'profile') and user.profile and user.profile.school:
                subjects = Subject.objects.filter(
                    classes__school=user.profile.school,
                    is_active=True
                ).distinct()
            else:
                subjects = Subject.objects.filter(is_active=True)
            
            for subject in subjects:
                subj_answers = answers.filter(
                    question__module_contents__chapter__module__subject=subject
                )
                s_total = subj_answers.count()
                s_correct = subj_answers.filter(is_correct=True).count()
                accuracy = self._calculate_accuracy(s_correct, s_total)
                
                if s_total > 0 and accuracy < 60:
                    weak_areas.append({
                        'area_type': 'subject',
                        'area_id': str(subject.id),
                        'area_name': subject.name,
                        'total_questions': s_total,
                        'correct_answers': s_correct,
                        'accuracy': accuracy,
                        'severity': 'high' if accuracy < 40 else 'medium',
                        'recommendation': f"Practice more questions in {subject.name}. Focus on fundamental concepts."
                    })
            
            for difficulty in ['easy', 'medium', 'hard']:
                diff_answers = answers.filter(question__difficulty_level=difficulty)
                d_total = diff_answers.count()
                d_correct = diff_answers.filter(is_correct=True).count()
                accuracy = self._calculate_accuracy(d_correct, d_total)
                
                threshold = {'easy': 80, 'medium': 60, 'hard': 40}
                if d_total > 0 and accuracy < threshold[difficulty]:
                    weak_areas.append({
                        'area_type': 'difficulty',
                        'area_id': None,
                        'area_name': f"{difficulty.capitalize()} Questions",
                        'total_questions': d_total,
                        'correct_answers': d_correct,
                        'accuracy': accuracy,
                        'severity': 'high' if accuracy < threshold[difficulty] - 20 else 'medium',
                        'recommendation': f"Work on {difficulty} level questions. Your accuracy is below expected."
                    })
            
            hots_answers = answers.filter(question__is_hots=True)
            h_total = hots_answers.count()
            h_correct = hots_answers.filter(is_correct=True).count()
            hots_accuracy = self._calculate_accuracy(h_correct, h_total)
            
            if h_total > 0 and hots_accuracy < 50:
                weak_areas.append({
                    'area_type': 'hots',
                    'area_id': None,
                    'area_name': 'Higher Order Thinking Skills (HOTS)',
                    'total_questions': h_total,
                    'correct_answers': h_correct,
                    'accuracy': hots_accuracy,
                    'severity': 'high' if hots_accuracy < 30 else 'medium',
                    'recommendation': "Focus on analytical and application-based questions."
                })
            
            weak_areas.sort(key=lambda x: (0 if x['severity'] == 'high' else 1, x['accuracy']))
            
            return success(
                data={
                    'student_id': str(student.id),
                    'total_weak_areas': len(weak_areas),
                    'weak_areas': weak_areas
                },
                message="Weak areas identified successfully"
            )
            
        except Exception as e:
            api_logger.error(f"Error in student_weak_areas: {str(e)}", exc_info=True)
            return validation_error({"error": f"Failed to identify weak areas: {str(e)}"})
    
    @action(detail=False, methods=['get'], url_path='student/leaderboard')
    def student_leaderboard(self, request):
        """
        Get class leaderboard for the logged-in student.
        """
        try:
            user = request.user
            api_logger.info(f"Student leaderboard requested by {user.username}")
            
            from .models import Student
            from gyaan_buddy.subjects.models import Answer
            
            if not hasattr(user, 'profile') or not hasattr(user.profile, 'student'):
                return validation_error({"error": "Student profile not found"})
            
            student = user.profile.student
            
            if not student.class_instance:
                return validation_error({"error": "Student not enrolled in any class"})
            
            classmates = Student.objects.filter(
                class_instance=student.class_instance,
                is_deleted=False
            ).select_related('user_profile__account', 'level').order_by('-total_exp')
            
            leaderboard = []
            my_rank = None
            
            for rank, classmate in enumerate(classmates, 1):
                account = classmate.user_profile.account
                
                answers = Answer.objects.filter(user=account)
                total = answers.count()
                correct = answers.filter(is_correct=True).count()
                
                entry = {
                    'rank': rank,
                    'student_id': str(classmate.id),
                    'student_name': account.get_full_name() or account.username,
                    'username': account.username,
                    'profile_picture': classmate.user_profile.profile_picture.url if classmate.user_profile.profile_picture else None,
                    'total_exp': classmate.total_exp,
                    'level': classmate.level.name if classmate.level else 1,
                    'questions_attempted': total,
                    'accuracy': self._calculate_accuracy(correct, total),
                    'is_me': classmate.id == student.id
                }
                leaderboard.append(entry)
                
                if classmate.id == student.id:
                    my_rank = rank
            
            return success(
                data={
                    'class_id': str(student.class_instance.id),
                    'class_name': student.class_instance.name,
                    'total_students': len(leaderboard),
                    'my_rank': my_rank,
                    'leaderboard': leaderboard[:50]
                },
                message="Leaderboard retrieved successfully"
            )
            
        except Exception as e:
            api_logger.error(f"Error in student_leaderboard: {str(e)}", exc_info=True)
            return validation_error({"error": f"Failed to retrieve leaderboard: {str(e)}"})
    
    
    @action(detail=False, methods=['get'], url_path='mission/(?P<mission_id>[^/.]+)')
    def mission_analytics(self, request, mission_id=None):
        """
        Get detailed analytics for a specific mission.
        """
        try:
            user = request.user
            api_logger.info(f"Mission analytics requested by {user.username} for mission {mission_id}")

            from gyaan_buddy.subjects.models import Answer
            
            try:
                mission = Mission.objects.select_related(
                    'account', 'subject', 'module', 'module_chapter'
                ).get(id=mission_id, is_deleted=False)
            except Mission.DoesNotExist:
                return not_found({"error": "Mission not found"})
            
            try:
                mission_progress = mission.progress
            except UserMissionProgress.DoesNotExist:
                mission_progress = None
            
            status = mission_progress.status if mission_progress else 'not_started'
            is_started = status != 'not_started'
            is_completed = status == 'completed'
            is_in_progress = status == 'in_progress'
            
            exp_earned = mission_progress.exp_earned if mission_progress else 0
            
            question_stats = []
            mission_questions = mission.mission_questions.all().order_by('order')
            account = mission.account
            
            for mq in mission_questions:
                question = mq.question
                
                q_answer = Answer.objects.filter(
                    question=question,
                    user=account
                ).first()
                
                question_stats.append({
                    'question_id': str(question.id),
                    'question_text': question.question_text[:100] + '...' if len(question.question_text) > 100 else question.question_text,
                    'order': mq.order,
                    'difficulty': question.difficulty_level,
                    'question_type': normalize_question_type(question.question_type),
                    'is_hots': question.is_hots,
                    'answered': q_answer is not None,
                    'is_correct': q_answer.is_correct if q_answer else None,
                    'tries': q_answer.tries if q_answer else 0,
                    'exp_points': question.exp_points
                })
            
            total_questions = len(question_stats)
            answered_count = sum(1 for q in question_stats if q['answered'])
            correct_count = sum(1 for q in question_stats if q['is_correct'])
            
            return success(
                data={
                    'mission_id': str(mission.id),
                    'mission_title': mission.title,
                    'mission_date': mission.mission_date.isoformat(),
                    'account_username': account.username,
                    'account_name': account.get_full_name() or account.username,
                    'subject_name': mission.subject.name if mission.subject else None,
                    'module_name': mission.module.name if mission.module else None,
                    'chapter_name': mission.module_chapter.name if mission.module_chapter else None,
                    'summary': {
                        'status': status,
                        'is_started': is_started,
                        'is_completed': is_completed,
                        'is_in_progress': is_in_progress,
                        'total_questions': total_questions,
                        'questions_answered': answered_count,
                        'correct_answers': correct_count,
                        'accuracy': self._calculate_accuracy(correct_count, answered_count),
                        'completion_percentage': self._calculate_accuracy(answered_count, total_questions),
                        'exp_earned': exp_earned,
                        'started_at': mission_progress.started_at.isoformat() if mission_progress and mission_progress.started_at else None,
                        'completed_at': mission_progress.completed_at.isoformat() if mission_progress and mission_progress.completed_at else None
                    },
                    'question_analytics': question_stats
                },
                message="Mission analytics retrieved successfully"
            )
            
        except Exception as e:
            api_logger.error(f"Error in mission_analytics: {str(e)}", exc_info=True)
            return validation_error({"error": f"Failed to retrieve analytics: {str(e)}"})
    
    
    @action(detail=False, methods=['get'], url_path='competition/(?P<competition_id>[^/.]+)')
    def competition_analytics(self, request, competition_id=None):
        """
        Get detailed analytics for a specific competition.
        """
        try:
            user = request.user
            api_logger.info(f"Competition analytics requested by {user.username} for competition {competition_id}")
            
            from gyaan_buddy.subjects.models import Answer
            
            try:
                competition = Competition.objects.select_related(
                    'subject', 'chapter', 'created_by'
                ).get(id=competition_id, is_deleted=False)
            except Competition.DoesNotExist:
                return not_found({"error": "Competition not found"})
            
            progress = UserCompetitionProgress.objects.filter(
                competition=competition
            ).select_related('account__profile')
            
            total_participants = progress.count()
            completed_participants = progress.filter(status='completed').count()
            
            completed_progress = progress.filter(status='completed')
            score_stats = completed_progress.aggregate(
                avg_score=Avg('score'),
                max_score=Max('score'),
                min_score=models.Min('score'),
                avg_time=Avg('time_taken'),
                avg_exp=Avg('exp_earned')
            )
            
            leaderboard = []
            for rank, p in enumerate(progress.order_by('-score', 'time_taken'), 1):
                account = p.account
                
                comp_question_ids = competition.competition_questions.values_list('question_id', flat=True)
                participant_answers = Answer.objects.filter(
                    user=account,
                    question_id__in=comp_question_ids
                )
                p_total = participant_answers.count()
                p_correct = participant_answers.filter(is_correct=True).count()
                
                leaderboard.append({
                    'rank': rank,
                    'user_id': str(account.id),
                    'user_name': account.get_full_name() or account.username,
                    'username': account.username,
                    'profile_picture': account.profile.profile_picture.url if hasattr(account, 'profile') and account.profile.profile_picture else None,
                    'score': p.score,
                    'time_taken_seconds': p.time_taken,
                    'time_taken_formatted': f"{p.time_taken // 60}m {p.time_taken % 60}s" if p.time_taken else None,
                    'accuracy': self._calculate_accuracy(p_correct, p_total),
                    'exp_earned': p.exp_earned,
                    'status': p.status
                })
            
            question_stats = []
            for cq in competition.competition_questions.all().order_by('order'):
                question = cq.question
                
                participant_ids = progress.values_list('account_id', flat=True)
                q_answers = Answer.objects.filter(
                    question=question,
                    user_id__in=participant_ids
                )
                q_total = q_answers.count()
                q_correct = q_answers.filter(is_correct=True).count()
                
                question_stats.append({
                    'question_id': str(question.id),
                    'question_text': question.question_text[:100] + '...' if len(question.question_text) > 100 else question.question_text,
                    'order': cq.order,
                    'points': cq.points,
                    'difficulty': question.difficulty_level,
                    'total_attempts': q_total,
                    'correct_attempts': q_correct,
                    'accuracy': self._calculate_accuracy(q_correct, q_total)
                })
            
            return success(
                data={
                    'competition_id': str(competition.id),
                    'competition_title': competition.title,
                    'competition_code': competition.code,
                    'subject_name': competition.subject.name if competition.subject else None,
                    'chapter_name': competition.chapter.title if competition.chapter else None,
                    'status': competition.status,
                    'total_time_minutes': competition.total_time,
                    'summary': {
                        'total_participants': total_participants,
                        'completed': completed_participants,
                        'in_progress': progress.filter(status='in_progress').count(),
                        'avg_score': round(score_stats['avg_score'] or 0, 2),
                        'max_score': score_stats['max_score'] or 0,
                        'min_score': score_stats['min_score'] or 0,
                        'avg_time_seconds': round(score_stats['avg_time'] or 0, 0),
                        'avg_exp_earned': round(score_stats['avg_exp'] or 0, 2),
                        'total_questions': len(question_stats)
                    },
                    'leaderboard': leaderboard[:100],
                    'question_analytics': question_stats
                },
                message="Competition analytics retrieved successfully"
            )
            
        except Exception as e:
            api_logger.error(f"Error in competition_analytics: {str(e)}", exc_info=True)
            return validation_error({"error": f"Failed to retrieve analytics: {str(e)}"})
    
    
    @action(detail=False, methods=['get'], url_path='admin/school-overview')
    def admin_school_overview(self, request):
        """
        Get school-wide analytics overview.
        """
        try:
            user = request.user
            api_logger.info(f"Admin school overview requested by {user.username}")
            
            from .models import Student
            from gyaan_buddy.subjects.models import Answer
            
            if not hasattr(user, 'profile') or not user.profile.school:
                return validation_error({"error": "User not associated with a school"})
            
            school = user.profile.school
            
            total_students = Account.objects.filter(
                profile__school=school,
                profile__user_type='student',
                is_active=True
            ).count()
            
            total_teachers = Account.objects.filter(
                profile__school=school,
                profile__user_type='teacher',
                is_active=True
            ).count()
            
            total_classes = Class.objects.filter(school=school, is_active=True).count()
            total_grades = Grade.objects.filter(school=school, is_active=True).count()
            total_subjects = Subject.objects.filter(classes__school=school, is_active=True).distinct().count()
            
            students = Student.objects.filter(
                user_profile__school=school,
                is_deleted=False
            )
            
            student_stats = students.aggregate(
                avg_exp=Avg('total_exp'),
                avg_level=Avg('level__name')
            )
            
            student_account_ids = students.values_list('user_profile__account_id', flat=True)
            answers = Answer.objects.filter(user_id__in=student_account_ids)
            total_answers = answers.count()
            correct_answers = answers.filter(is_correct=True).count()
            
            modules_completed = UserModuleProgress.objects.filter(
                account_id__in=student_account_ids,
                status='completed'
            ).count()
            
            today = timezone.now().date()
            active_missions = Mission.objects.filter(
                class_group__school=school,
                mission_date__gte=today,
                is_deleted=False,
                is_active=True
            ).count()
            
            active_competitions = Competition.objects.filter(
                subject__classes__school=school,
                status__in=['not_started', 'in_progress'],
                is_deleted=False,
                is_active=True
            ).distinct().count()
            
            grades = Grade.objects.filter(school=school, is_active=True)
            grade_breakdown = []
            
            for grade in grades:
                grade_students = students.filter(class_instance__grade=grade)
                g_count = grade_students.count()
                g_avg_exp = grade_students.aggregate(avg=Avg('total_exp'))['avg'] or 0
                
                grade_breakdown.append({
                    'grade_id': str(grade.id),
                    'grade_name': grade.name,
                    'student_count': g_count,
                    'class_count': Class.objects.filter(grade=grade, is_active=True).count(),
                    'avg_exp': round(g_avg_exp, 2)
                })
            
            return success(
                data={
                    'school_id': str(school.id),
                    'school_name': school.name,
                    'counts': {
                        'total_students': total_students,
                        'total_teachers': total_teachers,
                        'total_classes': total_classes,
                        'total_grades': total_grades,
                        'total_subjects': total_subjects
                    },
                    'student_stats': {
                        'avg_exp': round(student_stats['avg_exp'] or 0, 2),
                        'avg_level': round(student_stats['avg_level'] or 1, 1)
                    },
                    'activity_stats': {
                        'total_questions_attempted': total_answers,
                        'correct_answers': correct_answers,
                        'overall_accuracy': self._calculate_accuracy(correct_answers, total_answers),
                        'modules_completed': modules_completed
                    },
                    'active_items': {
                        'active_missions': active_missions,
                        'active_competitions': active_competitions
                    },
                    'grade_breakdown': grade_breakdown
                },
                message="School overview retrieved successfully"
            )
            
        except Exception as e:
            api_logger.error(f"Error in admin_school_overview: {str(e)}", exc_info=True)
            return validation_error({"error": f"Failed to retrieve analytics: {str(e)}"})
    
    @action(detail=False, methods=['get'], url_path='admin/grade/(?P<grade_id>[^/.]+)')
    def admin_grade_analytics(self, request, grade_id=None):
        """
        Get detailed analytics for a specific grade.
        """
        try:
            user = request.user
            api_logger.info(f"Admin grade analytics requested by {user.username} for grade {grade_id}")
            
            from .models import Student
            from gyaan_buddy.subjects.models import Answer
            
            try:
                grade = Grade.objects.get(id=grade_id, is_active=True)
            except Grade.DoesNotExist:
                return not_found({"error": "Grade not found"})
            
            classes = Class.objects.filter(grade=grade, is_active=True)
            
            class_performance = []
            
            for class_instance in classes:
                students = Student.objects.filter(
                    class_instance=class_instance,
                    is_deleted=False
                )
                student_count = students.count()
                
                student_account_ids = list(students.values_list('user_profile__account_id', flat=True))
                
                answers = Answer.objects.filter(user_id__in=student_account_ids)
                total_answers = answers.count()
                correct_answers = answers.filter(is_correct=True).count()
                
                student_stats = students.aggregate(
                    avg_exp=Avg('total_exp'),
                    avg_level=Avg('level__name')
                )
                
                top_student = students.order_by('-total_exp').first()
                
                class_performance.append({
                    'class_id': str(class_instance.id),
                    'class_name': class_instance.name,
                    'student_count': student_count,
                    'questions_attempted': total_answers,
                    'correct_answers': correct_answers,
                    'accuracy': self._calculate_accuracy(correct_answers, total_answers),
                    'avg_exp': round(student_stats['avg_exp'] or 0, 2),
                    'avg_level': round(student_stats['avg_level'] or 1, 1),
                    'top_performer': {
                        'name': top_student.user_profile.account.get_full_name() or top_student.user_profile.account.username if top_student else None,
                        'exp': top_student.total_exp if top_student else None
                    } if top_student else None,
                    'class_teacher': class_instance.class_teacher.user_profile.account.get_full_name() if class_instance.class_teacher else None
                })
            
            class_performance.sort(key=lambda x: x['accuracy'], reverse=True)
            
            all_students = Student.objects.filter(
                class_instance__grade=grade,
                is_deleted=False
            )
            all_student_ids = list(all_students.values_list('user_profile__account_id', flat=True))
            all_answers = Answer.objects.filter(user_id__in=all_student_ids)
            
            return success(
                data={
                    'grade_id': str(grade.id),
                    'grade_name': grade.name,
                    'overall_stats': {
                        'total_classes': len(classes),
                        'total_students': all_students.count(),
                        'total_questions_attempted': all_answers.count(),
                        'correct_answers': all_answers.filter(is_correct=True).count(),
                        'overall_accuracy': self._calculate_accuracy(
                            all_answers.filter(is_correct=True).count(),
                            all_answers.count()
                        )
                    },
                    'class_performance': class_performance
                },
                message="Grade analytics retrieved successfully"
            )
            
        except Exception as e:
            api_logger.error(f"Error in admin_grade_analytics: {str(e)}", exc_info=True)
            return validation_error({"error": f"Failed to retrieve analytics: {str(e)}"})
    
    @action(detail=False, methods=['get'], url_path='admin/class/(?P<class_id>[^/.]+)')
    def admin_class_analytics(self, request, class_id=None):
        """
        Get detailed analytics for a specific class (admin view).
        """
        try:
            user = request.user
            api_logger.info(f"Admin class analytics requested by {user.username} for class {class_id}")
            
            from .models import Student, Teacher
            from gyaan_buddy.subjects.models import Answer
            
            try:
                class_instance = Class.objects.select_related('grade', 'class_teacher').get(id=class_id, is_active=True)
            except Class.DoesNotExist:
                return not_found({"error": "Class not found"})
            
            students = Student.objects.filter(
                class_instance=class_instance,
                is_deleted=False
            ).select_related('user_profile__account', 'level')
            
            student_account_ids = list(students.values_list('user_profile__account_id', flat=True))
            
            answers = Answer.objects.filter(user_id__in=student_account_ids).select_related('question')
            total_answers = answers.count()
            correct_answers = answers.filter(is_correct=True).count()
            
            subjects = Subject.objects.filter(classes=class_instance, is_active=True)
            subject_performance = []
            
            for subject in subjects:
                subj_answers = answers.filter(
                    question__module_contents__chapter__module__subject=subject
                )
                s_total = subj_answers.count()
                s_correct = subj_answers.filter(is_correct=True).count()
                
                teacher_assignment = Teacher.objects.filter(
                    class_instance=class_instance,
                    subject=subject
                ).select_related('teacher__user_profile__account').first()
                
                subject_performance.append({
                    'subject_id': str(subject.id),
                    'subject_name': subject.name,
                    'questions_attempted': s_total,
                    'correct': s_correct,
                    'accuracy': self._calculate_accuracy(s_correct, s_total),
                    'teacher': teacher_assignment.teacher.user_profile.account.get_full_name() if teacher_assignment else None
                })
            
            student_rankings = []
            for student in students:
                s_answers = answers.filter(user_id=student.user_profile.account_id)
                s_total = s_answers.count()
                s_correct = s_answers.filter(is_correct=True).count()
                
                student_rankings.append({
                    'student_id': str(student.id),
                    'student_name': student.user_profile.account.get_full_name() or student.user_profile.account.username,
                    'roll_number': student.roll_number,
                    'total_exp': student.total_exp,
                    'level': student.level.name if student.level else 1,
                    'questions_attempted': s_total,
                    'correct': s_correct,
                    'accuracy': self._calculate_accuracy(s_correct, s_total)
                })
            
            student_rankings.sort(key=lambda x: x['total_exp'], reverse=True)
            
            return success(
                data={
                    'class_id': str(class_instance.id),
                    'class_name': class_instance.name,
                    'grade_name': class_instance.grade.name if class_instance.grade else None,
                    'class_teacher': class_instance.class_teacher.user_profile.account.get_full_name() if class_instance.class_teacher else None,
                    'overall_stats': {
                        'total_students': len(students),
                        'total_questions_attempted': total_answers,
                        'correct_answers': correct_answers,
                        'accuracy': self._calculate_accuracy(correct_answers, total_answers),
                        'avg_exp': round(students.aggregate(avg=Avg('total_exp'))['avg'] or 0, 2)
                    },
                    'subject_performance': subject_performance,
                    'student_rankings': student_rankings
                },
                message="Class analytics retrieved successfully"
            )
            
        except Exception as e:
            api_logger.error(f"Error in admin_class_analytics: {str(e)}", exc_info=True)
            return validation_error({"error": f"Failed to retrieve analytics: {str(e)}"})
    
    @action(detail=False, methods=['get'], url_path='answer-trends')
    def answer_trends(self, request):
        """
        Get answer trends over time for the current user or a specific student.
        """
        try:
            user = request.user
            student_id = request.query_params.get('student_id')
            days = int(request.query_params.get('days', 30))
            
            api_logger.info(f"Answer trends requested by {user.username}, student_id={student_id}, days={days}")
            
            from gyaan_buddy.subjects.models import Answer
            from django.db.models.functions import TruncDate
            
            if student_id and self._get_user_type(user) in ['teacher', 'administrator']:
                from .models import Student
                try:
                    student = Student.objects.get(id=student_id)
                    target_user_id = student.user_profile.account_id
                except Student.DoesNotExist:
                    return not_found({"error": "Student not found"})
            else:
                target_user_id = user.id
            
            end_date = timezone.now().date()
            start_date = end_date - timezone.timedelta(days=days)
            
            trends = Answer.objects.filter(
                user_id=target_user_id,
                created_at__date__gte=start_date,
                created_at__date__lte=end_date
            ).annotate(
                date=TruncDate('created_at')
            ).values('date').annotate(
                total_answers=Count('id'),
                correct_answers=Count('id', filter=Q(is_correct=True)),
                exp_earned=models.Sum(models.F('current_Exp') - models.F('prev_exp'))
            ).order_by('date')
            
            trend_data = []
            for t in trends:
                trend_data.append({
                    'date': t['date'].isoformat(),
                    'total_answers': t['total_answers'],
                    'correct_answers': t['correct_answers'],
                    'accuracy': self._calculate_accuracy(t['correct_answers'], t['total_answers']),
                    'exp_earned': t['exp_earned'] or 0
                })
            
            return success(
                data={
                    'start_date': start_date.isoformat(),
                    'end_date': end_date.isoformat(),
                    'total_days': days,
                    'trends': trend_data
                },
                message="Answer trends retrieved successfully"
            )
            
        except Exception as e:
            api_logger.error(f"Error in answer_trends: {str(e)}", exc_info=True)
            return validation_error({"error": f"Failed to retrieve trends: {str(e)}"})



class TestViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing Tests and their progress.
    
    Provides endpoints for:
    - Listing all tests (filtered by class for students)
    - Retrieving test details with user's progress
    - Creating tests (for teachers/admins)
    - Managing test progress
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action == 'create':
            return TestCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return TestUpdateSerializer
        elif self.action == 'list':
            return TestListSerializer
        elif self.action in ['my_tests', 'retrieve']:
            return TestWithProgressSerializer
        return TestSerializer
    
    def get_queryset(self):
        """
        Return tests based on user type:
        - Students: Tests for their class
        - Teachers: Tests for their classes
        - Admins: All tests
        """
        user = self.request.user

        queryset = Test.objects.select_related(
            'class_group', 'subject'
        ).prefetch_related(
            'module_chapters__module', 'module_chapters__module_chapter',
            'class_groups'
        ).filter(is_deleted=False)

        user_type = None
        try:
            if hasattr(user, 'profile'):
                user_type = user.profile.user_type
        except (UserProfile.DoesNotExist, AttributeError):
            pass

        if user_type == 'student':
            try:
                student = user.profile.student
                if student and student.class_instance:
                    from django.db.models import Q
                    queryset = queryset.filter(
                        Q(class_group=student.class_instance) |
                        Q(class_groups=student.class_instance)
                    ).distinct()
                else:
                    queryset = queryset.none()
            except (Student.DoesNotExist, AttributeError):
                queryset = queryset.none()
        elif user_type == 'teacher':
            try:
                queryset = queryset.filter(created_by=user)
            except (AttributeError, Exception):
                queryset = queryset.none()

        subject_id = self.request.query_params.get('subject', None)
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)

        class_id = self.request.query_params.get('class', None)
        if class_id:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(class_group_id=class_id) | Q(class_groups__id=class_id)
            ).distinct()
        
        module_id = self.request.query_params.get('module', None)
        if module_id:
            queryset = queryset.filter(module_chapters__module_id=module_id).distinct()
        
        return queryset.order_by('-test_datetime')
    
    def get_permissions(self):
        if self.action in ['create', 'destroy', 'update', 'partial_update', 'students_performance', 'remove_questions', 'add_questions', 'add_question']:
            return [TeacherAdminPermission()]
        return [permissions.IsAuthenticated()]

    def list(self, request, *args, **kwargs):
        """List all tests."""
        api_logger.info(f"Tests list requested by {request.user.username}")
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return success(data=serializer.data, message="Tests retrieved successfully")
        serializer = self.get_serializer(queryset, many=True)
        return success(data=serializer.data, message="Tests retrieved successfully")
    
    def retrieve(self, request, *args, **kwargs):
        """Retrieve a specific test with user's progress."""
        instance = self.get_object()
        serializer = self.get_serializer(instance, context={'request': request})
        return success(data=serializer.data, message="Test retrieved successfully")
    
    def create(self, request, *args, **kwargs):
        """Create a new test. When multiple class_groups are provided, one test per class is created."""
        api_logger.info(f"Test creation by {request.user.username}")
        serializer = self.get_serializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            test = serializer.save()
            # When multiple classes were passed, serializer creates one test per class and attaches list here.
            created_tests = getattr(serializer, '_created_tests', None)
            if created_tests:
                tests_qs = Test.objects.filter(pk__in=[t.pk for t in created_tests]).prefetch_related(
                    'class_groups', 'module_chapters__module', 'module_chapters__module_chapter'
                ).select_related('class_group', 'subject').order_by('created_at')
                tests_data = TestSerializer(tests_qs, many=True).data
                self._send_test_created_notifications(list(tests_qs))
                return created(
                    data=tests_data,
                    message=f"Tests created successfully for {len(created_tests)} class(es)."
                )
            # Single test: return one object (unchanged shape for existing clients).
            test_refreshed = Test.objects.prefetch_related(
                'class_groups', 'module_chapters__module', 'module_chapters__module_chapter'
            ).select_related('class_group', 'subject').get(pk=test.pk)
            self._send_test_created_notifications([test_refreshed])
            return created(
                data=TestSerializer(test_refreshed).data,
                message="Test created successfully"
            )
        return validation_error(serializer.errors)

    def _send_test_created_notifications(self, tests):
        """Send push notifications to all students in the class(es) of the given tests."""
        try:
            from gyaan_buddy.utils.firebase_notifications import firebase_notification_service
            for test in tests:
                assigned_classes = test.get_assigned_classes()
                for cls in assigned_classes:
                    students = Account.objects.filter(
                        profile__student__class_instance=cls,
                        profile__user_type='student',
                        is_active=True,
                        is_deleted=False,
                    )
                    if not students.exists():
                        continue
                    title = "📝 New Test Upcoming!"
                    body = f"{test.subject.name} test on {test.test_datetime.strftime('%-d %B %Y')}—start preparing now! 💪"
                    data = {
                        'type': 'test_created',
                        'test_id': str(test.id),
                        'subject_name': test.subject.name,
                        'test_datetime': str(test.test_datetime),
                        'class_id': str(cls.id),
                        'class_name': cls.name,
                        'action': 'open_test',
                    }
                    results = firebase_notification_service.send_notification_to_multiple_users(
                        list(students), title, body, data,
                        notification_type='test',
                        triggered_by='auto',
                    )
                    api_logger.info(f"Test created notification sent for test {test.id} to class '{cls.name}': {results}")
        except Exception as notif_error:
            api_logger.error(f"Failed to send test created notifications: {str(notif_error)}")
    
    def update(self, request, *args, **kwargs):
        """Update a test."""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            test = serializer.save()
            return success(data=TestSerializer(test).data, message="Test updated successfully")
        return validation_error(serializer.errors)
    
    def destroy(self, request, *args, **kwargs):
        """Soft delete a test."""
        instance = self.get_object()
        instance.soft_delete()
        api_logger.info(f"Test {instance.id} deleted by {request.user.username}")
        return success(data=None, message="Test deleted successfully")
    
    @action(detail=False, methods=['get'], url_path='my-tests')
    def my_tests(self, request):
        """Get all tests for the user's class with their progress."""
        api_logger.info(f"My tests requested by {request.user.username}")

        user = request.user
        
        class_instance = None
        try:
            if hasattr(user, 'profile') and hasattr(user.profile, 'student'):
                student = user.profile.student
                if student and student.class_instance:
                    class_instance = student.class_instance
        except (UserProfile.DoesNotExist, Student.DoesNotExist, AttributeError):
            pass
        
        if not class_instance:
            return success(data=[], message="No class assigned")

        from django.db.models import Q
        queryset = Test.objects.select_related(
            'class_group', 'subject'
        ).prefetch_related(
            'module_chapters__module', 'module_chapters__module_chapter',
            'class_groups'
        ).filter(
            Q(class_group=class_instance) | Q(class_groups=class_instance),
            is_deleted=False
        ).distinct().order_by('-test_datetime')

        serializer = TestWithProgressSerializer(queryset, many=True, context={'request': request})
        return success(data=serializer.data, message="Your tests retrieved successfully")
    
    @action(detail=True, methods=['post'], url_path='start')
    def start_test(self, request, pk=None):
        """Start a test - creates or updates progress."""
        test = self.get_object()
        
        progress, created = UserTestProgress.objects.get_or_create(
            account=request.user,
            test=test,
            defaults={
                'status': 'in_progress',
                'total_questions': test.questions.count(),
                'started_at': timezone.now()
            }
        )
        
        if not created and progress.status == 'not_started':
            progress.status = 'in_progress'
            progress.started_at = timezone.now()
            progress.save()
        
        serializer = TestWithProgressSerializer(test, context={'request': request})
        return success(data=serializer.data, message="Test started successfully")
    
    @action(detail=True, methods=['post'], url_path='complete')
    def complete_test(self, request, pk=None):
        """Complete a test."""
        test = self.get_object()
        
        try:
            progress = UserTestProgress.objects.get(account=request.user, test=test)
        except UserTestProgress.DoesNotExist:
            return validation_error({"error": "Test has not been started yet"})
        
        if progress.status == 'completed':
            return validation_error({"error": "Test has already been completed"})
        
        progress.status = 'completed'
        progress.completed_at = timezone.now()
        progress.percentage = 100
        progress.save()
        
        serializer = TestWithProgressSerializer(test, context={'request': request})
        return success(data=serializer.data, message="Test completed successfully")
    
    @action(detail=True, methods=['get'], url_path='questions')
    def get_questions(self, request, pk=None):
        """Get all questions for a test (only active questions so removed/inactive ones are not given)."""
        test = self.get_object()
        
        # Remove all answer entries for this user for this test's questions
        from gyaan_buddy.subjects.models import Answer
        Answer.objects.filter(user=request.user, test=test).delete()
        
        questions = test.questions.filter(is_active=True).prefetch_related('options')
        tq_order = {tq.question_id: tq.order for tq in test.test_questions.all().order_by('order')}
        questions = sorted(questions, key=lambda q: tq_order.get(q.id, 999))
        
        questions_data = []
        for question in questions:
            options_data = [{
                'id': str(option.id),
                'option_text': option.option_text,
                'order': option.order,
                'is_correct': option.is_correct,
            } for option in question.options.all()]
            
            questions_data.append({
                'id': str(question.id),
                'question_text': question.question_text,
                'image': question.image.url if question.image else None,
                'question_type': normalize_question_type(question.question_type),
                'exp_points': question.exp_points,
                'difficulty_level': question.difficulty_level,
                'explanation': question.explanation,
                'hint': question.hint,
                'is_hots': question.is_hots,
                'options': options_data,
            })
        
        return success(data=questions_data, message=f"Retrieved {len(questions_data)} questions")
    
    @action(detail=True, methods=['post'], url_path='remove-questions')
    def remove_questions(self, request, pk=None):
        """
        Remove given questions from the test and mark them inactive (is_active=False) if ai_generated.
        Body: { "question_ids": [uuid, ...] } - IDs of questions to remove from the test.
        Those questions are deleted from TestQuestion and, if ai_generated, set Question.is_active=False
        so they are not given elsewhere.
        """
        test = self.get_object()
        question_ids = request.data.get('question_ids', [])
        if not isinstance(question_ids, list):
            question_ids = [question_ids] if question_ids else []
        question_ids = [str(q).strip() for q in question_ids if q]
        if not question_ids:
            return success(
                data={'removed_count': 0, 'deactivated_count': 0},
                message="No questions to remove"
            )
        removed = TestQuestion.objects.filter(test=test, question_id__in=question_ids)
        removed_count = removed.count()
        removed.delete()
        deactivated = Question.objects.filter(
            id__in=question_ids,
            ai_generated=True,
            is_active=True
        )
        deactivated_count = deactivated.count()
        deactivated.update(is_active=False)
        api_logger.info(
            f"Test {test.id}: removed {removed_count} questions from test, "
            f"deactivated {deactivated_count} ai_generated questions"
        )
        return success(
            data={
                'removed_count': removed_count,
                'deactivated_count': deactivated_count,
            },
            message=f"Removed {removed_count} question(s) from test; {deactivated_count} marked inactive."
        )

    @action(detail=True, methods=['post'], url_path='add-questions')
    def add_questions(self, request, pk=None):
        """
        Add existing questions to the test by question IDs.
        Body: { "question_ids": [uuid, ...] } - IDs of questions to add to the test.
        """
        test = self.get_object()
        question_ids = request.data.get('question_ids', [])
        if not isinstance(question_ids, list):
            question_ids = [question_ids] if question_ids else []
        question_ids = [str(q).strip() for q in question_ids if q]
        if not question_ids:
            return validation_error({"error": "question_ids (array) is required and must not be empty"})
        questions = list(Question.objects.filter(id__in=question_ids, is_active=True))
        if len(questions) != len(question_ids):
            found_ids = {str(q.id) for q in questions}
            missing = [qid for qid in question_ids if qid not in found_ids]
            return validation_error({"error": f"Some questions not found or inactive: {missing}"})
        max_order = TestQuestion.objects.filter(test=test).aggregate(max_order=Max('order'))['max_order'] or 0
        next_order = max_order + 1
        added = 0
        for question in questions:
            if TestQuestion.objects.filter(test=test, question=question).exists():
                continue
            TestQuestion.objects.create(test=test, question=question, order=next_order)
            next_order += 1
            added += 1
        api_logger.info(f"Test {test.id}: added {added} question(s) from question_ids")
        return success(
            data={'added_count': added, 'total_requested': len(question_ids)},
            message=f"Added {added} question(s) to test."
        )

    @action(detail=True, methods=['post'], url_path='add-question')
    def add_question(self, request, pk=None):
        """
        Create a new question and add it to the test (manual question creation).
        Body: question_text, question_type, difficulty_level, optional: exp_points, explanation, options [{ option_text, is_correct, order }].
        """
        test = self.get_object()
        question_data = request.data.copy()
        options_data = question_data.pop('options', [])
        required_fields = ['question_text', 'question_type']
        for field in required_fields:
            if field not in question_data:
                return validation_error({"error": f"Missing required field: {field}"})
        question_type = question_data.get('question_type', 'mcq_single')
        if question_type not in ['mcq_single', 'mcq_multiple', 'short_answer', 'rearrange']:
            return validation_error({"error": "question_type must be one of: mcq_single, mcq_multiple, short_answer, rearrange"})
        difficulty_level = question_data.get('difficulty_level', 'medium')
        if difficulty_level not in ['easy', 'medium', 'hard']:
            difficulty_level = 'medium'
        try:
            question = Question.objects.create(
                question_text=question_data.get('question_text'),
                question_type=question_type,
                difficulty_level=difficulty_level,
                level=int(question_data.get('level', 1)),
                exp_points=question_data.get('exp_points', 10),
                explanation=question_data.get('explanation', ''),
                hint=question_data.get('hint', ''),
                is_active=True,
                ai_generated=False,
                created_by=request.user,
            )
        except Exception as e:
            api_logger.error(f"Error creating question for test: {str(e)}", exc_info=True)
            return validation_error({"error": f"Failed to create question: {str(e)}"})
        if options_data and question_type in ['mcq_single', 'mcq_multiple', 'rearrange']:
            for idx, option_data in enumerate(options_data):
                Option.objects.create(
                    question=question,
                    option_text=option_data.get('option_text', ''),
                    is_correct=option_data.get('is_correct', False),
                    order=option_data.get('order', idx + 1),
                )
        max_order = TestQuestion.objects.filter(test=test).aggregate(max_order=Max('order'))['max_order'] or 0
        TestQuestion.objects.create(test=test, question=question, order=max_order + 1)
        api_logger.info(f"Question created and added to test {test.id} (question id: {question.id})")
        from gyaan_buddy.subjects.serializers import QuestionSerializer
        question_serializer = QuestionSerializer(question, context={'request': request})
        return created(
            data={'question': question_serializer.data, 'test_id': str(test.id)},
            message="Question created and added to test successfully.",
        )

    @action(detail=True, methods=['post'], url_path='check-answer')
    def check_answer(self, request, pk=None):
        """Check an answer for a test question and update progress."""
        test = self.get_object()
        
        question_id = request.data.get('question_id')
        is_correct = request.data.get('is_correct', False)
        tries = request.data.get('tries', 1)
        
        if not question_id:
            return validation_error({"question_id": "Question ID is required"})
        
        try:
            question = Question.objects.get(id=question_id)
        except Question.DoesNotExist:
            return validation_error({"question_id": "Question not found"})
        
        if not test.questions.filter(id=question_id).exists():
            return validation_error({"question_id": "Question does not belong to this test"})
        
        from gyaan_buddy.subjects.models import Answer
        existing_answer = Answer.objects.filter(
            user=request.user,
            question=question,
            test=test,
        ).first()

        progress, created = UserTestProgress.objects.get_or_create(
            account=request.user,
            test=test,
            defaults={
                'status': 'in_progress',
                'total_questions': test.questions.count(),
                'started_at': timezone.now()
            }
        )
        
        if created or progress.status == 'not_started':
            progress.status = 'in_progress'
            progress.started_at = progress.started_at or timezone.now()
        
        if existing_answer:
            # Re-answering same question: update progress by swapping correct/wrong
            was_correct = existing_answer.is_correct
            if was_correct and not is_correct:
                progress.correct_answers = max(0, progress.correct_answers - 1)
                progress.wrong_answers += 1
            elif not was_correct and is_correct:
                progress.wrong_answers = max(0, progress.wrong_answers - 1)
                progress.correct_answers += 1
            # else same outcome (correct->correct or wrong->wrong), no change to counts
            # Do not increment questions_attempted (one question = one slot)
        else:
            progress.questions_attempted += 1
            if is_correct:
                progress.correct_answers += 1
            else:
                progress.wrong_answers += 1
        
        was_already_correct = existing_answer is not None and existing_answer.is_correct
        if not was_already_correct and is_correct and tries == 1:
            exp_earned = 2
        elif not was_already_correct and is_correct and tries == 2:
            exp_earned = 1
        else:
            exp_earned = 0
        if exp_earned > 0:
            progress.exp_earned += exp_earned

            if hasattr(request.user, 'profile') and hasattr(request.user.profile, 'student'):
                request.user.profile.student.add_exp(exp_earned)
                api_logger.info(f"User {request.user.username} earned {exp_earned} exp from test (tries: {tries})")
        
        progress.current_question = question
        progress.score = progress.correct_answers * 10
        progress.percentage = min(100, round((progress.correct_answers / progress.total_questions) * 100) if progress.total_questions else 0)
        
        if progress.questions_attempted >= progress.total_questions:
            progress.status = 'completed'
            progress.completed_at = timezone.now()
        
        progress.save()
        
        answer_text = request.data.get('answer', '') or (str(is_correct))
        Answer.objects.update_or_create(
            user=request.user,
            question=question,
            test=test,
            defaults={
                'is_correct': is_correct,
                'answer': answer_text,
                'tries': tries,
                'from_mission': False,
            },
        )
        
        return success(
            data={
                'is_correct': is_correct,
                'exp_earned': exp_earned,
                'total_exp_earned': progress.exp_earned,
                'questions_attempted': progress.questions_attempted,
                'total_questions': progress.total_questions,
                'correct_answers': progress.correct_answers,
                'wrong_answers': progress.wrong_answers,
                'percentage': progress.percentage,
                'score': progress.score,
                'status': progress.status,
                'is_last': progress.questions_attempted >= progress.total_questions,
            },
            message="Answer checked successfully"
        )

    @action(detail=True, methods=['get'], url_path='students-performance')
    def students_performance(self, request, pk=None):
        """
        Get test class students table: average %, correct answers, wrong answers, total questions.
        For teachers/admins: returns all students in the test's class with their attempt stats.
        """
        test = self.get_object()
        total_questions = test.questions.count()
        assigned_classes = getattr(test, 'get_assigned_classes', lambda: [])() or []
        if not assigned_classes and getattr(test, 'class_group', None):
            assigned_classes = [test.class_group]
        if not assigned_classes:
            return success(
                data={
                    'test_id': str(test.id),
                    'test_name': getattr(test, 'name', None) or 'Test',
                    'class_name': '',
                    'class_groups': [],
                    'subject_name': getattr(test.subject, 'name', '') if test.subject else '',
                    'total_questions': total_questions,
                    'average_percentage': 0,
                    'students': [],
                    'questions': [],
                },
                message="Test students performance retrieved successfully",
            )
        class_names = ', '.join(c.name for c in assigned_classes)
        subject_name = test.subject.name if test.subject else ''
        test_name = getattr(test, 'name', None) or f"{subject_name} - {class_names}"

        class_ids = [c.id for c in assigned_classes]
        students_in_class = Student.objects.filter(
            class_instance_id__in=class_ids,
            is_deleted=False
        ).select_related('user_profile__account', 'class_instance')

        progress_list = UserTestProgress.objects.filter(test=test).select_related('account')
        progress_by_account = {p.account_id: p for p in progress_list}

        from gyaan_buddy.subjects.models import Answer
        test_question_rows = list(
            TestQuestion.objects.filter(test=test).select_related('question').order_by('order')
        )
        test_question_ids = [tq.question_id for tq in test_question_rows if tq.question_id]

        students_data = []
        for student in students_in_class:
            try:
                if not getattr(student, 'user_profile', None):
                    continue
                account = student.user_profile.account
                if not account:
                    continue
                # Derive stats from Answer table (one row per question per user per test)
                student_answers = Answer.objects.filter(
                    test=test, user=account, question_id__in=test_question_ids
                ).select_related('question')
                answer_by_question = {a.question_id: a for a in student_answers}
                correct_answers = sum(1 for a in student_answers if a.is_correct)
                wrong_answers = len(student_answers) - correct_answers
                questions_attempted = len(student_answers)
                total = total_questions
                score = correct_answers * 10
                percentage = round((correct_answers / total * 100)) if total else 0
                percentage = min(100, percentage)
                progress = progress_by_account.get(account.id)
                status = progress.status if progress else ('completed' if questions_attempted >= total_questions else ('in_progress' if questions_attempted > 0 else 'not_started'))

                try:
                    user_name = (account.get_full_name() or '').strip() or account.username or '—'
                except Exception:
                    user_name = getattr(account, 'username', '—') or '—'
                answers_list = []
                for tq in test_question_rows:
                    q = tq.question
                    if not q or not getattr(q, 'is_active', True):
                        continue
                    a = answer_by_question.get(q.id)
                    answers_list.append({
                        'question_title': (q.question_text or '')[:300],
                        'is_correct': a.is_correct if a else None,
                    })

                student_class_name = student.class_instance.name if getattr(student, 'class_instance', None) else class_names
                students_data.append({
                    'user_id': str(account.id),
                    'user_name': user_name,
                    'roll_number': getattr(student, 'roll_number', None) or '—',
                    'class_name': student_class_name,
                    'percentage': percentage,
                    'correct_answer_count': correct_answers,
                    'wrong_answer_count': wrong_answers,
                    'questions_attempted': questions_attempted,
                    'total_questions': total,
                    'score': score,
                    'status': status,
                    'pass': percentage >= 40,
                    'answers': answers_list,
                })
            except Exception as e:
                api_logger.warning("Skipping student in students_performance: %s", e, exc_info=True)
                continue

        # Calculate average_percentage only for students who have given the test
        # average_percentage = total correct answers / total answers (for students who attempted)
        students_who_attempted = [s for s in students_data if s['questions_attempted'] > 0]
        if students_who_attempted:
            total_correct = sum(s['correct_answer_count'] for s in students_who_attempted)
            total_answers = sum(s['questions_attempted'] for s in students_who_attempted)
            avg_percentage = round((total_correct / total_answers * 100)) if total_answers > 0 else 0
        else:
            avg_percentage = 0

        from gyaan_buddy.subjects.models import Answer, ModuleContent, ChapterHOTS
        test_question_rows = TestQuestion.objects.filter(test=test).select_related('question').order_by('order')
        questions_data = []
        for order, tq in enumerate(test_question_rows, start=1):
            q = tq.question
            if not q or not q.is_active:
                continue
            module_chapter_name = ''
            mc = ModuleContent.objects.filter(question=q, is_deleted=False).select_related('chapter__module').first()
            if mc and mc.chapter:
                module_chapter_name = f"{mc.chapter.module.name} - {mc.chapter.title}"
            else:
                ch = ChapterHOTS.objects.filter(question=q).select_related('chapter__module').first()
                if ch and ch.chapter:
                    module_chapter_name = f"{ch.chapter.module.name} - {ch.chapter.title}"
            total_attempted = 0
            correct_count = 0
            try:
                total_attempted = Answer.objects.filter(test=test, question=q).count()
                correct_count = Answer.objects.filter(test=test, question=q, is_correct=True).count()
            except Exception as e:
                from django.db.utils import ProgrammingError
                if isinstance(e, ProgrammingError):
                    total_attempted = 0
                    correct_count = 0
                else:
                    raise
            correct_percentage = round(correct_count / total_attempted * 100) if total_attempted else 0
            sub_topic = module_chapter_name or '—'
            questions_data.append({
                'question_id': str(q.id),
                'order': order,
                'question_title': (q.question_text or '')[:300],
                'sub_topic': sub_topic,
                'module_chapter_name': module_chapter_name or '—',
                'level': getattr(q, 'level', None),
                'difficulty': getattr(q, 'difficulty_level', None) or '—',
                'correct_percentage': correct_percentage,
                'students_attempted': total_attempted,
                'students_correct': correct_count,
            })

        payload = {
            'test_id': str(test.id),
            'test_name': test_name,
            'class_name': class_names,
            'subject_name': subject_name,
            'total_questions': total_questions,
            'average_percentage': avg_percentage,
            'students': students_data,
            'questions': questions_data,
        }
        return success(data=payload, message="Test students performance retrieved successfully")

    @action(detail=True, methods=['get'], url_path='my-report')
    def my_report(self, request, pk=None):
        """
        Detailed report for the logged-in student on a specific test.
        This is intended for student/parent dashboard usage.
        """
        test = self.get_object()
        user = request.user

        if not hasattr(user, 'profile') or user.profile.user_type != 'student':
            return forbidden("This report is only available for student accounts")

        from gyaan_buddy.subjects.models import Answer

        question_rows = list(
            TestQuestion.objects.filter(test=test)
            .select_related('question')
            .order_by('order')
        )
        total_questions = len(question_rows)

        # Student's answers for this test
        answer_qs = Answer.objects.filter(test=test, user=user).select_related('question')
        answer_by_question = {a.question_id: a for a in answer_qs}
        correct_answers = sum(1 for a in answer_qs if a.is_correct)
        wrong_answers = len(answer_qs) - correct_answers
        attempted = len(answer_qs)
        percentage = round((correct_answers / total_questions) * 100) if total_questions else 0

        # Class average for students assigned to this test
        assigned_classes = getattr(test, 'get_assigned_classes', lambda: [])() or []
        if not assigned_classes and getattr(test, 'class_group', None):
            assigned_classes = [test.class_group]
        class_ids = [c.id for c in assigned_classes]
        class_students = Student.objects.filter(
            class_instance_id__in=class_ids,
            is_deleted=False
        ).select_related('user_profile__account')
        class_account_ids = [s.user_profile.account_id for s in class_students if getattr(s, 'user_profile', None)]
        class_answers = Answer.objects.filter(
            test=test,
            user_id__in=class_account_ids,
            question_id__in=[q.question_id for q in question_rows if q.question_id],
        )
        class_total_attempts = class_answers.count()
        class_correct = class_answers.filter(is_correct=True).count()
        class_average = round((class_correct / class_total_attempts) * 100) if class_total_attempts else 0

        topic_performance = []
        for row in question_rows:
            q = row.question
            if not q or not q.is_active:
                continue

            ans = answer_by_question.get(q.id)
            if not ans:
                status = 'Not Attempted'
            elif ans.is_correct and ans.tries == 1:
                status = 'Strong'
            elif ans.is_correct and ans.tries == 2:
                status = 'Good'
            elif ans.is_correct:
                status = 'Struggled'
            else:
                status = 'Weak'

            topic_performance.append({
                'topic': (q.question_text or '')[:120] or f'Question {row.order}',
                'status': status,
                'is_correct': ans.is_correct if ans else None,
                'tries': ans.tries if ans else 0,
            })

        perf_label = 'Excellent' if percentage >= 85 else 'Good' if percentage >= 70 else 'Needs Improvement' if percentage >= 45 else 'Weak'
        classes_name = ', '.join(c.name for c in assigned_classes)
        test_name = getattr(test, 'name', None) or f"{test.subject.name if test.subject else 'Test'}"

        payload = {
            'test_id': str(test.id),
            'test_name': test_name,
            'subject_name': test.subject.name if test.subject else '',
            'class_name': classes_name,
            'test_datetime': test.test_datetime,
            'score_percent': percentage,
            'score': correct_answers,
            'out_of': total_questions,
            'questions_attempted': attempted,
            'correct_answers': correct_answers,
            'wrong_answers': wrong_answers,
            'class_average_percent': class_average,
            'performance_label': perf_label,
            'topic_performance': topic_performance,
        }
        return success(data=payload, message="My test report retrieved successfully")
