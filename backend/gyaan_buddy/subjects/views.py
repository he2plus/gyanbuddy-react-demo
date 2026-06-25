import json
import logging
import os
from django.core.files import File
from django.utils import timezone
from django.db.models import Max, Avg, Min, F, Q, Count
from rest_framework import viewsets, permissions, parsers, status
from rest_framework.decorators import action, api_view, permission_classes as drf_permission_classes
from rest_framework.response import Response
from .models import Subject, Module, Question, Option, Theory, ModuleContent, ModuleChapter, ChapterHOTS, Answer, ManualVerificationAnswer
from .serializers import (
    SubjectSerializer, ModuleSerializer, ModuleWithProgressSerializer,
    QuestionSerializer, OptionSerializer, ModuleContentSerializer, ModuleChapterSerializer,
    TheorySerializer
)
from gyaan_buddy.utils.response_utils import success, created, validation_error
from gyaan_buddy.users.models import UserChapterProgress, UserModuleProgress, Mission, MissionQuestion, Test, TestQuestion
from .helpers import normalize_question_type
logger = logging.getLogger('gyaan_buddy.subjects')
api_logger = logging.getLogger('gyaan_buddy.api')

# ─── Mobile-safe formatting rules injected into every AI question prompt ───────
# The mobile app has NO LaTeX / MathML renderer. All math must be plain Unicode.
MOBILE_FORMAT_RULES = """
⚠️ CRITICAL FORMATTING — MOBILE APP DISPLAY (READ BEFORE WRITING ANYTHING):
This content is displayed DIRECTLY on a mobile app. The app has NO LaTeX engine,
NO MathML renderer, and NO markdown math support. Violations will show as broken
symbols like \\frac{}{} or \\sqrt{} on screen.

STRICTLY FORBIDDEN — never output these:
  • LaTeX commands: \\frac{}{} \\sqrt{} \\int \\sum \\alpha \\beta \\pi \\times \\cdot \\infty \\sin \\cos \\log \\lim
  • Dollar delimiters: $...$ or $$...$$ or \\(...\\) or \\[...\\]
  • Caret for power: x^2 or x^{2}  — use superscript ² instead
  • Underscore for subscript: H_2O or x_{1} — use subscript ₂ instead
  • Backslash before any word (\\alpha, \\theta, \\Delta, \\Sigma, etc.)

USE THESE UNICODE SYMBOLS INSTEAD:
  Powers      : ⁰ ¹ ² ³ ⁴ ⁵  (e.g. x², cm², r³, 10²)
  Subscripts  : ₀ ₁ ₂ ₃ ₄  (e.g. H₂O, CO₂, O₂, x₁, aₙ)
  Roots       : √ ∛  (e.g. √2, √(x+1), ∛8)
  Fractions   : write as a/b  (e.g. 1/2, (x+1)/(x-1), 3/4)
  Greek       : α β γ δ ε ζ θ λ μ ν π ρ σ τ φ ω Δ Γ Λ Σ Ω
  Operators   : × ÷ ± ≤ ≥ ≠ ≈ ∝ ∞ ∫ ∑ ∏ ∂ ∇
  Logic/Set   : ∈ ∉ ⊂ ⊃ ∩ ∪ ∧ ∨ ¬
  Arrows      : → ← ↔ ⇒ ⇔ ↑ ↓
  Other       : ° · … ₓ

CONVERSION EXAMPLES (follow exactly):
  ❌ \\frac{1}{2}        ✅ 1/2
  ❌ \\frac{a+b}{c}      ✅ (a+b)/c
  ❌ x^2 + y^2           ✅ x² + y²
  ❌ x^{2} + y^{2}       ✅ x² + y²
  ❌ \\sqrt{x}           ✅ √x
  ❌ \\sqrt{x^2+y^2}     ✅ √(x²+y²)
  ❌ H_2O or H_{2}O      ✅ H₂O
  ❌ CO_2                ✅ CO₂
  ❌ \\pi r^2            ✅ πr²
  ❌ \\alpha + \\beta    ✅ α + β
  ❌ \\int_0^1 f(x)dx    ✅ ∫₀¹ f(x)dx  or  ∫ f(x)dx from 0 to 1
  ❌ a \\times b          ✅ a × b
  ❌ \\sin(\\theta)       ✅ sin(θ)
  ❌ \\Delta x            ✅ Δx
  ❌ 90^{\\circ}          ✅ 90°
  ❌ 2.5 \\times 10^3     ✅ 2.5 × 10³

Every field — question_text, option_text, hint, explanation — must pass this rule.
"""


def add_wrong_question_to_mission(user, question, chapter):
    """
    Add a wrongly answered question to the user's mission and find 2 more similar questions.
    
    Args:
        user: The user who answered incorrectly
        question: The question that was answered incorrectly
        chapter: The chapter the question belongs to
    
    Returns:
        dict with mission info and questions added
    """
    from datetime import date, timedelta
    
    try:
        module = chapter.module
        subject = module.subject
        
        tomorrow = date.today() + timedelta(days=1)
        mission, mission_created = Mission.objects.get_or_create(
            account=user,
            subject=subject,
            mission_date=tomorrow,
            defaults={'is_deleted': False}
        )
        
        questions_added = []
        
        max_order = MissionQuestion.objects.filter(mission=mission).aggregate(
            max_order=Max('order')
        )['max_order'] or 0
        current_order = max_order + 1
        
        mission_question, created = MissionQuestion.objects.get_or_create(
            mission=mission,
            question=question,
            defaults={
                'chapter': chapter,
                'order': current_order
            }
        )
        
        if created:
            questions_added.append(question.id)
            current_order += 1
            api_logger.info(f"Added wrong question {question.id} to mission {mission.id} for user {user.username}")
        else:
            api_logger.info(f"Question {question.id} already in mission {mission.id} for user {user.username}")
        
        answered_question_ids = Answer.objects.filter(user=user).values_list('question_id', flat=True)
        
        mission_question_ids = MissionQuestion.objects.filter(mission=mission).values_list('question_id', flat=True)
        
        similar_questions = Question.objects.filter(
            module_contents__chapter=chapter,
            module_contents__is_deleted=False,
            level=question.level,
            is_active=True,
            is_deleted=False
        ).exclude(
            id__in=answered_question_ids
        ).exclude(
            id__in=mission_question_ids
        ).exclude(
            id=question.id
        ).distinct()[:2]
        
        for similar_question in similar_questions:
            MissionQuestion.objects.create(
                mission=mission,
                question=similar_question,
                chapter=chapter,
                order=current_order
            )
            questions_added.append(similar_question.id)
            current_order += 1
            api_logger.info(f"Added similar question {similar_question.id} to mission {mission.id} for user {user.username}")
        
        return {
            'mission_id': str(mission.id),
            'mission_created': mission_created,
            'subject_id': str(subject.id),
            'subject_name': subject.name,
            'questions_added': questions_added,
            'total_mission_questions': mission.questions.count()
        }
        
    except Exception as e:
        api_logger.error(f"Error adding question to mission for user {user.username}: {str(e)}")
        return None


@api_view(['POST'])
@drf_permission_classes([permissions.IsAuthenticated])
def generate_ai_questions(request):
    """
    Generate AI questions using ChatGPT and create entries in the database.
    
    Expected Input:
    {
        "class_id": optional (UUID) - Class ID for filtering,
        "subject_id": required (UUID) - Subject ID,
        "module_id": required (UUID) - Module ID,
        "chapter_id": required (UUID) - Chapter ID where questions will be created,
        "subject_name": required (str) - Name of the subject for prompt context,
        "module_name": required (str) - Name of the module for prompt context,
        "chapter_name": required (str) - Name of the chapter for prompt context,
        "number_of_questions": required (int) - Number of questions to generate (e.g., 3, 5, 10, 15, 20),
        "level": required (int) - Difficulty level 1-5 where:
            1 = Basic, 2 = Easy, 3 = Medium, 4 = Hard, 5 = HOTS (Advanced),
        "question_type": optional (str) - Type of questions to generate:
            "mcq_single" (default) - Multiple choice with single correct answer,
            "mcq_multiple" - Multiple choice with multiple correct answers,
            "rearrange" - Re-arrange/ordering questions
    }
    
    Expected Output:
    {
        "success": true,
        "data": {
            "questions_created": int - Number of questions created,
            "module_content_created": int - Number of module content entries (levels 1-4),
            "hots_created": int - Number of HOTS entries (level 5),
            "level": int - The difficulty level used,
            "is_hots": bool - Whether HOTS questions were created,
            "chapter_id": str - UUID of the chapter,
            "questions": [
                {
                    "id": str (UUID),
                    "question_text": str,
                    "question_type": str (mcq_single|mcq_multiple|rearrange),
                    "difficulty_level": str (easy|medium|hard),
                    "level": int (1-5),
                    "explanation": str,
                    "exp_points": int,
                    "is_hots": bool,
                    "ai_generated": true,
                    "options": [
                        {
                            "id": str (UUID),
                            "option_text": str,
                            "is_correct": bool,
                            "order": int
                        }
                    ]
                }
            ]
        },
        "message": str
    }
    
    Creates:
    - Question entries (with ai_generated=True)
    - Option entries for each question
    - ModuleContent entries (if level 1-4)
    - ChapterHOTS entries (if level 5)
    """
    api_logger.info(f"AI question generation requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")
    
    try:
        subject_id = request.data.get('subject_id')
        module_id = request.data.get('module_id')
        chapter_id = request.data.get('chapter_id')
        subject_name = request.data.get('subject_name', '')
        module_name = request.data.get('module_name', '')
        chapter_name = request.data.get('chapter_name', '')
        number_of_questions = request.data.get('number_of_questions', 5)
        level = request.data.get('level', 1)
        question_type = request.data.get('question_type', 'mcq_single')
        
        valid_question_types = ['mcq_single', 'mcq_multiple', 'rearrange']
        if question_type not in valid_question_types:
            question_type = 'mcq_single'
        
        if not all([subject_id, module_id, chapter_id]):
            return validation_error({
                "error": "subject_id, module_id, and chapter_id are required"
            })
        
        try:
            number_of_questions = int(number_of_questions)
            level = int(level)
        except (ValueError, TypeError):
            return validation_error({
                "error": "number_of_questions and level must be integers"
            })
        
        if level < 1 or level > 5:
            return validation_error({
                "error": "level must be between 1 and 5"
            })
        
        try:
            chapter = ModuleChapter.objects.get(id=chapter_id)
        except ModuleChapter.DoesNotExist:
            return validation_error({
                "error": f"Chapter with ID {chapter_id} not found"
            })
        
        is_hots = level == 5
        
        level_to_difficulty = {
            1: 'easy',
            2: 'easy',
            3: 'medium',
            4: 'hard',
            5: 'hard'
        }
        difficulty = level_to_difficulty.get(level, 'medium')
        
        generated_questions = generate_questions_with_chatgpt(
            subject_name=subject_name,
            module_name=module_name,
            chapter_name=chapter_name,
            level=level,
            number_of_questions=number_of_questions,
            question_type=question_type
        )
        
        if not generated_questions:
            return validation_error({
                "error": "Failed to generate questions from AI. Please try again."
            })
        
        questions_created = 0
        module_content_created = 0
        hots_created = 0
        created_questions_list = []
        
        for q_data in generated_questions:
            try:
                question = Question.objects.create(
                    question_text=q_data.get('question_text', ''),
                    question_type=question_type,
                    difficulty_level=difficulty,
                    explanation=q_data.get('explanation', ''),
                    exp_points=q_data.get('exp_points', 10),
                    is_active=True,
                    is_hots=is_hots,
                    ai_generated=True,
                    level=level,
                    created_by=request.user
                )
                questions_created += 1
                
                options_data = q_data.get('options', [])
                created_options = []
                for idx, opt_data in enumerate(options_data):
                    if question_type == 'rearrange':
                        order = opt_data.get('correct_order')
                        if order is None or not isinstance(order, (int, float)):
                            order = idx + 1
                        order = max(1, int(order))
                    else:
                        order = idx + 1
                    option = Option.objects.create(
                        question=question,
                        option_text=opt_data.get('option_text', f'Option {idx + 1}'),
                        is_correct=opt_data.get('is_correct', False),
                        order=order
                    )
                    created_options.append({
                        'id': str(option.id),
                        'option_text': option.option_text,
                        'is_correct': option.is_correct,
                        'order': option.order
                    })
                
                created_questions_list.append({
                    'id': str(question.id),
                    'question_text': question.question_text,
                    'question_type': question.question_type,
                    'difficulty_level': question.difficulty_level,
                    'level': question.level,
                    'explanation': question.explanation,
                    'exp_points': question.exp_points,
                    'is_hots': question.is_hots,
                    'ai_generated': question.ai_generated,
                    'options': created_options
                })
                
                if is_hots:
                    max_order = ChapterHOTS.objects.filter(chapter=chapter).aggregate(
                        max_order=Max('order')
                    )['max_order'] or 0
                    
                    ChapterHOTS.objects.create(
                        chapter=chapter,
                        question=question,
                        order=max_order + 1,
                        created_by=request.user
                    )
                    hots_created += 1
                    
                    chapter.has_hots = True
                    chapter.save(update_fields=['has_hots'])
                else:
                    max_order = ModuleContent.objects.filter(chapter=chapter).aggregate(
                        max_order=Max('order')
                    )['max_order'] or 0
                    
                    ModuleContent.objects.create(
                        chapter=chapter,
                        content_type='question',
                        question=question,
                        order=max_order + 1,
                        created_by=request.user
                    )
                    module_content_created += 1
                
                api_logger.info(f"Created question {question.id} for chapter {chapter_id}")
                
            except Exception as e:
                api_logger.error(f"Error creating question: {str(e)}")
                continue
        
        api_logger.info(f"AI generation complete: {questions_created} questions, {module_content_created} module content, {hots_created} HOTS entries")
        
        return success(
            data={
                'questions_created': questions_created,
                'module_content_created': module_content_created,
                'hots_created': hots_created,
                'level': level,
                'is_hots': is_hots,
                'questions': created_questions_list,
                'chapter_id': str(chapter_id),
            },
            message=f"Successfully generated {questions_created} questions"
        )
        
    except Exception as e:
        api_logger.error(f"Error in AI question generation: {str(e)}")
        return validation_error({
            "error": f"Failed to generate questions: {str(e)}"
        })


@api_view(['POST'])
@drf_permission_classes([permissions.IsAuthenticated])
def deactivate_ai_questions(request):
    """
    Deactivate AI-generated questions whose IDs are in the provided list.

    Expected input:
    {
        "question_ids": [list of question IDs to mark inactive],
        "module_chapter_id": required (UUID of the module chapter)
    }

    This will mark all AI-generated questions whose IDs are in question_ids as inactive (is_active=False).
    """
    api_logger.info(f"Deactivate AI questions requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")

    try:
        question_ids = request.data.get('question_ids', [])
        module_chapter_id = request.data.get('module_chapter_id')

        if not module_chapter_id:
            return validation_error({
                "error": "module_chapter_id is required"
            })

        if not isinstance(question_ids, list):
            question_ids = [question_ids] if question_ids else []

        import uuid as _uuid
        valid_question_ids = []
        for qid in question_ids:
            try:
                _uuid.UUID(str(qid))
                valid_question_ids.append(qid)
            except (ValueError, AttributeError):
                pass
        question_ids = valid_question_ids

        try:
            chapter = ModuleChapter.objects.get(id=module_chapter_id)
        except ModuleChapter.DoesNotExist:
            return validation_error({
                "error": f"ModuleChapter with ID {module_chapter_id} not found"
            })
        
        ai_questions_to_deactivate = Question.objects.filter(
            ai_generated=True,
            is_active=True,
            id__in=question_ids,
            module_contents__chapter=chapter,
            module_contents__is_deleted=False
        ).distinct()

        ai_hots_questions_to_deactivate = Question.objects.filter(
            ai_generated=True,
            is_active=True,
            id__in=question_ids,
            chapter_hots__chapter=chapter
        ).distinct()

        all_questions_to_deactivate = (ai_questions_to_deactivate | ai_hots_questions_to_deactivate).distinct()

        count_to_deactivate = all_questions_to_deactivate.count()

        deactivated_ids = list(all_questions_to_deactivate.values_list('id', flat=True))

        all_questions_to_deactivate.update(is_active=False)

        api_logger.info(f"Deactivated {count_to_deactivate} AI-generated questions for chapter {module_chapter_id}. IDs: {deactivated_ids}")

        return success(
            data={
                'deactivated_count': count_to_deactivate,
                'deactivated_question_ids': [str(qid) for qid in deactivated_ids],
                'module_chapter_id': str(module_chapter_id),
            },
            message=f"Successfully deactivated {count_to_deactivate} AI-generated questions"
        )
        
    except Exception as e:
        api_logger.error(f"Error in deactivate AI questions: {str(e)}")
        return validation_error({
            "error": f"Failed to deactivate questions: {str(e)}"
        })


@api_view(['POST'])
@drf_permission_classes([permissions.IsAuthenticated])
def activate_ai_questions(request):
    """
    Activate AI-generated questions by ID.

    Expected input:
    {
        "question_ids": [list of question UUIDs to activate]
    }

    Marks the specified questions as is_active=True.
    Questions not in the list remain untouched.
    """
    api_logger.info(f"Activate AI questions requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")

    try:
        question_ids = request.data.get('question_ids', [])
        if not isinstance(question_ids, list):
            question_ids = [question_ids] if question_ids else []
        question_ids = [str(qid).strip() for qid in question_ids if qid]

        if not question_ids:
            return validation_error({"error": "question_ids is required and must not be empty"})

        updated_count = Question.objects.filter(id__in=question_ids, is_active=False).update(is_active=True)

        api_logger.info(f"Activated {updated_count} questions: {question_ids}")
        return success(
            data={
                'activated_count': updated_count,
                'question_ids': question_ids,
            },
            message=f"Successfully activated {updated_count} questions"
        )

    except Exception as e:
        api_logger.error(f"Error in activate AI questions: {str(e)}")
        return validation_error({"error": f"Failed to activate questions: {str(e)}"})


def generate_questions_with_chatgpt(subject_name, module_name, chapter_name, level, number_of_questions, question_type='mcq_single'):
    """
    Generate questions using ChatGPT API.
    
    Args:
        subject_name (str): Name of the subject
        module_name (str): Name of the module
        chapter_name (str): Name of the chapter
        level (int): Difficulty level (1-5, where 5 is HOTS)
        number_of_questions (int): Number of questions to generate
        question_type (str): Type of question to generate. Valid values:
            - 'mcq_single': Multiple Choice Question with single correct answer
            - 'mcq_multiple': Multiple Choice Question with multiple correct answers
            - 'rearrange': Re-arrange/ordering question
    
    Returns:
        list: A list of question dictionaries with the following structure:
        
        For mcq_single:
        {
            'question_text': str,
            'question_type': 'mcq_single',
            'explanation': str,
            'exp_points': int (default: 10),
            'options': [
                {'option_text': str, 'is_correct': bool},
                ...
            ]
        }
        
        For mcq_multiple:
        {
            'question_text': str (includes "Select all that apply"),
            'question_type': 'mcq_multiple',
            'explanation': str,
            'exp_points': int (default: 15),
            'options': [
                {'option_text': str, 'is_correct': bool},
                ...
            ]
        }
        
        For rearrange:
        {
            'question_text': str (describes what to arrange),
            'question_type': 'rearrange',
            'explanation': str,
            'exp_points': int (default: 20),
            'options': [
                {'option_text': str, 'is_correct': True, 'correct_order': int},
                ...
            ]
        }
    """
    try:
        import openai
        
        openai_api_key = os.environ.get('OPENAI_API_KEY')
        
        if not openai_api_key:
            api_logger.error("OPENAI_API_KEY not found in environment")
            raise Exception("OPENAI_API_KEY not found in environment")
        
        client = openai.OpenAI(api_key=openai_api_key)
        
        level_descriptions = {
            1: "Very basic and simple questions suitable for beginners. Focus on recall and basic understanding.",
            2: "Easy questions that test fundamental concepts and simple applications.",
            3: "Medium difficulty questions that require understanding and application of concepts.",
            4: "Hard questions that require analysis and deeper understanding of the topic.",
            5: "Higher Order Thinking Skills (HOTS) questions that require critical thinking, analysis, synthesis, and evaluation. These are challenging questions that test deep understanding."
        }
        
        level_desc = level_descriptions.get(level, level_descriptions[3])
        
        if question_type == 'mcq_single':
            prompt = f"""You are an expert educational content creator. Generate {number_of_questions} REAL, FACTUALLY ACCURATE multiple choice questions for students studying this topic.

Subject: {subject_name}
Module/Unit: {module_name}
Chapter/Topic: {chapter_name}
Difficulty Level: {level} ({level_desc})
Question Type: mcq_single (Multiple Choice - Single Correct Answer)
{MOBILE_FORMAT_RULES}
CONTENT REQUIREMENTS:
1. Questions MUST be based on REAL, FACTUAL content from this topic
2. All answers must be SCIENTIFICALLY/ACADEMICALLY ACCURATE and verifiable
3. Each question should have exactly 4 options (A, B, C, D)
4. Only ONE option should be correct
5. Include a detailed explanation for why the correct answer is correct
6. Incorrect options (distractors) should be plausible but clearly incorrect
7. Questions should test real understanding, not just memorization
8. Use proper terminology and vocabulary appropriate for the subject
9. DO NOT use placeholder text - write REAL questions with REAL content

Return the response in the following JSON format:
{{
    "questions": [
        {{
            "question_text": "The actual question text with real content",
            "question_type": "mcq_single",
            "options": [
                {{"option_text": "Actual option with real content", "is_correct": false}},
                {{"option_text": "Actual correct answer", "is_correct": true}},
                {{"option_text": "Actual plausible distractor", "is_correct": false}},
                {{"option_text": "Actual plausible distractor", "is_correct": false}}
            ],
            "explanation": "Detailed explanation of why the correct answer is correct and why others are wrong",
            "exp_points": 10
        }}
    ]
}}

Generate exactly {number_of_questions} questions. Respond only with valid JSON."""

        elif question_type == 'mcq_multiple':
            prompt = f"""You are an expert educational content creator. Generate {number_of_questions} REAL, FACTUALLY ACCURATE multiple choice questions (with multiple correct answers) for students studying this topic.

Subject: {subject_name}
Module/Unit: {module_name}
Chapter/Topic: {chapter_name}
Difficulty Level: {level} ({level_desc})
Question Type: mcq_multiple (Multiple Choice - Multiple Correct Answers)
{MOBILE_FORMAT_RULES}
CONTENT REQUIREMENTS:
1. Questions MUST be based on REAL, FACTUAL content from this topic
2. All answers must be SCIENTIFICALLY/ACADEMICALLY ACCURATE and verifiable
3. Each question should have exactly 4 options (A, B, C, D)
4. TWO or MORE options should be correct
5. Include a detailed explanation for why the correct answers are correct
6. Incorrect options (distractors) should be plausible but clearly incorrect
7. The question text should indicate that multiple answers may be correct (e.g., "Select all that apply")
8. Use proper terminology and vocabulary appropriate for the subject
9. DO NOT use placeholder text - write REAL questions with REAL content

Return the response in the following JSON format:
{{
    "questions": [
        {{
            "question_text": "Select all that apply: The actual question with real content",
            "question_type": "mcq_multiple",
            "options": [
                {{"option_text": "Actual correct answer 1", "is_correct": true}},
                {{"option_text": "Actual correct answer 2", "is_correct": true}},
                {{"option_text": "Actual plausible distractor", "is_correct": false}},
                {{"option_text": "Actual plausible distractor", "is_correct": false}}
            ],
            "explanation": "Detailed explanation of why the correct answers are correct",
            "exp_points": 15
        }}
    ]
}}

Generate exactly {number_of_questions} questions. Respond only with valid JSON."""

        elif question_type == 'rearrange':
            prompt = f"""You are an expert educational content creator. Generate {number_of_questions} REAL, FACTUALLY ACCURATE re-arrange/ordering questions for students studying this topic.

Subject: {subject_name}
Module/Unit: {module_name}
Chapter/Topic: {chapter_name}
Difficulty Level: {level} ({level_desc})
Question Type: rearrange (Re-arrange/Ordering Question)
{MOBILE_FORMAT_RULES}
CONTENT REQUIREMENTS:
1. Questions MUST be based on REAL, FACTUAL sequences from this topic
2. All sequences must be ACADEMICALLY ACCURATE and verifiable
3. Each question should ask the user to arrange items in the correct order
4. Provide 4-6 items that need to be arranged in sequence
5. The options should be given in SCRAMBLED order, with correct_order field indicating the CORRECT position (1, 2, 3, etc.)
6. Include a detailed explanation for why this is the correct sequence
7. Use proper terminology and vocabulary appropriate for the subject
8. DO NOT use placeholder text - write REAL content with REAL steps/items from the topic

Return the response in the following JSON format:
{{
    "questions": [
        {{
            "question_text": "Arrange the following [actual items] in the correct order:",
            "question_type": "rearrange",
            "options": [
                {{"option_text": "Actual third item/step from the topic", "is_correct": true, "correct_order": 3}},
                {{"option_text": "Actual first item/step from the topic", "is_correct": true, "correct_order": 1}},
                {{"option_text": "Actual fourth item/step from the topic", "is_correct": true, "correct_order": 4}},
                {{"option_text": "Actual second item/step from the topic", "is_correct": true, "correct_order": 2}}
            ],
            "explanation": "Detailed explanation of why this is the correct order based on the subject matter",
            "exp_points": 20
        }}
    ]
}}

Generate exactly {number_of_questions} questions. Respond only with valid JSON."""

        else:
            prompt = f"""You are an expert educational content creator. Generate {number_of_questions} REAL, FACTUALLY ACCURATE multiple choice questions for students studying this topic.

Subject: {subject_name}
Module/Unit: {module_name}
Chapter/Topic: {chapter_name}
Difficulty Level: {level} ({level_desc})
Question Type: mcq_single (Multiple Choice - Single Correct Answer)
{MOBILE_FORMAT_RULES}
CONTENT REQUIREMENTS:
1. Questions MUST be based on REAL, FACTUAL content from this topic
2. All answers must be SCIENTIFICALLY/ACADEMICALLY ACCURATE and verifiable
3. Each question should have exactly 4 options (A, B, C, D)
4. Only ONE option should be correct
5. Include a detailed explanation for why the correct answer is correct
6. Incorrect options (distractors) should be plausible but clearly incorrect
7. Use proper terminology and vocabulary appropriate for the subject
8. DO NOT use placeholder text - write REAL questions with REAL content

Return the response in the following JSON format:
{{
    "questions": [
        {{
            "question_text": "The actual question text with real content",
            "question_type": "mcq_single",
            "options": [
                {{"option_text": "Actual option with real content", "is_correct": false}},
                {{"option_text": "Actual correct answer", "is_correct": true}},
                {{"option_text": "Actual plausible distractor", "is_correct": false}},
                {{"option_text": "Actual plausible distractor", "is_correct": false}}
            ],
            "explanation": "Detailed explanation of why the correct answer is correct",
            "exp_points": 10
        }}
    ]
}}

Generate exactly {number_of_questions} questions. Respond only with valid JSON."""

        system_messages = {
            'mcq_single': f"You are an expert educational content creator. Output ONLY valid JSON. NEVER use LaTeX, caret (^), underscore subscript, or any math notation. Use Unicode symbols only (², √, π, ×, H₂O, etc.).\n{MOBILE_FORMAT_RULES}",
            'mcq_multiple': f"You are an expert educational content creator. Output ONLY valid JSON. NEVER use LaTeX, caret (^), underscore subscript, or any math notation. Use Unicode symbols only (², √, π, ×, H₂O, etc.).\n{MOBILE_FORMAT_RULES}",
            'rearrange': f"You are an expert educational content creator. Output ONLY valid JSON. NEVER use LaTeX, caret (^), underscore subscript, or any math notation. Use Unicode symbols only (², √, π, ×, H₂O, etc.).\n{MOBILE_FORMAT_RULES}",
        }
        
        system_message = system_messages.get(question_type, system_messages['mcq_single'])

        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {
                    "role": "system",
                    "content": system_message
                },
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            temperature=0.7,
            max_tokens=4000
        )
        
        response_text = response.choices[0].message.content.strip()
        
        try:
            if '```json' in response_text:
                response_text = response_text.split('```json')[1].split('```')[0].strip()
            elif '```' in response_text:
                response_text = response_text.split('```')[1].split('```')[0].strip()
            
            result = json.loads(response_text)
            questions = result.get('questions', [])
            
            for q in questions:
                q['question_type'] = question_type
            
            api_logger.info(f"Successfully generated {len(questions)} {question_type} questions from ChatGPT")
            return questions
            
        except json.JSONDecodeError as e:
            api_logger.error(f"Failed to parse ChatGPT response as JSON: {e}")
            api_logger.error(f"Response text: {response_text[:500]}")
            raise Exception(f"Failed to parse ChatGPT response as JSON: {e}")
            
    except ImportError as e:
        api_logger.error("openai package not installed")
        raise Exception("openai package not installed")
    except Exception as e:
        api_logger.error(f"Error calling ChatGPT API: {str(e)}")
        raise


def generate_mock_questions(subject_name, module_name, chapter_name, level, number_of_questions, question_type='mcq_single'):
    """
    Generate mock questions when ChatGPT is not available.
    This is a fallback for development/testing purposes.
    """
    api_logger.info(f"Generating {number_of_questions} mock {question_type} questions for {subject_name}/{module_name}/{chapter_name}")
    
    level_desc = {
        1: "Basic",
        2: "Easy",
        3: "Medium",
        4: "Hard",
        5: "HOTS"
    }.get(level, "Medium")
    
    questions = []
    for i in range(number_of_questions):
        if question_type == 'mcq_single':
            question = {
                "question_text": f"[AI Generated - {level_desc}] Question {i + 1} about {chapter_name} in {module_name} ({subject_name}): What is the key concept related to this topic?",
                "question_type": "mcq_single",
                "explanation": f"This is an AI-generated explanation for the correct answer regarding {chapter_name}. The correct option demonstrates understanding of the core concept.",
                "exp_points": 10 + (level * 2),
                "options": [
                    {"option_text": f"Option A - This is the correct answer for {chapter_name}", "is_correct": True},
                    {"option_text": f"Option B - This is an incorrect but plausible answer", "is_correct": False},
                    {"option_text": f"Option C - This is another distractor option", "is_correct": False},
                    {"option_text": f"Option D - This is the final distractor option", "is_correct": False}
                ]
            }
        elif question_type == 'mcq_multiple':
            question = {
                "question_text": f"[AI Generated - {level_desc}] Select all that apply: Question {i + 1} about {chapter_name} in {module_name} ({subject_name}): Which of the following are key concepts?",
                "question_type": "mcq_multiple",
                "explanation": f"This is an AI-generated explanation for the correct answers regarding {chapter_name}. Multiple options are correct.",
                "exp_points": 15 + (level * 2),
                "options": [
                    {"option_text": f"Option A - First correct answer for {chapter_name}", "is_correct": True},
                    {"option_text": f"Option B - Second correct answer", "is_correct": True},
                    {"option_text": f"Option C - This is an incorrect option", "is_correct": False},
                    {"option_text": f"Option D - This is another incorrect option", "is_correct": False}
                ]
            }
        elif question_type == 'rearrange':
            question = {
                "question_text": f"[AI Generated - {level_desc}] Question {i + 1}: Arrange the following steps in the correct order for {chapter_name}:",
                "question_type": "rearrange",
                "explanation": f"This is the correct sequence for understanding {chapter_name}.",
                "exp_points": 20 + (level * 2),
                "options": [
                    {"option_text": "Step 3: Apply the concept", "is_correct": True, "correct_order": 3},
                    {"option_text": "Step 1: Understand the basics", "is_correct": True, "correct_order": 1},
                    {"option_text": "Step 4: Evaluate results", "is_correct": True, "correct_order": 4},
                    {"option_text": "Step 2: Analyze the problem", "is_correct": True, "correct_order": 2}
                ]
            }
        else:
            question = {
                "question_text": f"[AI Generated - {level_desc}] Question {i + 1} about {chapter_name} in {module_name} ({subject_name}): What is the key concept?",
                "question_type": question_type,
                "explanation": f"This is an AI-generated explanation regarding {chapter_name}.",
                "exp_points": 10 + (level * 2),
                "options": [
                    {"option_text": f"Option A - Correct answer", "is_correct": True},
                    {"option_text": f"Option B - Incorrect", "is_correct": False},
                    {"option_text": f"Option C - Incorrect", "is_correct": False},
                    {"option_text": f"Option D - Incorrect", "is_correct": False}
                ]
            }
        questions.append(question)
    
    return questions


def generate_questions_with_gemini(subject_name, module_name, chapter_name, level, number_of_questions, question_type='mcq_single'):
    """
    Generate questions using Google Gemini API.
    
    Args:
        subject_name (str): Name of the subject
        module_name (str): Name of the module
        chapter_name (str): Name of the chapter
        level (int): Difficulty level (1-5, where 5 is HOTS)
        number_of_questions (int): Number of questions to generate
        question_type (str): Type of question to generate. Valid values:
            - 'mcq_single': Multiple Choice Question with single correct answer
            - 'mcq_multiple': Multiple Choice Question with multiple correct answers
            - 'rearrange': Re-arrange/ordering question
    
    Returns:
        list: A list of question dictionaries (same structure as ChatGPT version)
    """
    try:
        import google.generativeai as genai
        
        gemini_api_key = os.environ.get('GEMINI_API_KEY')
        
        if not gemini_api_key:
            api_logger.error("GEMINI_API_KEY not found in environment")
            raise Exception("GEMINI_API_KEY not found in environment")
        
        genai.configure(api_key=gemini_api_key)
        
        level_descriptions = {
            1: "Very basic and simple questions suitable for beginners. Focus on recall and basic understanding.",
            2: "Easy questions that test fundamental concepts and simple applications.",
            3: "Medium difficulty questions that require understanding and application of concepts.",
            4: "Hard questions that require analysis and deeper understanding of the topic.",
            5: "Higher Order Thinking Skills (HOTS) questions that require critical thinking, analysis, synthesis, and evaluation. These are challenging questions that test deep understanding."
        }
        
        level_desc = level_descriptions.get(level, level_descriptions[3])
        
        if question_type == 'mcq_single':
            prompt = f"""You are an expert educational content creator. Generate {number_of_questions} REAL, FACTUALLY ACCURATE multiple choice questions for students studying this topic.

Subject: {subject_name}
Module/Unit: {module_name}
Chapter/Topic: {chapter_name}
Difficulty Level: {level} ({level_desc})
Question Type: mcq_single (Multiple Choice - Single Correct Answer){class_instruction}

{MOBILE_FORMAT_RULES}

CONTENT REQUIREMENTS:
1. Questions MUST be based on REAL, FACTUAL content from this topic - use actual facts, formulas, definitions, and concepts
2. All answers must be SCIENTIFICALLY/ACADEMICALLY ACCURATE and verifiable
3. Each question should have exactly 4 options (A, B, C, D)
4. Only ONE option should be correct
5. Include a detailed explanation for why the correct answer is correct
6. Incorrect options (distractors) should be plausible but clearly incorrect based on the subject matter
7. Questions should test real understanding, not just memorization
8. Use proper terminology and vocabulary appropriate for the subject
9. DO NOT use placeholder text like "Option A" or "Question about X" - write REAL questions with REAL content

Return the response in the following JSON format:
{{
    "questions": [
        {{
            "question_text": "The actual question text with real content",
            "question_type": "mcq_single",
            "options": [
                {{"option_text": "Actual option with real content", "is_correct": false}},
                {{"option_text": "Actual correct answer", "is_correct": true}},
                {{"option_text": "Actual plausible distractor", "is_correct": false}},
                {{"option_text": "Actual plausible distractor", "is_correct": false}}
            ],
            "explanation": "Detailed explanation of why the correct answer is correct and why others are wrong",
            "exp_points": 10
        }}
    ]
}}

Generate exactly {number_of_questions} questions. Respond only with valid JSON, no markdown formatting."""

        elif question_type == 'mcq_multiple':
            prompt = f"""You are an expert educational content creator. Generate {number_of_questions} REAL, FACTUALLY ACCURATE multiple choice questions (with multiple correct answers) for students studying this topic.

Subject: {subject_name}
Module/Unit: {module_name}
Chapter/Topic: {chapter_name}
Difficulty Level: {level} ({level_desc})
Question Type: mcq_multiple (Multiple Choice - Multiple Correct Answers)

{MOBILE_FORMAT_RULES}

CONTENT REQUIREMENTS:
1. Questions MUST be based on REAL, FACTUAL content from this topic - use actual facts, formulas, definitions, and concepts
2. All answers must be SCIENTIFICALLY/ACADEMICALLY ACCURATE and verifiable
3. Each question should have exactly 4 options (A, B, C, D)
4. TWO or MORE options should be correct
5. Include a detailed explanation for why the correct answers are correct
6. Incorrect options (distractors) should be plausible but clearly incorrect based on the subject matter
7. The question text should indicate that multiple answers may be correct (e.g., "Select all that apply")
8. Use proper terminology and vocabulary appropriate for the subject
9. DO NOT use placeholder text - write REAL questions with REAL content

Return the response in the following JSON format:
{{
    "questions": [
        {{
            "question_text": "Select all that apply: The actual question with real content",
            "question_type": "mcq_multiple",
            "options": [
                {{"option_text": "Actual correct answer 1", "is_correct": true}},
                {{"option_text": "Actual correct answer 2", "is_correct": true}},
                {{"option_text": "Actual plausible distractor", "is_correct": false}},
                {{"option_text": "Actual plausible distractor", "is_correct": false}}
            ],
            "explanation": "Detailed explanation of why the correct answers are correct",
            "exp_points": 15
        }}
    ]
}}

Generate exactly {number_of_questions} questions. Respond only with valid JSON, no markdown formatting."""

        elif question_type == 'rearrange':
            prompt = f"""You are an expert educational content creator. Generate {number_of_questions} REAL, FACTUALLY ACCURATE re-arrange/ordering questions for students studying this topic.

Subject: {subject_name}
Module/Unit: {module_name}
Chapter/Topic: {chapter_name}
Difficulty Level: {level} ({level_desc})
Question Type: rearrange (Re-arrange/Ordering Question)

{MOBILE_FORMAT_RULES}

CONTENT REQUIREMENTS:
1. Questions MUST be based on REAL, FACTUAL sequences from this topic - use actual processes, historical events, scientific procedures, or logical sequences
2. All sequences must be ACADEMICALLY ACCURATE and verifiable
3. Each question should ask the user to arrange items in the correct order
4. Provide 4-6 items that need to be arranged in sequence
5. The options should be given in SCRAMBLED order, with correct_order field indicating the CORRECT position (1, 2, 3, etc.)
6. Include a detailed explanation for why this is the correct sequence
7. Use proper terminology and vocabulary appropriate for the subject
8. DO NOT use placeholder text like "Step 1", "Step 2" - write REAL content with REAL steps/items from the topic

Return the response in the following JSON format:
{{
    "questions": [
        {{
            "question_text": "Arrange the following [actual items] in the correct order:",
            "question_type": "rearrange",
            "options": [
                {{"option_text": "Actual third item/step from the topic", "is_correct": true, "correct_order": 3}},
                {{"option_text": "Actual first item/step from the topic", "is_correct": true, "correct_order": 1}},
                {{"option_text": "Actual fourth item/step from the topic", "is_correct": true, "correct_order": 4}},
                {{"option_text": "Actual second item/step from the topic", "is_correct": true, "correct_order": 2}}
            ],
            "explanation": "Detailed explanation of why this is the correct order based on the subject matter",
            "exp_points": 20
        }}
    ]
}}

Generate exactly {number_of_questions} questions. Respond only with valid JSON, no markdown formatting."""

        else:
            prompt = f"""You are an expert educational content creator. Generate {number_of_questions} REAL, FACTUALLY ACCURATE multiple choice questions for students studying this topic.

Subject: {subject_name}
Module/Unit: {module_name}
Chapter/Topic: {chapter_name}
Difficulty Level: {level} ({level_desc})
Question Type: mcq_single (Multiple Choice - Single Correct Answer){class_instruction}

{MOBILE_FORMAT_RULES}

CONTENT REQUIREMENTS:
1. Questions MUST be based on REAL, FACTUAL content from this topic - use actual facts, formulas, definitions, and concepts
2. All answers must be SCIENTIFICALLY/ACADEMICALLY ACCURATE and verifiable
3. Each question should have exactly 4 options (A, B, C, D)
4. Only ONE option should be correct
5. Include a detailed explanation for why the correct answer is correct
6. Incorrect options (distractors) should be plausible but clearly incorrect
7. Use proper terminology and vocabulary appropriate for the subject
8. DO NOT use placeholder text - write REAL questions with REAL content

Return the response in the following JSON format:
{{
    "questions": [
        {{
            "question_text": "The actual question text with real content",
            "question_type": "mcq_single",
            "options": [
                {{"option_text": "Actual option with real content", "is_correct": false}},
                {{"option_text": "Actual correct answer", "is_correct": true}},
                {{"option_text": "Actual plausible distractor", "is_correct": false}},
                {{"option_text": "Actual plausible distractor", "is_correct": false}}
            ],
            "explanation": "Detailed explanation of why the correct answer is correct",
            "exp_points": 10
        }}
    ]
}}

Generate exactly {number_of_questions} questions. Respond only with valid JSON, no markdown formatting."""

        system_instructions = {
            'mcq_single': f"{MOBILE_FORMAT_RULES}\nYou are an expert educational content creator specializing in creating curriculum-aligned assessment questions. You create high-quality, FACTUALLY ACCURATE, and pedagogically sound multiple choice questions with single correct answers. All content must be real, verifiable, and based on actual educational curriculum. Never use placeholder text. Always respond with valid JSON only.",
            'mcq_multiple': f"{MOBILE_FORMAT_RULES}\nYou are an expert educational content creator specializing in creating curriculum-aligned assessment questions. You create high-quality, FACTUALLY ACCURATE, and pedagogically sound multiple choice questions where multiple answers can be correct. All content must be real, verifiable, and based on actual educational curriculum. Never use placeholder text. Always respond with valid JSON only.",
            'rearrange': f"{MOBILE_FORMAT_RULES}\nYou are an expert educational content creator specializing in creating curriculum-aligned assessment questions. You create high-quality, FACTUALLY ACCURATE, and pedagogically sound sequencing/ordering questions. All content must be real, verifiable, and based on actual educational curriculum. Never use placeholder text. Always respond with valid JSON only."
        }
        
        system_instruction = system_instructions.get(question_type, system_instructions['mcq_single'])

        full_prompt = f"{system_instruction}\n\n{prompt}"
        
        model = genai.GenerativeModel("gemini-3-flash-preview")
        
        response = model.generate_content(
            full_prompt,
            generation_config=genai.GenerationConfig(
                temperature=0.7,
                max_output_tokens=8192,
            )
        )
        
        response_text = response.text.strip()
        
        try:
            if '```json' in response_text:
                response_text = response_text.split('```json')[1].split('```')[0].strip()
            elif '```' in response_text:
                response_text = response_text.split('```')[1].split('```')[0].strip()
            
            result = json.loads(response_text)
            questions = result.get('questions', [])
            
            for q in questions:
                q['question_type'] = question_type
            
            api_logger.info(f"Successfully generated {len(questions)} {question_type} questions from Gemini")
            return questions
            
        except json.JSONDecodeError as e:
            api_logger.error(f"Failed to parse Gemini response as JSON: {e}")
            api_logger.error(f"Response text: {response_text[:500]}")
            raise Exception(f"Failed to parse Gemini response as JSON: {e}")
            
    except ImportError as e:
        api_logger.error("google-generativeai package not installed")
        raise Exception("google-generativeai package not installed")
    except Exception as e:
        api_logger.error(f"Error calling Gemini API: {str(e)}")
        raise


def generate_questions_with_vertex_ai(subject_name, module_name, chapter_name, level, number_of_questions, question_type='mcq_single', add_image=False, use_matplot=False, class_context=None):
    """
    Generate questions using Google Vertex AI (Gemini models).
    
    Uses the gemini_client utility for Vertex AI integration.
    Requires GOOGLE_CLOUD_PROJECT and GOOGLE_APPLICATION_CREDENTIALS environment variables.
    
    Args:
        subject_name (str): Name of the subject
        module_name (str): Name of the module
        chapter_name (str): Name of the chapter
        level (int): Difficulty level (1-5, where 5 is HOTS)
        number_of_questions (int): Number of questions to generate
        question_type (str): Type of question to generate
        add_image (bool): Whether to generate and add images to questions
        use_matplot (bool): If True and add_image, use matplotlib code from Gemini (executed server-side) instead of Imagen
        class_context (str, optional): Target class/grade e.g. "Class 10" or "Grade 10" - used to tailor difficulty and vocabulary
    
    Returns:
        list: A list of question dictionaries
    """
    try:
        from gyaan_buddy.utils.gemini_client import gemini_generate_json
        
        level_descriptions = {
            1: "Very basic and simple questions suitable for beginners. Focus on recall and basic understanding.",
            2: "Easy questions that test fundamental concepts and simple applications.",
            3: "Medium difficulty questions that require understanding and application of concepts.",
            4: "Hard questions that require analysis and deeper understanding of the topic.",
            5: "Higher Order Thinking Skills (HOTS) questions that require critical thinking, analysis, synthesis, and evaluation. These are challenging questions that test deep understanding."
        }
        
        level_desc = level_descriptions.get(level, level_descriptions[3])
        
        class_instruction = ""
        if class_context and isinstance(class_context, str) and class_context.strip():
            class_instruction = f"""
TARGET AUDIENCE: These questions are for students of {class_context.strip()}. Use vocabulary, complexity, and depth appropriate for this class/grade level. Do not use content or terminology that is too advanced or too simplistic for this grade."""
        
        if question_type == 'mcq_single':
            image_instruction = ""
            image_json_field = ""
            if add_image and use_matplot:
                image_instruction = (
                    "\nIMAGE: For EACH question provide matplotlib_code — Python using only plt/np."
                    " The figure must be a BLANK SETUP DIAGRAM ONLY — it shows the visual problem"
                    " setup so the student can answer by looking at it."
                    " FORBIDDEN (never include these in the code):"
                    " (a) plt.title() or ax.set_title() — no question text as a title;"
                    " (b) plt.text() or ax.annotate() showing any answer value, measurement, or result"
                    " (e.g. do NOT write angle values like '90°', lengths like '5 cm', or computed results);"
                    " (c) option labels (A/B/C/D) as text annotations;"
                    " (d) any text that gives away the answer or hints at it."
                    " ALLOWED: vertex labels (single letters like A, B, C, D), axis labels, tick marks,"
                    " shape outlines, lines, curves, and given/known values that are part of the problem"
                    " statement (not the answer to be found)."
                    " The figure must look INCOMPLETE — the unknown value the question asks about must"
                    " NOT appear anywhere in the figure."
                    " No plt.savefig() or plt.show()."
                )
                image_json_field = (
                    ',\n            "matplotlib_code": "Python using plt/np ONLY.'
                    ' Draw ONLY the geometric setup: shapes with vertex labels (A,B,C,D),'
                    ' given side lengths or angles that are STATED in the question (not the answer).'
                    ' NEVER use plt.title(), NEVER annotate the answer value, NEVER show the correct'
                    ' measurement that the question asks the student to find.'
                    ' No plt.savefig() or plt.show()."'
                )
            elif add_image:
                image_instruction = "\nIMAGE (OPTIONAL): Only include image_description for questions that specifically require a visual diagram, chart, or figure. Most questions should NOT have an image. If included, provide a visual-only description (no text/labels). For image-based questions write question_text as 'Based on the figure above, ...'. Image must NOT imply the answer."
                image_json_field = ',\n            "image_description": "(OPTIONAL) Include ONLY if a visual diagram/chart/figure is essential to this specific question. Omit this field entirely for text-based questions."'

            prompt = f"""You are an expert educational content creator. Generate {number_of_questions} REAL, FACTUALLY ACCURATE multiple choice questions for students studying this topic.

Subject: {subject_name}
Module/Unit: {module_name}
Chapter/Topic: {chapter_name}
Difficulty Level: {level} ({level_desc})
Question Type: mcq_single (Multiple Choice - Single Correct Answer){class_instruction}
{MOBILE_FORMAT_RULES}
CONTENT REQUIREMENTS:
1. Questions MUST be based on REAL, FACTUAL content from this topic - use actual facts, formulas, definitions, and concepts
2. All answers must be SCIENTIFICALLY/ACADEMICALLY ACCURATE and verifiable
3. Each question should have exactly 4 options (A, B, C, D)
4. Only ONE option should be correct
5. Include a detailed explanation for why the correct answer is correct
6. Include a helpful hint that guides the student toward the answer without giving it away directly
7. Incorrect options (distractors) should be plausible but clearly incorrect based on the subject matter
8. Questions should test real understanding, not just memorization
9. Use proper terminology and vocabulary appropriate for the subject
10. DO NOT use placeholder text like "Option A" or "Question about X" - write REAL questions with REAL content{image_instruction}

Return the response in the following JSON format:
{{
    "questions": [
        {{
            "question_text": "The actual question text with real content",
            "question_type": "mcq_single",
            "options": [
                {{"option_text": "Actual option with real content", "is_correct": false}},
                {{"option_text": "Actual correct answer", "is_correct": true}},
                {{"option_text": "Actual plausible distractor", "is_correct": false}},
                {{"option_text": "Actual plausible distractor", "is_correct": false}}
            ],
            "hint": "A helpful clue that guides the student toward the correct answer without revealing it directly",
            "explanation": "Detailed explanation of why the correct answer is correct and why others are wrong",
            "exp_points": 10{image_json_field}
        }}
    ]
}}

Generate exactly {number_of_questions} questions. Respond only with valid JSON, no markdown formatting."""

        elif question_type == 'mcq_multiple':
            image_instruction = ""
            image_json_field = ""
            if add_image and use_matplot:
                image_instruction = (
                    "\nIMAGE: For each question provide matplotlib_code — Python using only plt/np."
                    " The figure must be a BLANK SETUP DIAGRAM ONLY — it shows the visual problem"
                    " setup so the student can answer by looking at it."
                    " FORBIDDEN (never include these in the code):"
                    " (a) plt.title() or ax.set_title() — no question text as a title;"
                    " (b) plt.text() or ax.annotate() showing any answer value, measurement, or result"
                    " (e.g. do NOT write angle values like '90°', lengths like '5 cm', or computed results);"
                    " (c) option labels (A/B/C/D) as text annotations;"
                    " (d) any text that gives away the answer or hints at it."
                    " ALLOWED: vertex labels (single letters like A, B, C, D), axis labels, tick marks,"
                    " shape outlines, lines, curves, and given/known values that are part of the problem"
                    " statement (not the answer to be found)."
                    " The figure must look INCOMPLETE — the unknown value the question asks about must"
                    " NOT appear anywhere in the figure."
                    " No plt.savefig() or plt.show()."
                )
                image_json_field = (
                    ',\n            "matplotlib_code": "Python using plt/np ONLY.'
                    ' Draw ONLY the geometric setup: shapes with vertex labels (A,B,C,D),'
                    ' given side lengths or angles that are STATED in the question (not the answer).'
                    ' NEVER use plt.title(), NEVER annotate the answer value, NEVER show the correct'
                    ' measurement that the question asks the student to find.'
                    ' No plt.savefig() or plt.show()."'
                )
            elif add_image:
                image_instruction = "\nIMAGE (OPTIONAL): Only include image_description for questions that specifically require a visual diagram, chart, or figure. Most questions should NOT have an image. If included, provide a visual-only description (no text/labels). For image-based questions write question_text as 'Based on the figure above, ...'. Image must NOT imply the answer."
                image_json_field = ',\n            "image_description": "(OPTIONAL) Include ONLY if a visual diagram/chart/figure is essential to this specific question. Omit this field entirely for text-based questions."'

            prompt = f"""You are an expert educational content creator. Generate {number_of_questions} REAL, FACTUALLY ACCURATE multiple choice questions (with multiple correct answers) for students studying this topic.

Subject: {subject_name}
Module/Unit: {module_name}
Chapter/Topic: {chapter_name}
Difficulty Level: {level} ({level_desc})
Question Type: mcq_multiple (Multiple Choice - Multiple Correct Answers){class_instruction}
{MOBILE_FORMAT_RULES}
CONTENT REQUIREMENTS:
1. Questions MUST be based on REAL, FACTUAL content from this topic - use actual facts, formulas, definitions, and concepts
2. All answers must be SCIENTIFICALLY/ACADEMICALLY ACCURATE and verifiable
3. Each question should have exactly 4 options (A, B, C, D)
4. TWO or MORE options should be correct
5. Include a detailed explanation for why the correct answers are correct
6. Include a helpful hint that guides the student toward the answers without giving them away directly
7. Incorrect options (distractors) should be plausible but clearly incorrect based on the subject matter
8. The question text should indicate that multiple answers may be correct (e.g., "Select all that apply")
9. Use proper terminology and vocabulary appropriate for the subject
10. DO NOT use placeholder text - write REAL questions with REAL content{image_instruction}

Return the response in the following JSON format:
{{
    "questions": [
        {{
            "question_text": "Select all that apply: The actual question with real content",
            "question_type": "mcq_multiple",
            "options": [
                {{"option_text": "Actual correct answer 1", "is_correct": true}},
                {{"option_text": "Actual correct answer 2", "is_correct": true}},
                {{"option_text": "Actual plausible distractor", "is_correct": false}},
                {{"option_text": "Actual plausible distractor", "is_correct": false}}
            ],
            "hint": "A helpful clue that guides the student toward the correct answers without revealing them directly",
            "explanation": "Detailed explanation of why the correct answers are correct",
            "exp_points": 15{image_json_field}
        }}
    ]
}}

Generate exactly {number_of_questions} questions. Respond only with valid JSON, no markdown formatting."""

        elif question_type == 'rearrange':
            image_instruction = ""
            image_json_field = ""
            if add_image and use_matplot:
                image_instruction = (
                    "\nIMAGE: For each question provide matplotlib_code — Python using only plt/np."
                    " The figure must be a BLANK SETUP DIAGRAM ONLY — it shows the visual problem"
                    " setup so the student can answer by looking at it."
                    " FORBIDDEN (never include these in the code):"
                    " (a) plt.title() or ax.set_title() — no question text as a title;"
                    " (b) plt.text() or ax.annotate() showing any answer value, measurement, or result"
                    " (e.g. do NOT write angle values like '90°', lengths like '5 cm', or computed results);"
                    " (c) option labels (A/B/C/D) as text annotations;"
                    " (d) any text that gives away the answer or hints at it."
                    " ALLOWED: vertex labels (single letters like A, B, C, D), axis labels, tick marks,"
                    " shape outlines, lines, curves, and given/known values that are part of the problem"
                    " statement (not the answer to be found)."
                    " The figure must look INCOMPLETE — the unknown value the question asks about must"
                    " NOT appear anywhere in the figure."
                    " No plt.savefig() or plt.show()."
                )
                image_json_field = (
                    ',\n            "matplotlib_code": "Python using plt/np ONLY.'
                    ' Draw ONLY the geometric setup: shapes with vertex labels (A,B,C,D),'
                    ' given side lengths or angles that are STATED in the question (not the answer).'
                    ' NEVER use plt.title(), NEVER annotate the answer value, NEVER show the correct'
                    ' measurement that the question asks the student to find.'
                    ' No plt.savefig() or plt.show()."'
                )
            elif add_image:
                image_instruction = "\nIMAGE (OPTIONAL): Only include image_description for questions that specifically require a visual diagram, chart, or figure. Most questions should NOT have an image. If included, provide a visual-only description (no text/labels). For image-based questions write question_text as 'Based on the figure above, ...'. Image must NOT imply the answer."
                image_json_field = ',\n            "image_description": "(OPTIONAL) Include ONLY if a visual diagram/chart/figure is essential to this specific question. Omit this field entirely for text-based questions."'

            prompt = f"""You are an expert educational content creator. Generate {number_of_questions} REAL, FACTUALLY ACCURATE re-arrange/ordering questions for students studying this topic.

Subject: {subject_name}
Module/Unit: {module_name}
Chapter/Topic: {chapter_name}
Difficulty Level: {level} ({level_desc})
Question Type: rearrange (Re-arrange/Ordering Question){class_instruction}
{MOBILE_FORMAT_RULES}
CONTENT REQUIREMENTS:
1. Questions MUST be based on REAL, FACTUAL sequences from this topic - use actual processes, historical events, scientific procedures, or logical sequences
2. All sequences must be ACADEMICALLY ACCURATE and verifiable
3. Each question should ask the user to arrange items in the correct order
4. Provide 4-6 items that need to be arranged in sequence
5. The options array must be in SCRAMBLED (random) order. Each option MUST have "correct_order": N where N is 1-based position in the CORRECT sequence. The system uses correct_order to grade the answer.
6. Include a detailed explanation for why this is the correct sequence
7. Include a helpful hint that guides the student toward the correct order without giving it away directly
8. Use proper terminology and vocabulary appropriate for the subject
9. DO NOT use placeholder text like "Step 1", "Step 2" - write REAL content with REAL steps/items from the topic{image_instruction}

Return the response in the following JSON format (correct_order is mandatory per option, 1-based):
{{
    "questions": [
        {{
            "question_text": "Arrange the following [actual items] in the correct order:",
            "question_type": "rearrange",
            "options": [
                {{"option_text": "Actual third item/step from the topic", "is_correct": true, "correct_order": 3}},
                {{"option_text": "Actual first item/step from the topic", "is_correct": true, "correct_order": 1}},
                {{"option_text": "Actual fourth item/step from the topic", "is_correct": true, "correct_order": 4}},
                {{"option_text": "Actual second item/step from the topic", "is_correct": true, "correct_order": 2}}
            ],
            "hint": "A helpful clue that guides the student toward the correct sequence without revealing it directly",
            "explanation": "Detailed explanation of why this is the correct order based on the subject matter",
            "exp_points": 20{image_json_field}
        }}
    ]
}}

Generate exactly {number_of_questions} questions. Respond only with valid JSON, no markdown formatting."""

        else:
            image_instruction = ""
            image_json_field = ""
            if add_image and use_matplot:
                image_instruction = (
                    "\nIMAGE: For each question provide matplotlib_code — Python using only plt/np."
                    " The figure must be a BLANK SETUP DIAGRAM ONLY — it shows the visual problem"
                    " setup so the student can answer by looking at it."
                    " FORBIDDEN (never include these in the code):"
                    " (a) plt.title() or ax.set_title() — no question text as a title;"
                    " (b) plt.text() or ax.annotate() showing any answer value, measurement, or result"
                    " (e.g. do NOT write angle values like '90°', lengths like '5 cm', or computed results);"
                    " (c) option labels (A/B/C/D) as text annotations;"
                    " (d) any text that gives away the answer or hints at it."
                    " ALLOWED: vertex labels (single letters like A, B, C, D), axis labels, tick marks,"
                    " shape outlines, lines, curves, and given/known values that are part of the problem"
                    " statement (not the answer to be found)."
                    " The figure must look INCOMPLETE — the unknown value the question asks about must"
                    " NOT appear anywhere in the figure."
                    " No plt.savefig() or plt.show()."
                )
                image_json_field = (
                    ',\n            "matplotlib_code": "Python using plt/np ONLY.'
                    ' Draw ONLY the geometric setup: shapes with vertex labels (A,B,C,D),'
                    ' given side lengths or angles that are STATED in the question (not the answer).'
                    ' NEVER use plt.title(), NEVER annotate the answer value, NEVER show the correct'
                    ' measurement that the question asks the student to find.'
                    ' No plt.savefig() or plt.show()."'
                )
            elif add_image:
                image_instruction = "\nIMAGE (OPTIONAL): Only include image_description for questions that specifically require a visual diagram, chart, or figure. Most questions should NOT have an image. If included, provide a visual-only description (no text/labels). For image-based questions write question_text as 'Based on the figure above, ...'. Image must NOT imply the answer."
                image_json_field = ',\n            "image_description": "(OPTIONAL) Include ONLY if a visual diagram/chart/figure is essential to this specific question. Omit this field entirely for text-based questions."'

            prompt = f"""You are an expert educational content creator. Generate {number_of_questions} REAL, FACTUALLY ACCURATE multiple choice questions for students studying this topic.

Subject: {subject_name}
Module/Unit: {module_name}
Chapter/Topic: {chapter_name}
Difficulty Level: {level} ({level_desc})
Question Type: mcq_single (Multiple Choice - Single Correct Answer){class_instruction}
{MOBILE_FORMAT_RULES}
CONTENT REQUIREMENTS:
1. Questions MUST be based on REAL, FACTUAL content from this topic - use actual facts, formulas, definitions, and concepts
2. All answers must be SCIENTIFICALLY/ACADEMICALLY ACCURATE and verifiable
3. Each question should have exactly 4 options (A, B, C, D)
4. Only ONE option should be correct
5. Include a detailed explanation for why the correct answer is correct
6. Include a helpful hint that guides the student toward the answer without giving it away directly
7. Incorrect options (distractors) should be plausible but clearly incorrect
8. Use proper terminology and vocabulary appropriate for the subject
9. DO NOT use placeholder text - write REAL questions with REAL content{image_instruction}

Return the response in the following JSON format:
{{
    "questions": [
        {{
            "question_text": "The actual question text with real content",
            "question_type": "mcq_single",
            "options": [
                {{"option_text": "Actual option with real content", "is_correct": false}},
                {{"option_text": "Actual correct answer", "is_correct": true}},
                {{"option_text": "Actual plausible distractor", "is_correct": false}},
                {{"option_text": "Actual plausible distractor", "is_correct": false}}
            ],
            "hint": "A helpful clue that guides the student toward the correct answer without revealing it directly",
            "explanation": "Detailed explanation of why the correct answer is correct",
            "exp_points": 10{image_json_field}
        }}
    ]
}}

Generate exactly {number_of_questions} questions. Respond only with valid JSON, no markdown formatting."""

        result = gemini_generate_json(prompt)
        questions = result.get('questions', [])
        
        for q in questions:
            q['question_type'] = question_type
        
        if add_image:
            from gyaan_buddy.utils.gemini_client import generate_image
            from gyaan_buddy.utils.matplotlib_executor import execute_matplotlib_code
            import tempfile
            import uuid
            import base64 as b64
            
            from gyaan_buddy.utils.matplotlib_executor import sanitize_matplotlib_code
            for q in questions:
                matplotlib_code_raw = q.get('matplotlib_code') or q.get('matplotlibCode')
                if use_matplot and matplotlib_code_raw:
                    matplotlib_code = sanitize_matplotlib_code((matplotlib_code_raw or '').strip())
                    if matplotlib_code:
                        try:
                            result = execute_matplotlib_code(matplotlib_code)
                            if 'image_base64' in result:
                                image_data = b64.b64decode(result['image_base64'])
                                temp_file = tempfile.NamedTemporaryFile(
                                    delete=False,
                                    suffix='.png',
                                    prefix=f'question_image_{uuid.uuid4().hex[:8]}_'
                                )
                                temp_file.write(image_data)
                                temp_file.close()
                                q['image_path'] = temp_file.name
                                api_logger.info(f"Generated matplotlib image for question: {temp_file.name}")
                            else:
                                api_logger.warning(f"Matplotlib execution failed for question: {result.get('error', 'Unknown')}")
                                q['image_path'] = None
                        except Exception as e:
                            api_logger.error(f"Error generating matplotlib image for question: {str(e)}")
                            q['image_path'] = None
                    else:
                        q['image_path'] = None
                    continue
                
                image_description = (q.get('image_description') or q.get('imageDescription') or '').strip()

                if not image_description:
                    api_logger.info("Question has no image_description — skipping image generation for this question")
                    q['image_path'] = None
                    continue

                try:
                    temp_file = tempfile.NamedTemporaryFile(
                        delete=False,
                        suffix='.png',
                        prefix=f'question_image_{uuid.uuid4().hex[:8]}_'
                    )
                    temp_file.close()
                    image_prompt = f"""Create a PURE VISUAL educational illustration. The image will be used as the basis for a question - students will answer based on what they see. So the image must NOT contain or imply the answer, hint, or any question text.

Visual description to draw: {image_description}

CRITICAL REQUIREMENTS:
- Draw ONLY the visual elements described above - NO TEXT, NO words, NO question, NO options, NO hint, NO explanation in the image
- DO NOT include labels, captions, or written content (except minimal axis labels if essential for a graph)
- The image must be question-specific: it shows only the concept/diagram/figure; the question will be asked separately about this image
- Clean, professional, educational style suitable for students
- Colorful but not distracting; modern flat design
- Do not illustrate the answer or give away the solution in the image"""

                    image_path = generate_image(
                        prompt=image_prompt,
                        output_path=temp_file.name,
                        aspect_ratio="1:1"
                    )

                    if image_path and os.path.exists(image_path):
                        q['image_path'] = image_path
                        api_logger.info(f"Generated image for question: {image_path}")
                    else:
                        api_logger.warning(f"Failed to generate image for question. Description: {image_description[:50]}")
                        q['image_path'] = None
                except Exception as e:
                    api_logger.error(f"Error generating image for question: {str(e)}")
                    q['image_path'] = None
        
        api_logger.info(f"Successfully generated {len(questions)} {question_type} questions from Vertex AI")
        return questions
            
    except ImportError as e:
        api_logger.error(f"Vertex AI import error: {e}")
        raise Exception(f"Vertex AI package not installed: {e}")
    except json.JSONDecodeError as e:
        api_logger.error(f"Failed to parse Vertex AI response as JSON: {e}")
        raise Exception(f"Failed to parse Vertex AI response as JSON: {e}")
    except Exception as e:
        api_logger.error(f"Error calling Vertex AI API: {str(e)}")
        raise


def generate_questions_with_vertex_ai_mix(subject_name, module_name, chapter_name, number_of_questions, add_image=False, use_matplot=False, class_context=None, allowed_question_types=None):
    """
    Generate a MIX of question types and difficulty levels using Vertex AI.
    Returns list of question dicts; each has question_type and level (1-5).
    class_context: optional string e.g. "Class 10" to tailor difficulty and vocabulary.
    allowed_question_types: optional list e.g. ['mcq_single', 'mcq_multiple'] to restrict types; if None, use all (mcq_single, mcq_multiple, rearrange).
    """
    try:
        from gyaan_buddy.utils.gemini_client import gemini_generate_json

        class_instruction = ""
        if class_context and isinstance(class_context, str) and class_context.strip():
            class_instruction = f"\nTARGET AUDIENCE: These questions are for students of {class_context.strip()}. Use vocabulary and complexity appropriate for this class/grade level."

        image_instruction = ""
        image_json_field = ""
        if add_image and use_matplot:
            image_instruction = "\nIMAGE: For EACH question provide matplotlib_code — Python using only plt/np to create one educational figure. No plt.savefig() or plt.show()."
            image_json_field = ',\n            "matplotlib_code": "Python code using only plt (matplotlib.pyplot) and np (numpy) to create one figure. No plt.savefig() or plt.show()."'
        elif add_image:
            image_instruction = "\nIMAGE (OPTIONAL): Only include image_description for questions that specifically require a visual diagram, chart, or figure. Most questions should NOT have an image. If included, provide a visual-only description (no text/labels). For image-based questions write question_text as 'Based on the figure above, ...'. Image must NOT imply the answer."
            image_json_field = ',\n            "image_description": "(OPTIONAL) Include ONLY if a visual diagram/chart/figure is essential to this specific question. Omit this field entirely for text-based questions."'

        valid_allowed = [t for t in (allowed_question_types or []) if t in ('mcq_single', 'mcq_multiple', 'rearrange')]
        if valid_allowed:
            types_str = ', '.join(valid_allowed)
            question_types_instruction = f"Use ONLY these question types and vary across them: {types_str}. Each question_type must be one of: {types_str}."
        else:
            question_types_instruction = "Include mcq_single (single correct), mcq_multiple (multiple correct), and rearrange (ordering). Vary the types across the questions."

        valid_allowed = [t for t in (allowed_question_types or []) if t in ('mcq_single', 'mcq_multiple', 'rearrange')]
        if valid_allowed:
            types_str = ', '.join(valid_allowed)
            question_types_instruction = f"Use ONLY these question types and vary across them: {types_str}. Each question_type must be one of: {types_str}."
        else:
            question_types_instruction = "Include mcq_single (single correct), mcq_multiple (multiple correct), and rearrange (ordering). Vary the types across the questions."

        prompt = f"""You are an expert educational content creator. Generate exactly {number_of_questions} REAL, FACTUALLY ACCURATE questions for students studying this topic.

Subject: {subject_name}
Module/Unit: {module_name}
Chapter/Topic: {chapter_name}{class_instruction}
{MOBILE_FORMAT_RULES}
BALANCED MIX (MUST follow):
- Question types: {question_types_instruction}
- Difficulty levels: distribute levels 1–5 EVENLY. Level 1=very easy, 2=easy, 3=medium, 4=hard, 5=HOTS.
  For {number_of_questions} questions spread levels evenly; at least one level 1 and one level 2.

CONTENT REQUIREMENTS:
1. Each question MUST have "question_type" and "level" (integer 1-5).
2. Questions MUST be based on REAL, FACTUAL content from this topic.
3. For mcq_single: 4 options, one correct. For mcq_multiple: 4 options, two or more correct. For rearrange: options with "correct_order" (1-based).
4. Include hint and explanation for each question.
5. Use proper terminology. DO NOT use placeholder text.{image_instruction}

Return ONLY valid JSON in this format (no markdown):
{{
    "questions": [
        {{
            "question_text": "The actual question text",
            "question_type": "mcq_single",
            "level": 3,
            "options": [
                {{"option_text": "Option text", "is_correct": false}},
                {{"option_text": "Correct answer", "is_correct": true}},
                {{"option_text": "Distractor", "is_correct": false}},
                {{"option_text": "Distractor", "is_correct": false}}
            ],
            "hint": "A helpful clue",
            "explanation": "Detailed explanation",
            "exp_points": 10{image_json_field}
        }}
    ]
}}

Generate exactly {number_of_questions} questions with a BALANCED mix of types and levels (levels 1–5 spread evenly). Respond only with valid JSON, no markdown."""

        result = gemini_generate_json(prompt)
        questions = result.get('questions', [])

        for q in questions:
            q['question_type'] = normalize_question_type(q.get('question_type', 'mcq_single'))
            q.setdefault('level', 3)
            try:
                q['level'] = int(q['level'])
            except (TypeError, ValueError):
                q['level'] = 3
            q['level'] = max(1, min(5, q['level']))

        if add_image:
            from gyaan_buddy.utils.gemini_client import generate_image
            from gyaan_buddy.utils.matplotlib_executor import execute_matplotlib_code
            import tempfile
            import uuid as uuid_mod
            import base64 as b64

            from gyaan_buddy.utils.matplotlib_executor import sanitize_matplotlib_code
            for q in questions:
                q['image_path'] = None
                matplotlib_code_raw = q.get('matplotlib_code') or q.get('matplotlibCode')
                if use_matplot and matplotlib_code_raw:
                    matplotlib_code = sanitize_matplotlib_code((matplotlib_code_raw or '').strip())
                    if matplotlib_code:
                        try:
                            exec_result = execute_matplotlib_code(matplotlib_code)
                            if 'image_base64' in exec_result:
                                image_data = b64.b64decode(exec_result['image_base64'])
                                temp_file = tempfile.NamedTemporaryFile(
                                    delete=False, suffix='.png',
                                    prefix=f'question_image_{uuid_mod.uuid4().hex[:8]}_'
                                )
                                temp_file.write(image_data)
                                temp_file.close()
                                q['image_path'] = temp_file.name
                            else:
                                image_description = (q.get('image_description') or q.get('imageDescription') or '').strip()
                                if image_description:
                                    try:
                                        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png', prefix=f'question_image_{uuid_mod.uuid4().hex[:8]}_')
                                        temp_file.close()
                                        image_prompt = f"Create a PURE VISUAL educational illustration. Image will be used for a question - do NOT include answer, hint, or question text. Visual description: {image_description}. NO TEXT in image, NO labels except minimal axis labels if needed."
                                        image_path = generate_image(prompt=image_prompt, output_path=temp_file.name, aspect_ratio="1:1")
                                        q['image_path'] = image_path if image_path and os.path.exists(image_path) else None
                                    except Exception:
                                        pass
                        except Exception as e:
                            api_logger.warning(f"Matplotlib image error: {e}")
                            image_description = (q.get('image_description') or q.get('imageDescription') or '').strip()
                            if image_description:
                                try:
                                    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png', prefix=f'question_image_{uuid_mod.uuid4().hex[:8]}_')
                                    temp_file.close()
                                    image_prompt = f"Create a PURE VISUAL educational illustration. Do NOT include answer, hint, or question text. Visual only: {image_description}. NO TEXT in image."
                                    image_path = generate_image(prompt=image_prompt, output_path=temp_file.name, aspect_ratio="1:1")
                                    q['image_path'] = image_path if image_path and os.path.exists(image_path) else None
                                except Exception:
                                    pass
                    continue

                image_description = (q.get('image_description') or q.get('imageDescription') or '').strip()
                if not image_description:
                    api_logger.info("Question has no image_description — skipping image generation for this question")
                    continue
                try:
                    temp_file = tempfile.NamedTemporaryFile(
                        delete=False, suffix='.png',
                        prefix=f'question_image_{uuid_mod.uuid4().hex[:8]}_'
                    )
                    temp_file.close()
                    image_prompt = f"""Create a PURE VISUAL educational illustration. The image will be used as the basis for a question - students answer based on what they see. The image must NOT contain or imply the answer, hint, or any question text.

Visual description to draw: {image_description}

CRITICAL: Draw ONLY the visual elements - NO TEXT, NO words, NO question, NO options, NO hint in the image. The image is question-specific: it shows only the concept/diagram; the question will be asked separately. Clean, professional, educational style. No labels except minimal axis labels if essential for a graph."""
                    image_path = generate_image(
                        prompt=image_prompt,
                        output_path=temp_file.name,
                        aspect_ratio="1:1"
                    )
                    q['image_path'] = image_path if image_path and os.path.exists(image_path) else None
                except Exception as e:
                    api_logger.warning(f"Image generation error: {e}")

        api_logger.info(f"Generated {len(questions)} mixed questions from Vertex AI")
        return questions

    except Exception as e:
        api_logger.error(f"Error in generate_questions_with_vertex_ai_mix: {str(e)}")
        raise


@api_view(['POST'])
@drf_permission_classes([permissions.IsAuthenticated])
def generate_ai_questions_vertex(request):
    """
    Generate AI questions using Google Vertex AI and create entries in the database.
    
    Expected Input (single chapter):
        subject_id, module_id, chapter_id, subject_name, module_name, chapter_name,
        number_of_questions, [level], [question_type], add_image, use_matplot, for_test, test_id
    Or (multi module/chapter):
        subject_id, subject_name, number_of_questions, module_chapters: [{ module_id, chapter_id, module_name, chapter_name }, ...],
        [level], [question_type], add_image, use_matplot, for_test, test_id
    When level and question_type are omitted: generates a MIX of question types and difficulty levels.
    """
    api_logger.info(f"Vertex AI question generation requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")
    
    try:
        subject_id = request.data.get('subject_id')
        module_id = request.data.get('module_id')
        chapter_id = request.data.get('chapter_id')
        subject_name = request.data.get('subject_name', '')
        module_name = request.data.get('module_name', '')
        chapter_name = request.data.get('chapter_name', '')
        module_chapters = request.data.get('module_chapters')
        number_of_questions = request.data.get('number_of_questions', 5)
        level = request.data.get('level')
        question_type = request.data.get('question_type')
        question_types = request.data.get('question_types')
        if isinstance(question_types, list):
            question_types = [qt for qt in question_types if qt in ['mcq_single', 'mcq_multiple', 'short_answer', 'rearrange']]
        else:
            question_types = None
        add_image = request.data.get('add_image', False)
        use_matplot = request.data.get('use_matplot', False)
        for_test = request.data.get('for_test', False)
        test_id = request.data.get('test_id')
        test_ids = request.data.get('test_ids')
        if test_ids is not None and not isinstance(test_ids, list):
            test_ids = [test_ids] if test_ids else []
        test_ids = [str(tid).strip() for tid in test_ids if tid] if test_ids else []
        
        if isinstance(for_test, str):
            for_test = for_test.lower() in ('true', '1', 'yes', 'on')
        
        test = None
        tests_to_add = []  # when for_test and multiple classes: list of Test instances to add questions to
        class_context = request.data.get('class_context')
        if for_test:
            if test_ids:
                try:
                    tests_to_add = list(
                        Test.objects.select_related('class_group').prefetch_related('class_groups').filter(
                            id__in=test_ids, is_deleted=False
                        )
                    )
                    if len(tests_to_add) != len(test_ids):
                        return validation_error({"error": "One or more test IDs in test_ids not found or inactive"})
                    test = tests_to_add[0]
                except Exception as e:
                    api_logger.warning(f"Failed to resolve test_ids: {e}")
                    return validation_error({"error": "Invalid test_ids"})
            elif test_id:
                try:
                    test = Test.objects.select_related('class_group').prefetch_related('class_groups').get(id=test_id, is_deleted=False)
                    tests_to_add = [test]
                except Test.DoesNotExist:
                    return validation_error({"error": f"Test with ID {test_id} not found"})
            else:
                return validation_error({"error": "test_id or test_ids is required when for_test is True"})
            if not class_context or not str(class_context).strip():
                assigned = getattr(test, 'get_assigned_classes', lambda: [])() or []
                if not assigned and getattr(test, 'class_group', None):
                    assigned = [test.class_group]
                if assigned:
                    c = assigned[0]
                    name = getattr(c, 'name', None) or ''
                    grade = getattr(c, 'grade', None)
                    grade_name = getattr(grade, 'name', None) if grade else None
                    if grade_name and name:
                        class_context = f"{name} (Grade {grade_name})"
                    elif name:
                        class_context = name
                    else:
                        class_context = None
        
        if isinstance(add_image, str):
            add_image = add_image.lower() in ('true', '1', 'yes', 'on')
        if isinstance(use_matplot, str):
            use_matplot = use_matplot.lower() in ('true', '1', 'yes', 'on')
        
        try:
            number_of_questions = int(number_of_questions)
        except (ValueError, TypeError):
            return validation_error({"error": "number_of_questions must be an integer"})
        
        # When level is not provided, use mix generator to get a spread of levels 1-5
        use_mix = (
            (level is None)
            or request.data.get('use_mix', False)
            or bool(for_test and (test_id or test_ids))
        )
        if not use_mix:
            level = int(level) if level is not None else 3
            if level < 1 or level > 5:
                level = 3
            if question_types and len(question_types) > 0:
                question_type = None
            else:
                question_type = question_type or 'mcq_single'
                if question_type not in ['mcq_single', 'mcq_multiple', 'rearrange']:
                    question_type = 'mcq_single'
        
        level_to_difficulty = {1: 'easy', 2: 'easy', 3: 'medium', 4: 'hard', 5: 'hard'}
        
        if module_chapters and isinstance(module_chapters, list) and len(module_chapters) > 0:
            chapters_to_process = []
            for item in module_chapters:
                mid = item.get('module_id') or item.get('module')
                cid = item.get('chapter_id') or item.get('chapter')
                mname = item.get('module_name', '')
                cname = item.get('chapter_name', '')
                if not mid or not cid:
                    continue
                try:
                    ch = ModuleChapter.objects.get(id=cid)
                    if str(ch.module_id) != str(mid):
                        continue
                    chapters_to_process.append({
                        'module_id': mid,
                        'chapter_id': cid,
                        'module_name': mname or ch.module.name,
                        'chapter_name': cname or ch.title,
                        'chapter': ch
                    })
                except ModuleChapter.DoesNotExist:
                    continue
            if not chapters_to_process:
                return validation_error({"error": "module_chapters: no valid module/chapter pairs found"})
            if not subject_id or not subject_name:
                return validation_error({"error": "subject_id and subject_name are required when using module_chapters"})
            per_chapter = number_of_questions // len(chapters_to_process)
            remainder = number_of_questions - (per_chapter * len(chapters_to_process))
        else:
            if not all([subject_id, module_id, chapter_id]):
                return validation_error({"error": "subject_id, module_id, and chapter_id are required (or use module_chapters)"})
            try:
                chapter = ModuleChapter.objects.get(id=chapter_id)
            except ModuleChapter.DoesNotExist:
                return validation_error({"error": f"Chapter with ID {chapter_id} not found"})
            chapters_to_process = [{
                'module_id': module_id,
                'chapter_id': chapter_id,
                'module_name': module_name or chapter.module.name,
                'chapter_name': chapter_name or chapter.title,
                'chapter': chapter
            }]
            per_chapter = number_of_questions
            remainder = 0
        
        questions_created = 0
        module_content_created = 0
        hots_created = 0
        test_questions_created = 0
        created_questions_list = []
        test_question_order_per_test = {}
        if for_test and tests_to_add:
            for t in tests_to_add:
                max_order = TestQuestion.objects.filter(test=t).aggregate(max_order=Max('order'))['max_order'] or 0
                test_question_order_per_test[t.id] = max_order + 1
        
        for idx, ch_info in enumerate(chapters_to_process):
            num_for_chapter = per_chapter + (1 if idx < remainder else 0)
            if num_for_chapter < 1:
                continue
            chapter = ch_info['chapter']
            mn = ch_info['module_name']
            cn = ch_info['chapter_name']
            
            if use_mix:
                generated_questions = generate_questions_with_vertex_ai_mix(
                    subject_name=subject_name,
                    module_name=mn,
                    chapter_name=cn,
                    number_of_questions=num_for_chapter,
                    add_image=add_image,
                    use_matplot=use_matplot,
                    class_context=class_context,
                    allowed_question_types=question_types if question_types else None
                )
            elif question_types and len(question_types) > 0:
                valid_types = [t for t in question_types if t in ['mcq_single', 'mcq_multiple', 'rearrange']]
                if not valid_types:
                    valid_types = ['mcq_single']
                per_type = num_for_chapter // len(valid_types)
                remainder = num_for_chapter % len(valid_types)
                generated_questions = []
                for i, qt in enumerate(valid_types):
                    n = per_type + (1 if i < remainder else 0)
                    if n < 1:
                        continue
                    batch = generate_questions_with_vertex_ai(
                        subject_name=subject_name,
                        module_name=mn,
                        chapter_name=cn,
                        level=3,
                        number_of_questions=n,
                        question_type=qt,
                        add_image=add_image,
                        use_matplot=use_matplot,
                        class_context=class_context
                    )
                    if batch:
                        generated_questions.extend(batch)
            else:
                generated_questions = generate_questions_with_vertex_ai(
                    subject_name=subject_name,
                    module_name=mn,
                    chapter_name=cn,
                    level=level,
                    number_of_questions=num_for_chapter,
                    question_type=question_type,
                    add_image=add_image,
                    use_matplot=use_matplot,
                    class_context=class_context
                )
            
            if not generated_questions:
                continue
            
            use_multi_types = question_types and len(question_types) > 0
            for q_data in generated_questions:
                try:
                    q_level = q_data.get('level', 3) if use_mix else level
                    q_type = q_data.get('question_type', 'mcq_single') if (use_mix or use_multi_types) else question_type
                    q_type = normalize_question_type(q_type)
                    try:
                        q_level = int(q_level)
                    except (TypeError, ValueError):
                        q_level = 3
                    q_level = max(1, min(5, q_level))
                    is_hots = q_level == 5
                    difficulty = level_to_difficulty.get(q_level, 'medium')
                    
                    image_path = q_data.get('image_path')
                    question_kwargs = {
                        'question_text': q_data.get('question_text', ''),
                        'question_type': q_type,
                        'difficulty_level': difficulty,
                        'explanation': q_data.get('explanation', ''),
                        'hint': q_data.get('hint', ''),
                        'exp_points': q_data.get('exp_points', 10),
                        'is_active': not for_test,  # inactive until user confirms via activate-questions
                        'is_hots': is_hots,
                        'ai_generated': True,
                        'level': q_level,
                        'created_by': request.user
                    }
                    
                    if image_path and os.path.exists(image_path):
                        try:
                            with open(image_path, 'rb') as f:
                                image_content = f.read()
                            from io import BytesIO
                            image_file = File(BytesIO(image_content), name=os.path.basename(image_path))
                            question_kwargs['image'] = image_file
                            question = Question.objects.create(**question_kwargs)
                            try:
                                os.unlink(image_path)
                            except Exception as e:
                                api_logger.warning(f"Failed to delete temporary image file {image_path}: {str(e)}")
                        except Exception as e:
                            api_logger.warning(f"Failed to process image file {image_path}: {str(e)}")
                            question = Question.objects.create(**question_kwargs)
                            try:
                                os.unlink(image_path)
                            except Exception:
                                pass
                    else:
                        question = Question.objects.create(**question_kwargs)
                    
                    questions_created += 1
                    
                    options_data = q_data.get('options', [])
                    created_options = []
                    for opt_idx, opt_data in enumerate(options_data):
                        if q_type == 'rearrange':
                            order = opt_data.get('correct_order')
                            if order is None or not isinstance(order, (int, float)):
                                order = opt_idx + 1
                            order = max(1, int(order))
                        else:
                            order = opt_idx + 1
                        option = Option.objects.create(
                            question=question,
                            option_text=opt_data.get('option_text', f'Option {opt_idx + 1}'),
                            is_correct=opt_data.get('is_correct', False),
                            order=order
                        )
                        created_options.append({
                            'id': str(option.id),
                            'option_text': option.option_text,
                            'is_correct': option.is_correct,
                            'order': option.order
                        })
                    
                    question_data = {
                        'id': str(question.id),
                        'question_text': question.question_text,
                        'question_type': normalize_question_type(question.question_type),
                        'difficulty_level': question.difficulty_level,
                        'level': question.level,
                        'hint': question.hint,
                        'explanation': question.explanation,
                        'exp_points': question.exp_points,
                        'is_hots': question.is_hots,
                        'ai_generated': question.ai_generated,
                        'options': created_options
                    }
                    
                    if question.image:
                        question_data['image'] = request.build_absolute_uri(question.image.url)
                    
                    created_questions_list.append(question_data)
                    
                    if is_hots:
                        max_order = ChapterHOTS.objects.filter(chapter=chapter).aggregate(
                            max_order=Max('order')
                        )['max_order'] or 0
                        ChapterHOTS.objects.create(
                            chapter=chapter,
                            question=question,
                            order=max_order + 1,
                            created_by=request.user
                        )
                        hots_created += 1
                        chapter.has_hots = True
                        chapter.save(update_fields=['has_hots'])
                    else:
                        max_order = ModuleContent.objects.filter(chapter=chapter).aggregate(
                            max_order=Max('order')
                        )['max_order'] or 0
                        ModuleContent.objects.create(
                            chapter=chapter,
                            content_type='question',
                            question=question,
                            order=max_order + 1,
                            created_by=request.user
                        )
                        module_content_created += 1
                    
                    if for_test and tests_to_add:
                        for t in tests_to_add:
                            try:
                                order = test_question_order_per_test.get(t.id, 1)
                                TestQuestion.objects.create(
                                    test=t,
                                    question=question,
                                    order=order
                                )
                                test_question_order_per_test[t.id] = order + 1
                                test_questions_created += 1
                                api_logger.info(f"Added question {question.id} to test {t.id} (order: {order})")
                            except Exception as e:
                                api_logger.warning(f"Failed to add question {question.id} to test {t.id}: {str(e)}")
                    
                    api_logger.info(f"Created question {question.id} for chapter {chapter.id} using Vertex AI")
                
                except Exception as e:
                    api_logger.error(f"Error creating question: {str(e)}")
                    continue
        
        api_logger.info(f"Vertex AI generation complete: {questions_created} questions, {module_content_created} module content, {hots_created} HOTS entries, {test_questions_created} test questions")
        
        response_data = {
            'questions_created': questions_created,
            'module_content_created': module_content_created,
            'hots_created': hots_created,
            'questions': created_questions_list,
        }
        if not use_mix:
            response_data['level'] = level
            response_data['is_hots'] = (level == 5)
        if chapters_to_process:
            response_data['chapter_ids'] = [str(c['chapter_id']) for c in chapters_to_process]
            if len(chapters_to_process) == 1:
                response_data['chapter_id'] = str(chapters_to_process[0]['chapter_id'])
        
        if for_test and tests_to_add:
            response_data['test_questions_created'] = test_questions_created
            response_data['test_id'] = str(tests_to_add[0].id)
            if len(tests_to_add) > 1:
                response_data['test_ids'] = [str(t.id) for t in tests_to_add]
        
        return success(
            data=response_data,
            message=f"Successfully generated {questions_created} questions using Vertex AI" + (f" and added {test_questions_created} to test(s)" if for_test and test_questions_created > 0 else "")
        )
        
    except Exception as e:
        api_logger.error(f"Error in Vertex AI question generation: {str(e)}")
        return validation_error({
            "error": f"Failed to generate questions: {str(e)}"
        })


@api_view(['POST'])
@drf_permission_classes([permissions.IsAuthenticated])
def generate_ai_questions_gemini(request):
    """
    Generate AI questions using Google Gemini and create entries in the database.
    
    Expected Input:
    {
        "class_id": optional (UUID) - Class ID for filtering,
        "subject_id": required (UUID) - Subject ID,
        "module_id": required (UUID) - Module ID,
        "chapter_id": required (UUID) - Chapter ID where questions will be created,
        "subject_name": required (str) - Name of the subject for prompt context,
        "module_name": required (str) - Name of the module for prompt context,
        "chapter_name": required (str) - Name of the chapter for prompt context,
        "number_of_questions": required (int) - Number of questions to generate (e.g., 3, 5, 10, 15, 20),
        "level": required (int) - Difficulty level 1-5 where:
            1 = Basic, 2 = Easy, 3 = Medium, 4 = Hard, 5 = HOTS (Advanced),
        "question_type": optional (str) - Type of questions to generate:
            "mcq_single" (default) - Multiple choice with single correct answer,
            "mcq_multiple" - Multiple choice with multiple correct answers,
            "rearrange" - Re-arrange/ordering questions
    }
    
    Expected Output:
    {
        "success": true,
        "data": {
            "questions_created": int - Number of questions created,
            "module_content_created": int - Number of module content entries (levels 1-4),
            "hots_created": int - Number of HOTS entries (level 5),
            "level": int - The difficulty level used,
            "is_hots": bool - Whether HOTS questions were created,
            "chapter_id": str - UUID of the chapter,
            "questions": [
                {
                    "id": str (UUID),
                    "question_text": str,
                    "question_type": str (mcq_single|mcq_multiple|rearrange),
                    "difficulty_level": str (easy|medium|hard),
                    "level": int (1-5),
                    "explanation": str,
                    "exp_points": int,
                    "is_hots": bool,
                    "ai_generated": true,
                    "options": [
                        {
                            "id": str (UUID),
                            "option_text": str,
                            "is_correct": bool,
                            "order": int
                        }
                    ]
                }
            ]
        },
        "message": str
    }
    
    Creates:
    - Question entries (with ai_generated=True)
    - Option entries for each question
    - ModuleContent entries (if level 1-4)
    - ChapterHOTS entries (if level 5)
    """
    api_logger.info(f"Gemini AI question generation requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")
    
    try:
        subject_id = request.data.get('subject_id')
        module_id = request.data.get('module_id')
        chapter_id = request.data.get('chapter_id')
        subject_name = request.data.get('subject_name', '')
        module_name = request.data.get('module_name', '')
        chapter_name = request.data.get('chapter_name', '')
        number_of_questions = request.data.get('number_of_questions', 5)
        level = request.data.get('level', 1)
        question_type = request.data.get('question_type', 'mcq_single')
        
        valid_question_types = ['mcq_single', 'mcq_multiple', 'rearrange']
        if question_type not in valid_question_types:
            question_type = 'mcq_single'
        
        if not all([subject_id, module_id, chapter_id]):
            return validation_error({
                "error": "subject_id, module_id, and chapter_id are required"
            })
        
        try:
            number_of_questions = int(number_of_questions)
            level = int(level)
        except (ValueError, TypeError):
            return validation_error({
                "error": "number_of_questions and level must be integers"
            })
        
        if level < 1 or level > 5:
            return validation_error({
                "error": "level must be between 1 and 5"
            })
        
        try:
            chapter = ModuleChapter.objects.get(id=chapter_id)
        except ModuleChapter.DoesNotExist:
            return validation_error({
                "error": f"Chapter with ID {chapter_id} not found"
            })
        
        is_hots = level == 5
        
        level_to_difficulty = {
            1: 'easy',
            2: 'easy',
            3: 'medium',
            4: 'hard',
            5: 'hard'
        }
        difficulty = level_to_difficulty.get(level, 'medium')
        
        generated_questions = generate_questions_with_gemini(
            subject_name=subject_name,
            module_name=module_name,
            chapter_name=chapter_name,
            level=level,
            number_of_questions=number_of_questions,
            question_type=question_type
        )
        
        if not generated_questions:
            return validation_error({
                "error": "Failed to generate questions from Gemini AI. Please try again."
            })
        
        questions_created = 0
        module_content_created = 0
        hots_created = 0
        created_questions_list = []
        
        for q_data in generated_questions:
            try:
                question = Question.objects.create(
                    question_text=q_data.get('question_text', ''),
                    question_type=question_type,
                    difficulty_level=difficulty,
                    explanation=q_data.get('explanation', ''),
                    exp_points=q_data.get('exp_points', 10),
                    is_active=True,
                    is_hots=is_hots,
                    ai_generated=True,
                    level=level,
                    created_by=request.user
                )
                questions_created += 1
                
                options_data = q_data.get('options', [])
                created_options = []
                for idx, opt_data in enumerate(options_data):
                    if question_type == 'rearrange':
                        order = opt_data.get('correct_order')
                        if order is None or not isinstance(order, (int, float)):
                            order = idx + 1
                        order = max(1, int(order))
                    else:
                        order = idx + 1
                    option = Option.objects.create(
                        question=question,
                        option_text=opt_data.get('option_text', f'Option {idx + 1}'),
                        is_correct=opt_data.get('is_correct', False),
                        order=order
                    )
                    created_options.append({
                        'id': str(option.id),
                        'option_text': option.option_text,
                        'is_correct': option.is_correct,
                        'order': option.order
                    })
                
                created_questions_list.append({
                    'id': str(question.id),
                    'question_text': question.question_text,
                    'question_type': question.question_type,
                    'difficulty_level': question.difficulty_level,
                    'level': question.level,
                    'explanation': question.explanation,
                    'exp_points': question.exp_points,
                    'is_hots': question.is_hots,
                    'ai_generated': question.ai_generated,
                    'options': created_options
                })
                
                if is_hots:
                    max_order = ChapterHOTS.objects.filter(chapter=chapter).aggregate(
                        max_order=Max('order')
                    )['max_order'] or 0
                    
                    ChapterHOTS.objects.create(
                        chapter=chapter,
                        question=question,
                        order=max_order + 1,
                        created_by=request.user
                    )
                    hots_created += 1
                    
                    chapter.has_hots = True
                    chapter.save(update_fields=['has_hots'])
                else:
                    max_order = ModuleContent.objects.filter(chapter=chapter).aggregate(
                        max_order=Max('order')
                    )['max_order'] or 0
                    
                    ModuleContent.objects.create(
                        chapter=chapter,
                        content_type='question',
                        question=question,
                        order=max_order + 1,
                        created_by=request.user
                    )
                    module_content_created += 1
                
                api_logger.info(f"Created question {question.id} for chapter {chapter_id} using Gemini")
                
            except Exception as e:
                api_logger.error(f"Error creating question: {str(e)}")
                continue
        
        api_logger.info(f"Gemini AI generation complete: {questions_created} questions, {module_content_created} module content, {hots_created} HOTS entries")
        
        return success(
            data={
                'questions_created': questions_created,
                'module_content_created': module_content_created,
                'hots_created': hots_created,
                'level': level,
                'is_hots': is_hots,
                'questions': created_questions_list,
                'chapter_id': str(chapter_id),
            },
            message=f"Successfully generated {questions_created} questions using Gemini"
        )
        
    except Exception as e:
        api_logger.error(f"Error in Gemini AI question generation: {str(e)}")
        return validation_error({
            "error": f"Failed to generate questions: {str(e)}"
        })


def generate_chapter_image_with_gemini(chapter_name: str, theory_text: str, api_key: str = None) -> dict:
    """
    Generate an educational image using Vertex AI Imagen based on chapter name and theory.
    
    Args:
        chapter_name (str): Name of the chapter
        theory_text (str): Theory content (600-650 characters)
        api_key (str, optional): Not used - kept for backward compatibility. 
                                 Vertex AI uses GOOGLE_APPLICATION_CREDENTIALS.
    
    Returns:
        dict: Contains 'image_base64', 'mime_type' on success, or 'error' on failure
    """
    try:
        from gyaan_buddy.utils.gemini_client import generate_chapter_image
        
        result = generate_chapter_image(
            chapter_name=chapter_name,
            theory_text=theory_text,
            aspect_ratio="16:9"
        )
        
        if "error" in result:
            api_logger.error(f"Error generating image with Vertex AI: {result['error']}")
            return result
        
        api_logger.info(f"Successfully generated image for chapter: {chapter_name}")
        return result
            
    except Exception as e:
        api_logger.error(f"Error generating image with Vertex AI: {str(e)}")
        return {"error": f"Failed to generate image: {str(e)}"}


@api_view(['POST'])
@drf_permission_classes([permissions.IsAuthenticated])
def generate_chapter_image(request):
    """
    Generate an educational image for a chapter using Google Gemini's Imagen model.
    
    Expected Input:
    {
        "chapter_name": required (str) - Name of the chapter (for context),
        "theory": required (str) - Theory content (600-650 characters) to visualize,
        "api_key": optional (str) - Gemini API key (uses environment variable if not provided)
    }
    
    Expected Output:
    {
        "success": true,
        "data": {
            "image_base64": str - Base64 encoded image data,
            "mime_type": str - MIME type of the image (image/png),
            "chapter_name": str - The chapter name used
        },
        "message": str
    }
    
    Error Output:
    {
        "success": false,
        "error": str - Error description
    }
    """
    api_logger.info(f"Chapter image generation requested by {request.user.username} (ID: {request.user.id})")
    
    try:
        chapter_name = request.data.get('chapter_name', '').strip()
        theory = request.data.get('theory', '').strip()
        api_key = request.data.get('api_key')
        
        if not chapter_name:
            return validation_error({
                "error": "chapter_name is required"
            })
        
        if not theory:
            return validation_error({
                "error": "theory is required"
            })
        
        if len(theory) < 100:
            return validation_error({
                "error": "theory should be at least 100 characters for meaningful image generation"
            })
        
        if len(theory) > 1000:
            api_logger.warning(f"Theory text too long ({len(theory)} chars), will be truncated to 650 chars")
        
        result = generate_chapter_image_with_gemini(
            chapter_name=chapter_name,
            theory_text=theory,
            api_key=api_key
        )
        
        if "error" in result:
            api_logger.error(f"Image generation failed: {result['error']}")
            return validation_error({
                "error": result["error"]
            })
        
        api_logger.info(f"Chapter image generated successfully for: {chapter_name}")
        
        return success(
            data={
                "image_base64": result["image_base64"],
                "mime_type": result["mime_type"],
                "chapter_name": chapter_name
            },
            message="Chapter image generated successfully"
        )
        
    except Exception as e:
        api_logger.error(f"Error in chapter image generation: {str(e)}")
        return validation_error({
            "error": f"Failed to generate chapter image: {str(e)}"
        })


@api_view(['POST'])
@drf_permission_classes([permissions.IsAuthenticated])
def execute_matplotlib_image(request):
    """
    Execute matplotlib code (e.g. provided by Vertex Gemini) and return the generated image.

    Expected Input:
    {
        "matplotlib_code": required (str) - Python code that uses matplotlib to create a figure.
                            Should use plt (matplotlib.pyplot) and optionally np (numpy).
                            The current figure is saved and returned as PNG.
    }

    Expected Output:
    {
        "success": true,
        "data": {
            "image_base64": str - Base64 encoded PNG image,
            "mime_type": "image/png"
        },
        "message": str
    }
    """
    api_logger.info(f"Matplotlib image execution requested by {request.user.username} (ID: {request.user.id})")
    try:
        matplotlib_code = request.data.get('matplotlib_code')
        if not matplotlib_code or not isinstance(matplotlib_code, str):
            return validation_error({"error": "matplotlib_code is required and must be a non-empty string"})
        from gyaan_buddy.utils.matplotlib_executor import execute_matplotlib_code
        result = execute_matplotlib_code(matplotlib_code.strip())
        if "error" in result:
            return validation_error({"error": result["error"]})
        return success(
            data={
                "image_base64": result["image_base64"],
                "mime_type": result.get("mime_type", "image/png"),
            },
            message="Matplotlib image generated successfully",
        )
    except Exception as e:
        api_logger.error(f"Error in matplotlib image execution: {str(e)}")
        return validation_error({"error": str(e)})


class SubjectViewSet(viewsets.ModelViewSet):
    """ViewSet for Subject model."""
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [parsers.JSONParser, parsers.MultiPartParser, parsers.FormParser]
    
    def get_queryset(self):
        """Return optimized queryset with related data, filtered by school.
        Teachers see only subjects they are assigned to via Teacher assignments.
        Students see only subjects they are enrolled in via StudentSubjectEnrollment.
        Supports ?class=<id> or ?class_instance=<id> to further filter by class."""
        from gyaan_buddy.users.models import Teacher as TeacherAssignment, StudentSubjectEnrollment
        user = self.request.user
        api_logger.info(f"[get_queryset] User: {user.username} (ID: {user.id})")

        def _filter_subjects_for_class(base_queryset, selected_class_id):
            # Prefer Module.class_instance as the authoritative class-subject mapping.
            via_modules = base_queryset.filter(
                modules__class_instance_id=selected_class_id,
                modules__is_active=True,
            ).distinct()
            if via_modules.exists():
                return via_modules
            return base_queryset.filter(classes__id=selected_class_id).distinct()

        queryset = Subject.objects.select_related(
            'school',
            'created_by',
        ).prefetch_related(
            'modules',
        ).filter(is_active=True)

        if hasattr(user, 'profile') and user.profile and user.profile.school:
            queryset = queryset.filter(school=user.profile.school)

            class_instance_id = (
                self.request.query_params.get('class_instance')
                or self.request.query_params.get('class')
            )

            if (user.profile.user_type == 'teacher'
                    and hasattr(user.profile, 'teacher_profile')):
                teacher_qs = TeacherAssignment.objects.filter(
                    teacher=user.profile.teacher_profile,
                    is_deleted=False,
                )
                if class_instance_id:
                    teacher_qs = teacher_qs.filter(class_instance_id=class_instance_id)
                assigned_subject_ids = teacher_qs.values_list('subject_id', flat=True).distinct()
                queryset = queryset.filter(id__in=assigned_subject_ids)
                api_logger.info(f"[get_queryset] Teacher {user.username}: filtered to {queryset.count()} assigned subjects (class={class_instance_id or 'all'})")

            elif (user.profile.user_type == 'student'
                    and hasattr(user.profile, 'student')):
                student = user.profile.student
                enrolled_subject_ids = list(
                    StudentSubjectEnrollment.objects.filter(
                        student=student,
                        is_active=True,
                    ).values_list('subject_id', flat=True).distinct()
                )
                student_class_id = class_instance_id or getattr(student, 'class_instance_id', None)

                if enrolled_subject_ids:
                    queryset = queryset.filter(id__in=enrolled_subject_ids)
                    if class_instance_id:
                        queryset = _filter_subjects_for_class(queryset, class_instance_id)
                    api_logger.info(f"[get_queryset] Student {user.username}: filtered to {queryset.count()} enrolled subjects")
                elif student_class_id:
                    # No explicit enrollment rows yet — fall back to the subjects
                    # offered to the student's class (via Module.class_instance) so
                    # the screen is never blank. Without this a student that was
                    # added without enrollment records sees zero subjects.
                    queryset = _filter_subjects_for_class(queryset, student_class_id)
                    api_logger.info(f"[get_queryset] Student {user.username}: no enrollments, fell back to {queryset.count()} class subjects (class={student_class_id})")
                else:
                    queryset = queryset.none()
                    api_logger.info(f"[get_queryset] Student {user.username}: no enrollments and no class — 0 subjects")

            else:
                # Principal / admin — prefer module-class mapping, then fall back to Subject.classes.
                if class_instance_id:
                    queryset = _filter_subjects_for_class(queryset, class_instance_id)
                    api_logger.info(f"[get_queryset] Admin {user.username}: filtered to {queryset.count()} subjects for class {class_instance_id}")

        return queryset
    
    def list(self, request, *args, **kwargs):
        """List subjects with logging."""
        api_logger.info(f"Subject list requested by {request.user.username} (ID: {request.user.id}) from {request.META.get('REMOTE_ADDR', 'unknown')}")
        
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        api_logger.info(f"Subject list returned {len(serializer.data)} subjects")
        
        return success(
            data=serializer.data,
            message="Subjects retrieved successfully"
        )
    
    def create(self, request, *args, **kwargs):
        """Create a subject with logging."""
        api_logger.info(f"Subject creation requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")
        
        classes_data = request.data.getlist('classes') if hasattr(request.data, 'getlist') else (
            request.data.get('classes', []) if isinstance(request.data.get('classes'), list) else 
            [request.data.get('classes')] if request.data.get('classes') else []
        )
        
        serializer_data = {}
        for key, value in request.data.items():
            if key != 'classes':
                serializer_data[key] = value
        
        for key, file_obj in request.FILES.items():
            if key != 'classes':
                serializer_data[key] = file_obj
        
        if 'is_active' in serializer_data:
            is_active_value = serializer_data['is_active']
            if isinstance(is_active_value, str):
                serializer_data['is_active'] = is_active_value.lower() in ('true', '1', 'yes', 'on')
            elif isinstance(is_active_value, bool):
                serializer_data['is_active'] = is_active_value
            else:
                serializer_data['is_active'] = True
        
        if 'name' in serializer_data and serializer_data['name']:
            serializer_data['name'] = str(serializer_data['name']).strip()
        if 'code' in serializer_data and serializer_data['code']:
            serializer_data['code'] = str(serializer_data['code']).strip()
        
        if 'description' in serializer_data:
            serializer_data['description'] = str(serializer_data['description']).strip() if serializer_data['description'] else ''
        
        from django.core.files.uploadedfile import UploadedFile
        logo_file = None
        if 'logo' in request.FILES:
            logo_file = request.FILES['logo']
        elif 'logo' in serializer_data:
            logo_file = serializer_data['logo']
        
        if logo_file is not None:
            if isinstance(logo_file, UploadedFile):
                if hasattr(logo_file, 'size') and logo_file.size == 0:
                    if 'logo' in serializer_data:
                        del serializer_data['logo']
                else:
                    serializer_data['logo'] = logo_file
            elif isinstance(logo_file, str):
                if 'logo' in serializer_data:
                    del serializer_data['logo']
            else:
                if 'logo' in serializer_data:
                    del serializer_data['logo']
        elif 'logo' in serializer_data:
            del serializer_data['logo']
        
        serializer = self.get_serializer(data=serializer_data)
        if serializer.is_valid():
            school = getattr(getattr(request.user, 'profile', None), 'school', None)
            if not school:
                return validation_error({"school": ["User has no school assigned."]})
            subject = serializer.save(created_by=request.user, school=school)

            if classes_data:
                from gyaan_buddy.users.models import Class
                class_objects = Class.objects.filter(id__in=classes_data, is_active=True, school=school)
                subject.classes.set(class_objects)
                api_logger.info(f"Assigned {class_objects.count()} classes to subject {subject.name}")

            api_logger.info(f"Subject created successfully: {subject.name} (ID: {subject.id}) by {request.user.username}")
            return created(
                data=serializer.data,
                message="Subject created successfully"
            )
        api_logger.warning(f"Subject creation failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)
    
    def update(self, request, *args, **kwargs):
        """Update a subject with logging."""
        subject_id = kwargs.get('pk')
        api_logger.info(f"Subject update requested by {request.user.username} (ID: {request.user.id}) for subject ID: {subject_id} - Data: {request.data}")
        
        classes_data = request.data.getlist('classes') if hasattr(request.data, 'getlist') else (
            request.data.get('classes', []) if isinstance(request.data.get('classes'), list) else 
            [request.data.get('classes')] if request.data.get('classes') else []
        )
        
        serializer_data = {}
        for key, value in request.data.items():
            if key != 'classes':
                serializer_data[key] = value
        
        for key, file_obj in request.FILES.items():
            if key != 'classes':
                serializer_data[key] = file_obj
        
        if 'is_active' in serializer_data:
            is_active_value = serializer_data['is_active']
            if isinstance(is_active_value, str):
                serializer_data['is_active'] = is_active_value.lower() in ('true', '1', 'yes', 'on')
            elif isinstance(is_active_value, bool):
                serializer_data['is_active'] = is_active_value
        
        if 'name' in serializer_data and serializer_data['name']:
            serializer_data['name'] = str(serializer_data['name']).strip()
        if 'code' in serializer_data and serializer_data['code']:
            serializer_data['code'] = str(serializer_data['code']).strip()
        
        if 'description' in serializer_data:
            serializer_data['description'] = str(serializer_data['description']).strip() if serializer_data['description'] else ''
        
        from django.core.files.uploadedfile import UploadedFile
        logo_file = None
        if 'logo' in request.FILES:
            logo_file = request.FILES['logo']
        elif 'logo' in serializer_data:
            logo_file = serializer_data['logo']
        
        if logo_file is not None:
            if isinstance(logo_file, UploadedFile):
                if hasattr(logo_file, 'size') and logo_file.size == 0:
                    if 'logo' in serializer_data:
                        del serializer_data['logo']
                else:
                    serializer_data['logo'] = logo_file
            elif isinstance(logo_file, str):
                if 'logo' in serializer_data:
                    del serializer_data['logo']
            else:
                if 'logo' in serializer_data:
                    del serializer_data['logo']
        elif 'logo' in serializer_data:
            del serializer_data['logo']
        
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=serializer_data, partial=partial)
        
        if serializer.is_valid():
            subject = serializer.save()
            
            if 'classes' in request.data or classes_data is not None:
                from gyaan_buddy.users.models import Class
                user_school = getattr(getattr(request.user, 'profile', None), 'school', None)
                if classes_data:
                    class_objects = Class.objects.filter(id__in=classes_data, is_active=True, school=user_school)
                    subject.classes.set(class_objects)
                    api_logger.info(f"Updated classes for subject {subject.name}: {class_objects.count()} classes")
                else:
                    subject.classes.clear()
                    api_logger.info(f"Cleared all classes for subject {subject.name}")
            
            api_logger.info(f"Subject updated successfully: {subject.name} (ID: {subject.id}) by {request.user.username}")
            return success(
                data=serializer.data,
                message="Subject updated successfully"
            )
        api_logger.warning(f"Subject update failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)
    
    @action(detail=True, methods=['get'], url_path='modules')
    def modules(self, request, pk=None):
        """List modules for a specific subject with user progress."""
        try:
            subject = self.get_object()
            api_logger.info(f"Modules requested for subject '{subject.name}' (ID: {subject.id}) by {request.user.username}")
            
            modules = Module.objects.filter(subject=subject, is_active=True).select_related('subject', 'created_by').order_by('order')

            initial_count = modules.count()
            api_logger.info(f"[modules] Total active modules for subject (before filtering): {initial_count}")

            has_profile = hasattr(request.user, 'profile')
            api_logger.info(f"[modules] User has profile: {has_profile}")

            # Filter by student's class if the user is a student
            user_class = None
            if has_profile:
                profile = request.user.profile
                if hasattr(profile, 'student') and profile.student.class_instance:
                    user_class = profile.student.class_instance
                    modules = modules.filter(class_instance=user_class)
                    api_logger.info(f"[modules] Filtered by student class '{user_class}' → {modules.count()} modules")
            
            
            serializer = ModuleWithProgressSerializer(modules, many=True, context={'request': request})
            
            api_logger.info(f"Returned {len(serializer.data)} modules with progress for subject '{subject.name}'")
            return success(
                data=serializer.data,
                message=f"Modules for {subject.name} retrieved successfully with user progress"
            )
        except Exception as e:
            api_logger.error(f"Error retrieving modules for subject {pk}: {str(e)}")
            return validation_error({"error": "Failed to retrieve modules for this subject"})

    @action(detail=False, methods=['post'], url_path='import_from_excel')
    def import_from_excel(self, request):
        """
        Import subjects, modules, and chapters from an Excel file.

        Columns: Subject | Module | Chapter | Icon (optional image URL)

        Body (JSON):
          class_id   – required, ID of the target Class
          excel_url  – optional, URL of the .xlsx file (defaults to the standard Chapters.xlsx)
          dry_run    – optional bool, preview without saving
        """
        import io
        import urllib.request
        from collections import defaultdict

        import openpyxl
        from django.db import transaction
        from gyaan_buddy.users.models import Class, Teacher as TeacherAssignment, TeacherProfile

        DEFAULT_EXCEL_URL = "https://storage.googleapis.com/gyaanbuddy-media/Chapters.xlsx"

        SUBJECT_DEFAULTS = {
            "maths":       ("MATH", "FF6B6B"),
            "math":        ("MATH", "FF6B6B"),
            "mathematics": ("MATH", "FF6B6B"),
            "science":     ("SCI",  "4ECDC4"),
            "english":     ("ENG",  "45B7D1"),
            "history":     ("HIST", "96CEB4"),
            "civics":      ("CIV",  "FFEAA7"),
            "geography":   ("GEO",  "DDA0DD"),
            "economics":   ("ECO",  "98FB98"),
        }

        def _subject_code_and_color(name):
            key = name.strip().lower()
            if key in SUBJECT_DEFAULTS:
                return SUBJECT_DEFAULTS[key]
            code = name.strip().upper().replace(" ", "")[:10]
            return code, "0DA6F2"

        class_id  = request.data.get('class_id')
        excel_url = request.data.get('excel_url') or DEFAULT_EXCEL_URL
        dry_run   = bool(request.data.get('dry_run', False))

        if not class_id:
            return validation_error({"class_id": "This field is required."})

        # Resolve school from requesting user
        user = request.user
        if not (hasattr(user, 'profile') and user.profile and user.profile.school):
            return validation_error({"error": "Your account is not linked to a school."})
        school = user.profile.school

        # Resolve class
        try:
            cls = Class.objects.get(id=class_id, school=school)
        except Class.DoesNotExist:
            return validation_error({"class_id": f"Class with id {class_id} not found in your school."})

        # Resolve teacher (if requesting user is a teacher)
        teacher_profile = None
        created_by = None
        if (hasattr(user, 'profile')
                and user.profile.user_type == 'teacher'
                and hasattr(user.profile, 'teacher_profile')):
            teacher_profile = user.profile.teacher_profile
            created_by = user

        # Fetch and parse Excel
        try:
            data = urllib.request.urlopen(excel_url, timeout=30).read()
        except Exception as exc:
            return validation_error({"excel_url": f"Failed to fetch Excel file: {exc}"})

        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        except Exception as exc:
            return validation_error({"excel_url": f"Failed to parse Excel file: {exc}"})

        ws = wb.active
        raw_sample = [list(row) for row in ws.iter_rows(values_only=True, max_row=3)]
        HEADER_KEYWORDS = {"subject", "module", "chapter", "icon", "title", "name"}
        rows = []
        for row in ws.iter_rows(values_only=True):
            subject_name  = str(row[0]).strip() if row[0] else ""
            module_name   = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            chapter_title = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            icon_url      = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            if subject_name.lower() in HEADER_KEYWORDS:
                continue  # skip header row
            if subject_name and module_name and chapter_title:
                rows.append((subject_name, module_name, chapter_title, icon_url))

        if not rows:
            return validation_error({"excel_url": "No valid rows found. Excel must have 3 columns: Subject, Module, Chapter."})

        # Build dry-run summary
        summary = defaultdict(lambda: defaultdict(list))
        for subject_name, module_name, chapter_title, icon_url in rows:
            summary[subject_name][module_name].append(chapter_title)

        if dry_run:
            preview = []
            for subj, modules in summary.items():
                preview.append({
                    "subject": subj,
                    "modules": [
                        {"module": mod, "chapter_count": len(chaps)}
                        for mod, chaps in modules.items()
                    ],
                })
            return success(
                data={"dry_run": True, "total_rows": len(rows), "preview": preview},
                message="Dry run complete — nothing was saved."
            )

        # Persist
        subject_created = subject_existing = 0
        module_created  = module_existing  = 0
        chapter_created = chapter_existing = 0
        new_chapters_to_process = []  # (chapter, module) pairs for post-transaction AI generation

        with transaction.atomic():
            subject_cache    = {}
            module_cache     = {}
            module_order_map = {}
            subject_order_map = {}

            next_subject_order = (
                Subject.objects.filter(school=school)
                .order_by("-order")
                .values_list("order", flat=True)
                .first()
            ) or 0

            for subject_name, module_name, chapter_title, icon_url in rows:
                if subject_name not in subject_cache:
                    code, color = _subject_code_and_color(subject_name)
                    next_subject_order += 1
                    subject, s_created = Subject.objects.get_or_create(
                        name__iexact=subject_name,
                        school=school,
                        defaults={
                            "name": subject_name,
                            "code": code,
                            "color": color,
                            "is_active": True,
                            "order": next_subject_order,
                            "created_by": created_by,
                        },
                    )
                    if not s_created:
                        next_subject_order -= 1
                    subject_cache[subject_name] = subject
                    if s_created:
                        subject_created += 1
                    else:
                        subject_existing += 1

                    cls.subjects.add(subject)

                    if teacher_profile:
                        already_assigned = TeacherAssignment.objects.filter(
                            class_instance=cls, subject=subject
                        ).exists()
                        if not already_assigned:
                            TeacherAssignment.objects.create(
                                teacher=teacher_profile,
                                class_instance=cls,
                                subject=subject,
                            )

                subject = subject_cache[subject_name]

                module_key = (subject_name, module_name)
                if module_key not in module_cache:
                    if subject.id not in subject_order_map:
                        subject_order_map[subject.id] = (
                            Module.objects.filter(subject=subject, class_instance=cls)
                            .order_by("-order")
                            .values_list("order", flat=True)
                            .first()
                        ) or 0
                    subject_order_map[subject.id] += 1
                    module, m_created = Module.objects.get_or_create(
                        name__iexact=module_name,
                        subject=subject,
                        class_instance=cls,
                        defaults={
                            "name": module_name,
                            "subject": subject,
                            "class_instance": cls,
                            "is_active": True,
                            "is_enabled": False,
                            "order": subject_order_map[subject.id],
                            "logo_url": icon_url or "",
                            "created_by": created_by,
                        },
                    )
                    if not m_created:
                        subject_order_map[subject.id] -= 1
                        if icon_url:
                            module.logo_url = icon_url
                            module.save(update_fields=["logo_url"])
                    module_cache[module_key] = module
                    if m_created:
                        module_created += 1
                    else:
                        module_existing += 1

                module = module_cache[module_key]

                if module.id not in module_order_map:
                    existing_max = (
                        ModuleChapter.objects.filter(module=module)
                        .order_by("-order")
                        .values_list("order", flat=True)
                        .first()
                    ) or 0
                    module_order_map[module.id] = existing_max

                module_order_map[module.id] += 1
                chapter, c_created = ModuleChapter.objects.get_or_create(
                    title__iexact=chapter_title,
                    module=module,
                    defaults={
                        "title": chapter_title,
                        "module": module,
                        "order": module_order_map[module.id],
                        "is_enabled": True,
                        "is_important": False,
                        "created_by": created_by,
                    },
                )
                if not c_created:
                    module_order_map[module.id] -= 1
                if c_created:
                    chapter_created += 1
                    new_chapters_to_process.append((chapter, module))
                else:
                    chapter_existing += 1
                    # Backfill missing theory/image for existing chapters
                    needs_theory = not (chapter.theory and chapter.theory.strip())
                    needs_image  = not chapter.logo
                    if needs_theory or needs_image:
                        new_chapters_to_process.append((chapter, module))

        # Generate theory + image for every newly created chapter (outside transaction)
        theory_generated = theory_failed = 0
        image_generated  = image_failed  = 0

        for chapter, module in new_chapters_to_process:
            matplotlib_code = None
            needs_theory = not (chapter.theory and chapter.theory.strip())
            needs_image  = not chapter.logo

            if needs_theory:
                try:
                    from gyaan_buddy.utils.gemini_client import generate_chapter_theory
                    subject_name = module.subject.name if module.subject else None
                    api_logger.info(f"Generating theory for chapter '{chapter.title}' (ID: {chapter.id})")
                    theory_result = generate_chapter_theory(
                        chapter_name=chapter.title,
                        subject=subject_name,
                        min_length=600,
                        max_length=650,
                    )
                    if 'error' not in theory_result:
                        chapter.theory = theory_result['theory_text']
                        chapter.save(update_fields=['theory'])
                        theory_generated += 1
                        matplotlib_code = theory_result.get('matplotlib_code')
                        needs_image = True  # always generate image after fresh theory
                    else:
                        api_logger.error(f"Theory generation failed for chapter {chapter.id}: {theory_result['error']}")
                        theory_failed += 1
                except Exception as exc:
                    api_logger.error(f"Theory generation error for chapter {chapter.id}: {exc}")
                    theory_failed += 1

            if needs_image and chapter.theory and chapter.theory.strip():
                try:
                    self._generate_and_save_chapter_image(chapter, matplotlib_code=matplotlib_code)
                    image_generated += 1
                except Exception as img_exc:
                    api_logger.error(f"Image generation failed for chapter {chapter.id}: {img_exc}")
                    image_failed += 1

        icon_debug = {
            mod_name: module_cache[(subj_name, mod_name)].logo_url
            for subj_name, mod_name, _, __ in rows
            for _ in [None] if (subj_name, mod_name) in module_cache
        }

        return success(
            data={
                "dry_run": False,
                "subjects": {"created": subject_created, "existing": subject_existing},
                "modules":  {"created": module_created,  "existing": module_existing},
                "chapters": {"created": chapter_created, "existing": chapter_existing},
                "theory_generation": {"success": theory_generated, "failed": theory_failed},
                "image_generation":  {"success": image_generated,  "failed": image_failed},
                "debug_module_logos": icon_debug,
                "debug_excel_sample": raw_sample,
            },
            message="Import complete."
        )


class ModuleViewSet(viewsets.ModelViewSet):
    """ViewSet for Module model."""
    queryset = Module.objects.all()
    serializer_class = ModuleSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [parsers.JSONParser, parsers.MultiPartParser, parsers.FormParser]
    
    def get_queryset(self):
        """Return optimized queryset with related data, scoped to user's school.
        Supports ?subject=<id> and ?class_instance=<id> query params.
        Teachers see only modules whose subject they are assigned to."""
        from gyaan_buddy.users.models import Teacher as TeacherAssignment
        user = self.request.user
        queryset = Module.objects.select_related(
            'subject',
            'class_instance',
        ).prefetch_related(
            'chapters',
        ).filter(is_active=True)

        if hasattr(user, 'profile') and user.profile and user.profile.school:
            queryset = queryset.filter(subject__school=user.profile.school)

            if (user.profile.user_type == 'teacher'
                    and hasattr(user.profile, 'teacher_profile')):
                assigned_subject_ids = TeacherAssignment.objects.filter(
                    teacher=user.profile.teacher_profile,
                    is_deleted=False,
                ).values_list('subject_id', flat=True).distinct()
                queryset = queryset.filter(subject_id__in=assigned_subject_ids)
        else:
            queryset = queryset.none()

        subject_id = self.request.query_params.get('subject')
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)

        class_instance_id = (
            self.request.query_params.get('class_instance')
            or self.request.query_params.get('class')
        )
        if class_instance_id:
            queryset = queryset.filter(class_instance_id=class_instance_id)

        return queryset.order_by('order', 'created_at')

    def list(self, request, *args, **kwargs):
        """List modules with logging."""
        api_logger.info(f"Module list requested by {request.user.username} (ID: {request.user.id}) from {request.META.get('REMOTE_ADDR', 'unknown')}")

        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        api_logger.info(f"Module list returned {len(serializer.data)} modules")

        return success(
            data=serializer.data,
            message="Modules retrieved successfully"
        )

    def create(self, request, *args, **kwargs):
        """Create a module with logging."""
        api_logger.info(f"Module creation requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")

        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            module = serializer.save(created_by=request.user)
            api_logger.info(f"Module created successfully: {module.name} (ID: {module.id}) by {request.user.username}")
            return created(
                data=serializer.data,
                message="Module created successfully"
            )
        api_logger.warning(f"Module creation failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)
    
    def update(self, request, *args, **kwargs):
        """Update a module with logging."""
        module_id = kwargs.get('pk')
        api_logger.info(f"Module update requested by {request.user.username} (ID: {request.user.id}) for module ID: {module_id} - Data: {request.data}")
        
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        
        if serializer.is_valid():
            module = serializer.save()
            api_logger.info(f"Module updated successfully: {module.name} (ID: {module.id}) by {request.user.username}")
            return success(
                data=serializer.data,
                message="Module updated successfully"
            )
        api_logger.warning(f"Module update failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)
    
    def partial_update(self, request, *args, **kwargs):
        """Partial update a module with logging."""
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)
    
    @action(detail=True, methods=['get'], url_path='module_chapters')
    def module_chapters(self, request, pk=None):
        """Get chapters for a specific module."""
        try:
            module = self.get_object()
            api_logger.info(f"Chapters requested for module '{module.name}' (ID: {module.id}) by {request.user.username}")
            
            chapters = ModuleChapter.objects.filter(
                module=module, 
                is_enabled=True,
                is_deleted=False
            ).select_related('module', 'created_by').prefetch_related('contents').order_by('order')
            
            chapter_progress = UserChapterProgress.objects.none()
            
            if chapters.exists():
                chapter_progress = UserChapterProgress.objects.filter(
                    account=request.user,
                    chapter__in=chapters,
                    status='in_progress'
                )

                if not chapter_progress.exists():
                    completed_chapter_ids = UserChapterProgress.objects.filter(
                        account=request.user,
                        chapter__in=chapters,
                        status='completed'
                    ).values_list('chapter_id', flat=True)
                    
                    available_chapter = chapters.exclude(id__in=completed_chapter_ids).first()
                    
                    if available_chapter:
                        progress, created = UserChapterProgress.objects.get_or_create(
                            account=request.user,
                            chapter=available_chapter,
                            defaults={'status': 'in_progress'}
                        )
                        if not created and progress.status != 'in_progress':
                            progress.status = 'in_progress'
                            progress.save()
                

            serializer = ModuleChapterSerializer(chapters, many=True, context={'request': request, 'chapter_progress': chapter_progress})

            
            api_logger.info(f"Returned {len(serializer.data)} chapters for module '{module.name}'")
            return success(
                data=serializer.data,
                message=f"Chapters for {module.name} retrieved successfully"
            )
        except Exception as e:
            api_logger.error(f"Error retrieving chapters for module {pk}: {str(e)}")
            return validation_error({"error": "Failed to retrieve chapters for this module"})

    @action(detail=True, methods=['post'], url_path='set_due')
    def set_due(self, request, pk=None):
        """Set or clear due_date on all active chapters of this module.

        Body (either field):
          { "is_due": true/false }           — sets today / clears due_date
          { "due_date": "YYYY-MM-DD" | null } — sets specific date / clears
        """
        module = self.get_object()

        due_date_raw = request.data.get('due_date', '__missing__')
        is_due = request.data.get('is_due')

        if due_date_raw != '__missing__':
            if due_date_raw is None:
                new_due_date = None
            else:
                from datetime import date
                try:
                    new_due_date = date.fromisoformat(str(due_date_raw))
                except ValueError:
                    return validation_error({"due_date": "Invalid date format. Use YYYY-MM-DD."})
        elif is_due is not None:
            new_due_date = timezone.now().date() if is_due else None
        else:
            return validation_error({"error": "Provide 'is_due' or 'due_date'."})

        was_already_enabled = module.is_enabled

        chapters = ModuleChapter.objects.filter(module=module, is_deleted=False)
        updated_count = chapters.update(due_date=new_due_date)

        if new_due_date is not None and not module.is_enabled:
            module.is_enabled = True
            module.save(update_fields=['is_enabled'])
            api_logger.info(f"Module '{module.name}' (ID: {module.id}) is_enabled set to True because due_date was set")
        elif new_due_date is None and module.is_enabled:
            all_cleared = not ModuleChapter.objects.filter(module=module, is_enabled=True, is_deleted=False, due_date__isnull=False).exists()
            if all_cleared:
                module.is_enabled = False
                module.save(update_fields=['is_enabled'])
                api_logger.info(f"Module '{module.name}' (ID: {module.id}) is_enabled set to False because all chapter due_dates were cleared")

        api_logger.info(
            f"Module '{module.name}' (ID: {module.id}) chapters due_date set to "
            f"{new_due_date} by {request.user.username} ({updated_count} chapters updated)"
        )

        serializer = self.get_serializer(module, context={'request': request})
        return success(data=serializer.data, message="Module due date updated successfully")


class ModuleChapterViewSet(viewsets.ModelViewSet):
    """ViewSet for ModuleChapter model."""
    queryset = ModuleChapter.objects.all()
    serializer_class = ModuleChapterSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Return optimized queryset with related data, scoped to user's school.
        Teachers see only chapters whose subject they are assigned to."""
        from gyaan_buddy.users.models import Teacher as TeacherAssignment
        user = self.request.user
        queryset = ModuleChapter.objects.select_related(
            'module',
            'created_by',
        ).prefetch_related(
            'contents',
        ).filter(is_enabled=True, is_deleted=False)

        if hasattr(user, 'profile') and user.profile and user.profile.school:
            queryset = queryset.filter(module__subject__school=user.profile.school)

            if (user.profile.user_type == 'teacher'
                    and hasattr(user.profile, 'teacher_profile')):
                assigned_subject_ids = TeacherAssignment.objects.filter(
                    teacher=user.profile.teacher_profile,
                    is_deleted=False,
                ).values_list('subject_id', flat=True).distinct()
                queryset = queryset.filter(module__subject_id__in=assigned_subject_ids)
        else:
            queryset = queryset.none()

        return queryset
    
    def list(self, request, *args, **kwargs):
        """List chapters with logging."""
        api_logger.info(f"Chapter list requested by {request.user.username} (ID: {request.user.id}) from {request.META.get('REMOTE_ADDR', 'unknown')}")
        
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True, context={'request': request})
        api_logger.info(f"Chapter list returned {len(serializer.data)} chapters")
        
        return success(
            data=serializer.data,
            message="Chapters retrieved successfully"
        )
    
    def retrieve(self, request, *args, **kwargs):
        """Retrieve a single chapter with logging."""
        api_logger.info(f"Chapter retrieval requested by {request.user.username} (ID: {request.user.id}) for chapter ID: {kwargs.get('pk')}")
        
        instance = self.get_object()
        serializer = self.get_serializer(instance, context={'request': request})
        api_logger.info(f"Chapter retrieved successfully: {instance.title} (ID: {instance.id})")
        
        return success(
            data=serializer.data,
            message="Chapter retrieved successfully"
        )
    
    def create(self, request, *args, **kwargs):
        """Create a chapter with logging and auto-generate image from theory."""
        api_logger.info(f"Chapter creation requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")
        
        data = {}
        for key, value in request.data.items():
            data[key] = value
        
        for key, file_obj in request.FILES.items():
            data[key] = file_obj
        
        theory_provided = 'theory' in data and data.get('theory', '').strip()
        if theory_provided:
            theory_preview = str(data['theory'])[:100] if data['theory'] else 'None'
            api_logger.info(f"Theory field present in request data: '{theory_preview}...' (length: {len(str(data['theory'])) if data['theory'] else 0})")
        else:
            api_logger.info("Theory field not provided or empty - will auto-generate using Vertex AI Gemini")
        
        module_id = data.get('module')
        module = None
        
        if module_id:
            try:
                module = Module.objects.select_related('subject').get(id=module_id)
            except Module.DoesNotExist:
                api_logger.warning(f"Module {module_id} not found")
            except Exception as e:
                api_logger.error(f"Error fetching module: {str(e)}")

        if not theory_provided:
            chapter_name = data.get('title', '')
            if chapter_name and module:
                try:
                    from gyaan_buddy.utils.gemini_client import generate_chapter_theory
                    
                    subject_name = module.subject.name if module.subject else None
                    api_logger.info(f"Generating theory for chapter '{chapter_name}' using Vertex AI Gemini (subject: {subject_name})")
                    
                    theory_result = generate_chapter_theory(
                        chapter_name=chapter_name,
                        subject=subject_name,
                        min_length=600,
                        max_length=650
                    )
                    
                    if 'error' in theory_result:
                        api_logger.error(f"Failed to generate theory: {theory_result['error']}")
                    else:
                        data['theory'] = theory_result['theory_text']
                        data['_matplotlib_code'] = theory_result.get('matplotlib_code')
                        api_logger.info(
                            f"Successfully generated theory ({theory_result['character_count']} characters)"
                            f"{', with diagram' if theory_result.get('matplotlib_code') else ''}"
                        )
                except Exception as e:
                    api_logger.error(f"Error generating theory with Vertex AI: {str(e)}")
                    import traceback
                    api_logger.error(f"Traceback: {traceback.format_exc()}")
            else:
                api_logger.warning(f"Cannot generate theory: chapter_name='{chapter_name}', module={'found' if module else 'not found'}")
        
        if module:
            try:
                requested_order = data.get('order')
                if requested_order:
                    order = int(requested_order)
                else:
                    order = (ModuleChapter.objects.filter(
                        module=module,
                        is_deleted=False
                    ).aggregate(Max('order'))['order__max'] or 0) + 1

                # Increment until a free slot is found
                while ModuleChapter.objects.filter(
                    module=module, order=order, is_deleted=False
                ).exists():
                    order += 1

                if str(order) != str(requested_order):
                    api_logger.info(f"Order {requested_order} taken, using next available order: {order}")
                data['order'] = order
            except Exception as e:
                api_logger.error(f"Error handling order field: {str(e)}")

        serializer = None
        for _attempt in range(10):
            serializer = self.get_serializer(data=data, context={'request': request})
            if serializer.is_valid():
                break
            errors = serializer.errors
            non_field = errors.get('non_field_errors', [])
            if non_field and any('unique set' in str(e) for e in non_field) and module:
                data['order'] = int(data.get('order', 1)) + 1
                api_logger.info(f"Order conflict at validation, retrying with order: {data['order']}")
                continue
            api_logger.warning(f"Chapter creation failed - Errors: {errors}")
            return validation_error(errors)
        else:
            api_logger.warning(f"Chapter creation failed after retries - Errors: {serializer.errors}")
            return validation_error(serializer.errors)

        chapter = serializer.save(created_by=request.user)
        api_logger.info(f"Chapter created successfully: {chapter.title} (ID: {chapter.id}) by {request.user.username}")

        if chapter.due_date is not None:
            module = chapter.module
            if not module.is_enabled:
                module.is_enabled = True
                module.save(update_fields=['is_enabled'])
                api_logger.info(f"Module '{module.name}' (ID: {module.id}) is_enabled set to True because chapter '{chapter.title}' was created with due_date")

        matplotlib_code = data.get('_matplotlib_code')
        if chapter.theory and chapter.theory.strip():
            api_logger.info(f"Theory available for chapter '{chapter.title}' — generating image")
            try:
                self._generate_and_save_chapter_image(chapter, matplotlib_code=matplotlib_code)
                api_logger.info(f"Image generated successfully for chapter: {chapter.title} (ID: {chapter.id})")
            except Exception as e:
                api_logger.error(f"Image generation failed for chapter {chapter.id}: {str(e)}")
                import traceback
                api_logger.error(traceback.format_exc())
        else:
            api_logger.info(f"No theory for chapter '{chapter.title}' — skipping image generation")

        serializer = self.get_serializer(chapter, context={'request': request})
        return created(
            data=serializer.data,
            message="Chapter created successfully"
        )
    
    def update(self, request, *args, **kwargs):
        """Update a chapter with logging. Theory and image are updated directly from payload."""
        chapter_id = kwargs.get('pk')
        api_logger.info(f"Chapter update requested by {request.user.username} (ID: {request.user.id}) for chapter ID: {chapter_id} - Data: {request.data}")

        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        old_due_date = instance.due_date

        serializer = self.get_serializer(instance, data=request.data, partial=partial, context={'request': request})

        if serializer.is_valid():
            chapter = serializer.save()
            api_logger.info(f"Chapter updated successfully: {chapter.title} (ID: {chapter.id}) by {request.user.username}")

            new_due_date = chapter.due_date

            module = chapter.module
            if new_due_date is not None:
                if not module.is_enabled:
                    module.is_enabled = True
                    module.save(update_fields=['is_enabled'])
                    api_logger.info(f"Module '{module.name}' (ID: {module.id}) is_enabled set to True because chapter '{chapter.title}' due_date was set")
            elif new_due_date is None and old_due_date is not None and module.is_enabled:
                all_cleared = not ModuleChapter.objects.filter(module=module, is_enabled=True, is_deleted=False, due_date__isnull=False).exists()
                if all_cleared:
                    module.is_enabled = False
                    module.save(update_fields=['is_enabled'])
                    api_logger.info(f"Module '{module.name}' (ID: {module.id}) is_enabled set to False because all chapter due_dates were cleared")

            if new_due_date is not None and old_due_date is None:
                try:
                    from django.contrib.auth import get_user_model
                    from gyaan_buddy.utils.firebase_notifications import firebase_notification_service
                    User = get_user_model()
                    class_instance_id = chapter.module.class_instance_id
                    if class_instance_id:
                        students = User.objects.filter(
                            profile__student__class_instance_id=class_instance_id,
                            profile__user_type='student',
                            is_active=True,
                            is_deleted=False,
                        )
                        if students.exists():
                            _day = new_due_date.day
                            _suffix = 'th' if 11 <= _day <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(_day % 10, 'th')
                            _date_str = f"{_day}{_suffix} {new_due_date.strftime('%B')}"
                            title = f"New {chapter.module.subject.name} assignment alert!"
                            body = f"{chapter.title} is live, complete it before {_date_str} to stay on track"
                            data = {
                                'type': 'chapter_due',
                                'chapter_id': str(chapter.id),
                                'chapter_title': chapter.title,
                                'module_id': str(chapter.module_id),
                                'module_name': chapter.module.name,
                                'subject_name': chapter.module.subject.name,
                                'due_date': str(new_due_date),
                                'action': 'open_chapter',
                            }
                            results = firebase_notification_service.send_notification_to_multiple_users(
                                list(students), title, body, data,
                                notification_type='module',
                                triggered_by='auto',
                            )
                            api_logger.info(f"Chapter due notification sent for '{chapter.title}': {results}")
                except Exception as notif_error:
                    api_logger.error(f"Failed to send chapter due notification for chapter {chapter.id}: {str(notif_error)}")

            return success(
                data=serializer.data,
                message="Chapter updated successfully"
            )
        api_logger.warning(f"Chapter update failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)
    
    def _generate_and_save_chapter_image(self, chapter, matplotlib_code=None):
        """
        Generate an image and save it to the chapter's logo field.

        If matplotlib_code is provided (extracted from the theory diagram), execute it
        to produce the image. Otherwise fall back to Imagen via generate_chapter_image_with_gemini.

        Args:
            chapter: ModuleChapter instance with title and theory
            matplotlib_code: Optional Python matplotlib code string from theory generation
        """
        import base64
        from django.core.files.base import ContentFile

        api_logger.info(f"Generating image for chapter: {chapter.title} (ID: {chapter.id})")

        image_base64 = None

        # Prefer matplotlib code from theory (same flow as question image generation)
        if matplotlib_code and matplotlib_code.strip():
            api_logger.info(f"Using matplotlib diagram from theory for chapter: {chapter.title}")
            from gyaan_buddy.utils.matplotlib_executor import execute_matplotlib_code
            result = execute_matplotlib_code(matplotlib_code)
            if "error" not in result:
                image_base64 = result["image_base64"]
                api_logger.info(f"Matplotlib diagram executed successfully for chapter: {chapter.title}")
            else:
                api_logger.warning(
                    f"Matplotlib execution failed for chapter {chapter.id}: {result['error']} "
                    f"— falling back to Imagen"
                )

        # Fall back to Imagen if no matplotlib code or execution failed
        if not image_base64:
            # theory_text = chapter.theory[:650] if len(chapter.theory) > 650 else chapter.theory
            # result = generate_chapter_image_with_gemini(
            #     chapter_name=chapter.title,
            #     theory_text=theory_text
            # )
            # if "error" in result:
            #     api_logger.error(f"Image generation failed for chapter {chapter.id}: {result['error']}")
            #     raise Exception(result["error"])
            # image_base64 = result["image_base64"]

            api_logger.warning(f"No image generated for chapter {chapter.id} — Imagen fallback disabled and matplotlib unavailable/failed")
            return

        image_data = base64.b64decode(image_base64)
        image_filename = f"chapter_{chapter.id}_generated.png"
        chapter.logo.save(image_filename, ContentFile(image_data), save=True)
        api_logger.info(f"Image saved for chapter: {chapter.title} (ID: {chapter.id})")
    
    def destroy(self, request, *args, **kwargs):
        """Delete (soft delete) a chapter with logging."""
        chapter_id = kwargs.get('pk')
        api_logger.info(f"Chapter deletion requested by {request.user.username} (ID: {request.user.id}) for chapter ID: {chapter_id}")
        
        try:
            instance = self.get_object()
            instance.is_deleted = True
            instance.save()
            api_logger.info(f"Chapter soft deleted successfully: {instance.title} (ID: {instance.id}) by {request.user.username}")
            return success(
                data={},
                message="Chapter deleted successfully"
            )
        except Exception as e:
            api_logger.error(f"Error deleting chapter {chapter_id}: {str(e)}")
            return validation_error({"error": f"Failed to delete chapter: {str(e)}"})
    
    @action(detail=True, methods=['get'], url_path='get_next_content')
    def get_next_content(self, request, pk=None):
        """Get the next content item for a specific chapter."""
        try:
            chapter = self.get_object()
            
            current_content_id = request.query_params.get('id')
            
            if current_content_id:
                try:
                    from .helpers import handle_next_content_request
                    return handle_next_content_request(
                        request, chapter, current_content_id, 
                        api_logger, success, validation_error, 
                        ModuleContentSerializer
                    )
                except Exception as e:
                    api_logger.error(f"Error in handle_next_content_request: {str(e)}")
                    return validation_error({"error": f"Failed to process next content request: {str(e)}"})
            else:
                try:
                    from .helpers import handle_first_content_request
                    return handle_first_content_request(
                        request, chapter, api_logger, success, 
                        ModuleContentSerializer
                    )
                except Exception as e:
                    api_logger.error(f"Error in handle_first_content_request: {str(e)}")
                    return validation_error({"error": f"Failed to process first content request: {str(e)}"})
                
        except Exception as e:
            api_logger.error(f"Error retrieving next content for chapter {pk}: {str(e)}")
            return validation_error({"error": f"Failed to retrieve next content: {str(e)}"})
    
    @action(detail=True, methods=['get'], url_path='module_questions')
    def module_questions(self, request, pk=None):
        """Get random questions for a specific chapter, limited by max_questions."""
        try:
            chapter = self.get_object()
            
            api_logger.info(f"Chapter questions requested for chapter '{chapter.title}' (ID: {chapter.id}) by {request.user.username}")
            
            max_questions = min(chapter.max_questions, 100)

            chapter_contents = ModuleContent.objects.filter(
                chapter=chapter,
                content_type='question',
                is_deleted=False,
                question__is_active=True,
                question__is_deleted=False
            ).select_related(
                'chapter', 'question', 'created_by'
            ).prefetch_related(
                'question__options'
            ).order_by('question__level')[:max_questions]

            questions = [content.question for content in chapter_contents if content.question]
            
            questions_serializer = QuestionSerializer(questions, many=True)
            
            api_logger.info(f"Returned {len(questions)} random questions (max: {max_questions}) for chapter '{chapter.title}'")
            return success(
                data=questions_serializer.data,
                message=f"Questions for chapter '{chapter.title}' retrieved successfully"
            )
            
        except Exception as e:
            api_logger.error(f"Error retrieving chapter questions for chapter {pk}: {str(e)}")
            return validation_error({"error": f"Failed to retrieve chapter questions: {str(e)}"})


    
    @action(detail=True, methods=['post'], url_path='generate_learn_mode')
    def generate_learn_mode(self, request, pk=None):
        """Generate (or regenerate) theory and diagram image for an existing chapter."""
        try:
            chapter = self.get_object()
            api_logger.info(
                f"generate_learn_mode requested for chapter '{chapter.title}' (ID: {chapter.id}) "
                f"by {request.user.username}"
            )

            module = chapter.module
            subject_name = module.subject.name if module and module.subject else None

            from gyaan_buddy.utils.gemini_client import generate_chapter_theory
            theory_result = generate_chapter_theory(
                chapter_name=chapter.title,
                subject=subject_name,
                min_length=600,
                max_length=650,
            )

            if 'error' in theory_result:
                api_logger.error(f"Theory generation failed for chapter {chapter.id}: {theory_result['error']}")
                return validation_error({"error": f"Theory generation failed: {theory_result['error']}"})

            chapter.theory = theory_result['theory_text']
            chapter.save(update_fields=['theory'])
            api_logger.info(
                f"Theory saved for chapter '{chapter.title}' ({theory_result['character_count']} chars)"
            )

            matplotlib_code = theory_result.get('matplotlib_code')
            try:
                self._generate_and_save_chapter_image(chapter, matplotlib_code=matplotlib_code)
                api_logger.info(f"Image generated for chapter '{chapter.title}' (ID: {chapter.id})")
            except Exception as img_err:
                api_logger.error(f"Image generation failed for chapter {chapter.id}: {str(img_err)}")

            serializer = self.get_serializer(chapter, context={'request': request})
            return success(
                data=serializer.data,
                message="Theory and image generated successfully",
            )

        except Exception as e:
            api_logger.error(f"generate_learn_mode failed for chapter {pk}: {str(e)}")
            return validation_error({"error": f"Failed to generate learn mode: {str(e)}"})

    @action(detail=True, methods=['get'], url_path='exp_statistics')
    def exp_statistics(self, request, pk=None):
        """Get per-user exp gain statistics (average, max, min) for all questions in a modulechapter."""
        try:
            chapter = self.get_object()
            
            api_logger.info(f"Exp statistics requested for chapter '{chapter.title}' (ID: {chapter.id}) by {request.user.username}")
            
            chapter_questions = Question.objects.filter(
                module_contents__chapter=chapter,
                module_contents__content_type='question',
                module_contents__is_deleted=False
            ).distinct()
            
            if not chapter_questions.exists():
                api_logger.info(f"No questions found in chapter '{chapter.title}'")
                return success(
                    data={
                        'chapter_id': str(chapter.id),
                        'chapter_title': chapter.title,
                        'statistics': []
                    },
                    message="No questions found in this chapter"
                )
            
            answers = Answer.objects.filter(
                question__in=chapter_questions
            ).annotate(
                exp_gain=F('current_Exp') - F('prev_exp')
            ).select_related('user', 'question')
            
            user_stats = answers.values('user__id', 'user__username').annotate(
                avg_exp_gain=Avg('exp_gain'),
                max_exp_gain=Max('exp_gain'),
                min_exp_gain=Min('exp_gain'),
                total_answers=Count('id')
            ).order_by('-avg_exp_gain')
            
            statistics = []
            for stat in user_stats:
                statistics.append({
                    'user_id': str(stat['user__id']),
                    'username': stat['user__username'],
                    'average_exp_gain': round(float(stat['avg_exp_gain']), 2) if stat['avg_exp_gain'] is not None else 0,
                    'max_exp_gain': stat['max_exp_gain'] if stat['max_exp_gain'] is not None else 0,
                    'min_exp_gain': stat['min_exp_gain'] if stat['min_exp_gain'] is not None else 0,
                    'total_answers': stat['total_answers']
                })
            
            api_logger.info(f"Exp statistics calculated for {len(statistics)} users in chapter '{chapter.title}'")
            return success(
                data={
                    'chapter_id': str(chapter.id),
                    'chapter_title': chapter.title,
                    'total_questions': chapter_questions.count(),
                    'statistics': statistics
                },
                message=f"Exp statistics retrieved successfully for chapter '{chapter.title}'"
            )
            
        except Exception as e:
            api_logger.error(f"Error retrieving exp statistics for chapter {pk}: {str(e)}")
            return validation_error({"error": f"Failed to retrieve exp statistics: {str(e)}"})
    
    @action(detail=True, methods=['get', 'post'], url_path='module_content')
    def module_content(self, request, pk=None):
        """Get or create module content (question or theory) for a specific chapter."""
        try:
            chapter = self.get_object()
            
            if request.method == 'GET':
                api_logger.info(f"Module content retrieval requested for chapter '{chapter.title}' (ID: {chapter.id}) by {request.user.username}")
                
                module_contents = ModuleContent.objects.filter(
                    chapter=chapter,
                    is_deleted=False,
                ).exclude(
                    content_type='question',
                    question__is_deleted=True,
                ).exclude(
                    content_type='question',
                    question__is_active=False,
                ).select_related(
                    'question', 'theory', 'created_by'
                ).prefetch_related(
                    'question__options'
                ).order_by('order')
                
                serializer = ModuleContentSerializer(module_contents, many=True, context={'request': request})
                api_logger.info(f"Module content retrieved successfully: {len(serializer.data)} items for chapter '{chapter.title}'")
                return success(
                    data=serializer.data,
                    message="Module content retrieved successfully"
                )
            
            elif request.method == 'POST':
                api_logger.info(f"Module content creation requested for chapter '{chapter.title}' (ID: {chapter.id}) by {request.user.username} - Data: {request.data}")
                
                content_type = request.data.get('type') or request.data.get('content_type')
                if not content_type:
                    return validation_error({"error": "Field 'type' is required"})
                
                if content_type not in ['question', 'theory']:
                    return validation_error({"error": "Field 'type' must be either 'question' or 'theory'"})
                
                if content_type == 'question':
                    question_id = request.data.get('question')
                    if not question_id:
                        return validation_error({"error": "Field 'question' is required when type is 'question'"})
                    
                    try:
                        question = Question.objects.get(id=question_id, is_deleted=False)
                    except Question.DoesNotExist:
                        return validation_error({"error": f"Question with ID {question_id} not found"})
                    
                    max_order = ModuleContent.objects.filter(
                        chapter=chapter,
                        is_deleted=False
                    ).aggregate(max_order=Max('order'))['max_order'] or 0
                    
                    module_content = ModuleContent.objects.create(
                        chapter=chapter,
                        content_type='question',
                        question=question,
                        order=max_order + 1,
                        created_by=request.user
                    )
                    
                    serializer = ModuleContentSerializer(module_content, context={'request': request})
                    api_logger.info(f"Module content created successfully: question (ID: {module_content.id}) for chapter '{chapter.title}'")
                    return created(
                        data=serializer.data,
                        message="Module content created successfully"
                    )
                
                elif content_type == 'theory':
                    theory_id = request.data.get('theory')
                    if not theory_id:
                        return validation_error({"error": "Field 'theory' is required when type is 'theory'"})
                    
                    try:
                        theory = Theory.objects.get(id=theory_id, is_deleted=False)
                    except Theory.DoesNotExist:
                        return validation_error({"error": f"Theory with ID {theory_id} not found"})
                    
                    max_order = ModuleContent.objects.filter(
                        chapter=chapter,
                        is_deleted=False
                    ).aggregate(max_order=Max('order'))['max_order'] or 0
                    
                    module_content = ModuleContent.objects.create(
                        chapter=chapter,
                        content_type='theory',
                        theory=theory,
                        order=max_order + 1,
                        created_by=request.user
                    )
                    
                    serializer = ModuleContentSerializer(module_content, context={'request': request})
                    api_logger.info(f"Module content created successfully: theory (ID: {module_content.id}) for chapter '{chapter.title}'")
                    return created(
                        data=serializer.data,
                        message="Module content created successfully"
                    )
                
        except Exception as e:
            api_logger.error(f"Error handling module content for chapter {pk}: {str(e)}")
            return validation_error({"error": f"Failed to handle module content: {str(e)}"})
    
    @action(detail=True, methods=['get'], url_path='hots_questions')
    def hots_questions(self, request, pk=None):
        """Get all HOTS questions for a specific chapter."""
        try:
            if True:
                return success(data=[], message="HOTS questions retrieved successfully")

            chapter = self.get_object()
            api_logger.info(f"HOTS questions requested for chapter '{chapter.title}' (ID: {chapter.id}) by {request.user.username}")

            chapter_hots = ChapterHOTS.objects.filter(
                chapter=chapter
            ).select_related(
                'question', 'created_by'
            ).prefetch_related(
                'question__options'
            ).order_by('order')
            
            questions = [chapter_hot.question for chapter_hot in chapter_hots if chapter_hot.question and not chapter_hot.question.is_deleted and chapter_hot.question.is_active][:3]

            questions_serializer = QuestionSerializer(questions, many=True, context={'request': request})

            api_logger.info(f"Returned {len(questions)} HOTS questions for chapter '{chapter.title}'")
            return success(
                data=questions_serializer.data,
                message=f"HOTS questions for chapter '{chapter.title}' retrieved successfully"
            )
            
        except Exception as e:
            api_logger.error(f"Error retrieving HOTS questions for chapter {pk}: {str(e)}")
            return validation_error({"error": f"Failed to retrieve HOTS questions: {str(e)}"})


class QuestionViewSet(viewsets.ModelViewSet):
    """ViewSet for Question model."""
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Return optimized queryset with related data, filtered by user's school."""
        user = self.request.user
        queryset = Question.objects.select_related(
            'created_by',
        ).prefetch_related(
            'options',
        ).filter(is_deleted=False)

        # check_answer must resolve any question the user was served, regardless of creator's school
        if self.action == 'check_answer':
            return queryset

        if hasattr(user, 'profile') and user.profile and user.profile.school:
            queryset = queryset.filter(created_by__profile__school=user.profile.school)
        else:
            queryset = queryset.none()

        if self.action == 'list':
            queryset = queryset.filter(is_active=True)

        return queryset
    
    def list(self, request, *args, **kwargs):
        """List questions with logging."""
        api_logger.info(f"Question list requested by {request.user.username} (ID: {request.user.id}) from {request.META.get('REMOTE_ADDR', 'unknown')}")
        
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        api_logger.info(f"Question list returned {len(serializer.data)} questions")
        
        return success(
            data=serializer.data,
            message="Questions retrieved successfully"
        )
    
    def create(self, request, *args, **kwargs):
        """Create a question with options handling."""
        api_logger.info(f"Question creation requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")
        
        question_data = request.data.copy()
        options_data = question_data.pop('options', None)
        chapter_id = question_data.pop('chapter_id', None)
        question_data.setdefault('difficulty_level', 'medium')

        if isinstance(chapter_id, list) and len(chapter_id) > 0:
            chapter_id = chapter_id[0]

        if options_data is not None:
            if isinstance(options_data, str):
                try:
                    options_data = json.loads(options_data)
                except (json.JSONDecodeError, ValueError) as e:
                    api_logger.warning(f"Failed to parse options JSON string: {e}")
                    options_data = None
            elif isinstance(options_data, list):
                parsed_options = []
                for item in options_data:
                    if isinstance(item, str):
                        try:
                            parsed = json.loads(item)
                            if isinstance(parsed, list):
                                parsed_options.extend(parsed)
                            else:
                                parsed_options.append(parsed)
                        except (json.JSONDecodeError, ValueError) as e:
                            api_logger.warning(f"Failed to parse option JSON string: {e}")
                            continue
                    elif isinstance(item, dict):
                        parsed_options.append(item)
                options_data = parsed_options if parsed_options else None
        
        level = question_data.get('level', 1)
        if isinstance(level, str):
            try:
                level = int(level)
            except ValueError:
                level = 1
        if level in [4, 5]:
            question_data['is_hots'] = True
        
        serializer = self.get_serializer(data=question_data)
        if serializer.is_valid():
            question = serializer.save()
            api_logger.info(f"Question created successfully: {question.question_text[:50]}... (ID: {question.id}) by {request.user.username}")
            
            is_hots = question.is_hots
            
            if is_hots and chapter_id:
                try:
                    max_hots_order = ChapterHOTS.objects.filter(chapter_id=chapter_id).aggregate(
                        max_order=Max('order')
                    )['max_order'] or 0
                    
                    chapter_hots = ChapterHOTS.objects.create(
                        chapter_id=chapter_id,
                        question=question,
                        order=max_hots_order + 1,
                        created_by=request.user
                    )
                    
                    ModuleChapter.objects.filter(id=chapter_id).update(has_hots=True)
                    
                    api_logger.info(f"ChapterHOTS entry created for question {question.id} in chapter {chapter_id}")
                except Exception as e:
                    api_logger.warning(f"Failed to create ChapterHOTS entry for question {question.id}: {str(e)}")
            
            elif chapter_id:
                try:
                    max_order = ModuleContent.objects.filter(chapter_id=chapter_id).aggregate(
                        max_order=Max('order')
                    )['max_order'] or 0
                    
                    module_content = ModuleContent.objects.create(
                        chapter_id=chapter_id,
                        content_type='question',
                        question=question,
                        order=max_order + 1,
                        created_by=request.user
                    )
                    api_logger.info(f"ModuleContent entry created for question {question.id} in chapter {chapter_id}")
                except Exception as e:
                    api_logger.warning(f"Failed to create ModuleContent entry for question {question.id}: {str(e)}")
            
            if options_data is not None and question.question_type in ['mcq_single', 'mcq_multiple', 'rearrange']:
                created_options = []
                errors = []

                for idx, option_data in enumerate(options_data):
                    option_text = option_data.get('option_text')
                    if not option_text:
                        errors.append(f"Option {idx + 1}: 'option_text' is required")
                        continue

                    is_correct = option_data.get('is_correct', False)
                    order = option_data.get('order', idx + 1)

                    try:
                        option = Option.objects.create(
                            question=question,
                            option_text=option_text,
                            is_correct=is_correct,
                            order=order
                        )
                        created_options.append(option)
                    except Exception as e:
                        errors.append(f"Option {idx + 1}: {str(e)}")

                if errors:
                    api_logger.warning(f"Options creation completed with errors for question {question.id}: {errors}")

                correct_count = sum(1 for opt in created_options if opt.is_correct)
                warning_message = None

                if question.question_type == 'mcq_single':
                    if correct_count == 0:
                        warning_message = "Warning: MCQ Single choice questions should have exactly 1 correct answer. No correct answers found."
                    elif correct_count > 1:
                        warning_message = f"Warning: MCQ Single choice questions should have exactly 1 correct answer. Found {correct_count} correct answers."
                elif question.question_type == 'mcq_multiple':
                    if correct_count == 0:
                        warning_message = "Warning: MCQ Multiple choice questions should have at least 1 correct answer. No correct answers found."

                api_logger.info(f"Successfully created {len(created_options)} options for question {question.id}")

                if warning_message:
                    api_logger.warning(f"Options validation warning for question {question.id}: {warning_message}")

            updated_serializer = self.get_serializer(question)
            response_data = updated_serializer.data

            if options_data is not None and question.question_type in ['mcq_single', 'mcq_multiple', 'rearrange']:
                correct_count = sum(1 for opt in question.options.all() if opt.is_correct)
                if question.question_type == 'mcq_single' and correct_count != 1:
                    response_data['warning'] = f"Warning: MCQ Single choice questions should have exactly 1 correct answer. Found {correct_count} correct answers."
                elif question.question_type == 'mcq_multiple' and correct_count == 0:
                    response_data['warning'] = "Warning: MCQ Multiple choice questions should have at least 1 correct answer. No correct answers found."
                if errors:
                    response_data['option_errors'] = errors

            return created(
                data=response_data,
                message="Question created successfully"
            )
        api_logger.warning(f"Question creation failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)
    
    def update(self, request, *args, **kwargs):
        """Update a question with options handling."""
        try:
            instance = self.get_object()
        except Exception as e:
            api_logger.error(f"Question update get_object failed: {str(e)}", exc_info=True)
            raise
        api_logger.info(f"Question update requested by {request.user.username} (ID: {request.user.id}) for question {instance.id} - Data: {request.data}")

        question_data = request.data.copy()
        options_data = question_data.pop('options', None)
        
        if options_data is not None:
            if isinstance(options_data, str):
                try:
                    options_data = json.loads(options_data)
                except (json.JSONDecodeError, ValueError) as e:
                    api_logger.warning(f"Failed to parse options JSON string: {e}")
                    options_data = None
            elif isinstance(options_data, list):
                parsed_options = []
                for item in options_data:
                    if isinstance(item, str):
                        try:
                            parsed = json.loads(item)
                            if isinstance(parsed, list):
                                parsed_options.extend(parsed)
                            else:
                                parsed_options.append(parsed)
                        except (json.JSONDecodeError, ValueError) as e:
                            api_logger.warning(f"Failed to parse option JSON string: {e}")
                            continue
                    elif isinstance(item, dict):
                        parsed_options.append(item)
                options_data = parsed_options if parsed_options else None
        
        serializer = self.get_serializer(instance, data=question_data, partial=kwargs.get('partial', False))
        if serializer.is_valid():
            try:
                question = serializer.save()
            except Exception as e:
                api_logger.error(f"Question update serializer.save() failed for question {instance.id}: {str(e)}", exc_info=True)
                from rest_framework.response import Response
                from rest_framework import status
                return Response({'error': f'Failed to save question: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            api_logger.info(f"Question updated successfully: {question.question_text[:50]}... (ID: {question.id})")
            
            if options_data is not None:
                if question.question_type in ['mcq_single', 'mcq_multiple', 'rearrange']:
                    existing_options = Option.objects.filter(question=question)
                    existing_count = existing_options.count()
                    existing_options.delete()
                    api_logger.info(f"Deleted {existing_count} existing options for question {question.id}")

                    created_options = []
                    errors = []

                    for idx, option_data in enumerate(options_data):
                        option_text = option_data.get('option_text')
                        if not option_text:
                            errors.append(f"Option {idx + 1}: 'option_text' is required")
                            continue

                        is_correct = option_data.get('is_correct', False)
                        order = option_data.get('order', idx + 1)

                        try:
                            option = Option.objects.create(
                                question=question,
                                option_text=option_text,
                                is_correct=is_correct,
                                order=order
                            )
                            created_options.append(option)
                        except Exception as e:
                            errors.append(f"Option {idx + 1}: {str(e)}")

                    if errors:
                        api_logger.warning(f"Options update completed with errors for question {question.id}: {errors}")

                    correct_count = sum(1 for opt in created_options if opt.is_correct)
                    warning_message = None

                    if question.question_type == 'mcq_single':
                        if correct_count == 0:
                            warning_message = "Warning: MCQ Single choice questions should have exactly 1 correct answer. No correct answers found."
                        elif correct_count > 1:
                            warning_message = f"Warning: MCQ Single choice questions should have exactly 1 correct answer. Found {correct_count} correct answers."
                    elif question.question_type == 'mcq_multiple':
                        if correct_count == 0:
                            warning_message = "Warning: MCQ Multiple choice questions should have at least 1 correct answer. No correct answers found."

                    api_logger.info(f"Successfully created {len(created_options)} options for question {question.id}")

                    if warning_message:
                        api_logger.warning(f"Options validation warning for question {question.id}: {warning_message}")
                else:
                    existing_options = Option.objects.filter(question=question)
                    if existing_options.exists():
                        existing_count = existing_options.count()
                        existing_options.delete()
                        api_logger.info(f"Deleted {existing_count} existing options for question {question.id} (question type doesn't require options)")

            updated_serializer = self.get_serializer(question)
            response_data = updated_serializer.data

            if options_data is not None and question.question_type in ['mcq_single', 'mcq_multiple', 'rearrange']:
                correct_count = sum(1 for opt in question.options.all() if opt.is_correct)
                if question.question_type == 'mcq_single' and correct_count != 1:
                    response_data['warning'] = f"Warning: MCQ Single choice questions should have exactly 1 correct answer. Found {correct_count} correct answers."
                elif question.question_type == 'mcq_multiple' and correct_count == 0:
                    response_data['warning'] = "Warning: MCQ Multiple choice questions should have at least 1 correct answer. No correct answers found."
                if errors:
                    response_data['option_errors'] = errors

            return success(
                data=response_data,
                message="Question updated successfully"
            )
        
        api_logger.warning(f"Question update failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)

    def destroy(self, request, *args, **kwargs):
        """Soft delete a question (set is_deleted=True)."""
        question_id = kwargs.get('pk')
        api_logger.info(f"Question deletion requested by {request.user.username} (ID: {request.user.id}) for question ID: {question_id}")
        try:
            instance = Question.objects.get(id=question_id)
            instance.is_deleted = True
            instance.is_active = False
            instance.save(update_fields=['is_deleted', 'is_active'])
            api_logger.info(f"Question soft deleted successfully: {instance.question_text[:50]}... (ID: {instance.id}) by {request.user.username}")
            return success(data={}, message="Question deleted successfully")
        except Question.DoesNotExist:
            return validation_error({"error": "Question not found."})
        except Exception as e:
            api_logger.error(f"Error deleting question {question_id}: {str(e)}")
            return validation_error({"error": f"Failed to delete question: {str(e)}"})

    @action(detail=False, methods=['post'], url_path='ai/generate')
    def ai_generate(self, request):
        """Generate AI questions."""
        api_logger.info(f"AI question generation requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")
        
        subject = request.data.get('subject', 'Mathematics')
        grade = request.data.get('grade', '10')
        difficulty = request.data.get('difficulty', 'Medium')
        question_type = request.data.get('type', 'Multiple Choice')
        count = request.data.get('count', 5)
        topics = request.data.get('topics', [])
        
        generated_questions = []
        
        subject_obj = Subject.objects.filter(name__icontains=subject, is_active=True).first()
        modules = Module.objects.filter(subject=subject_obj, is_active=True) if subject_obj else []
        
        for i in range(count):
            module = modules[i % len(modules)] if modules else None
            question_text = f"What is the main concept covered in {module.name if module else subject}?"
            
            question_data = {
                'question_text': question_text,
                'english_text': f"(AI Generated Question {i+1} for {subject} Grade {grade})",
                'question_type': question_type.lower().replace(' ', '_'),
                'difficulty_level': difficulty.lower(),
                'explanation': f"This question tests understanding of {module.name if module else subject} concepts for Grade {grade} students.",
                'exp_points': 10,
                'subject_id': subject_obj.id if subject_obj else None,
                'module_id': module.id if module else None,
                'options': [
                    {'option_text': 'Option A - Correct Answer', 'is_correct': True, 'order': 1},
                    {'option_text': 'Option B - Incorrect', 'is_correct': False, 'order': 2},
                    {'option_text': 'Option C - Incorrect', 'is_correct': False, 'order': 3},
                    {'option_text': 'Option D - Incorrect', 'is_correct': False, 'order': 4},
                ]
            }
            generated_questions.append(question_data)
        
        api_logger.info(f"AI generated {len(generated_questions)} questions for {request.user.username}")
        return success(
            data={
                'questions': generated_questions,
                'metadata': {
                    'subject': subject,
                    'grade': grade,
                    'difficulty': difficulty,
                    'type': question_type,
                    'count': count,
                    'topics': topics,
                    'generated_at': timezone.now().isoformat()
                }
            },
            message=f"Successfully generated {count} AI questions"
        )
    
    @action(detail=True, methods=['post'], url_path='options')
    def create_options(self, request, pk=None):
        """Create options for a specific question."""
        try:
            question = self.get_object()
            api_logger.info(f"Options creation requested for question (ID: {question.id}) by {request.user.username} - Data: {request.data}")
            
            options_data = request.data.get('options', [])
            if not isinstance(options_data, list):
                return validation_error({"error": "Field 'options' must be a list"})
            
            if len(options_data) == 0:
                return validation_error({"error": "At least one option is required"})
            
            if question.question_type not in ['mcq_single', 'mcq_multiple']:
                return validation_error({"error": f"Options can only be added to MCQ questions. This question is of type: {question.question_type}"})
            
            created_options = []
            errors = []
            
            for idx, option_data in enumerate(options_data):
                option_text = option_data.get('option_text')
                if not option_text:
                    errors.append(f"Option {idx + 1}: 'option_text' is required")
                    continue
                
                is_correct = option_data.get('is_correct', False)
                order = option_data.get('order', idx + 1)
                
                if Option.objects.filter(question=question, option_text=option_text).exists():
                    errors.append(f"Option {idx + 1}: Duplicate option text '{option_text}'")
                    continue
                
                try:
                    option = Option.objects.create(
                        question=question,
                        option_text=option_text,
                        is_correct=is_correct,
                        order=order
                    )
                    created_options.append(option)
                except Exception as e:
                    errors.append(f"Option {idx + 1}: {str(e)}")
            
            if errors:
                if created_options:
                    serializer = OptionSerializer(created_options, many=True)
                    api_logger.warning(f"Options creation completed with errors: {errors}")
                    return validation_error({
                        "errors": errors,
                        "created_options": serializer.data,
                        "message": "Some options were created, but there were errors"
                    })
                else:
                    api_logger.warning(f"Options creation failed - Errors: {errors}")
                    return validation_error({"errors": errors})
            
            new_correct_count = sum(1 for opt in created_options if opt.is_correct)
            
            existing_options = Option.objects.filter(question=question).exclude(id__in=[opt.id for opt in created_options])
            existing_correct_count = sum(1 for opt in existing_options if opt.is_correct)
            
            total_correct_count = existing_correct_count + new_correct_count
            
            warning_message = None
            if question.question_type == 'mcq_single':
                if total_correct_count == 0:
                    warning_message = "Warning: MCQ Single choice questions should have exactly 1 correct answer. No correct answers found. You can mark an option as correct later."
                elif total_correct_count > 1:
                    warning_message = f"Warning: MCQ Single choice questions should have exactly 1 correct answer. Found {total_correct_count} correct answers."
            elif question.question_type == 'mcq_multiple':
                if total_correct_count == 0:
                    warning_message = "Warning: MCQ Multiple choice questions should have at least 1 correct answer. No correct answers found. You can mark options as correct later."
            
            serializer = OptionSerializer(created_options, many=True)
            api_logger.info(f"Successfully created {len(created_options)} options for question (ID: {question.id})")
            
            response_data = serializer.data
            message = f"Successfully created {len(created_options)} options"
            
            if warning_message:
                response_data = {
                    "options": serializer.data,
                    "warning": warning_message
                }
                api_logger.warning(f"Options created with warning for question {question.id}: {warning_message}")
            
            return created(data=response_data, message=message)
            
        except Exception as e:
            api_logger.error(f"Error creating options for question {pk}: {str(e)}")
            return validation_error({"error": f"Failed to create options: {str(e)}"})
    
    @action(detail=False, methods=['get'], url_path='ai/generated')
    def ai_generated(self, request):
        """Get AI generated questions."""
        api_logger.info(f"AI generated questions list requested by {request.user.username} (ID: {request.user.id})")
        
        queryset = Question.objects.filter(
            created_by=request.user,
            is_active=True,
            is_deleted=False
        ).select_related('created_by').prefetch_related('options')
        
        subject = request.query_params.get('subject', None)
        if subject:
            pass
        
        difficulty = request.query_params.get('difficulty', None)
        if difficulty:
            queryset = queryset.filter(difficulty_level=difficulty.lower())
        
        question_type = request.query_params.get('type', None)
        if question_type:
            queryset = queryset.filter(question_type=question_type.lower().replace(' ', '_'))
        
        serializer = QuestionSerializer(queryset, many=True)
        api_logger.info(f"AI generated questions list returned {len(serializer.data)} questions")
        
        return success(
            data=serializer.data,
            message="AI generated questions retrieved successfully"
        )
    
    @action(detail=False, methods=['post'], url_path='ai/save')
    def ai_save(self, request):
        """Save AI generated questions to question bank."""
        api_logger.info(f"AI question save requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")
        
        questions_data = request.data.get('questions', [])
        saved_questions = []
        
        for question_data in questions_data:
            question = Question.objects.create(
                question_text=question_data.get('question_text', ''),
                question_type=question_data.get('question_type', 'mcq_single'),
                difficulty_level=question_data.get('difficulty_level', 'medium'),
                explanation=question_data.get('explanation', ''),
                exp_points=question_data.get('exp_points', 10),
                created_by=request.user
            )
            
            options_data = question_data.get('options', [])
            for option_data in options_data:
                Option.objects.create(
                    question=question,
                    option_text=option_data.get('option_text', ''),
                    is_correct=option_data.get('is_correct', False),
                    order=option_data.get('order', 1)
                )
            
            saved_questions.append(QuestionSerializer(question).data)
        
        api_logger.info(f"AI saved {len(saved_questions)} questions for {request.user.username}")
        return success(
            data=saved_questions,
            message=f"Successfully saved {len(saved_questions)} AI questions"
        )

    @action(detail=True, methods=['patch'], url_path='check')
    def check_answer(self, request, *args, **kwargs):
        """Check answer with logging."""
        question = self.get_object()
        api_logger.info(f"Answer check requested by {request.user.username} (ID: {request.user.id}) for question {question.id} - Data: {request.data}")
        
        # Resolve the user's class so all ModuleContent lookups use the correct
        # class-specific module. Without this, with multiple classes per subject,
        # filter(...).first() returns an arbitrary class's chapter.
        _user_class = None
        try:
            _user_class = request.user.profile.student.class_instance
        except Exception:
            pass

        def _get_module_content(q):
            qs = ModuleContent.objects.filter(question=q, is_deleted=False)
            if _user_class:
                qs = qs.filter(chapter__module__class_instance=_user_class)
            return qs.first()

        answer_id = request.data.get('answer_id')
        answer_text = request.data.get('answer_text')
        answer_ids = request.data.get('answer_ids')
        tries = request.data.get('tries', 1)
        is_last = request.data.get('is_last', False)
        
        tries = max(1, int(tries)) if tries else 1
        
        api_logger.info(f"Answer check - Question ID: {question.id}, Answer ID: {answer_id}, Answer Text: {answer_text}, Answer IDs: {answer_ids}, Tries: {tries}, Is Last: {is_last}")
        
        prev_exp = 0
        if hasattr(request.user, 'profile') and request.user.profile:
            prev_exp = request.user.profile.total_exp

        already_correct = Answer.objects.filter(
            user=request.user, question=question, is_correct=True
        ).exists()

        answer_text_to_store = ""
        is_correct = False
        
        if answer_text:
            answer_text_to_store = answer_text
        elif answer_ids:
            if isinstance(answer_ids, list):
                option_texts = []
                for ans_id in answer_ids:
                    try:
                        option = Option.objects.get(id=ans_id, question=question)
                        option_texts.append(option.option_text)
                    except Option.DoesNotExist:
                        api_logger.warning(f"Option with ID {ans_id} not found for question {question.id}")
                answer_text_to_store = ", ".join(option_texts) if option_texts else ",".join([str(a) for a in answer_ids])
            else:
                try:
                    option = Option.objects.get(id=answer_ids, question=question)
                    answer_text_to_store = option.option_text
                except Option.DoesNotExist:
                    api_logger.warning(f"Option with ID {answer_ids} not found for question {question.id}")
        elif answer_id:
            try:
                option = Option.objects.get(id=answer_id, question=question)
                answer_text_to_store = option.option_text
            except Option.DoesNotExist:
                api_logger.warning(f"Option with ID {answer_id} not found for question {question.id}")
        
        if question.question_type in ['mcq_single', 'mcq_multiple']:
            if answer_ids:
                if isinstance(answer_ids, list):
                    selected_options = Option.objects.filter(id__in=answer_ids, question=question)
                    correct_options = question.options.filter(is_correct=True)
                    
                    selected_ids = set(selected_options.values_list('id', flat=True))
                    correct_ids = set(correct_options.values_list('id', flat=True))
                    is_correct = selected_ids == correct_ids and len(selected_ids) > 0
                else:
                    try:
                        selected_option = Option.objects.get(id=answer_ids, question=question)
                        is_correct = selected_option.is_correct
                    except Option.DoesNotExist:
                        is_correct = False
            elif answer_id:
                try:
                    selected_option = Option.objects.get(id=answer_id, question=question)
                    is_correct = selected_option.is_correct
                except Option.DoesNotExist:
                    is_correct = False
        elif question.question_type == 'short_answer':
            is_correct = True
            
            if answer_text_to_store:
                ManualVerificationAnswer.objects.create(
                    user=request.user,
                    question=question,
                    answer=answer_text_to_store
                )
                api_logger.info(f"Short answer saved to ManualVerificationAnswer for user {request.user.username} and question {question.id}")
        
        if not already_correct and is_correct and tries == 1:
            base_exp = 2
        elif not already_correct and is_correct and tries == 2:
            base_exp = 1
        else:
            base_exp = 0
        exp_multiplier = 1.0
        final_exp = base_exp

        if final_exp > 0:
            if hasattr(request.user, 'profile'):
                request.user.profile.add_exp(final_exp)
                api_logger.info(f"User {request.user.username} earned {final_exp} exp (tries: {tries})")
            else:
                api_logger.warning(f"User {request.user.username} has no profile, cannot add exp")
        else:
            api_logger.info(f"User {request.user.username} earned 0 exp (correct: {is_correct}, tries: {tries}, already_correct: {already_correct})")
        
        current_Exp = prev_exp
        if hasattr(request.user, 'profile') and request.user.profile:
            current_Exp = request.user.profile.total_exp
        
        _chapter_for_answer = None
        _module_content = _get_module_content(question)
        if _module_content:
            _chapter_for_answer = _module_content.chapter

        # Answer has PARTIAL unique constraints: one practice row per
        # (user, question) WHERE test IS NULL, plus one row per (user, question,
        # test) for test answers. So a user can legitimately have a practice
        # answer AND test answer(s) for the same question. This is the PRACTICE
        # check endpoint, so it must scope the lookup to the practice row
        # (test IS NULL). The old get_or_create(user, question) ignored `test`
        # and matched multiple rows → MultipleObjectsReturned (HTTP 500), which
        # silently broke the whole quiz for any student who had also answered the
        # question inside a test. Never touch test answers here.
        answer = (
            Answer.objects.filter(user=request.user, question=question, test__isnull=True)
            .order_by('created_at')
            .first()
        )
        created = answer is None
        if created:
            answer = Answer(user=request.user, question=question)  # test stays NULL

        answer.answer = answer_text_to_store
        answer.is_correct = is_correct
        answer.tries = tries
        answer.prev_exp = prev_exp
        answer.current_Exp = current_Exp
        if created:
            answer.chapter = _chapter_for_answer
        elif answer.chapter is None and _chapter_for_answer is not None:
            answer.chapter = _chapter_for_answer
        answer.save()
        api_logger.info(
            f"{'Created' if created else 'Updated'} practice answer for user {request.user.username} and question {question.id}"
        )
        
        mission_info = None
        if not is_correct:
            module_content_for_mission = _get_module_content(question)
            if module_content_for_mission:
                chapter_for_mission = module_content_for_mission.chapter
                mission_info = add_wrong_question_to_mission(request.user, question, chapter_for_mission)
                if mission_info:
                    api_logger.info(f"Mission updated for user {request.user.username} - Mission ID: {mission_info['mission_id']}, Questions added: {mission_info['questions_added']}")
                    
                    if mission_info.get('mission_created', False):
                        try:
                            from gyaan_buddy.utils.firebase_notifications import firebase_notification_service
                            
                            title = "New Mission Available!"
                            body = f"You have a new mission for {mission_info['subject_name']}. Complete it to strengthen your knowledge!"
                            
                            data = {
                                'type': 'mission_created',
                                'mission_id': mission_info['mission_id'],
                                'subject_id': mission_info['subject_id'],
                                'subject_name': mission_info['subject_name'],
                                'action': 'open_mission'
                            }
                            
                            firebase_notification_service.send_notification_to_user(
                                user=request.user,
                                title=title,
                                body=body,
                                data=data,
                                notification_type='mission',
                                triggered_by='auto'
                            )
                            api_logger.info(f"Mission created notification sent to user {request.user.username} for mission {mission_info['mission_id']}")
                        except Exception as notif_error:
                            api_logger.error(f"Failed to send mission created notification to user {request.user.username}: {str(notif_error)}")
                else:
                    api_logger.warning(f"Failed to add question to mission for user {request.user.username}")
        
        chapter_progress_updated = False
        module_progress_updated = False
        
        if is_last:
            module_content = _get_module_content(question)
            if module_content:
                chapter = module_content.chapter
                
                try:
                    from .helpers import update_chapter_progress
                    user_chapter_progress, created = update_chapter_progress(
                        request.user, 
                        chapter, 
                        percentage=100, 
                        status='completed'
                    )
                    user_chapter_progress.current_question = question
                    user_chapter_progress.save(update_fields=['current_question', 'status', 'percentage'])
                    chapter_progress_updated = True
                    api_logger.info(f"Chapter '{chapter.title}' marked as completed for user {request.user.username} - Percentage: 100%, Status: {user_chapter_progress.status}")
                except Exception as e:
                    api_logger.error(f"Error updating chapter progress for user {request.user.username}: {str(e)}")
                    chapter_progress_updated = False
                
                if chapter_progress_updated:
                    module = chapter.module
                    try:
                        from .helpers import update_module_progress
                        user_module_progress, module_created, module_percentage = update_module_progress(request.user, module)
                        user_module_progress.current_question = question
                        user_module_progress.save(update_fields=['current_question'])
                        module_progress_updated = True
                        api_logger.info(f"Module '{module.name}' progress updated for user {request.user.username} - Status: {user_module_progress.status}, Percentage: {user_module_progress.percentage}%")
                    except Exception as e:
                        api_logger.error(f"Error updating module progress for user {request.user.username}: {str(e)}")
                        module_progress_updated = False

                    # Pre-advance next chapter to in_progress so GET module_chapters
                    # always returns the correct state even if called before this response returns.
                    try:
                        next_chapter = ModuleChapter.objects.filter(
                            module=chapter.module,
                            is_enabled=True,
                            is_deleted=False,
                            order__gt=chapter.order
                        ).order_by('order').first()
                        if next_chapter:
                            next_progress, next_created = UserChapterProgress.objects.get_or_create(
                                account=request.user,
                                chapter=next_chapter,
                                defaults={'status': 'in_progress'}
                            )
                            if not next_created and next_progress.status != 'in_progress' and next_progress.status != 'completed':
                                next_progress.status = 'in_progress'
                                next_progress.save(update_fields=['status'])
                            api_logger.info(f"Next chapter '{next_chapter.title}' set to in_progress for user {request.user.username} (created={next_created})")
                    except Exception as e:
                        api_logger.error(f"Error pre-advancing next chapter for user {request.user.username}: {str(e)}")
        else:
            module_content = _get_module_content(question)
            if module_content:
                chapter = module_content.chapter
                module = chapter.module
                
                try:
                    from .helpers import calculate_chapter_progress_percentage, update_chapter_progress
                    current_percentage = calculate_chapter_progress_percentage(chapter, module_content)
                    
                    user_chapter_progress, created = update_chapter_progress(
                        request.user, 
                        chapter, 
                        percentage=current_percentage, 
                        status='in_progress'
                    )
                    user_chapter_progress.current_question = question
                    user_chapter_progress.save(update_fields=['current_question', 'status', 'percentage'])
                    chapter_progress_updated = True
                    api_logger.info(f"Chapter '{chapter.title}' progress updated for user {request.user.username} - Percentage: {current_percentage}%, Status: {user_chapter_progress.status}")
                except Exception as e:
                    api_logger.error(f"Error updating chapter progress for user {request.user.username}: {str(e)}")
                    chapter_progress_updated = False
                
                if chapter_progress_updated:
                    try:
                        from .helpers import update_module_progress
                        user_module_progress, module_created, module_percentage = update_module_progress(request.user, module)
                        user_module_progress.current_question = question
                        user_module_progress.save(update_fields=['current_question'])
                        module_progress_updated = True
                        api_logger.info(f"Module '{module.name}' progress updated for user {request.user.username} - Status: {user_module_progress.status}, Percentage: {user_module_progress.percentage}%")
                    except Exception as e:
                        api_logger.error(f"Error updating module progress for user {request.user.username}: {str(e)}")
                        module_progress_updated = False
        
        api_logger.info(f"Answer check completed for user {request.user.username} - Question: {question.id}, Answer ID: {answer_id}, Tries: {tries}, Exp earned: {final_exp}, Chapter progress updated: {chapter_progress_updated}, Module progress updated: {module_progress_updated}")
        
        chapter_progress_info = {}
        module_progress_info = {}
        
        if chapter_progress_updated:
            try:
                module_content = _get_module_content(question)
                if module_content:
                    chapter = module_content.chapter
                    user_chapter_progress = UserChapterProgress.objects.filter(
                        account=request.user, 
                        chapter=chapter
                    ).first()
                    if user_chapter_progress:
                        chapter_progress_info = {
                            'chapter_id': chapter.id,
                            'chapter_title': chapter.title,
                            'status': user_chapter_progress.status,
                            'percentage': user_chapter_progress.percentage
                        }
            except Exception as e:
                api_logger.error(f"Error getting chapter progress info: {str(e)}")
        
        if module_progress_updated:
            try:
                module_content = _get_module_content(question)
                if module_content:
                    module = module_content.chapter.module
                    user_module_progress = UserModuleProgress.objects.filter(
                        account=request.user, 
                        module=module
                    ).first()
                    if user_module_progress:
                        module_progress_info = {
                            'module_id': module.id,
                            'module_name': module.name,
                            'status': user_module_progress.status,
                            'percentage': user_module_progress.percentage
                        }
            except Exception as e:
                api_logger.error(f"Error getting module progress info: {str(e)}")
        
        return success(
            data={
                'question_id': question.id,
                'answer_id': answer_id,
                'answer_text': answer_text_to_store,
                'is_correct': is_correct,
                'tries': tries,
                'base_exp': base_exp,
                'exp_multiplier': round(exp_multiplier, 2),
                'exp_earned': final_exp,
                'prev_exp': prev_exp,
                'current_Exp': current_Exp,
                'total_user_exp': request.user.profile.total_exp if hasattr(request.user, 'profile') and request.user.profile else 0,
                'is_last': is_last,
                'chapter_progress_updated': chapter_progress_updated,
                'module_progress_updated': module_progress_updated,
                'chapter_progress': chapter_progress_info,
                'module_progress': module_progress_info,
                'mission_info': mission_info,
                'answer_saved': True,
                'message': f'Answer check completed. Earned {final_exp} experience points.'
            },
            message="Answer check completed successfully"
        )


class OptionViewSet(viewsets.ModelViewSet):
    """ViewSet for Option model."""
    queryset = Option.objects.all()
    serializer_class = OptionSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Return optimized queryset with related data, filtered by user's school."""
        user = self.request.user
        queryset = Option.objects.select_related(
            'question',
        ).filter(question__is_active=True, question__is_deleted=False)
        
        if hasattr(user, 'profile') and user.profile and user.profile.school:
            queryset = queryset.filter(question__created_by__profile__school=user.profile.school).distinct()
        
        return queryset
    
    def list(self, request, *args, **kwargs):
        """List options with logging."""
        api_logger.info(f"Option list requested by {request.user.username} (ID: {request.user.id}) from {request.META.get('REMOTE_ADDR', 'unknown')}")
        
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        api_logger.info(f"Option list returned {len(serializer.data)} options")
        
        return success(
            data=serializer.data,
            message="Options retrieved successfully"
        )
    
    def create(self, request, *args, **kwargs):
        """Create an option with logging."""
        api_logger.info(f"Option creation requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")
        
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            option = serializer.save()
            api_logger.info(f"Option created successfully: {option.option_text[:50]}... (ID: {option.id}) by {request.user.username}")
            return created(
                data=serializer.data,
                message="Option created successfully"
            )
        api_logger.warning(f"Option creation failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)


class TheoryViewSet(viewsets.ModelViewSet):
    """ViewSet for Theory model."""
    queryset = Theory.objects.all()
    serializer_class = TheorySerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Return optimized queryset with related data, filtered by user's school."""
        user = self.request.user
        queryset = Theory.objects.select_related(
            'created_by',
        ).filter(is_deleted=False)
        
        if hasattr(user, 'profile') and user.profile and user.profile.school:
            queryset = queryset.filter(created_by__profile__school=user.profile.school).distinct()
        
        return queryset
    
    def list(self, request, *args, **kwargs):
        """List theories with logging."""
        api_logger.info(f"Theory list requested by {request.user.username} (ID: {request.user.id}) from {request.META.get('REMOTE_ADDR', 'unknown')}")
        
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        api_logger.info(f"Theory list returned {len(serializer.data)} theories")
        
        return success(
            data=serializer.data,
            message="Theories retrieved successfully"
        )
    
    def create(self, request, *args, **kwargs):
        """Create a theory with logging."""
        api_logger.info(f"Theory creation requested by {request.user.username} (ID: {request.user.id}) - Data: {request.data}")
        
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            theory = serializer.save()
            api_logger.info(f"Theory created successfully: {theory.title} (ID: {theory.id}) by {request.user.username}")
            return created(
                data=serializer.data,
                message="Theory created successfully"
            )
        api_logger.warning(f"Theory creation failed - Errors: {serializer.errors}")
        return validation_error(serializer.errors)


# ── Assessment Generator Views ─────────────────────────────────────────────────

import hashlib
import re
import httpx
from django.conf import settings as django_settings
from rest_framework.views import APIView
from gyaan_buddy.users.views import TeacherAdminPermission
from gyaan_buddy.users.models import Class
from .models import PdfReference, AssessmentSession, QuestionModification
from .serializers import PdfReferenceSerializer

_AI_SERVICE_URL = getattr(django_settings, 'AI_SERVICE_URL', 'http://localhost:8001')
_AI_SERVICE_TIMEOUT = 240  # seconds — generation can take time


def _get_gcs_client():
    from google.cloud import storage
    return storage.Client()


class PdfUploadView(APIView):
    """POST /api/chapter-pdf/upload/ — Upload PDF to GCS and create PdfReference."""
    parser_classes = [parsers.MultiPartParser]
    permission_classes = [permissions.IsAuthenticated, TeacherAdminPermission]

    def post(self, request):
        chapter_id = request.data.get('chapter_id')
        pdf_file = request.FILES.get('file')

        if not chapter_id or not pdf_file:
            return validation_error({'detail': 'chapter_id and file are required.'})

        try:
            chapter = Module.objects.select_related('subject', 'class_instance').get(pk=chapter_id)
        except Module.DoesNotExist:
            return Response({'detail': 'Chapter not found.'}, status=status.HTTP_404_NOT_FOUND)

        # SHA-256 dedup
        pdf_bytes = pdf_file.read()
        file_hash = hashlib.sha256(pdf_bytes).hexdigest()
        existing = PdfReference.objects.filter(chapter=chapter, file_hash=file_hash, is_active=True).first()
        if existing:
            if existing.embedding_status == 'FAILED':
                # Previous upload failed at embedding — re-trigger without re-uploading to GCS
                existing.embedding_status = 'PENDING'
                existing.save(update_fields=['embedding_status'])
                try:
                    from .tasks import process_pdf
                    process_pdf.delay(str(existing.id))
                except Exception as exc:
                    logger.warning(f"Failed to re-enqueue process_pdf for {existing.id}: {exc}")
                return success(
                    data=PdfReferenceSerializer(existing).data,
                    message='Previous embedding failed — re-queued for processing.',
                )
            return success(
                data=PdfReferenceSerializer(existing).data,
                message='PDF already exists (duplicate detected by file hash).',
            )

        # Build GCS path
        class_id = str(chapter.class_instance_id) if chapter.class_instance_id else 'unknown'
        subject_id = str(chapter.subject_id)
        chapter_id_str = str(chapter.id)
        safe_name = pdf_file.name.replace(' ', '_')
        gcs_path = f"pdfs/{class_id}/{subject_id}/{chapter_id_str}/{safe_name}"

        # Upload to GCS
        try:
            bucket_name = django_settings.GS_BUCKET_NAME
            gcs = _get_gcs_client()
            bucket = gcs.bucket(bucket_name)
            blob = bucket.blob(gcs_path)
            blob.upload_from_string(pdf_bytes, content_type='application/pdf')
        except Exception as exc:
            logger.error(f"GCS upload failed for chapter {chapter_id}: {exc}")
            return Response({'detail': 'GCS upload failed.'}, status=status.HTTP_502_BAD_GATEWAY)

        pdf_ref = PdfReference.objects.create(
            chapter=chapter,
            file_name=safe_name,
            gcs_path=f"gs://{bucket_name}/{gcs_path}",
            file_hash=file_hash,
            embedding_status='PENDING',
        )

        # Trigger async embedding pipeline
        try:
            from .tasks import process_pdf
            process_pdf.delay(str(pdf_ref.id))
        except Exception as exc:
            logger.warning(f"Failed to enqueue process_pdf task for {pdf_ref.id}: {exc}")

        return created(
            data=PdfReferenceSerializer(pdf_ref).data,
            message='PDF uploaded successfully. Embedding in progress.',
        )


class PdfListView(APIView):
    """GET /api/chapter-pdf/?chapter_id=<uuid> — List all PDFs, grouped by chapter.
    If chapter_id is provided, returns only PDFs for that chapter."""
    permission_classes = [permissions.IsAuthenticated, TeacherAdminPermission]

    def get(self, request):
        chapter_id = request.query_params.get('chapter_id')
        qs = PdfReference.objects.filter(is_active=True).select_related('chapter').order_by('chapter_id', '-created_at')
        if chapter_id:
            qs = qs.filter(chapter_id=chapter_id)

        grouped = {}
        for pdf in qs:
            cid = str(pdf.chapter_id)
            if cid not in grouped:
                grouped[cid] = {
                    'chapter_id': cid,
                    'chapter_name': pdf.chapter.name if pdf.chapter else '',
                    'pdfs': [],
                }
            grouped[cid]['pdfs'].append(PdfReferenceSerializer(pdf).data)

        return success(data=list(grouped.values()))


class PdfDeleteView(APIView):
    """DELETE /api/chapter-pdf/<pdf_id>/ — Soft delete: sets is_active=False and deactivates Qdrant vectors."""
    permission_classes = [permissions.IsAuthenticated, TeacherAdminPermission]

    def delete(self, request, pdf_id):
        from .tasks import delete_pdf_embeddings
        try:
            pdf_ref = PdfReference.objects.get(pk=pdf_id, is_active=True)
        except PdfReference.DoesNotExist:
            return Response({'detail': 'PDF not found.'}, status=status.HTTP_404_NOT_FOUND)

        if pdf_ref.is_default:
            return Response({'detail': 'Cannot delete the default PDF.'}, status=status.HTTP_400_BAD_REQUEST)

        pdf_ref.is_active = False
        pdf_ref.save(update_fields=['is_active'])

        # Deactivate Qdrant vectors so they are excluded from context retrieval
        delete_pdf_embeddings.delay(str(pdf_ref.id))

        return success(message='PDF deleted.')


class PdfReactivateView(APIView):
    """PATCH /api/chapter-pdf/<pdf_id>/reactivate/ — Restore a soft-deleted PDF and its Qdrant vectors."""
    permission_classes = [permissions.IsAuthenticated, TeacherAdminPermission]

    def patch(self, request, pdf_id):
        from .tasks import reactivate_pdf_embeddings
        try:
            pdf_ref = PdfReference.objects.get(pk=pdf_id)
        except PdfReference.DoesNotExist:
            return Response({'detail': 'PDF not found.'}, status=status.HTTP_404_NOT_FOUND)

        if pdf_ref.is_active:
            return success(message='PDF is already active.')

        pdf_ref.is_active = True
        pdf_ref.save(update_fields=['is_active'])

        # Reactivate Qdrant vectors so they are included in context retrieval again
        reactivate_pdf_embeddings.delay(str(pdf_ref.id))

        return success(message='PDF reactivated.')


class PdfPermanentDeleteView(APIView):
    """DELETE /api/chapter-pdf/<pdf_id>/permanent/ — Hard delete: PostgreSQL + Qdrant + GCS."""
    permission_classes = [permissions.IsAuthenticated, TeacherAdminPermission]

    def delete(self, request, pdf_id):
        import httpx
        try:
            pdf_ref = PdfReference.objects.get(pk=pdf_id)
        except PdfReference.DoesNotExist:
            return Response({'detail': 'PDF not found.'}, status=status.HTTP_404_NOT_FOUND)

        gcs_path = pdf_ref.gcs_path

        # 1. Hard delete Qdrant vectors
        try:
            resp = httpx.delete(
                f"{_AI_SERVICE_URL}/ai/embed/{pdf_id}/permanent",
                timeout=30.0,
            )
            resp.raise_for_status()
        except Exception as exc:
            logger.warning(f"Qdrant hard delete failed for pdf {pdf_id}: {exc}")

        # 2. Delete from GCS
        if gcs_path:
            try:
                bucket_name = django_settings.GS_BUCKET_NAME
                gcs = _get_gcs_client()
                bucket = gcs.bucket(bucket_name)
                blob_name = gcs_path.replace(f"gs://{bucket_name}/", "", 1)
                bucket.blob(blob_name).delete()
            except Exception as exc:
                logger.warning(f"GCS delete failed for pdf {pdf_id} ({gcs_path}): {exc}")

        # 3. Hard delete from PostgreSQL
        pdf_ref.delete()

        return success(message='PDF permanently deleted.')


class PdfDownloadView(APIView):
    """GET /api/chapter-pdf/<pdf_id>/download/ — Stream PDF from GCS."""
    permission_classes = [permissions.IsAuthenticated, TeacherAdminPermission]

    def get(self, request, pdf_id):
        try:
            pdf_ref = PdfReference.objects.get(pk=pdf_id, is_active=True)
        except PdfReference.DoesNotExist:
            return Response({'detail': 'PDF not found.'}, status=status.HTTP_404_NOT_FOUND)

        gcs_path = pdf_ref.gcs_path
        try:
            from django.http import HttpResponse
            bucket_name = django_settings.GS_BUCKET_NAME
            gcs = _get_gcs_client()
            bucket = gcs.bucket(bucket_name)
            blob_name = gcs_path.replace(f"gs://{bucket_name}/", "", 1)
            blob = bucket.blob(blob_name)
            pdf_bytes = blob.download_as_bytes()
        except Exception as exc:
            logger.error(f"GCS download failed for pdf {pdf_id}: {exc}")
            return Response({'detail': 'Failed to download PDF.'}, status=status.HTTP_502_BAD_GATEWAY)

        from django.http import HttpResponse
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{pdf_ref.file_name}"'
        return response


class AssessmentGenerateView(APIView):
    """POST /api/assessment/generate — Create session and synchronously call ai-service."""
    permission_classes = [permissions.IsAuthenticated, TeacherAdminPermission]

    def post(self, request):
        data = request.data
        required = ['class_ref', 'subject', 'chapter', 'topic', 'num_questions']
        missing = [f for f in required if not data.get(f)]
        if missing:
            return validation_error({'detail': f"Missing fields: {', '.join(missing)}"})

        # Fetch all related objects
        try:
            class_obj = Class.objects.select_related('grade').get(pk=data['class_ref'])
        except Class.DoesNotExist:
            return Response({'detail': 'Class not found.'}, status=status.HTTP_404_NOT_FOUND)

        try:
            chapter = Module.objects.select_related('subject').get(pk=data['chapter'])
        except Module.DoesNotExist:
            return Response({'detail': 'Chapter not found.'}, status=status.HTTP_404_NOT_FOUND)

        try:
            topic = ModuleChapter.objects.get(pk=data['topic'], module=chapter)
        except ModuleChapter.DoesNotExist:
            return Response({'detail': 'Topic not found or does not belong to this chapter.'}, status=status.HTTP_404_NOT_FOUND)

        # Extract grade level number from grade name (e.g. "Grade 8" → 8)
        grade_level = 8
        if class_obj.grade and class_obj.grade.name:
            match = re.search(r'\d+', class_obj.grade.name)
            if match:
                grade_level = int(match.group())

        subject_name = chapter.subject.name if chapter.subject else ''

        # Create session record
        session = AssessmentSession.objects.create(
            user=request.user,
            class_ref=class_obj,
            subject_id=data['subject'],
            chapter=chapter,
            topic_name=topic.title,
            num_questions_requested=int(data['num_questions']),
            status='GENERATING',
        )

        # Context: use this topic's theory as fallback for ai-service
        theory_text = topic.theory or ''

        # Fetch existing question stems for this chapter to avoid duplicates
        existing_stems = list(
            Question.objects.filter(
                module_contents__chapter=topic,
                is_active=True,
                is_deleted=False,
            ).values_list('question_text', flat=True).distinct()
        )

        payload = {
            'session_id': session.session_id,
            'chapter_id': str(chapter.id),
            'subject': subject_name,
            'chapter': chapter.name,
            'topic': topic.title,
            'num_questions': int(data['num_questions']),
            'grade_level': grade_level,
            'context_text': theory_text,
            'existing_question_stems': existing_stems,
        }

        try:
            resp = httpx.post(
                f"{_AI_SERVICE_URL}/ai/generate",
                json=payload,
                timeout=_AI_SERVICE_TIMEOUT,
            )
            resp.raise_for_status()
            ai_result = resp.json()
        except httpx.TimeoutException:
            session.status = 'FAILED'
            session.save(update_fields=['status'])
            return Response({'detail': 'AI service timed out.'}, status=status.HTTP_504_GATEWAY_TIMEOUT)
        except Exception as exc:
            session.status = 'FAILED'
            session.save(update_fields=['status'])
            logger.error(f"AI service call failed for session {session.session_id}: {exc}")
            return Response({'detail': 'AI service error.'}, status=status.HTTP_502_BAD_GATEWAY)

        questions = ai_result.get('questions', [])
        session.status = 'COMPLETED'
        session.num_questions_returned = len(questions)
        session.ai_model_used = ai_result.get('model_used', '')
        session.generation_time_ms = ai_result.get('generation_time_ms')
        session.save(update_fields=['status', 'num_questions_returned', 'ai_model_used', 'generation_time_ms'])

        level_to_difficulty = {1: 'easy', 2: 'easy', 3: 'medium', 4: 'hard', 5: 'hard'}
        created_questions_list = []

        for q_data in questions:
            try:
                q_level = q_data.get('difficulty_level', 3)
                try:
                    q_level = int(q_level)
                except (TypeError, ValueError):
                    q_level = 3
                q_level = max(1, min(5, q_level))
                is_hots = q_level == 5
                difficulty = level_to_difficulty.get(q_level, 'medium')
                q_type = normalize_question_type(q_data.get('question_type', 'mcq_single'))

                question = Question.objects.create(
                    question_text=q_data.get('question_text', ''),
                    question_type=q_type,
                    difficulty_level=difficulty,
                    explanation=q_data.get('explanation', ''),
                    hint=q_data.get('hint', ''),
                    exp_points=q_data.get('exp_points', 10),
                    is_active=True,
                    is_hots=is_hots,
                    ai_generated=True,
                    level=q_level,
                    created_by=request.user,
                )

                options_data = q_data.get('options', [])
                created_options = []
                for opt_idx, opt_data in enumerate(options_data):
                    if q_type == 'rearrange':
                        order = opt_data.get('correct_order')
                        if order is None or not isinstance(order, (int, float)):
                            order = opt_idx + 1
                        order = max(1, int(order))
                    else:
                        order = opt_idx + 1
                    option = Option.objects.create(
                        question=question,
                        option_text=opt_data.get('option_text', f'Option {opt_idx + 1}'),
                        is_correct=opt_data.get('is_correct', False),
                        order=order,
                    )
                    created_options.append({
                        'id': str(option.id),
                        'option_text': option.option_text,
                        'is_correct': option.is_correct,
                        'order': option.order,
                    })

                if is_hots:
                    max_order = ChapterHOTS.objects.filter(chapter=topic).aggregate(
                        max_order=Max('order')
                    )['max_order'] or 0
                    ChapterHOTS.objects.create(
                        chapter=topic,
                        question=question,
                        order=max_order + 1,
                        created_by=request.user,
                    )
                    topic.has_hots = True
                    topic.save(update_fields=['has_hots'])
                else:
                    max_order = ModuleContent.objects.filter(chapter=topic).aggregate(
                        max_order=Max('order')
                    )['max_order'] or 0
                    ModuleContent.objects.create(
                        chapter=topic,
                        content_type='question',
                        question=question,
                        order=max_order + 1,
                        created_by=request.user,
                    )

                created_questions_list.append({
                    'id': str(question.id),
                    'question_text': question.question_text,
                    'question_type': normalize_question_type(question.question_type),
                    'difficulty_level': q_level,
                    'exp_points': question.exp_points,
                    'hint': question.hint,
                    'explanation': question.explanation,
                    'is_hots': question.is_hots,
                    'ai_generated': question.ai_generated,
                    'options': created_options,
                })
                logger.info(f"Assessment: created question {question.id} for topic {topic.id} (session {session.session_id})")

            except Exception as exc:
                logger.error(f"Assessment: error saving question for session {session.session_id}: {exc}")
                continue

        return success(data={
            'session_id': session.session_id,
            'questions': created_questions_list,
            'num_questions': len(created_questions_list),
            'model_used': session.ai_model_used,
            'generation_time_ms': session.generation_time_ms,
            'warning': ai_result.get('warning'),
        })

class AssessmentModifyView(APIView):
    """POST /api/assessment/modify — Modify a single question via ai-service."""
    permission_classes = [permissions.IsAuthenticated, TeacherAdminPermission]

    def post(self, request):
        data = request.data
        session_id = data.get('session_id')
        question = data.get('question')  # the selected question JSON
        modification_type = data.get('modification_type', 'CUSTOM')
        instruction = data.get('instruction', '')

        if not session_id or not question:
            return validation_error({'detail': 'session_id and question are required.'})

        try:
            session = AssessmentSession.objects.select_related('subject', 'chapter', 'class_ref__grade').get(session_id=session_id)
        except AssessmentSession.DoesNotExist:
            return Response({'detail': 'Session not found.'}, status=status.HTTP_404_NOT_FOUND)

        subject_name = session.subject.name if session.subject else ''
        chapter_name = session.chapter.name if session.chapter else ''

        grade_level = 8
        if session.class_ref and session.class_ref.grade:
            match = re.search(r'\d+', session.class_ref.grade.name)
            if match:
                grade_level = int(match.group())

        payload = {
            'session_id': session_id,
            'question': question,
            'modification_type': modification_type,
            'instruction': instruction,
            'grade_level': grade_level,
            'subject': subject_name,
            'chapter': chapter_name,
            'topic': session.topic_name or '',
            'chapter_id': str(session.chapter_id) if session.chapter_id else '',
        }

        try:
            resp = httpx.post(
                f"{_AI_SERVICE_URL}/ai/modify",
                json=payload,
                timeout=_AI_SERVICE_TIMEOUT,
            )
            resp.raise_for_status()
            ai_result = resp.json()
        except httpx.TimeoutException:
            return Response({'detail': 'AI service timed out.'}, status=status.HTTP_504_GATEWAY_TIMEOUT)
        except Exception as exc:
            logger.error(f"AI modify call failed for session {session_id}: {exc}")
            return Response({'detail': 'AI service error.'}, status=status.HTTP_502_BAD_GATEWAY)

        modified_question = ai_result.get('question', {})

        # Upload image_base64 to GCS and replace with public URL
        image_key = None
        image_base64 = modified_question.pop('image_base64', None)
        if image_base64:
            try:
                import base64
                import uuid
                image_bytes = base64.b64decode(image_base64)
                bucket_name = django_settings.GS_BUCKET_NAME
                gcs = _get_gcs_client()
                bucket = gcs.bucket(bucket_name)
                image_key = f"assessment-images/{session_id}/{uuid.uuid4().hex}.png"
                blob = bucket.blob(image_key)
                blob.upload_from_string(image_bytes, content_type='image/png')
                modified_question['image_url'] = f"https://storage.googleapis.com/{bucket_name}/{image_key}"
            except Exception as exc:
                logger.error(f"Failed to upload assessment image for session {session_id}: {exc}")

        # Update the Question DB entry if a question_id is available
        db_question = None
        question_id = data.get('question_id') or question.get('id')
        if question_id:
            try:
                _DIFFICULTY_MAP = {1: 'easy', 2: 'medium', 3: 'hard'}
                db_question = Question.objects.get(pk=question_id)

                update_fields = []
                if 'question_text' in modified_question:
                    db_question.question_text = modified_question['question_text']
                    update_fields.append('question_text')
                if 'hint' in modified_question:
                    db_question.hint = modified_question['hint'] or ''
                    update_fields.append('hint')
                if 'explanation' in modified_question:
                    db_question.explanation = modified_question['explanation'] or ''
                    update_fields.append('explanation')
                if 'exp_points' in modified_question:
                    db_question.exp_points = modified_question['exp_points']
                    update_fields.append('exp_points')
                dl = modified_question.get('difficulty_level')
                if dl is not None:
                    db_question.difficulty_level = _DIFFICULTY_MAP.get(dl, dl) if isinstance(dl, int) else dl
                    update_fields.append('difficulty_level')
                if image_key:
                    db_question.image = image_key
                    update_fields.append('image')

                if update_fields:
                    db_question.save(update_fields=update_fields)

                # Recreate options from AI response
                if 'options' in modified_question:
                    db_question.options.all().delete()
                    for i, opt in enumerate(modified_question['options'], start=1):
                        Option.objects.create(
                            question=db_question,
                            option_text=opt.get('option_text', ''),
                            is_correct=opt.get('is_correct', False),
                            order=opt.get('order', i),
                        )

                # Refresh from DB so serializer picks up new option IDs
                db_question.refresh_from_db()
            except Question.DoesNotExist:
                logger.warning(f"Question {question_id} not found for update in session {session_id}")
            except Exception as exc:
                logger.error(f"Failed to update question {question_id} in session {session_id}: {exc}")
                db_question = None

        # Log modification for audit
        QuestionModification.objects.create(
            session=session,
            question_id=question.get('id', ''),
            modification_type=modification_type,
            user_instruction=instruction,
            original_snapshot=question,
            modified_snapshot=modified_question,
        )

        if db_question is not None:
            return success(data={'question': QuestionSerializer(db_question).data})
        return success(data={'question': modified_question})


# ── Question Bank ──────────────────────────────────────────────────────────────

from rest_framework.pagination import PageNumberPagination
from .serializers import QuestionBankSerializer


class QuestionBankPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100


class QuestionBankView(APIView):
    """
    GET /question-bank/

    Query params:
        topic       (required) — ModuleChapter title to filter questions by
        chapter     (optional) — Module title to narrow results further
        level       (optional) — Question level (1-5)
        page        (optional) — Page number (default 1)
        page_size   (optional) — Page size (default 20, max 100)
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        topic = request.query_params.get('topic', '').strip()
        chapter = request.query_params.get('chapter', '').strip()
        level = request.query_params.get('level', '').strip()

        if not topic:
            return validation_error({'topic': 'This field is required.'})

        filters = {
            'module_contents__chapter__title__iexact': topic,
        }
        if chapter:
            filters['module_contents__chapter__module__name__iexact'] = chapter

        queryset = Question.objects.select_related(
            'created_by__profile',
        ).prefetch_related(
            'options',
        ).filter(**filters).distinct()

        if level:
            try:
                queryset = queryset.filter(level=int(level))
            except ValueError:
                return validation_error({'level': 'Must be an integer between 1 and 5.'})

        paginator = QuestionBankPagination()
        page = paginator.paginate_queryset(queryset, request)
        serializer = QuestionBankSerializer(
            page if page is not None else queryset,
            many=True,
            context={'request': request},
        )

        if page is not None:
            from gyaan_buddy.utils.response_utils import ResponseUtils
            return ResponseUtils.paginated_response(
                data=serializer.data,
                count=paginator.page.paginator.count,
                next_url=paginator.get_next_link(),
                previous_url=paginator.get_previous_link(),
                message="Question bank retrieved successfully",
            )

        return success(data=serializer.data, message="Question bank retrieved successfully")


class QuestionBankAddToChapterView(APIView):
    """
    POST /question-bank/add-to-chapter/

    Body:
        chapter_id   (required) — UUID of the chapter to add questions to
        question_ids (required) — list of question UUIDs to add
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        chapter_id = request.data.get('chapter_id')
        question_ids = request.data.get('question_ids', [])

        if not chapter_id:
            return validation_error({'chapter_id': 'This field is required.'})
        if not question_ids or not isinstance(question_ids, list):
            return validation_error({'question_ids': 'A non-empty list of question IDs is required.'})

        try:
            chapter = ModuleChapter.objects.get(id=chapter_id)
        except (ModuleChapter.DoesNotExist, Exception):
            return validation_error({'chapter_id': 'Chapter not found.'})

        # Questions already in this chapter (including soft-deleted, to allow reactivation)
        existing_module_contents = {
            str(mc.question_id): mc
            for mc in ModuleContent.objects.filter(chapter=chapter, content_type='question')
        }

        # Use all records (including soft-deleted) to avoid unique_together conflicts on 'order'
        max_order_result = ModuleContent.objects.filter(
            chapter=chapter
        ).aggregate(Max('order'))
        max_order = max_order_result['order__max'] or 0

        added = 0
        skipped = 0
        for qid in question_ids:
            existing_mc = existing_module_contents.get(str(qid))
            if existing_mc:
                if not existing_mc.is_deleted:
                    if existing_mc.question and not existing_mc.question.is_active:
                        existing_mc.question.is_active = True
                        existing_mc.question.save(update_fields=['is_active'])
                    skipped += 1
                    continue
                # Reactivate the soft-deleted entry (ModuleContent has no is_active field)
                existing_mc.is_deleted = False
                existing_mc.save(update_fields=['is_deleted'])
                if existing_mc.question and not existing_mc.question.is_active:
                    existing_mc.question.is_active = True
                    existing_mc.question.save(update_fields=['is_active'])
                added += 1
                continue
            try:
                question = Question.objects.get(id=qid, is_deleted=False)
                if not question.is_active:
                    question.is_active = True
                    question.save(update_fields=['is_active'])
                max_order += 1
                ModuleContent.objects.create(
                    chapter=chapter,
                    content_type='question',
                    question=question,
                    order=max_order,
                    created_by=request.user,
                )
                added += 1
            except (Question.DoesNotExist, Exception):
                pass

        return success(
            data={'added': added, 'skipped': skipped},
            message=f'{added} question(s) added to the assignment.'
        )
